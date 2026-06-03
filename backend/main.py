from fastapi import FastAPI, HTTPException, UploadFile, File, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List
import psycopg2
import psycopg2.extras
import os
import uuid
import shutil
import hashlib
import hmac
import secrets
import base64
import json
import time

_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if not _line or _line.startswith("#") or "=" not in _line:
                continue
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip().strip('"').strip("'"))

YANDEX_API_KEY = os.getenv("YANDEX_API_KEY", "")
YANDEX_FOLDER_ID = os.getenv("YANDEX_FOLDER_ID", "")
VK_TOKEN = os.getenv("VK_TOKEN", "")
AUTH_SECRET = os.getenv("AUTH_SECRET") or (os.getenv("DB_PASSWORD", "password") + "|stroyka-auth")
AUTH_TOKEN_TTL_SECONDS = int(os.getenv("AUTH_TOKEN_TTL_SECONDS", "86400"))

def _env_list(name: str, default: list[str]) -> list[str]:
    raw = os.getenv(name, "")
    if not raw:
        return default
    return [x.strip() for x in raw.split(",") if x.strip()]

CORS_ORIGINS = _env_list("CORS_ORIGINS", [
    "https://stroyka26.pro",
    "https://www.stroyka26.pro",
    "http://stroyka26.pro",
    "http://www.stroyka26.pro",
    "http://147.45.237.127",
    "https://147.45.237.127",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
])

DB_CONFIG = {
    "dbname": os.getenv("DB_NAME", "stroyka"),
    "user": os.getenv("DB_USER", "nikolas"),
    "password": os.getenv("DB_PASSWORD", "password"),
    "host": os.getenv("DB_HOST", "localhost"),
    "port": os.getenv("DB_PORT", "5432"),
}

DEFAULT_MATERIAL_NORMS = [
    {"ruleKey":"plaster_mix","name":"Штукатурная смесь","work":["штукатур"],"blockWork":["демонтаж","разбор"],"material":["штукатур","ротбанд","гипсов"],"workUnit":"м2","materialUnit":"кг","qtyPerUnit":8.5,"thicknessBaseMm":10,"defaultThicknessMm":10,"label":"штукатурная смесь 8.5 кг/м2 на 10 мм"},
    {"ruleKey":"screed_mix","name":"Смесь для стяжки","work":["стяжк","наливн"],"blockWork":["демонтаж","разбор"],"material":["пескобетон","смесь","стяжк"],"workUnit":"м2","materialUnit":"кг","qtyPerUnit":18,"thicknessBaseMm":10,"defaultThicknessMm":40,"label":"смесь для стяжки 18 кг/м2 на 10 мм"},
    {"ruleKey":"putty_mix","name":"Шпаклевка","work":["шпаклев","шпатлев"],"blockWork":["демонтаж","разбор"],"material":["шпаклев","шпатлев"],"workUnit":"м2","materialUnit":"кг","qtyPerUnit":1.2,"label":"шпаклевка 1.2 кг/м2"},
    {"ruleKey":"tile_glue","name":"Плиточный клей","work":["плитк","керамогранит","облицов"],"blockWork":["демонтаж","разбор"],"material":["клей","плиточ"],"workUnit":"м2","materialUnit":"кг","qtyPerUnit":4.5,"label":"плиточный клей 4.5 кг/м2"},
    {"ruleKey":"tile_grout","name":"Затирка","work":["плитк","керамогранит","облицов"],"blockWork":["демонтаж","разбор"],"material":["затир"],"workUnit":"м2","materialUnit":"кг","qtyPerUnit":0.3,"label":"затирка 0.3 кг/м2"},
    {"ruleKey":"paint","name":"Краска","work":["окраск","покраск"],"blockWork":["демонтаж","разбор"],"material":["краск","эмал"],"workUnit":"м2","materialUnit":"л","qtyPerUnit":0.2,"label":"краска 0.2 л/м2"},
    {"ruleKey":"primer","name":"Грунтовка","work":["грунтов","окраск","шпаклев","шпатлев"],"blockWork":["демонтаж","разбор"],"material":["грунтов"],"workUnit":"м2","materialUnit":"л","qtyPerUnit":0.12,"label":"грунтовка 0.12 л/м2"},
    {"ruleKey":"gkl_sheet","name":"ГКЛ лист","work":["гипсокарт","гкл","обшив"],"blockWork":["демонтаж","разбор"],"material":["гипсокарт","гкл","лист"],"workUnit":"м2","materialUnit":"шт","qtyPerUnit":0.35,"label":"ГКЛ 0.35 листа/м2"},
    {"ruleKey":"gkl_profile","name":"Профиль ГКЛ","work":["гипсокарт","гкл","обшив"],"blockWork":["демонтаж","разбор"],"material":["профиль"],"workUnit":"м2","materialUnit":"м","qtyPerUnit":1.6,"label":"профиль 1.6 м/м2"},
    {"ruleKey":"gkl_screws","name":"Саморезы ГКЛ","work":["гипсокарт","гкл","обшив"],"blockWork":["демонтаж","разбор"],"material":["саморез"],"workUnit":"м2","materialUnit":"шт","qtyPerUnit":20,"label":"саморезы 20 шт/м2"},
    {"ruleKey":"cable_line","name":"Кабель / провод","work":["кабел","провод","проклад"],"blockWork":["демонтаж","разбор"],"material":["кабель","провод","utp","ftp","f-utp","u-utp","кпс","ксвв","кспв","ввг","nym"],"workUnit":"м","materialUnit":"м","qtyPerUnit":1.05,"label":"кабель 1.05 м на 1 м трассы"},
    {"ruleKey":"cable_protection","name":"Гофра / кабель-канал","work":["кабел","провод","проклад"],"blockWork":["демонтаж","разбор"],"material":["гофр","труба пнд","кабель-канал","кабель канал"],"workUnit":"м","materialUnit":"м","qtyPerUnit":1.05,"label":"защита кабеля 1.05 м на 1 м трассы"},
    {"ruleKey":"cable_channel_box","name":"Короб / кабель-канал","work":["короб","кабель-канал","кабель канал","проклад"],"blockWork":["демонтаж","разбор"],"material":["короб","кабель-канал","кабель канал"],"workUnit":"м","materialUnit":"м","qtyPerUnit":1.05,"label":"короб/кабель-канал 1.05 м на 1 м трассы"},
    {"ruleKey":"junction_box","name":"Коробка ответвительная","work":["коробка ответв","распаечн","распределительн короб"],"blockWork":["демонтаж","разбор"],"material":["коробка ответв","распаечн","распределительн короб"],"workUnit":"шт","materialUnit":"шт","qtyPerUnit":1,"label":"коробка 1 шт на точку"},
    {"ruleKey":"luminaire","name":"Светильник / табло","work":["светильник","табло","транспарант"],"blockWork":["демонтаж","разбор"],"material":["светильник","табло","транспарант"],"workUnit":"шт","materialUnit":"шт","qtyPerUnit":1,"label":"светильник/табло 1 шт на точку"},
    {"ruleKey":"socket_switch","name":"Розетка / выключатель","work":["розет","выключател","штепсель"],"blockWork":["демонтаж","разбор"],"material":["розет","выключател","штепсель"],"workUnit":"шт","materialUnit":"шт","qtyPerUnit":1,"label":"механизм 1 шт на точку"},
    {"ruleKey":"electric_panel","name":"Щит / шкаф распределительный","work":["щит","шкаф","распределительн"],"blockWork":["демонтаж","разбор"],"material":["щит","шкаф","бокс распредел"],"workUnit":"шт","materialUnit":"шт","qtyPerUnit":1,"label":"щит/шкаф 1 шт на точку"},
    {"ruleKey":"pipe_pp","name":"Труба полипропиленовая","work":["трубопровод","водоснаб","отоплен"],"blockWork":["демонтаж","разбор"],"material":["труба полипропилен","трубы полипропилен","полипропилен","ppr","pprc","ppr-c"],"workUnit":"м","materialUnit":"м","qtyPerUnit":1.03,"label":"труба 1.03 м на 1 м трассы"},
    {"ruleKey":"pipe_fittings","name":"Фитинги трубопроводов","work":["соединен","сварк","узл","трубопровод"],"blockWork":["демонтаж","разбор"],"material":["муфта","угольник","угол","тройник","переход","кран","фитинг"],"workUnit":"соединений","materialUnit":"шт","qtyPerUnit":1,"label":"фитинг 1 шт на соединение"},
    {"ruleKey":"pipe_clamps","name":"Крепление трубопроводов","work":["трубопровод","водоснаб","отоплен"],"blockWork":["демонтаж","разбор"],"material":["хомут","креплен","кронштейн"],"workUnit":"м","materialUnit":"шт","qtyPerUnit":1.2,"label":"крепление трубы 1.2 шт/м"},
    {"ruleKey":"pipe_insulation","name":"Изоляция труб","work":["изоляц","трубопровод"],"blockWork":["демонтаж","разбор"],"material":["изоляц","энергофлекс","пенополиэтилен"],"workUnit":"м","materialUnit":"м","qtyPerUnit":1.05,"label":"изоляция трубы 1.05 м/м"},
    {"ruleKey":"thermal_insulation_board","name":"Теплоизоляция плитная/рулонная","work":["изоляция изделиями","теплоизоляц","изоляц"],"blockWork":["демонтаж","разбор"],"material":["пенополиэтилен","минераловат","теплоизоляц","изовер","технониколь","вата"],"workUnit":"м2","materialUnit":"м2","qtyPerUnit":1.05,"label":"теплоизоляция 1.05 м2/м2"},
    {"ruleKey":"radiator_device","name":"Радиатор отопления","work":["радиатор"],"blockWork":["демонтаж","разбор"],"material":["радиатор"],"workUnit":"шт","materialUnit":"шт","qtyPerUnit":1,"label":"радиатор 1 шт на прибор"},
    {"ruleKey":"radiator_mount","name":"Крепление радиатора","work":["радиатор"],"blockWork":["демонтаж","разбор"],"material":["кронштейн","креплен","крепеж"],"workUnit":"шт","materialUnit":"шт","qtyPerUnit":4,"label":"крепление радиатора 4 шт на прибор"},
    {"ruleKey":"plaster_mesh","name":"Штукатурная сетка","work":["сетка","штукатур"],"blockWork":["демонтаж","разбор"],"material":["сетка"],"workUnit":"м2","materialUnit":"м2","qtyPerUnit":1.1,"label":"штукатурная сетка 1.1 м2/м2"},
    {"ruleKey":"plaster_beacon","name":"Маячный профиль","work":["маяк","маяч"],"blockWork":["демонтаж","разбор"],"material":["маяк","маяч","профиль маяч"],"workUnit":"м2","materialUnit":"м","qtyPerUnit":0.85,"label":"маячный профиль 0.85 м/м2"},
    {"ruleKey":"linoleum_sheet","name":"Линолеум","work":["линолеум"],"blockWork":["демонтаж","разбор"],"material":["линолеум"],"workUnit":"м2","materialUnit":"м2","qtyPerUnit":1.02,"label":"линолеум 1.02 м2/м2"},
    {"ruleKey":"linoleum_glue","name":"Клей для линолеума","work":["линолеум"],"blockWork":["демонтаж","разбор"],"material":["клей","мастик"],"workUnit":"м2","materialUnit":"кг","qtyPerUnit":0.5,"label":"клей для линолеума 0.5 кг/м2"},
    {"ruleKey":"pvc_plinth","name":"Плинтус","work":["плинтус"],"blockWork":["демонтаж","разбор"],"material":["плинтус"],"workUnit":"м","materialUnit":"м","qtyPerUnit":1.03,"label":"плинтус 1.03 м/м"},
    {"ruleKey":"plywood_subfloor","name":"Фанера / основание пола","work":["фанер","основания полов","оснований полов"],"blockWork":["демонтаж","разбор"],"material":["фанер"],"workUnit":"м2","materialUnit":"м3","qtyPerUnit":0.025,"label":"фанера 0.025 м3/м2 основания"},
    {"ruleKey":"osb_subfloor","name":"OSB / древесная плита","work":["фанер","основания полов","оснований полов"],"blockWork":["демонтаж","разбор"],"material":["osb","осп","древесноструж","ориентированной стружкой"],"workUnit":"м2","materialUnit":"м2","qtyPerUnit":2.04,"label":"OSB/плита 2.04 м2/м2 основания"},
    {"ruleKey":"wood_frame","name":"Брус каркаса","work":["каркас","брус"],"blockWork":["демонтаж","разбор"],"material":["брус"],"workUnit":"м3","materialUnit":"м3","qtyPerUnit":1.05,"label":"брус 1.05 м3/м3 каркаса"},
    {"ruleKey":"door_block","name":"Дверной блок","work":["дверн","двер"],"blockWork":["демонтаж","разбор","снятие"],"material":["блок двер","дверн блок","дверь","полотно"],"workUnit":"м2","materialUnit":"м2","qtyPerUnit":1,"label":"дверной блок 1 м2/м2"},
    {"ruleKey":"windowsill","name":"Подоконник","work":["подокон"],"blockWork":["демонтаж","разбор"],"material":["подокон"],"workUnit":"м","materialUnit":"м","qtyPerUnit":1,"label":"подоконник 1 м/м"},
    {"ruleKey":"brick_masonry","name":"Кирпич","work":["кладк"],"blockWork":["демонтаж","разбор"],"material":["кирпич"],"workUnit":"м2","materialUnit":"шт","qtyPerUnit":51,"label":"кирпич 51 шт/м2 кладки в 1/2 кирпича"},
    {"ruleKey":"concrete","name":"Бетон","work":["бетон"],"blockWork":["демонтаж","разбор"],"material":["бетон"],"workUnit":"м3","materialUnit":"м3","qtyPerUnit":1,"label":"бетон 1 м3/м3"},
]

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

PASSWORD_HASH_PREFIX = "pbkdf2_sha256"
PASSWORD_HASH_ITERATIONS = 260000

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), PASSWORD_HASH_ITERATIONS)
    return f"{PASSWORD_HASH_PREFIX}${PASSWORD_HASH_ITERATIONS}${salt}${digest.hex()}"

def verify_password(password: str, stored: str) -> bool:
    if not stored:
        return False
    parts = stored.split("$")
    if len(parts) == 4 and parts[0] == PASSWORD_HASH_PREFIX:
        try:
            iterations = int(parts[1])
            salt = parts[2]
            expected = parts[3]
            digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations).hex()
            return hmac.compare_digest(digest, expected)
        except Exception:
            return False
    # Backward compatibility for existing plaintext passwords. Successful login upgrades it.
    return hmac.compare_digest(password, stored)

def is_legacy_password(stored: str) -> bool:
    return bool(stored) and not stored.startswith(PASSWORD_HASH_PREFIX + "$")

def public_user(row: dict, include_token: bool = False) -> dict:
    user = dict(row)
    user.pop("password", None)
    user.pop("reset_token", None)
    user.pop("reset_token_expires", None)
    if include_token:
        user["authToken"] = create_auth_token(dict(row))
    return user

def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("utf-8").rstrip("=")

def _b64url_decode(data: str) -> bytes:
    padding = "=" * (-len(data) % 4)
    return base64.urlsafe_b64decode((data + padding).encode("utf-8"))

def create_auth_token(user: dict) -> str:
    payload = {
        "id": user.get("id"),
        "email": user.get("email") or "",
        "role": user.get("role") or "",
        "name": user.get("name") or "",
        "exp": int(time.time()) + AUTH_TOKEN_TTL_SECONDS,
    }
    body = _b64url(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
    sig = hmac.new(AUTH_SECRET.encode("utf-8"), body.encode("utf-8"), hashlib.sha256).digest()
    return body + "." + _b64url(sig)

def verify_auth_token(token: str) -> dict:
    try:
        body, sig = token.split(".", 1)
        expected = _b64url(hmac.new(AUTH_SECRET.encode("utf-8"), body.encode("utf-8"), hashlib.sha256).digest())
        if not hmac.compare_digest(sig, expected):
            raise ValueError("bad signature")
        payload = json.loads(_b64url_decode(body).decode("utf-8"))
        if int(payload.get("exp") or 0) < int(time.time()):
            raise ValueError("expired")
        return payload
    except Exception:
        raise HTTPException(status_code=401, detail="Сессия недействительна. Войдите заново.")

def get_current_user(authorization: Optional[str] = Header(default=None)) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Требуется вход в систему")
    payload = verify_auth_token(authorization.split(" ", 1)[1].strip())
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE id=%s", (payload.get("id"),))
    user = cur.fetchone()
    cur.close(); conn.close()
    if not user:
        raise HTTPException(status_code=401, detail="Пользователь не найден")
    return public_user(user, include_token=True)

def require_roles(*roles: str):
    allowed = set(roles)
    def _dep(user: dict = Depends(get_current_user)) -> dict:
        if user.get("role") not in allowed:
            raise HTTPException(status_code=403, detail="Недостаточно прав")
        return user
    return _dep

LEADERSHIP_ROLES = ("директор", "зам_директора")
FINANCE_ROLES = ("директор", "зам_директора", "бухгалтер")
WAREHOUSE_ROLES = ("директор", "зам_директора", "кладовщик", "снабженец", "прораб")
SUPPLY_ROLES = ("директор", "зам_директора", "снабженец", "кладовщик", "прораб", "мастер", "субподрядчик", "поставщик", "бухгалтер")
SUPPLY_INTERNAL_ROLES = ("директор", "зам_директора", "снабженец", "кладовщик", "прораб", "бухгалтер")
PROJECT_WRITE_ROLES = ("директор", "зам_директора", "прораб", "главный_инженер", "сметчик")
PROJECT_DOCUMENT_ROLES = ("директор", "зам_директора", "бухгалтер", "прораб", "главный_инженер", "сметчик", "мастер", "субподрядчик", "кладовщик", "снабженец", "технадзор", "заказчик", "стройконтроль")
PROJECT_DOCUMENT_WRITE_ROLES = ("директор", "зам_директора", "бухгалтер", "прораб", "главный_инженер", "сметчик", "технадзор", "стройконтроль")
CONTRACT_ROLES = ("директор", "зам_директора", "бухгалтер", "прораб", "главный_инженер", "сметчик", "мастер", "субподрядчик")
JOURNAL_WRITE_ROLES = ("директор", "зам_директора", "прораб", "главный_инженер", "сметчик", "кладовщик", "снабженец", "мастер", "субподрядчик", "технадзор", "стройконтроль")
STAFF_MANAGE_ROLES = ("директор", "зам_директора", "бухгалтер")
PRICELIST_MANAGE_ROLES = ("директор", "зам_директора", "прораб", "главный_инженер", "сметчик")
OWN_EXPENSE_ROLES = ("директор", "зам_директора", "бухгалтер", "прораб", "главный_инженер", "сметчик", "мастер", "субподрядчик", "кладовщик", "снабженец")
SUPPLIER_INVOICE_VIEW_ROLES = ("директор", "зам_директора", "бухгалтер", "прораб", "кладовщик", "снабженец", "поставщик")

def user_project_names(user: dict) -> list[str]:
    names = []
    if user.get("projectName"):
        names.append(user.get("projectName"))
    if user.get("project_name"):
        names.append(user.get("project_name"))
    ap = user.get("assignedProjects", user.get("assigned_projects", []))
    try:
        if isinstance(ap, str):
            ap = json.loads(ap)
    except Exception:
        ap = []
    if isinstance(ap, list):
        names.extend([str(x) for x in ap if x])
    return sorted(set([x for x in names if x]))

def current_supplier_id(cur, user: dict):
    cur.execute("SELECT id FROM suppliers WHERE user_id=%s OR LOWER(email)=LOWER(%s) OR name=%s LIMIT 1",
                (user.get("id"), user.get("email") or "", user.get("name") or ""))
    row = cur.fetchone()
    if not row:
        return None
    try:
        return row["id"]
    except Exception:
        return row[0]

def can_see_all_company_data(user: dict) -> bool:
    return user.get("role") in ("директор", "зам_директора", "бухгалтер", "главный_инженер", "сметчик")

def visible_project_names(user: dict) -> Optional[List[str]]:
    if can_see_all_company_data(user):
        return None
    return user_project_names(user)

def has_project_access(user: dict, project_name: str) -> bool:
    allowed = visible_project_names(user)
    return allowed is None or bool(project_name and project_name in allowed)

def require_project_access(user: dict, project_name: str):
    if not has_project_access(user, project_name):
        raise HTTPException(status_code=403, detail="Нет доступа к объекту")

def hidden_work_all_signed(signed_customer="", signed_supervisor="", signed_contractor="", signed_subcontractor="") -> bool:
    return all(str(v or "").strip() for v in (
        signed_customer, signed_supervisor, signed_contractor, signed_subcontractor
    ))

def hidden_work_effective_status(status="", signed_customer="", signed_supervisor="", signed_contractor="", signed_subcontractor="") -> str:
    manual_status = (status or "").strip()
    if manual_status:
        return manual_status
    if hidden_work_all_signed(signed_customer, signed_supervisor, signed_contractor, signed_subcontractor):
        return "Подписан"
    return "Черновик"

def require_row_project_access(cur, table: str, row_id: int, user: dict, project_column: str = "project_name"):
    cur.execute(f"SELECT {project_column} FROM {table} WHERE id=%s", (row_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    project_name = row[0] if not isinstance(row, dict) else row.get(project_column)
    require_project_access(user, project_name or "")

def project_name_from_payload(cur, data: dict) -> str:
    project_name = data.get("projectName") or data.get("project_name") or data.get("project") or ""
    if project_name:
        return project_name
    project_id = data.get("projectId") or data.get("project_id")
    if not project_id:
        raise HTTPException(status_code=400, detail="Не указан объект")
    cur.execute("SELECT name FROM projects WHERE id=%s", (project_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Объект не найден")
    project_name = row[0] if not isinstance(row, dict) else row.get("name", "")
    if not project_name:
        raise HTTPException(status_code=400, detail="Не указан объект")
    return project_name

def require_checklist_access(cur, checklist_id: int, user: dict):
    cur.execute("SELECT project_name FROM project_checklists WHERE id=%s", (checklist_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Чек-лист не найден")
    project_name = row[0] if not isinstance(row, dict) else row.get("project_name")
    require_project_access(user, project_name or "")

def require_checklist_item_access(cur, item_id: int, user: dict):
    cur.execute("""SELECT pc.project_name
                   FROM checklist_items ci
                   JOIN project_checklists pc ON pc.id = ci.checklist_id
                   WHERE ci.id=%s""", (item_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Пункт чек-листа не найден")
    project_name = row[0] if not isinstance(row, dict) else row.get("project_name")
    require_project_access(user, project_name or "")

def require_room_access(cur, room_id: int, user: dict):
    cur.execute("SELECT project FROM rooms WHERE id=%s", (room_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Помещение не найдено")
    project_name = row[0] if not isinstance(row, dict) else row.get("project")
    require_project_access(user, project_name or "")

def require_room_child_access(cur, table: str, row_id: int, user: dict):
    cur.execute(f"""SELECT r.project
                    FROM {table} child
                    JOIN rooms r ON r.id = child.room_id
                    WHERE child.id=%s""", (row_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    project_name = row[0] if not isinstance(row, dict) else row.get("project")
    require_project_access(user, project_name or "")

def require_tool_access(cur, tool_id: int, user: dict):
    role = user.get("role")
    cur.execute("SELECT project, location, master_id, master_name FROM tools WHERE id=%s", (tool_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Инструмент не найден")
    if can_see_all_company_data(user) or role in ("кладовщик", "снабженец"):
        return
    project = (row[0] if not isinstance(row, dict) else row.get("project")) or ""
    location = (row[1] if not isinstance(row, dict) else row.get("location")) or ""
    master_id = row[2] if not isinstance(row, dict) else row.get("master_id")
    master_name = (row[3] if not isinstance(row, dict) else row.get("master_name")) or ""
    if role == "прораб":
        allowed = user_project_names(user)
        if not project or project in allowed or location in allowed:
            return
    if role in ("мастер", "субподрядчик"):
        if str(master_id or "") == str(user.get("id") or "") or master_name == (user.get("name") or ""):
            return
    raise HTTPException(status_code=403, detail="Нет доступа к инструменту")

def require_inventory_access(cur, inventory_id: int, user: dict):
    cur.execute("SELECT project FROM inventory WHERE id=%s", (inventory_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Инвентаризация не найдена")
    if can_see_all_company_data(user) or user.get("role") in ("кладовщик", "снабженец"):
        return
    project_name = row[0] if not isinstance(row, dict) else row.get("project")
    require_project_access(user, project_name or "")

def require_brigade_item_access(cur, item_id: int, user: dict):
    cur.execute("""SELECT bc.project_name
                   FROM brigade_contract_items bci
                   JOIN brigade_contracts bc ON bc.id = bci.contract_id
                   WHERE bci.id=%s""", (item_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    require_project_access(user, row[0] or "")

def require_worker_brigade_contract_access(cur, contract_id: int, user: dict):
    if user.get("role") not in ("мастер", "субподрядчик"):
        return
    cur.execute("SELECT contractor_id, brigade_name FROM brigade_contracts WHERE id=%s", (contract_id,))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    contractor_id = row[0] if not isinstance(row, dict) else row.get("contractor_id")
    brigade_name = row[1] if not isinstance(row, dict) else row.get("brigade_name")
    user_id = user.get("id")
    user_name = user.get("name") or ""
    if str(contractor_id or "") != str(user_id or "") and (brigade_name or "") != user_name:
        raise HTTPException(status_code=403, detail="Нет доступа к наряду")

def recalc_brigade_contract_total(cur, contract_id: int):
    cur.execute("""UPDATE brigade_contracts bc
                   SET total_amount = COALESCE((
                       SELECT SUM(COALESCE(quantity,0)*COALESCE(price_brigade,0))
                       FROM brigade_contract_items
                       WHERE contract_id=%s
                   ),0)
                   WHERE bc.id=%s""", (contract_id, contract_id))

def get_db():
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = True
    return conn

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            client VARCHAR(255),
            status VARCHAR(100),
            budget FLOAT DEFAULT 0,
            deadline VARCHAR(50),
            progress INT DEFAULT 0,
            tasks TEXT[] DEFAULT '{}',
            pricelist_id INT
        );
        CREATE TABLE IF NOT EXISTS clients (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            phone VARCHAR(100),
            email VARCHAR(255),
            status VARCHAR(100),
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS materials (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            unit VARCHAR(50),
            quantity FLOAT DEFAULT 0,
            price FLOAT DEFAULT 0,
            min_quantity FLOAT DEFAULT 0,
            project VARCHAR(255),
            category VARCHAR(255)
        );
        CREATE TABLE IF NOT EXISTS warehouse_history (
            id SERIAL PRIMARY KEY,
            material VARCHAR(255),
            type VARCHAR(50),
            quantity FLOAT,
            date VARCHAR(50),
            project VARCHAR(255),
            issued_to VARCHAR(255),
            issued_by VARCHAR(255),
            date_time VARCHAR(100)
        );
        CREATE TABLE IF NOT EXISTS staff (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            role VARCHAR(100),
            phone VARCHAR(100),
            salary FLOAT DEFAULT 0,
            project VARCHAR(255),
            pay_type VARCHAR(50) DEFAULT 'оклад'
        );
        CREATE TABLE IF NOT EXISTS piecework (
            id SERIAL PRIMARY KEY,
            staff_id VARCHAR(50),
            description VARCHAR(255),
            unit VARCHAR(50),
            quantity FLOAT,
            price_per_unit FLOAT,
            total FLOAT,
            project VARCHAR(255),
            date VARCHAR(50),
            comment TEXT,
            photo_url VARCHAR(255)
        );
        ALTER TABLE piecework ADD COLUMN IF NOT EXISTS work_journal_id INT;
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            email VARCHAR(255) UNIQUE,
            password VARCHAR(255),
            role VARCHAR(100)
        );
        ALTER TABLE users ADD COLUMN IF NOT EXISTS project_id INT;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS project_name VARCHAR(255);
        ALTER TABLE users ADD COLUMN IF NOT EXISTS reset_token VARCHAR(20);
        ALTER TABLE users ADD COLUMN IF NOT EXISTS reset_token_expires TIMESTAMP;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS failed_login_count INT DEFAULT 0;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS locked_until TIMESTAMP;
        CREATE TABLE IF NOT EXISTS messages (
            id SERIAL PRIMARY KEY,
            chat_type VARCHAR(50) DEFAULT 'company',
            project_id INT,
            author_id INT,
            author_name VARCHAR(255),
            author_role VARCHAR(100),
            text TEXT,
            photo_url VARCHAR(500),
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE messages ADD COLUMN IF NOT EXISTS read_by JSONB DEFAULT '[]'::jsonb;
        ALTER TABLE own_expenses ADD COLUMN IF NOT EXISTS category VARCHAR(50) DEFAULT 'other';
        ALTER TABLE users ADD COLUMN IF NOT EXISTS assigned_projects JSONB DEFAULT '[]'::jsonb;
        ALTER TABLE supplier_invoices ADD COLUMN IF NOT EXISTS paid_amount NUMERIC(14,2) DEFAULT 0;
        ALTER TABLE supplier_invoices ADD COLUMN IF NOT EXISTS offer_id INT;
        ALTER TABLE supplier_invoices ADD COLUMN IF NOT EXISTS request_id INT;
        ALTER TABLE supplier_invoices ADD COLUMN IF NOT EXISTS payment_terms VARCHAR(100);
        ALTER TABLE supplier_invoices ADD COLUMN IF NOT EXISTS material_name VARCHAR(255);
        CREATE TABLE IF NOT EXISTS tb_journal (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            master_name VARCHAR(255),
            instructor VARCHAR(255),
            instruction_type VARCHAR(100),
            program TEXT,
            instruction_text TEXT,
            participants_json TEXT,
            photo_url VARCHAR(500),
            date DATE,
            ai_filled BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE tb_journal ADD COLUMN IF NOT EXISTS program TEXT;
        ALTER TABLE tb_journal ADD COLUMN IF NOT EXISTS instruction_text TEXT;
        ALTER TABLE tb_journal ADD COLUMN IF NOT EXISTS participants_json TEXT;
        ALTER TABLE tb_journal ADD COLUMN IF NOT EXISTS photo_url VARCHAR(500);
        ALTER TABLE tb_journal ADD COLUMN IF NOT EXISTS ai_filled BOOLEAN DEFAULT FALSE;
        CREATE TABLE IF NOT EXISTS warranty_defects (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            description TEXT,
            found_at DATE,
            reported_by VARCHAR(255),
            reporter_phone VARCHAR(50),
            status VARCHAR(50) DEFAULT 'Открыт',
            assigned_to VARCHAR(255),
            fix_notes TEXT,
            fixed_at DATE,
            photo_url TEXT,
            severity VARCHAR(50),
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS warranty_start_date DATE;
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS warranty_end_date DATE;
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS warranty_contact VARCHAR(255);
        CREATE TABLE IF NOT EXISTS expense_reports (
            id SERIAL PRIMARY KEY,
            employee_id INT,
            employee_name VARCHAR(255),
            project_name VARCHAR(255),
            report_type VARCHAR(50) DEFAULT 'Авансовый отчёт',
            purpose TEXT,
            total_amount NUMERIC(14,2),
            issued_amount NUMERIC(14,2),
            spent_amount NUMERIC(14,2),
            balance NUMERIC(14,2),
            items_json TEXT,
            photo_url TEXT,
            date_from DATE,
            date_to DATE,
            status VARCHAR(50) DEFAULT 'На утверждении',
            approved_by VARCHAR(255),
            approved_at DATE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS supplier_invoices (
            id SERIAL PRIMARY KEY,
            supplier_id INT,
            supplier_name VARCHAR(255),
            project_name VARCHAR(255),
            invoice_number VARCHAR(100),
            invoice_date DATE,
            amount NUMERIC(14,2),
            vat_amount NUMERIC(14,2),
            description TEXT,
            file_url TEXT,
            photo_url TEXT,
            status VARCHAR(50) DEFAULT 'На утверждении',
            approved_by VARCHAR(255),
            approved_at DATE,
            paid_at DATE,
            paid_by VARCHAR(255),
            paid_note TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS inspection_orders (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            order_number VARCHAR(100),
            body VARCHAR(100),
            inspector VARCHAR(255),
            description TEXT,
            recommendations TEXT,
            deadline DATE,
            status VARCHAR(50) DEFAULT 'Открыто',
            photo_url TEXT,
            file_url TEXT,
            date DATE,
            response TEXT,
            response_date DATE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS document_versions (
            id SERIAL PRIMARY KEY,
            document_type VARCHAR(100),
            document_id INT,
            version_label VARCHAR(50),
            snapshot_json TEXT,
            changed_by VARCHAR(255),
            change_reason TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            user_id INT,
            user_name VARCHAR(255),
            user_role VARCHAR(100),
            action VARCHAR(100),
            entity_type VARCHAR(100),
            entity_id INT,
            description TEXT,
            project_name VARCHAR(255),
            ip VARCHAR(50),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS pricelists (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            description TEXT,
            for_who VARCHAR(255),
            coefficient FLOAT DEFAULT 1.0
        );
        CREATE TABLE IF NOT EXISTS pricelist_items (
            id SERIAL PRIMARY KEY,
            pricelist_id INT,
            name VARCHAR(255),
            unit VARCHAR(50),
            price FLOAT DEFAULT 0,
            category VARCHAR(255),
            specialization VARCHAR(255)
        );
        CREATE TABLE IF NOT EXISTS invite_codes (
            id SERIAL PRIMARY KEY,
            code VARCHAR(20) UNIQUE,
            role VARCHAR(100),
            used BOOLEAN DEFAULT FALSE
        );
        CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            phone VARCHAR(100),
            email VARCHAR(255),
            specialization VARCHAR(255),
            category VARCHAR(255),
            rating FLOAT DEFAULT 5.0,
            status VARCHAR(100)
        );
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS inn VARCHAR(50);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS kpp VARCHAR(50);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS ogrn VARCHAR(50);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS legal_address TEXT;
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS actual_address TEXT;
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS bank VARCHAR(255);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS bik VARCHAR(50);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS account VARCHAR(50);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS kor_account VARCHAR(50);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS director_name VARCHAR(255);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS director_position VARCHAR(255);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS contract_url VARCHAR(500);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS contract_number VARCHAR(100);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS contract_date DATE;
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS license_url VARCHAR(500);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS price_url VARCHAR(500);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS website VARCHAR(255);
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS notes TEXT;
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS user_id INT;
        ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS registered_at TIMESTAMP;
        ALTER TABLE invite_codes ADD COLUMN IF NOT EXISTS supplier_id INT;
        ALTER TABLE invite_codes ADD COLUMN IF NOT EXISTS preset_name VARCHAR(255);
        ALTER TABLE invite_codes ADD COLUMN IF NOT EXISTS preset_category VARCHAR(255);
        ALTER TABLE invite_codes ADD COLUMN IF NOT EXISTS created_by VARCHAR(255);
        ALTER TABLE invite_codes ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW();
        ALTER TABLE invite_codes ADD COLUMN IF NOT EXISTS expires_at TIMESTAMP;
        CREATE TABLE IF NOT EXISTS supplier_documents (
            id SERIAL PRIMARY KEY,
            supplier_id INT,
            doc_type VARCHAR(100),
            title VARCHAR(255),
            file_url VARCHAR(500),
            status VARCHAR(50) DEFAULT 'Загружен',
            signed_at DATE,
            expires_at DATE,
            notes TEXT,
            uploaded_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS supply_requests (
            id SERIAL PRIMARY KEY,
            material_name VARCHAR(255),
            quantity FLOAT,
            unit VARCHAR(50),
            project VARCHAR(255),
            created_by VARCHAR(255),
            date VARCHAR(50),
            status VARCHAR(100) DEFAULT 'Новая',
            notes TEXT,
            selected_suppliers INT[] DEFAULT '{}'
        );
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS requested_by_role VARCHAR(50);
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS requested_by_id INT;
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS urgency VARCHAR(50) DEFAULT 'обычная';
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS prorab_id INT;
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS prorab_name VARCHAR(255);
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS prorab_confirmed_at TIMESTAMP;
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS director_id INT;
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS director_name VARCHAR(255);
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS director_approved_at TIMESTAMP;
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS reject_reason TEXT;
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS category VARCHAR(100);
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW();
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS items_json TEXT;
        -- Multi-tenancy подготовка (см. ONBOARDING — план «единый кабинет поставщика»)
        -- Пока не используется в коде, но колонки добавлены чтобы при миграции не делать ALTER большим таблицам
        CREATE TABLE IF NOT EXISTS companies (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            short_name VARCHAR(255),
            inn VARCHAR(50),
            kpp VARCHAR(50),
            ogrn VARCHAR(50),
            legal_address TEXT,
            phone VARCHAR(100),
            email VARCHAR(255),
            logo_url VARCHAR(500),
            subdomain VARCHAR(100),
            plan VARCHAR(50) DEFAULT 'demo',
            active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        INSERT INTO companies (id, name, short_name, plan) VALUES (1, 'СтройКа', 'СтройКа', 'pro')
        ON CONFLICT (id) DO NOTHING;
        -- SaaS-биллинг: расширение companies для управления подписками
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS contact_name VARCHAR(255);
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS contact_phone VARCHAR(100);
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS contact_email VARCHAR(255);
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS trial_until DATE;
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS plan_expires_at DATE;
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS monthly_fee NUMERIC(10,2);
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS payment_status VARCHAR(50) DEFAULT 'active';
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS suspended_at TIMESTAMP;
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS suspended_reason TEXT;
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS max_projects INT;
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS max_users INT;
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS last_active_at TIMESTAMP;
        ALTER TABLE companies ADD COLUMN IF NOT EXISTS notes TEXT;
        -- Оплаты от компаний-клиентов (SaaS-выручка)
        CREATE TABLE IF NOT EXISTS company_payments (
            id SERIAL PRIMARY KEY,
            company_id INT NOT NULL,
            amount NUMERIC(10,2),
            payment_date DATE,
            method VARCHAR(50),
            invoice_number VARCHAR(100),
            status VARCHAR(50) DEFAULT 'paid',
            period_start DATE,
            period_end DATE,
            notes TEXT,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        -- Заявки на демо с лендинга
        CREATE TABLE IF NOT EXISTS demo_requests (
            id SERIAL PRIMARY KEY,
            company_name VARCHAR(255),
            contact_name VARCHAR(255),
            phone VARCHAR(100),
            email VARCHAR(255),
            employees_count VARCHAR(50),
            projects_count VARCHAR(50),
            comment TEXT,
            source VARCHAR(100),
            status VARCHAR(50) DEFAULT 'Новая',
            assigned_company_id INT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            processed_at TIMESTAMP
        );
        -- Ставим company_id на ключевые таблицы — пока всем 1 (текущая компания)
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE users ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE supply_requests ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE supplier_invoices ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE estimates ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE materials ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE warehouse_main ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE brigade_contracts ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE interim_acts ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        ALTER TABLE interim_acts ADD COLUMN IF NOT EXISTS scan_url TEXT;
        ALTER TABLE brigade_contracts ADD COLUMN IF NOT EXISTS act_scan_url TEXT;
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS archived BOOLEAN DEFAULT FALSE;
        ALTER TABLE projects ADD COLUMN IF NOT EXISTS archived_at TIMESTAMP;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS company_id INT DEFAULT 1;
        -- Поставщики — глобальные. company_id у них означает «первичный клиент» (необязательно)
        -- Связи компания-поставщик через company_supplier_links (договор и рейтинг свой у каждой пары)
        CREATE TABLE IF NOT EXISTS company_supplier_links (
            id SERIAL PRIMARY KEY,
            company_id INT NOT NULL,
            supplier_id INT NOT NULL,
            contract_url VARCHAR(500),
            contract_number VARCHAR(100),
            contract_date DATE,
            rating FLOAT,
            status VARCHAR(50) DEFAULT 'Активный',
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(company_id, supplier_id)
        );
        -- Подписки поставщиков (биллинг — пока заготовка)
        CREATE TABLE IF NOT EXISTS supplier_subscriptions (
            id SERIAL PRIMARY KEY,
            supplier_id INT NOT NULL,
            plan VARCHAR(50) DEFAULT 'free',
            valid_until DATE,
            monthly_fee NUMERIC(10,2) DEFAULT 0,
            payment_method VARCHAR(50),
            active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS supplier_offers (
            id SERIAL PRIMARY KEY,
            request_id INT,
            supplier_id INT,
            price_per_unit FLOAT,
            total_price FLOAT,
            delivery_days INT,
            notes TEXT,
            status VARCHAR(100) DEFAULT 'Ожидает'
        );
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS payment_terms VARCHAR(100) DEFAULT 'Постоплата';
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS vat_included BOOLEAN DEFAULT TRUE;
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS pdf_url VARCHAR(500);
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS valid_until DATE;
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS supplier_message TEXT;
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS requested_at TIMESTAMP DEFAULT NOW();
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS responded_at TIMESTAMP;
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS ai_recommended BOOLEAN DEFAULT FALSE;
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS delivery_status VARCHAR(100);
        ALTER TABLE supplier_offers ADD COLUMN IF NOT EXISTS items_kp_json TEXT;
        ALTER TABLE supplier_invoices ADD COLUMN IF NOT EXISTS delivery_allowed BOOLEAN DEFAULT FALSE;
        CREATE TABLE IF NOT EXISTS supply_deliveries (
            id SERIAL PRIMARY KEY,
            offer_id INT,
            request_id INT,
            supplier_id INT,
            supplier_name VARCHAR(255),
            project VARCHAR(255),
            material_name VARCHAR(255),
            planned_quantity NUMERIC(14,4),
            shipped_quantity NUMERIC(14,4),
            received_quantity NUMERIC(14,4) DEFAULT 0,
            unit VARCHAR(50),
            price_per_unit NUMERIC(14,2),
            total_price NUMERIC(14,2),
            status VARCHAR(100) DEFAULT 'Готовится к отгрузке',
            waybill_number VARCHAR(100),
            waybill_date DATE,
            vehicle_number VARCHAR(100),
            driver_name VARCHAR(255),
            document_url TEXT,
            photo_url TEXT,
            shipped_at TIMESTAMP,
            received_at TIMESTAMP,
            received_by VARCHAR(255),
            quality_status VARCHAR(100) DEFAULT 'Не проверено',
            quality_notes TEXT,
            shortage_quantity NUMERIC(14,4) DEFAULT 0,
            ai_check_result TEXT,
            claim_id INT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS supply_claims (
            id SERIAL PRIMARY KEY,
            delivery_id INT,
            request_id INT,
            offer_id INT,
            supplier_id INT,
            project VARCHAR(255),
            material_name VARCHAR(255),
            claim_type VARCHAR(100),
            description TEXT,
            expected_quantity NUMERIC(14,4),
            received_quantity NUMERIC(14,4),
            shortage_quantity NUMERIC(14,4),
            photo_url TEXT,
            status VARCHAR(100) DEFAULT 'Открыта',
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW(),
            resolved_at TIMESTAMP,
            resolution TEXT
        );
        CREATE TABLE IF NOT EXISTS supply_history (
            id SERIAL PRIMARY KEY,
            supplier_id INT,
            material_name VARCHAR(255),
            quantity FLOAT,
            unit VARCHAR(50),
            price_per_unit FLOAT,
            total_price FLOAT,
            project VARCHAR(255),
            date VARCHAR(50),
            status VARCHAR(100),
            confirmed_by VARCHAR(255)
        );
        ALTER TABLE supply_history ADD COLUMN IF NOT EXISTS request_id INT;
        ALTER TABLE supply_history ADD COLUMN IF NOT EXISTS delivery_id INT;
        CREATE TABLE IF NOT EXISTS supplier_catalog (
            id SERIAL PRIMARY KEY,
            supplier_id INT,
            supplier_name VARCHAR(255),
            material_name VARCHAR(255),
            unit VARCHAR(50) DEFAULT 'шт',
            price FLOAT DEFAULT 0,
            min_quantity FLOAT DEFAULT 1,
            delivery_days INT DEFAULT 3,
            in_stock BOOLEAN DEFAULT TRUE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS supply_request_templates (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            category VARCHAR(100),
            items_json TEXT,
            created_by VARCHAR(255),
            created_by_id INT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS material_transfers (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            from_location VARCHAR(255),
            to_person VARCHAR(255),
            to_person_role VARCHAR(100),
            material_name VARCHAR(255),
            quantity FLOAT,
            unit VARCHAR(50) DEFAULT 'шт',
            transfer_date DATE,
            signed BOOLEAN DEFAULT FALSE,
            signed_at TIMESTAMP,
            notes TEXT,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS brigade_contracts (
            id SERIAL PRIMARY KEY,
            project_id INT,
            project_name VARCHAR(255),
            brigade_name VARCHAR(255),
            contractor_type VARCHAR(50),
            contractor_id INT,
            total_amount NUMERIC(14,2) DEFAULT 0,
            status VARCHAR(50) DEFAULT 'Черновик',
            signed_at DATE,
            notes TEXT,
            pricelist_id INT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS brigade_contract_items (
            id SERIAL PRIMARY KEY,
            contract_id INT,
            estimate_section VARCHAR(255),
            description VARCHAR(255),
            unit VARCHAR(50),
            quantity FLOAT DEFAULT 0,
            price_smeta NUMERIC(14,2) DEFAULT 0,
            price_brigade NUMERIC(14,2) DEFAULT 0,
            done_quantity FLOAT DEFAULT 0,
            status VARCHAR(50) DEFAULT 'Не начато'
        );
        CREATE TABLE IF NOT EXISTS brigade_payments (
            id SERIAL PRIMARY KEY,
            contract_id INT,
            amount NUMERIC(14,2) DEFAULT 0,
            paid_by VARCHAR(255),
            paid_date DATE,
            note TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS brigade_acts (
            id SERIAL PRIMARY KEY,
            contract_id INT,
            project_name VARCHAR(255),
            brigade_name VARCHAR(255),
            period_from VARCHAR(50),
            period_to VARCHAR(50),
            total_amount NUMERIC(14,2) DEFAULT 0,
            status VARCHAR(50) DEFAULT 'Черновик',
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS project_documents (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            side VARCHAR(20) DEFAULT 'customer',
            doc_type VARCHAR(100),
            number VARCHAR(100),
            doc_date DATE,
            counterparty VARCHAR(255),
            sign_status VARCHAR(40) DEFAULT 'Не подписан',
            scan_url TEXT,
            amount NUMERIC(14,2) DEFAULT 0,
            notes TEXT,
            uploaded_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS project_measurements (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            source_type VARCHAR(100) DEFAULT 'Фактический обмер',
            doc_type VARCHAR(100) DEFAULT 'Обмер',
            title VARCHAR(255),
            file_url TEXT,
            status VARCHAR(50) DEFAULT 'Черновик',
            rooms_created INT DEFAULT 0,
            notes TEXT,
            uploaded_by VARCHAR(255),
            reviewed_by VARCHAR(255),
            reviewed_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS idx_project_measurements_project ON project_measurements(project_name);
        CREATE INDEX IF NOT EXISTS idx_project_measurements_status ON project_measurements(status);
        CREATE TABLE IF NOT EXISTS measurement_room_drafts (
            id SERIAL PRIMARY KEY,
            measurement_id INT,
            project_name VARCHAR(255),
            name VARCHAR(255),
            floor INT DEFAULT 1,
            liter VARCHAR(100),
            room_type VARCHAR(100) DEFAULT 'Комната',
            floor_area FLOAT DEFAULT 0,
            wall_area FLOAT DEFAULT 0,
            ceiling_area FLOAT DEFAULT 0,
            height FLOAT DEFAULT 0,
            windows INT DEFAULT 0,
            doors INT DEFAULT 0,
            notes TEXT,
            status VARCHAR(50) DEFAULT 'Черновик ИИ',
            created_by VARCHAR(255),
            accepted_room_id INT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS idx_measurement_room_drafts_project ON measurement_room_drafts(project_name);
        CREATE INDEX IF NOT EXISTS idx_measurement_room_drafts_measurement ON measurement_room_drafts(measurement_id);
        CREATE INDEX IF NOT EXISTS idx_measurement_room_drafts_status ON measurement_room_drafts(status);
        CREATE TABLE IF NOT EXISTS project_letters (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            side VARCHAR(20) DEFAULT 'customer',
            direction VARCHAR(20) DEFAULT 'outgoing',
            subject VARCHAR(500),
            body TEXT,
            counterparty VARCHAR(255),
            letter_date DATE,
            file_url TEXT,
            author VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS estimates (
            id SERIAL PRIMARY KEY,
            project_id INT,
            project_name VARCHAR(255),
            name VARCHAR(255),
            version VARCHAR(50),
            sections_json TEXT,
            status VARCHAR(50) DEFAULT 'Черновик',
            is_template BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS project_payments (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            amount NUMERIC(14,2) DEFAULT 0,
            note TEXT,
            date VARCHAR(50),
            added_by VARCHAR(255),
            paid_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS accountable_payments (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            given_to VARCHAR(255),
            given_to_id INT,
            amount NUMERIC(14,2) DEFAULT 0,
            payment_method VARCHAR(50),
            purpose TEXT,
            date VARCHAR(50),
            added_by VARCHAR(255),
            status VARCHAR(50) DEFAULT 'Выдан',
            spent_amount NUMERIC(14,2) DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS accountable_expenses (
            id SERIAL PRIMARY KEY,
            payment_id INT,
            project_name VARCHAR(255),
            description TEXT,
            amount NUMERIC(14,2) DEFAULT 0,
            photo_url TEXT,
            date VARCHAR(50),
            added_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS own_expenses (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            employee_name VARCHAR(255),
            employee_id INT,
            description TEXT,
            amount NUMERIC(14,2) DEFAULT 0,
            photo_url TEXT,
            date VARCHAR(50),
            category VARCHAR(100),
            status VARCHAR(50) DEFAULT 'Ожидает',
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS expenses (
            id SERIAL PRIMARY KEY,
            project VARCHAR(255),
            category VARCHAR(100),
            amount NUMERIC(14,2) DEFAULT 0,
            note TEXT,
            date VARCHAR(50),
            added_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS unexpected_works (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            description TEXT,
            unit VARCHAR(50),
            quantity FLOAT DEFAULT 0,
            price NUMERIC(14,2) DEFAULT 0,
            total NUMERIC(14,2) DEFAULT 0,
            added_by VARCHAR(255),
            added_by_role VARCHAR(50),
            status VARCHAR(50) DEFAULT 'На рассмотрении',
            approved_by VARCHAR(255),
            approved_at VARCHAR(50),
            notes TEXT,
            photo_url TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS change_type VARCHAR(80) DEFAULT 'Работа вне сметы';
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS estimate_id INT;
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS section_name VARCHAR(255);
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS estimate_item_name TEXT;
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS base_quantity NUMERIC(14,4);
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS new_required_quantity NUMERIC(14,4);
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS delta_quantity NUMERIC(14,4);
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS included_in_estimate_id INT;
        ALTER TABLE unexpected_works ADD COLUMN IF NOT EXISTS reason TEXT;
        UPDATE unexpected_works SET change_type='Работа вне сметы' WHERE change_type IS NULL OR change_type='';
        CREATE TABLE IF NOT EXISTS prescriptions (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            number VARCHAR(100),
            issued_by VARCHAR(255),
            issued_by_role VARCHAR(50),
            violation TEXT,
            deadline VARCHAR(50),
            responsible VARCHAR(255),
            status VARCHAR(50) DEFAULT 'Открыто',
            photo_url TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS project_stages (
            id SERIAL PRIMARY KEY,
            project_id INT,
            project_name VARCHAR(255),
            name VARCHAR(255),
            status VARCHAR(50) DEFAULT 'Не начат',
            start_date VARCHAR(50),
            end_date VARCHAR(50),
            progress INT DEFAULT 0,
            responsible VARCHAR(255),
            notes TEXT,
            order_num INT DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS project_checklists (
            id SERIAL PRIMARY KEY,
            project_id INT,
            project_name VARCHAR(255),
            name VARCHAR(255),
            template VARCHAR(100),
            status VARCHAR(50) DEFAULT 'Активен',
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS checklist_items (
            id SERIAL PRIMARY KEY,
            checklist_id INT,
            name VARCHAR(255),
            checked BOOLEAN DEFAULT FALSE,
            checked_by VARCHAR(255),
            checked_at VARCHAR(50),
            order_num INT DEFAULT 0
        );
        ALTER TABLE checklist_items ADD COLUMN IF NOT EXISTS checklist_id INT;
        ALTER TABLE checklist_items ADD COLUMN IF NOT EXISTS name VARCHAR(255);
        ALTER TABLE checklist_items ADD COLUMN IF NOT EXISTS checked BOOLEAN DEFAULT FALSE;
        ALTER TABLE checklist_items ADD COLUMN IF NOT EXISTS checked_by VARCHAR(255);
        ALTER TABLE checklist_items ADD COLUMN IF NOT EXISTS checked_at VARCHAR(50);
        ALTER TABLE checklist_items ADD COLUMN IF NOT EXISTS order_num INT DEFAULT 0;
        CREATE TABLE IF NOT EXISTS company_requisites (
            id SERIAL PRIMARY KEY,
            full_name VARCHAR(255),
            short_name VARCHAR(255),
            inn VARCHAR(20),
            kpp VARCHAR(20),
            ogrn VARCHAR(20),
            legal_address TEXT,
            actual_address TEXT,
            phone VARCHAR(50),
            email VARCHAR(255),
            director_name VARCHAR(255),
            director_position VARCHAR(100),
            basis VARCHAR(255),
            bank_name VARCHAR(255),
            bik VARCHAR(20),
            rs VARCHAR(40),
            ks VARCHAR(40)
        );
        CREATE TABLE IF NOT EXISTS company_documents (
            id SERIAL PRIMARY KEY,
            company_id INT,
            name VARCHAR(255),
            doc_type VARCHAR(100),
            file_url TEXT,
            expires_at VARCHAR(50),
            uploaded_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS project_chat (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            author_id INT,
            author_name VARCHAR(255),
            author_role VARCHAR(50),
            text TEXT,
            photo_url TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS room_doors (
            id SERIAL PRIMARY KEY,
            room_id INT,
            name VARCHAR(255),
            width FLOAT DEFAULT 0,
            height FLOAT DEFAULT 0,
            door_type VARCHAR(100),
            door_purpose VARCHAR(100),
            reveal_depth FLOAT DEFAULT 0,
            reveal_material VARCHAR(100),
            order_num INT DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS room_windows (
            id SERIAL PRIMARY KEY,
            room_id INT,
            name VARCHAR(255),
            width FLOAT DEFAULT 0,
            height FLOAT DEFAULT 0,
            window_type VARCHAR(100),
            reveal_depth FLOAT DEFAULT 0,
            reveal_material VARCHAR(100),
            order_num INT DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS warehouse_invoices (
            id SERIAL PRIMARY KEY,
            number VARCHAR(100),
            date VARCHAR(50),
            supplier_id INT,
            supplier_name VARCHAR(255),
            accepted_by VARCHAR(255),
            location VARCHAR(255),
            project VARCHAR(255),
            vat BOOLEAN DEFAULT FALSE,
            items TEXT,
            total_base NUMERIC(14,2) DEFAULT 0,
            total_vat NUMERIC(14,2) DEFAULT 0,
            total_with_vat NUMERIC(14,2) DEFAULT 0,
            status VARCHAR(50) DEFAULT 'Принято',
            added_by VARCHAR(255),
            photo_url TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS warehouses (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            city VARCHAR(255),
            address TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS salary_payments (
            id SERIAL PRIMARY KEY,
            staff_id INT,
            staff_name VARCHAR(255),
            month VARCHAR(7),
            amount NUMERIC(14,2) DEFAULT 0,
            paid_by VARCHAR(255),
            paid_date VARCHAR(50),
            note TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS crm_leads (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            phone VARCHAR(50),
            email VARCHAR(255),
            source VARCHAR(255),
            budget NUMERIC(14,2) DEFAULT 0,
            notes TEXT,
            stage VARCHAR(50) DEFAULT 'Новый',
            created_by VARCHAR(255),
            created_at VARCHAR(50)
        );
        CREATE TABLE IF NOT EXISTS work_journal (
            id SERIAL PRIMARY KEY,
            master_id INT,
            master_name VARCHAR(255),
            project VARCHAR(255),
            description VARCHAR(255),
            unit VARCHAR(50),
            quantity FLOAT,
            price_per_unit FLOAT,
            total FLOAT,
            date VARCHAR(50),
            status VARCHAR(100) DEFAULT 'На проверке',
            comment TEXT,
            photo_url VARCHAR(255),
            confirmed_by VARCHAR(255),
            confirmed_at VARCHAR(50)
        );
        CREATE TABLE IF NOT EXISTS master_profiles (
            id SERIAL PRIMARY KEY,
            user_id INT UNIQUE,
            full_name VARCHAR(255),
            passport VARCHAR(255),
            inn VARCHAR(50),
            contract_type VARCHAR(50),
            bank_account VARCHAR(255),
            bank_name VARCHAR(255),
            phone VARCHAR(100),
            specialization VARCHAR(255),
            ogrnip VARCHAR(50),
            profile_completed BOOLEAN DEFAULT TRUE
        );
        CREATE TABLE IF NOT EXISTS contracts (
            id SERIAL PRIMARY KEY,
            master_id INT,
            master_name VARCHAR(255),
            contract_type VARCHAR(50),
            contract_number VARCHAR(100),
            project VARCHAR(255),
            start_date VARCHAR(50),
            end_date VARCHAR(50)
        );
        CREATE TABLE IF NOT EXISTS interim_acts (
            id SERIAL PRIMARY KEY,
            master_id INT,
            master_name VARCHAR(255),
            project VARCHAR(255),
            period_start VARCHAR(50),
            period_end VARCHAR(50),
            total_amount FLOAT DEFAULT 0,
            paid_amount FLOAT DEFAULT 0,
            contract_id INT,
            status VARCHAR(100) DEFAULT 'Новый'
        );
        CREATE TABLE IF NOT EXISTS timesheet (
            id SERIAL PRIMARY KEY,
            staff_id INT,
            day VARCHAR(10)
        );
        CREATE TABLE IF NOT EXISTS rooms (
            id SERIAL PRIMARY KEY,
            project VARCHAR(255),
            name VARCHAR(255),
            floor INT DEFAULT 1,
            liter VARCHAR(100),
            room_type VARCHAR(100) DEFAULT 'Комната',
            floor_area FLOAT DEFAULT 0,
            wall_area FLOAT DEFAULT 0,
            ceiling_area FLOAT DEFAULT 0,
            height FLOAT DEFAULT 0,
            ceiling_type VARCHAR(100),
            wall_material VARCHAR(100),
            floor_material VARCHAR(100),
            windows INT DEFAULT 0,
            doors INT DEFAULT 0,
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS room_works (
            id SERIAL PRIMARY KEY,
            room_id INT,
            project VARCHAR(255),
            room_name VARCHAR(255),
            master_id INT,
            master_name VARCHAR(255),
            description VARCHAR(255),
            surface VARCHAR(100),
            unit VARCHAR(50),
            quantity FLOAT,
            price_per_unit FLOAT DEFAULT 0,
            total FLOAT DEFAULT 0,
            date VARCHAR(50),
            status VARCHAR(100) DEFAULT 'На проверке',
            photo_url VARCHAR(255),
            confirmed_by VARCHAR(255)
        );
        CREATE TABLE IF NOT EXISTS tools (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            inventory_number VARCHAR(100),
            cost FLOAT DEFAULT 0,
            status VARCHAR(100) DEFAULT 'На складе',
            location VARCHAR(100) DEFAULT 'Основной склад',
            project VARCHAR(255),
            master_id INT,
            master_name VARCHAR(255),
            issue_type VARCHAR(100),
            photo_url VARCHAR(255),
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS tool_history (
            id SERIAL PRIMARY KEY,
            tool_id INT,
            tool_name VARCHAR(255),
            action VARCHAR(100),
            from_location VARCHAR(255),
            to_location VARCHAR(255),
            master_name VARCHAR(255),
            project VARCHAR(255),
            issue_type VARCHAR(100),
            condition VARCHAR(100),
            date VARCHAR(50),
            created_by VARCHAR(255)
        );
        CREATE TABLE IF NOT EXISTS warehouse_main (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255),
            unit VARCHAR(50),
            quantity FLOAT DEFAULT 0,
            price FLOAT DEFAULT 0,
            min_quantity FLOAT DEFAULT 0,
            category VARCHAR(255)
        );
        CREATE TABLE IF NOT EXISTS warehouse_movements (
            id SERIAL PRIMARY KEY,
            material_name VARCHAR(255),
            from_location VARCHAR(255),
            to_location VARCHAR(255),
            quantity FLOAT,
            unit VARCHAR(50),
            date VARCHAR(50),
            created_by VARCHAR(255),
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS inventory (
            id SERIAL PRIMARY KEY,
            project VARCHAR(255),
            date VARCHAR(50),
            created_by VARCHAR(255),
            status VARCHAR(100) DEFAULT 'Открыта',
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS inventory_items (
            id SERIAL PRIMARY KEY,
            inventory_id INT,
            material_name VARCHAR(255),
            unit VARCHAR(50),
            expected FLOAT,
            actual FLOAT,
            difference FLOAT,
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS pd_consents (
            id SERIAL PRIMARY KEY,
            user_id INT UNIQUE,
            signed_at VARCHAR(100),
            scan_url VARCHAR(255),
            uploaded_by VARCHAR(255)
        );
        CREATE TABLE IF NOT EXISTS project_ai_summary (
            project_name VARCHAR(255) PRIMARY KEY,
            payload_hash VARCHAR(64),
            summary TEXT,
            updated_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS ai_findings (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            finding_type VARCHAR(100) DEFAULT 'rule',
            category VARCHAR(100) DEFAULT 'Общее',
            severity VARCHAR(50) DEFAULT 'Проверить',
            title TEXT,
            description TEXT,
            source VARCHAR(100) DEFAULT 'system_rules',
            linked_entity_type VARCHAR(100),
            linked_entity_id VARCHAR(100),
            suggested_action TEXT,
            assigned_role VARCHAR(100),
            assigned_to VARCHAR(255),
            status VARCHAR(50) DEFAULT 'Новое',
            dedupe_key VARCHAR(255),
            created_by VARCHAR(255),
            created_by_id INT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS idx_ai_findings_project ON ai_findings(project_name);
        CREATE INDEX IF NOT EXISTS idx_ai_findings_status ON ai_findings(status);
        CREATE INDEX IF NOT EXISTS idx_ai_findings_dedupe ON ai_findings(project_name, dedupe_key);
        CREATE TABLE IF NOT EXISTS ai_tasks (
            id SERIAL PRIMARY KEY,
            finding_id INT,
            project_name VARCHAR(255),
            title TEXT,
            description TEXT,
            assigned_role VARCHAR(100),
            assigned_to VARCHAR(255),
            status VARCHAR(50) DEFAULT 'Новое',
            due_date DATE,
            accepted_by VARCHAR(255),
            accepted_at TIMESTAMP,
            closed_by VARCHAR(255),
            closed_at TIMESTAMP,
            action_label VARCHAR(255),
            action_payload TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS idx_ai_tasks_project ON ai_tasks(project_name);
        CREATE INDEX IF NOT EXISTS idx_ai_tasks_status ON ai_tasks(status);
        CREATE INDEX IF NOT EXISTS idx_ai_tasks_finding ON ai_tasks(finding_id);
        CREATE TABLE IF NOT EXISTS material_norms (
            id SERIAL PRIMARY KEY,
            rule_key VARCHAR(100) UNIQUE,
            name VARCHAR(255),
            work_keywords TEXT,
            block_work_keywords TEXT,
            material_keywords TEXT,
            work_unit VARCHAR(50),
            material_unit VARCHAR(50),
            qty_per_unit NUMERIC(14,4) DEFAULT 0,
            thickness_base_mm NUMERIC(14,4),
            default_thickness_mm NUMERIC(14,4),
            label TEXT,
            active BOOLEAN DEFAULT TRUE,
            updated_by VARCHAR(255),
            updated_at TIMESTAMP DEFAULT NOW(),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS material_norm_overrides (
            id SERIAL PRIMARY KEY,
            base_norm_id INT,
            project_name VARCHAR(255),
            estimate_id INT,
            section_name TEXT,
            work_name TEXT,
            material_name TEXT,
            work_keywords TEXT,
            block_work_keywords TEXT,
            material_keywords TEXT,
            work_unit VARCHAR(50),
            material_unit VARCHAR(50),
            qty_per_unit NUMERIC(14,4) DEFAULT 0,
            thickness_base_mm NUMERIC(14,4),
            default_thickness_mm NUMERIC(14,4),
            label TEXT,
            reason TEXT,
            active BOOLEAN DEFAULT TRUE,
            updated_by VARCHAR(255),
            updated_at TIMESTAMP DEFAULT NOW(),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS idx_material_norm_overrides_project ON material_norm_overrides(project_name);
        CREATE INDEX IF NOT EXISTS idx_material_norm_overrides_estimate ON material_norm_overrides(estimate_id);
        CREATE TABLE IF NOT EXISTS material_norm_suggestions (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            suggestion_type VARCHAR(100) DEFAULT 'without_norm',
            status VARCHAR(50) DEFAULT 'Новая',
            severity VARCHAR(50) DEFAULT 'Проверить',
            work_name TEXT,
            section_name TEXT,
            work_unit VARCHAR(50),
            material_name TEXT,
            material_unit VARCHAR(50),
            current_norm_id INT,
            current_qty_per_unit NUMERIC(14,4),
            suggested_qty_per_unit NUMERIC(14,4),
            sample_count INT DEFAULT 0,
            actual_qty NUMERIC(14,4) DEFAULT 0,
            norm_qty NUMERIC(14,4) DEFAULT 0,
            confidence NUMERIC(5,2) DEFAULT 0,
            reason TEXT,
            ai_summary TEXT,
            work_keywords TEXT,
            material_keywords TEXT,
            block_work_keywords TEXT,
            label TEXT,
            source VARCHAR(100) DEFAULT 'rules',
            dedupe_key VARCHAR(255),
            created_by VARCHAR(255),
            applied_norm_id INT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );
        CREATE INDEX IF NOT EXISTS idx_material_norm_suggestions_project ON material_norm_suggestions(project_name);
        CREATE INDEX IF NOT EXISTS idx_material_norm_suggestions_status ON material_norm_suggestions(status);
        CREATE INDEX IF NOT EXISTS idx_material_norm_suggestions_dedupe ON material_norm_suggestions(dedupe_key);
        ALTER TABLE rooms ADD COLUMN IF NOT EXISTS floor INT DEFAULT 1;
        ALTER TABLE rooms ADD COLUMN IF NOT EXISTS liter VARCHAR(100);
        ALTER TABLE rooms ADD COLUMN IF NOT EXISTS room_type VARCHAR(100) DEFAULT 'Комната';
        ALTER TABLE rooms ADD COLUMN IF NOT EXISTS height FLOAT DEFAULT 0;
        ALTER TABLE rooms ADD COLUMN IF NOT EXISTS ceiling_type VARCHAR(100);
        ALTER TABLE rooms ADD COLUMN IF NOT EXISTS wall_material VARCHAR(100);
        ALTER TABLE rooms ADD COLUMN IF NOT EXISTS floor_material VARCHAR(100);
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS materials_used TEXT;
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS estimate_id INT;
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS section_name VARCHAR(255);
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS responsible_itr VARCHAR(255);
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS weather VARCHAR(255);
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS time_start VARCHAR(10);
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS time_end VARCHAR(10);
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS hidden_work BOOLEAN DEFAULT FALSE;
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS quality_status VARCHAR(50);
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS normatives TEXT;
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS project_docs TEXT;
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS ai_filled BOOLEAN DEFAULT FALSE;
        ALTER TABLE work_journal ADD COLUMN IF NOT EXISTS unexpected_work_id INT;
        CREATE TABLE IF NOT EXISTS material_inspection_journal (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            invoice_id INT,
            material_name VARCHAR(255),
            unit VARCHAR(50),
            quantity NUMERIC(14,4),
            supplier VARCHAR(255),
            batch_number VARCHAR(100),
            passport_number VARCHAR(100),
            certificate_number VARCHAR(100),
            test_protocol_number VARCHAR(100),
            visual_inspection_result VARCHAR(50),
            remarks TEXT,
            inspector_name VARCHAR(255),
            received_at DATE,
            inspected_at DATE,
            inspected BOOLEAN DEFAULT FALSE,
            normatives TEXT,
            ai_filled BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE material_inspection_journal ADD COLUMN IF NOT EXISTS delivery_id INT;
        CREATE TABLE IF NOT EXISTS supervisor_acts (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            act_number VARCHAR(100),
            act_type VARCHAR(100),
            description TEXT,
            findings TEXT,
            recommendations TEXT,
            issued_by VARCHAR(255),
            issued_by_role VARCHAR(50),
            date DATE,
            photo_url TEXT,
            file_url TEXT,
            status VARCHAR(50) DEFAULT 'Открыт',
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE prescriptions ADD COLUMN IF NOT EXISTS photo_url TEXT;
        CREATE TABLE IF NOT EXISTS cable_journal (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            invoice_id INT,
            cable_brand VARCHAR(255),
            cross_section NUMERIC(8,2),
            cores_count INT,
            length_received NUMERIC(10,2),
            length_installed NUMERIC(10,2),
            drum_number VARCHAR(100),
            manufacturer VARCHAR(255),
            supplier VARCHAR(255),
            certificate_number VARCHAR(100),
            passport_number VARCHAR(100),
            insulation_before NUMERIC(8,2),
            insulation_after NUMERIC(8,2),
            installation_location TEXT,
            installation_method VARCHAR(255),
            installed_at DATE,
            received_at DATE,
            responsible_itr VARCHAR(255),
            normatives TEXT,
            ai_filled BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE cable_journal ADD COLUMN IF NOT EXISTS delivery_id INT;
        ALTER TABLE cable_journal ADD COLUMN IF NOT EXISTS cable_type VARCHAR(100);
        ALTER TABLE estimates ADD COLUMN IF NOT EXISTS status VARCHAR(50) DEFAULT 'Черновик';
        ALTER TABLE estimates ADD COLUMN IF NOT EXISTS is_template BOOLEAN DEFAULT FALSE;
        ALTER TABLE estimates ADD COLUMN IF NOT EXISTS smeta_type VARCHAR(50) DEFAULT 'Заказчик';
        ALTER TABLE estimates ADD COLUMN IF NOT EXISTS work_package VARCHAR(100) DEFAULT 'Основная';
        UPDATE estimates SET smeta_type='Заказчик' WHERE smeta_type IS NULL OR smeta_type='';
        UPDATE estimates SET work_package='Основная' WHERE work_package IS NULL OR work_package='';
        CREATE TABLE IF NOT EXISTS estimate_versions (
            id SERIAL PRIMARY KEY,
            estimate_id INT NOT NULL,
            version_label VARCHAR(100),
            sections_json TEXT,
            total NUMERIC(14,2) DEFAULT 0,
            comment TEXT,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS estimate_chat_messages (
            id SERIAL PRIMARY KEY,
            estimate_id INT NOT NULL,
            role VARCHAR(20) NOT NULL,
            content TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE brigade_contracts ADD COLUMN IF NOT EXISTS pricelist_id INT;
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS last_name VARCHAR(100);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS first_name VARCHAR(100);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS middle_name VARCHAR(100);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS birth_date DATE;
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS citizenship VARCHAR(50);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS address TEXT;
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS photo_url VARCHAR(255);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS email_work VARCHAR(255);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS email_personal VARCHAR(255);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS phone_extra VARCHAR(50);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS passport_series VARCHAR(10);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS passport_number VARCHAR(20);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS passport_issued_by VARCHAR(255);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS passport_issued_date DATE;
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS inn VARCHAR(20);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS snils VARCHAR(20);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS specialization VARCHAR(100);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS category VARCHAR(50);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS employment_type VARCHAR(50);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS hired_date DATE;
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS fired_date DATE;
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS status VARCHAR(50);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS brigade VARCHAR(100);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS bank_account VARCHAR(50);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS bank_name VARCHAR(255);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS bank_bik VARCHAR(20);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS bank_corr VARCHAR(50);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS ogrnip VARCHAR(20);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS card_number VARCHAR(20);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS signature_url VARCHAR(255);
        ALTER TABLE staff ADD COLUMN IF NOT EXISTS notes TEXT;
        ALTER TABLE pricelist_items ADD COLUMN IF NOT EXISTS item_type VARCHAR(20);
        CREATE TABLE IF NOT EXISTS hidden_works_acts (
            id SERIAL PRIMARY KEY,
            project_name VARCHAR(255),
            estimate_id INT,
            act_number VARCHAR(100),
            work_name TEXT,
            section_name VARCHAR(255),
            brigade VARCHAR(255),
            quantity NUMERIC(14,4),
            unit VARCHAR(50),
            price_per_unit NUMERIC(14,2),
            total NUMERIC(14,2),
            work_date DATE,
            materials_used TEXT,
            project_docs TEXT,
            conclusion TEXT,
            signed_customer VARCHAR(255),
            signed_supervisor VARCHAR(255),
            signed_contractor VARCHAR(255),
            signed_subcontractor VARCHAR(255),
            signed_customer_at DATE,
            signed_supervisor_at DATE,
            signed_contractor_at DATE,
            signed_subcontractor_at DATE,
            photos TEXT,
            certificates TEXT,
            city VARCHAR(100),
            status VARCHAR(50) DEFAULT 'Черновик',
            comments TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS signed_customer_at DATE;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS signed_supervisor_at DATE;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS signed_contractor_at DATE;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS signed_subcontractor_at DATE;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS photos TEXT;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS certificates TEXT;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS city VARCHAR(100);
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS ai_filled BOOLEAN DEFAULT FALSE;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS paid_status VARCHAR(50) DEFAULT 'Не оплачен';
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS paid_amount NUMERIC(14,2);
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS paid_at DATE;
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS paid_by VARCHAR(255);
        ALTER TABLE hidden_works_acts ADD COLUMN IF NOT EXISTS paid_note TEXT;
        CREATE TABLE IF NOT EXISTS staff_documents (
            id SERIAL PRIMARY KEY,
            staff_id INT NOT NULL,
            doc_type VARCHAR(50) NOT NULL,
            title VARCHAR(255),
            file_url VARCHAR(500),
            status VARCHAR(50) DEFAULT 'действует',
            signed_at DATE,
            expires_at DATE,
            notes TEXT,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        );
    """)
    seed_users = [
        ('Директор', 'admin@stroyka.ru', 'admin123', 'директор'),
        ('Бухгалтер', 'buh@stroyka.ru', 'buh123', 'бухгалтер'),
        ('Прораб', 'prorab@stroyka.ru', 'prorab123', 'прораб'),
        ('Мастер', 'master@stroyka.ru', 'master123', 'мастер'),
    ]
    for seed_name, seed_email, seed_password, seed_role in seed_users:
        cur.execute("""
            INSERT INTO users (name, email, password, role)
            VALUES (%s,%s,%s,%s)
            ON CONFLICT (email) DO NOTHING
        """, (seed_name, seed_email, hash_password(seed_password), seed_role))
    for rule in DEFAULT_MATERIAL_NORMS:
        cur.execute("""
            INSERT INTO material_norms (
                rule_key, name, work_keywords, block_work_keywords, material_keywords,
                work_unit, material_unit, qty_per_unit, thickness_base_mm,
                default_thickness_mm, label, active
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
            ON CONFLICT (rule_key) DO NOTHING
        """, (
            rule.get("ruleKey", ""),
            rule.get("name", ""),
            json.dumps(rule.get("work", []), ensure_ascii=False),
            json.dumps(rule.get("blockWork", []), ensure_ascii=False),
            json.dumps(rule.get("material", []), ensure_ascii=False),
            rule.get("workUnit", ""),
            rule.get("materialUnit", ""),
            rule.get("qtyPerUnit", 0),
            rule.get("thicknessBaseMm"),
            rule.get("defaultThicknessMm"),
            rule.get("label", ""),
        ))
    conn.close()

init_db()
class ProjectModel(BaseModel):
    name: str
    client: str = ""
    status: str = "Планирование"
    budget: float = 0
    deadline: str = ""
    progress: int = 0
    tasks: List[str] = []
    pricelistId: Optional[int] = None
    floors: int = 1
    liters: str = ""

class ClientModel(BaseModel):
    name: str
    phone: str = ""
    email: str = ""
    status: str = "Активный"
    notes: str = ""

class MaterialModel(BaseModel):
    name: str
    unit: str = "шт"
    quantity: float = 0
    price: float = 0
    minQuantity: float = 0
    project: str = ""
    category: str = ""

class WarehouseMainModel(BaseModel):
    name: str
    unit: str = "шт"
    quantity: float = 0
    price: float = 0
    minQuantity: float = 0
    category: str = ""

class WarehouseHistoryModel(BaseModel):
    material: str
    type: str
    quantity: float
    date: str
    project: str = ""
    issuedTo: str = ""
    issuedBy: str = ""
    dateTime: str = ""

class WarehouseMovementModel(BaseModel):
    materialName: str
    fromLocation: str
    toLocation: str
    quantity: float
    unit: str = ""
    date: str = ""
    createdBy: str = ""
    notes: str = ""

class StaffModel(BaseModel):
    name: str
    role: str = ""
    phone: str = ""
    salary: float = 0
    project: str = ""
    payType: str = "оклад"
    lastName: Optional[str] = ""
    firstName: Optional[str] = ""
    middleName: Optional[str] = ""
    birthDate: Optional[str] = ""
    citizenship: Optional[str] = ""
    address: Optional[str] = ""
    photoUrl: Optional[str] = ""
    emailWork: Optional[str] = ""
    emailPersonal: Optional[str] = ""
    phoneExtra: Optional[str] = ""
    passportSeries: Optional[str] = ""
    passportNumber: Optional[str] = ""
    passportIssuedBy: Optional[str] = ""
    passportIssuedDate: Optional[str] = ""
    inn: Optional[str] = ""
    snils: Optional[str] = ""
    specialization: Optional[str] = ""
    category: Optional[str] = ""
    employmentType: Optional[str] = ""
    hiredDate: Optional[str] = ""
    firedDate: Optional[str] = ""
    status: Optional[str] = "Активен"
    brigade: Optional[str] = ""
    bankAccount: Optional[str] = ""
    bankName: Optional[str] = ""
    bankBik: Optional[str] = ""
    bankCorr: Optional[str] = ""
    ogrnip: Optional[str] = ""
    cardNumber: Optional[str] = ""
    signatureUrl: Optional[str] = ""
    notes: Optional[str] = ""

class PieceworkModel(BaseModel):
    staffId: str
    description: str
    unit: str = "м2"
    quantity: float
    pricePerUnit: float
    total: float
    project: str = ""
    date: str = ""
    comment: str = ""
    photoUrl: str = ""
    workJournalId: Optional[int] = None

class UserModel(BaseModel):
    name: str
    email: str
    password: str = ""
    role: str = "прораб"
    projectId: str = ""
    projectName: str = ""

class LoginModel(BaseModel):
    email: str
    password: str

class PricelistModel(BaseModel):
    name: str
    description: str = ""
    forWho: str = ""
    coefficient: float = 1.0

class PricelistItemModel(BaseModel):
    pricelistId: int
    name: str
    unit: str = "м2"
    price: float = 0
    category: str = ""
    specialization: str = ""

class InviteCodeModel(BaseModel):
    role: str

class RegisterModel(BaseModel):
    name: str
    email: str
    password: str
    code: str

class SupplierModel(BaseModel):
    name: str
    phone: str = ""
    email: str = ""
    specialization: str = ""
    category: str = ""
    rating: float = 5.0
    status: str = "Активный"

class SupplyRequestModel(BaseModel):
    materialName: str = ""
    quantity: float = 0
    unit: str = "шт"
    project: str = ""
    createdBy: str = ""
    date: str = ""
    notes: str = ""
    selectedSuppliers: List[int] = []
    requestedByRole: str = ""
    requestedById: Optional[int] = None
    urgency: str = "обычная"
    category: str = ""
    # Многопозиционная заявка: массив объектов {materialName, quantity, unit}
    items: List[dict] = []

class SupplierOfferModel(BaseModel):
    requestId: int
    supplierId: int
    pricePerUnit: float
    totalPrice: float
    deliveryDays: int = 0
    notes: str = ""

class SupplyHistoryModel(BaseModel):
    supplierId: int
    materialName: str
    quantity: float
    unit: str = ""
    pricePerUnit: float
    totalPrice: float
    project: str = ""
    date: str = ""
    status: str = "Ожидает поставки"

class MaterialNormModel(BaseModel):
    ruleKey: str = ""
    name: str = ""
    work: List[str] = []
    blockWork: List[str] = []
    material: List[str] = []
    workUnit: str = "м2"
    materialUnit: str = "кг"
    qtyPerUnit: float = 0
    thicknessBaseMm: Optional[float] = None
    defaultThicknessMm: Optional[float] = None
    label: str = ""
    active: bool = True

class MaterialNormOverrideModel(BaseModel):
    baseNormId: Optional[int] = None
    projectName: str = ""
    estimateId: Optional[int] = None
    sectionName: str = ""
    workName: str = ""
    materialName: str = ""
    work: List[str] = []
    blockWork: List[str] = []
    material: List[str] = []
    workUnit: str = "м2"
    materialUnit: str = "кг"
    qtyPerUnit: float = 0
    thicknessBaseMm: Optional[float] = None
    defaultThicknessMm: Optional[float] = None
    label: str = ""
    reason: str = ""
    active: bool = True

class MaterialNormSuggestionUpdateModel(BaseModel):
    status: str = ""

class WorkJournalModel(BaseModel):
    masterId: int
    masterName: str
    project: str
    description: str
    unit: str
    quantity: float
    pricePerUnit: float = 0
    total: float = 0
    date: str
    comment: str = ""
    photoUrl: str = ""
    materialsUsed: list = []
    estimateId: int | None = None
    sectionName: str = ""
    responsibleItr: str = ""
    weather: str = ""
    timeStart: str = ""
    timeEnd: str = ""
    hiddenWork: bool = False
    qualityStatus: str = ""
    normatives: str = ""
    projectDocs: str = ""

class MasterProfileModel(BaseModel):
    userId: int
    fullName: str
    passport: str = ""
    inn: str = ""
    contractType: str = "ГПХ"
    bankAccount: str = ""
    bankName: str = ""
    phone: str = ""
    specialization: str = ""
    ogrnip: str = ""

class ContractModel(BaseModel):
    masterId: int
    masterName: str
    contractType: str = "ГПХ"
    contractNumber: str
    project: str
    startDate: str = ""
    endDate: str = ""

class InterimActModel(BaseModel):
    masterId: int
    masterName: str
    project: str
    periodStart: str
    periodEnd: str
    totalAmount: float = 0
    paidAmount: float = 0
    contractId: Optional[int] = None

class TimesheetModel(BaseModel):
    staffId: int
    day: str

class CopyPricelistModel(BaseModel):
    name: str

class RoomModel(BaseModel):
    floor: int = 1
    liter: str = ''
    roomType: str = 'Комната'
    project: str
    name: str
    floorArea: float = 0
    wallArea: float = 0
    ceilingArea: float = 0
    height: float = 0
    ceilingType: str = "Простой"
    wallMaterial: str = "Штукатурка"
    floorMaterial: str = "Стяжка"
    windows: int = 0
    doors: int = 0
    notes: str = ""

class RoomWorkModel(BaseModel):
    roomId: int
    project: str
    roomName: str
    masterId: int
    masterName: str
    description: str
    surface: str = ""
    unit: str = "м2"
    quantity: float = 0
    pricePerUnit: float = 0
    total: float = 0
    date: str = ""
    photoUrl: str = ""

class AiFindingModel(BaseModel):
    projectName: str = ""
    findingType: str = "manual"
    category: str = "Общее"
    severity: str = "Проверить"
    title: str = ""
    description: str = ""
    source: str = "manual"
    linkedEntityType: str = ""
    linkedEntityId: str = ""
    suggestedAction: str = ""
    assignedRole: str = ""
    assignedTo: str = ""
    status: str = "Новое"
    dedupeKey: str = ""

class AiTaskModel(BaseModel):
    findingId: Optional[int] = None
    projectName: str = ""
    title: str = ""
    description: str = ""
    assignedRole: str = ""
    assignedTo: str = ""
    status: str = "Новое"
    dueDate: str = ""
    actionLabel: str = ""
    actionPayload: str = ""

class ToolModel(BaseModel):
    name: str
    inventoryNumber: str = ""
    cost: float = 0
    status: str = "На складе"
    location: str = "Основной склад"
    project: str = ""
    masterId: Optional[int] = None
    masterName: str = ""
    issueType: str = ""
    photoUrl: str = ""
    notes: str = ""

class ToolHistoryModel(BaseModel):
    toolId: int
    toolName: str
    action: str
    fromLocation: str = ""
    toLocation: str = ""
    masterName: str = ""
    project: str = ""
    issueType: str = ""
    condition: str = ""
    date: str = ""
    createdBy: str = ""

class InventoryModel(BaseModel):
    project: str
    date: str
    createdBy: str
    notes: str = ""

class InventoryItemModel(BaseModel):
    inventoryId: int
    materialName: str
    unit: str
    expected: float
    actual: float
    difference: float
    notes: str = ""

class PdConsentModel(BaseModel):
    userId: int
    signedAt: str = ""
    scanUrl: str = ""
    uploadedBy: str = ""

@app.post("/login")
def login(data: LoginModel):
    from datetime import datetime as _dt, timedelta as _td
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # Проверяем не заблокирован ли пользователь
    cur.execute("SELECT id, failed_login_count, locked_until FROM users WHERE email=%s", (data.email,))
    user_check = cur.fetchone()
    if user_check and user_check.get('locked_until'):
        if user_check['locked_until'] > _dt.now():
            mins_left = int((user_check['locked_until'] - _dt.now()).total_seconds() / 60) + 1
            conn.close()
            raise HTTPException(status_code=429, detail="Аккаунт временно заблокирован после 5 неверных попыток. Попробуйте через "+str(mins_left)+" мин.")
        else:
            # Срок блокировки истёк — сбрасываем
            cur.execute("UPDATE users SET failed_login_count=0, locked_until=NULL WHERE id=%s", (user_check['id'],))
            conn.commit()
    # Логинимся: пароль проверяем в Python, чтобы поддержать хеши и старые plaintext-пароли.
    cur.execute("SELECT * FROM users WHERE LOWER(email)=LOWER(%s)", (data.email,))
    user = cur.fetchone()
    if not user or not verify_password(data.password, user.get("password") or ""):
        # Увеличиваем счётчик неудачных попыток
        if user_check:
            new_count = (user_check.get('failed_login_count') or 0) + 1
            if new_count >= 5:
                lock_until = _dt.now() + _td(minutes=15)
                cur.execute("UPDATE users SET failed_login_count=%s, locked_until=%s WHERE id=%s", (new_count, lock_until, user_check['id']))
                conn.commit()
                conn.close()
                raise HTTPException(status_code=429, detail="Превышено количество попыток. Аккаунт заблокирован на 15 минут.")
            else:
                cur.execute("UPDATE users SET failed_login_count=%s WHERE id=%s", (new_count, user_check['id']))
                conn.commit()
        conn.close()
        raise HTTPException(status_code=401, detail="Неверный email или пароль")
    # Успех — сбрасываем счётчик
    if is_legacy_password(user.get("password") or ""):
        new_hash = hash_password(data.password)
        cur.execute("UPDATE users SET password=%s WHERE id=%s", (new_hash, user['id']))
        user["password"] = new_hash
    cur.execute("UPDATE users SET failed_login_count=0, locked_until=NULL WHERE id=%s", (user['id'],))
    conn.commit()
    conn.close()
    log_audit(user_name=user.get("name",""), user_role=user.get("role",""),
              action="login", entity_type="user", entity_id=user['id'],
              description="Успешный вход в систему")
    return public_user(user, include_token=True)

@app.post("/password-reset-request")
def password_reset_request(data: dict):
    """Генерирует 6-значный код для восстановления пароля. Действителен 30 минут."""
    from datetime import datetime as _dt, timedelta as _td
    import random
    email = (data.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Введите email")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM users WHERE LOWER(email)=%s", (email,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        # Не палим существование email для безопасности
        return {"ok": True, "message": "Если такой email зарегистрирован, код отправлен"}
    code = str(random.randint(100000, 999999))
    expires = _dt.now() + _td(minutes=30)
    cur.execute("UPDATE users SET reset_token=%s, reset_token_expires=%s WHERE id=%s", (code, expires, row[0]))
    conn.commit()
    cur.close(); conn.close()
    # В реальной системе тут отправляется email/SMS. Пока выводим в ответе для админа (в логах сервера)
    print(f"PASSWORD RESET CODE for {email}: {code} (valid 30 min)")
    log_audit(user_name=row[1] or "—", user_role="—",
              action="password_reset_request", entity_type="user", entity_id=row[0],
              description="Запрошен код восстановления пароля")
    response = {"ok": True, "message": "Код отправлен на email (или его выдаст администратор из логов)"}
    if os.getenv("SHOW_RESET_CODE", "").lower() in ("1", "true", "yes"):
        response["_devCode"] = code
    return response

@app.post("/password-reset")
def password_reset(data: dict):
    """Меняет пароль по коду восстановления."""
    from datetime import datetime as _dt
    email = (data.get("email") or "").strip().lower()
    code = (data.get("code") or "").strip()
    new_password = data.get("newPassword") or ""
    if not email or not code or not new_password:
        raise HTTPException(status_code=400, detail="Заполните email, код и новый пароль")
    if len(new_password) < 5:
        raise HTTPException(status_code=400, detail="Пароль должен быть не короче 5 символов")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, reset_token, reset_token_expires FROM users WHERE LOWER(email)=%s", (email,))
    row = cur.fetchone()
    if not row or not row[1] or row[1] != code:
        cur.close(); conn.close()
        raise HTTPException(status_code=401, detail="Неверный код или email")
    if row[2] and row[2] < _dt.now():
        cur.close(); conn.close()
        raise HTTPException(status_code=401, detail="Код истёк — запросите новый")
    cur.execute("UPDATE users SET password=%s, reset_token=NULL, reset_token_expires=NULL, failed_login_count=0, locked_until=NULL WHERE id=%s",
                (hash_password(new_password), row[0]))
    conn.commit()
    cur.close(); conn.close()
    log_audit(user_name="—", user_role="—",
              action="password_reset", entity_type="user", entity_id=row[0],
              description="Сброс пароля через код восстановления")
    return {"ok": True, "message": "Пароль изменён, можно входить"}

@app.post("/register")
def register(data: dict):
    from datetime import datetime
    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    password = data.get("password") or ""
    if not code or not name or not email or not password:
        raise HTTPException(status_code=400, detail="Заполните имя, email, пароль и код приглашения")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM invite_codes WHERE code=%s AND used=FALSE", (code,))
    invite = cur.fetchone()
    if not invite:
        conn.close()
        raise HTTPException(status_code=400, detail="Неверный или использованный код")
    role = invite['role']
    # Проверка срока действия
    if invite.get('expires_at') and invite['expires_at'] < datetime.now():
        conn.close()
        raise HTTPException(status_code=400, detail="Код истёк — попросите новую ссылку")
    try:
        cur.execute("INSERT INTO users (name,email,password,role) VALUES (%s,%s,%s,%s) RETURNING *",
                    (name, email, hash_password(password), role))
        user = cur.fetchone()
        # Если регистрируется поставщик — создаём/связываем suppliers row
        if role == 'поставщик':
            company_name = data.get("companyName") or name
            supplier_id = invite.get('supplier_id')
            if supplier_id:
                # Привязываем к существующей компании
                cur.execute(
                    "UPDATE suppliers SET phone=COALESCE(%s,phone), email=COALESCE(%s,email), "
                    "inn=COALESCE(%s,inn), kpp=COALESCE(%s,kpp), ogrn=COALESCE(%s,ogrn), "
                    "legal_address=COALESCE(%s,legal_address), bank=COALESCE(%s,bank), "
                    "bik=COALESCE(%s,bik), account=COALESCE(%s,account), "
                    "director_name=COALESCE(%s,director_name), "
                    "user_id=%s, registered_at=NOW() WHERE id=%s",
                    (data.get("phone"), email, data.get("inn"), data.get("kpp"),
                     data.get("ogrn"), data.get("legalAddress"), data.get("bank"),
                     data.get("bik"), data.get("account"), data.get("directorName"),
                     user['id'], supplier_id))
            else:
                # Создаём новую компанию
                cur.execute(
                    "INSERT INTO suppliers (name, phone, email, category, specialization, "
                    "inn, kpp, ogrn, legal_address, bank, bik, account, director_name, "
                    "status, rating, user_id, registered_at) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())",
                    (company_name, data.get("phone",""), email,
                     invite.get('preset_category') or data.get("category",""),
                     data.get("specialization",""),
                     data.get("inn"), data.get("kpp"), data.get("ogrn"),
                     data.get("legalAddress"), data.get("bank"), data.get("bik"),
                     data.get("account"), data.get("directorName"),
                     'Активный', 5.0, user['id']))
        cur.execute("UPDATE invite_codes SET used=TRUE WHERE code=%s", (code,))
        conn.close()
        return public_user(user, include_token=True)
    except HTTPException:
        conn.close()
        raise
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/projects")
def get_projects(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    allowed_projects = visible_project_names(current_user)
    if allowed_projects is None:
        cur.execute("SELECT id,name,client,status,budget,deadline,progress,tasks,pricelist_id as \"pricelistId\",floors,liters,warranty_start_date as \"warrantyStartDate\",warranty_end_date as \"warrantyEndDate\",warranty_contact as \"warrantyContact\",COALESCE(archived,false) as archived,archived_at as \"archivedAt\" FROM projects")
    elif not allowed_projects:
        cur.close(); conn.close()
        return []
    else:
        cur.execute("SELECT id,name,client,status,budget,deadline,progress,tasks,pricelist_id as \"pricelistId\",floors,liters,warranty_start_date as \"warrantyStartDate\",warranty_end_date as \"warrantyEndDate\",warranty_contact as \"warrantyContact\",COALESCE(archived,false) as archived,archived_at as \"archivedAt\" FROM projects WHERE name = ANY(%s)", (allowed_projects,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/projects")
def create_project(p: ProjectModel, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO projects (name,client,status,budget,deadline,progress,tasks,pricelist_id,floors,liters) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id,name,client,status,budget,deadline,progress,tasks,pricelist_id as \"pricelistId\",floors,liters",
                (p.name,p.client,p.status,p.budget,p.deadline,p.progress,p.tasks,p.pricelistId,p.floors,p.liters))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/projects/{id}")
def update_project(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    fields_map = [
        ('name','name'),('client','client'),('status','status'),('budget','budget'),
        ('deadline','deadline'),('progress','progress'),('tasks','tasks'),
        ('pricelistId','pricelist_id'),('floors','floors'),('liters','liters'),
        ('warrantyStartDate','warranty_start_date'),
        ('warrantyEndDate','warranty_end_date'),
        ('warrantyContact','warranty_contact'),
        ('archived','archived'),
        ('archivedAt','archived_at'),
    ]
    sets, vals = [], []
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            v = data[js_key]
            if db_col in ('warranty_start_date','warranty_end_date','deadline','archived_at') and not v:
                v = None
            vals.append(v)
    if not sets:
        return {"ok": True}
    vals.append(id)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE projects SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/projects/{id}")
def delete_project(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM projects WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/clients")
def get_clients(_current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "менеджер_crm"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM clients")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/clients")
def create_client(c: ClientModel, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO clients (name,phone,email,status,notes) VALUES (%s,%s,%s,%s,%s) RETURNING *",
                (c.name,c.phone,c.email,c.status,c.notes))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/clients/{id}")
def update_client(id: int, c: ClientModel, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE clients SET name=%s,phone=%s,email=%s,status=%s,notes=%s WHERE id=%s",
                (c.name,c.phone,c.email,c.status,c.notes,id))
    conn.close()
    return {"ok": True}

@app.delete("/clients/{id}")
def delete_client(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM clients WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/materials")
def get_materials(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    base = "SELECT id,name,unit,quantity,price,min_quantity as \"minQuantity\",project,category FROM materials"
    projects = user_project_names(current_user)
    if current_user.get("role") in WAREHOUSE_ROLES or can_see_all_company_data(current_user):
        cur.execute(base)
    elif projects:
        cur.execute(base + " WHERE project = ANY(%s)", (projects,))
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/materials")
def create_material(m: MaterialModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO materials (name,unit,quantity,price,min_quantity,project,category) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id,name,unit,quantity,price,min_quantity as \"minQuantity\",project,category",
                (m.name,m.unit,m.quantity,m.price,m.minQuantity,m.project,m.category))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/materials/{id}")
def update_material(id: int, m: MaterialModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE materials SET name=%s,unit=%s,quantity=%s,price=%s,min_quantity=%s,project=%s,category=%s WHERE id=%s",
                (m.name,m.unit,m.quantity,m.price,m.minQuantity,m.project,m.category,id))
    conn.close()
    return {"ok": True}

@app.delete("/materials/{id}")
def delete_material(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM materials WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/warehouse-main")
def get_warehouse_main(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in WAREHOUSE_ROLES and not can_see_all_company_data(current_user):
        return []
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id,name,unit,quantity,price,min_quantity as \"minQuantity\",category FROM warehouse_main")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/warehouse-main")
def create_warehouse_main(m: WarehouseMainModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO warehouse_main (name,unit,quantity,price,min_quantity,category) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id,name,unit,quantity,price,min_quantity as \"minQuantity\",category",
                (m.name,m.unit,m.quantity,m.price,m.minQuantity,m.category))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/warehouse-main/{id}")
def update_warehouse_main(id: int, m: WarehouseMainModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE warehouse_main SET name=%s,unit=%s,quantity=%s,price=%s,min_quantity=%s,category=%s WHERE id=%s",
                (m.name,m.unit,m.quantity,m.price,m.minQuantity,m.category,id))
    conn.close()
    return {"ok": True}

@app.delete("/warehouse-main/{id}")
def delete_warehouse_main(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM warehouse_main WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/warehouse-movements")
def get_warehouse_movements(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if current_user.get("role") == "прораб":
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,material_name as \"materialName\",from_location as \"fromLocation\",to_location as \"toLocation\",quantity,unit,date,created_by as \"createdBy\",notes FROM warehouse_movements WHERE from_location = ANY(%s) OR to_location = ANY(%s) ORDER BY id DESC", (projects, projects))
    elif current_user.get("role") in WAREHOUSE_ROLES or current_user.get("role") in FINANCE_ROLES:
        cur.execute("SELECT id,material_name as \"materialName\",from_location as \"fromLocation\",to_location as \"toLocation\",quantity,unit,date,created_by as \"createdBy\",notes FROM warehouse_movements ORDER BY id DESC")
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/warehouse-movements")
def create_warehouse_movement(m: WarehouseMovementModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO warehouse_movements (material_name,from_location,to_location,quantity,unit,date,created_by,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (m.materialName,m.fromLocation,m.toLocation,m.quantity,m.unit,m.date,m.createdBy,m.notes))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.get("/warehouse-history")
def get_warehouse_history(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if current_user.get("role") == "прораб":
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,material,type,quantity,date,project,issued_to as \"issuedTo\",issued_by as \"issuedBy\",date_time as \"dateTime\" FROM warehouse_history WHERE project = ANY(%s) ORDER BY id DESC", (projects,))
    elif current_user.get("role") in WAREHOUSE_ROLES or current_user.get("role") in FINANCE_ROLES:
        cur.execute("SELECT id,material,type,quantity,date,project,issued_to as \"issuedTo\",issued_by as \"issuedBy\",date_time as \"dateTime\" FROM warehouse_history ORDER BY id DESC")
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/warehouse-history")
def create_warehouse_history(h: WarehouseHistoryModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_to,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (h.material,h.type,h.quantity,h.date,h.project,h.issuedTo,h.issuedBy,h.dateTime))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.delete("/warehouse-history/{id}")
def delete_warehouse_history(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM warehouse_history WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

STAFF_COLUMNS = """id, name, role, phone, salary, project, pay_type as "payType",
    last_name as "lastName", first_name as "firstName", middle_name as "middleName",
    birth_date as "birthDate", citizenship, address, photo_url as "photoUrl",
    email_work as "emailWork", email_personal as "emailPersonal", phone_extra as "phoneExtra",
    passport_series as "passportSeries", passport_number as "passportNumber",
    passport_issued_by as "passportIssuedBy", passport_issued_date as "passportIssuedDate",
    inn, snils, specialization, category,
    employment_type as "employmentType", hired_date as "hiredDate", fired_date as "firedDate",
    status, brigade, bank_account as "bankAccount", bank_name as "bankName",
    bank_bik as "bankBik", bank_corr as "bankCorr", ogrnip, card_number as "cardNumber",
    signature_url as "signatureUrl", notes"""

def _staff_tuple(s):
    def d(v):
        return v if v else None
    return (s.name, s.role, s.phone, s.salary, s.project, s.payType,
            s.lastName or None, s.firstName or None, s.middleName or None,
            d(s.birthDate), s.citizenship or None, s.address or None, s.photoUrl or None,
            s.emailWork or None, s.emailPersonal or None, s.phoneExtra or None,
            s.passportSeries or None, s.passportNumber or None,
            s.passportIssuedBy or None, d(s.passportIssuedDate),
            s.inn or None, s.snils or None, s.specialization or None, s.category or None,
            s.employmentType or None, d(s.hiredDate), d(s.firedDate),
            s.status or "Активен", s.brigade or None,
            s.bankAccount or None, s.bankName or None, s.bankBik or None, s.bankCorr or None,
            s.ogrnip or None, s.cardNumber or None, s.signatureUrl or None, s.notes or None)

STAFF_INSERT_COLS = """name, role, phone, salary, project, pay_type,
    last_name, first_name, middle_name, birth_date, citizenship, address, photo_url,
    email_work, email_personal, phone_extra,
    passport_series, passport_number, passport_issued_by, passport_issued_date,
    inn, snils, specialization, category,
    employment_type, hired_date, fired_date, status, brigade,
    bank_account, bank_name, bank_bik, bank_corr, ogrnip, card_number,
    signature_url, notes"""
STAFF_PLACEHOLDERS = ",".join(["%s"] * 37)

@app.get("/staff")
def get_staff(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in STAFF_MANAGE_ROLES and current_user.get("role") not in ("прораб", "главный_инженер"):
        return []
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT " + STAFF_COLUMNS + " FROM staff ORDER BY id")
    rows = cur.fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        for k in ("birthDate", "passportIssuedDate", "hiredDate", "firedDate"):
            d[k] = str(d[k]) if d.get(k) else ""
        for k in list(d.keys()):
            if d[k] is None:
                d[k] = ""
        result.append(d)
    return result

@app.post("/staff")
def create_staff(s: StaffModel, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO staff (" + STAFF_INSERT_COLS + ") VALUES (" + STAFF_PLACEHOLDERS + ") RETURNING id", _staff_tuple(s))
    new_id = cur.fetchone()[0]
    conn.commit(); conn.close()
    return {"id": new_id, "ok": True}

@app.put("/staff/{id}")
def update_staff(id: int, s: StaffModel, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""UPDATE staff SET name=%s, role=%s, phone=%s, salary=%s, project=%s, pay_type=%s,
        last_name=%s, first_name=%s, middle_name=%s, birth_date=%s, citizenship=%s, address=%s, photo_url=%s,
        email_work=%s, email_personal=%s, phone_extra=%s,
        passport_series=%s, passport_number=%s, passport_issued_by=%s, passport_issued_date=%s,
        inn=%s, snils=%s, specialization=%s, category=%s,
        employment_type=%s, hired_date=%s, fired_date=%s, status=%s, brigade=%s,
        bank_account=%s, bank_name=%s, bank_bik=%s, bank_corr=%s, ogrnip=%s, card_number=%s,
        signature_url=%s, notes=%s WHERE id=%s""", _staff_tuple(s) + (id,))
    conn.commit(); conn.close()
    return {"ok": True}

@app.delete("/staff/{id}")
def delete_staff(id: int, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM staff WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/staff/{staff_id}/profile")
def get_staff_profile(staff_id: int, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES, "прораб", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM staff WHERE id=%s", (staff_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Сотрудник не найден")
    staff_name = row[1] or ""

    cur.execute("SELECT id, doc_type, title, file_url, status, signed_at, expires_at, notes, created_at FROM staff_documents WHERE staff_id=%s ORDER BY id DESC", (staff_id,))
    custom = [{"id": r[0], "docType": r[1], "title": r[2] or "", "fileUrl": r[3] or "", "status": r[4] or "", "signedAt": str(r[5]) if r[5] else "", "expiresAt": str(r[6]) if r[6] else "", "notes": r[7] or "", "createdAt": str(r[8])} for r in cur.fetchall()]

    # Match user by name (legacy linkage)
    cur.execute("SELECT id FROM users WHERE name=%s LIMIT 1", (staff_name,))
    u = cur.fetchone()
    user_id = u[0] if u else None

    contracts_list = []
    if user_id is not None:
        cur.execute("SELECT id, contract_number, project, start_date, end_date, signed_at, status FROM contracts WHERE master_id=%s ORDER BY id DESC", (user_id,))
        for r in cur.fetchall():
            contracts_list.append({"id": r[0], "contractNumber": r[1] or "", "project": r[2] or "", "startDate": str(r[3]) if r[3] else "", "endDate": str(r[4]) if r[4] else "", "signedAt": str(r[5]) if r[5] else "", "status": r[6] or ""})

    acts_list = []
    if user_id is not None:
        cur.execute("SELECT id, act_number, project, period_from, period_to, total_amount, paid_amount, status, created_at FROM interim_acts WHERE master_id=%s ORDER BY id DESC", (user_id,))
        for r in cur.fetchall():
            acts_list.append({"id": r[0], "actNumber": r[1] or "", "project": r[2] or "", "periodFrom": str(r[3]) if r[3] else "", "periodTo": str(r[4]) if r[4] else "", "totalAmount": float(r[5] or 0), "paidAmount": float(r[6] or 0), "status": r[7] or "", "createdAt": str(r[8])})

    pd_consents = []
    if user_id is not None:
        cur.execute("SELECT id, signed_at, scan_url, uploaded_by FROM pd_consents WHERE user_id=%s ORDER BY id DESC", (user_id,))
        for r in cur.fetchall():
            pd_consents.append({"id": r[0], "signedAt": r[1] or "", "scanUrl": r[2] or "", "uploadedBy": r[3] or ""})

    tb_entries = []
    cur.execute("SELECT id, project_name, instructor, instruction_type, date FROM tb_journal WHERE master_name=%s ORDER BY id DESC LIMIT 20", (staff_name,))
    for r in cur.fetchall():
        tb_entries.append({"id": r[0], "projectName": r[1] or "", "instructor": r[2] or "", "instructionType": r[3] or "", "date": str(r[4]) if r[4] else ""})

    works = []
    cur.execute("SELECT project, description, quantity, unit, total, date, status FROM work_journal WHERE master_name=%s ORDER BY id DESC LIMIT 50", (staff_name,))
    for r in cur.fetchall():
        works.append({"project": r[0] or "", "description": r[1] or "", "quantity": float(r[2] or 0), "unit": r[3] or "", "total": float(r[4] or 0), "date": str(r[5]) if r[5] else "", "status": r[6] or ""})

    cur.close(); conn.close()
    return {
        "staffId": staff_id,
        "staffName": staff_name,
        "userId": user_id,
        "customDocuments": custom,
        "contracts": contracts_list,
        "acts": acts_list,
        "pdConsents": pd_consents,
        "tbJournal": tb_entries,
        "workJournal": works,
    }

@app.post("/staff/{staff_id}/documents")
def add_staff_document(staff_id: int, data: dict, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO staff_documents (staff_id, doc_type, title, file_url, status, signed_at, expires_at, notes, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id""",
        (staff_id, data.get("docType","другое"), data.get("title",""), data.get("fileUrl","") or None,
         data.get("status","действует"), data.get("signedAt") or None, data.get("expiresAt") or None,
         data.get("notes",""), data.get("createdBy","")))
    new_id = cur.fetchone()[0]
    conn.commit(); cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@app.delete("/staff-documents/{doc_id}")
def delete_staff_document(doc_id: int, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM staff_documents WHERE id=%s", (doc_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@app.get("/piecework")
def get_piecework(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    allowed_projects = visible_project_names(current_user)
    cols = "id,staff_id as \"staffId\",description,unit,quantity,price_per_unit as \"pricePerUnit\",total,project,date,comment,photo_url as \"photoUrl\",work_journal_id as \"workJournalId\""
    if allowed_projects is None:
        cur.execute(f"SELECT {cols} FROM piecework ORDER BY id DESC")
    elif not allowed_projects:
        cur.execute(f"SELECT {cols} FROM piecework WHERE staff_id=%s ORDER BY id DESC", (str(current_user.get("id")),))
    else:
        cur.execute(f"SELECT {cols} FROM piecework WHERE project = ANY(%s) OR staff_id=%s ORDER BY id DESC", (allowed_projects, str(current_user.get("id"))))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/piecework")
def create_piecework(p: PieceworkModel, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    require_project_access(_current_user, p.project)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if p.workJournalId:
        cur.execute("SELECT id FROM piecework WHERE work_journal_id=%s LIMIT 1", (p.workJournalId,))
        existing = cur.fetchone()
        if existing:
            conn.close()
            return dict(existing)
    cur.execute("INSERT INTO piecework (staff_id,description,unit,quantity,price_per_unit,total,project,date,comment,photo_url,work_journal_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (p.staffId,p.description,p.unit,p.quantity,p.pricePerUnit,p.total,p.project,p.date,p.comment,p.photoUrl,p.workJournalId))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.delete("/piecework/{id}")
def delete_piecework(id: int, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES, "прораб", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "piecework", id, _current_user, "project")
    cur.execute("DELETE FROM piecework WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/users")
def get_users(current_user: dict = Depends(get_current_user)):
    import json as _j
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if current_user.get("role") in LEADERSHIP_ROLES or current_user.get("role") == "бухгалтер":
        cur.execute("SELECT id,name,email,role,project_id,project_name,assigned_projects FROM users")
    else:
        cur.execute("SELECT id,name,email,role,project_id,project_name,assigned_projects FROM users WHERE id=%s", (current_user.get("id"),))
    rows = cur.fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        ap = d.get('assigned_projects')
        try:
            if isinstance(ap, str): d['assignedProjects'] = _j.loads(ap)
            elif isinstance(ap, list): d['assignedProjects'] = ap
            else: d['assignedProjects'] = []
        except: d['assignedProjects'] = []
        d.pop('assigned_projects', None)
        d['projectId'] = d.pop('project_id', None)
        d['projectName'] = d.pop('project_name', '')
        out.append(d)
    return out

@app.put("/users/{id}/assigned-projects")
def update_assigned_projects(id: int, data: dict, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    """Назначить прорабу список проектов (массив имён)."""
    import json as _j
    projects_list = data.get('assignedProjects') or []
    if not isinstance(projects_list, list):
        return {"ok": False, "error": "assignedProjects must be array"}
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE users SET assigned_projects=%s::jsonb WHERE id=%s",
                (_j.dumps(projects_list), id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/users")
def create_user(u: UserModel, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("INSERT INTO users (name,email,password,role,project_id,project_name) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id,name,email,role,project_id,project_name",
                    (u.name,u.email,hash_password(u.password),u.role,int(u.projectId) if u.projectId else None,u.projectName or ""))
        row = cur.fetchone()
        conn.close()
        return dict(row)
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/users/{id}")
def update_user(id: int, u: UserModel, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    if u.password:
        cur.execute("UPDATE users SET name=%s,email=%s,password=%s,role=%s WHERE id=%s",
                    (u.name,u.email,hash_password(u.password),u.role,id))
    else:
        cur.execute("UPDATE users SET name=%s,email=%s,role=%s WHERE id=%s",
                    (u.name,u.email,u.role,id))
    conn.close()
    return {"ok": True}

@app.delete("/users/{id}")
def delete_user(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM users WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/pricelists")
def get_pricelists(_current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id,name,description,for_who as \"forWho\",coefficient FROM pricelists")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/pricelists")
def create_pricelist(p: PricelistModel, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO pricelists (name,description,for_who,coefficient) VALUES (%s,%s,%s,%s) RETURNING id,name,description,for_who as \"forWho\",coefficient",
                (p.name,p.description,p.forWho,p.coefficient))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/pricelists/{id}")
def update_pricelist(id: int, p: PricelistModel, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE pricelists SET name=%s,description=%s,for_who=%s,coefficient=%s WHERE id=%s",
                (p.name,p.description,p.forWho,p.coefficient,id))
    conn.close()
    return {"ok": True}

@app.delete("/pricelists/{id}")
def delete_pricelist(id: int, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM pricelists WHERE id=%s", (id,))
    cur.execute("DELETE FROM pricelist_items WHERE pricelist_id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.post("/pricelists/{id}/copy")
def copy_pricelist(id: int, data: CopyPricelistModel, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM pricelists WHERE id=%s", (id,))
    pl = cur.fetchone()
    cur.execute("INSERT INTO pricelists (name,description,for_who,coefficient) VALUES (%s,%s,%s,%s) RETURNING *",
                (data.name,pl['description'],pl['for_who'],pl['coefficient']))
    new_pl = cur.fetchone()
    cur.execute("SELECT * FROM pricelist_items WHERE pricelist_id=%s", (id,))
    items = cur.fetchall()
    for item in items:
        cur.execute("INSERT INTO pricelist_items (pricelist_id,name,unit,price,category,specialization) VALUES (%s,%s,%s,%s,%s,%s)",
                    (new_pl['id'],item['name'],item['unit'],item['price'],item['category'],item['specialization']))
    conn.close()
    return dict(new_pl)

@app.get("/pricelists/{id}/items")
def get_pricelist_items(id: int, _current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id,pricelist_id as \"pricelistId\",name,unit,price,category,specialization FROM pricelist_items WHERE pricelist_id=%s ORDER BY category,name", (id,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/pricelist-items")
def create_pricelist_item(item: PricelistItemModel, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO pricelist_items (pricelist_id,name,unit,price,category,specialization) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id,pricelist_id as \"pricelistId\",name,unit,price,category,specialization",
                (item.pricelistId,item.name,item.unit,item.price,item.category,item.specialization))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/pricelist-items/{id}")
def update_pricelist_item(id: int, item: PricelistItemModel, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE pricelist_items SET name=%s,unit=%s,price=%s,category=%s,specialization=%s WHERE id=%s",
                (item.name,item.unit,item.price,item.category,item.specialization,id))
    conn.close()
    return {"ok": True}

@app.delete("/pricelist-items/{id}")
def delete_pricelist_item(id: int, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM pricelist_items WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/invite-codes")
def get_invite_codes(_current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "system_owner"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM invite_codes ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/invite-codes")
def create_invite_code(data: dict, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "system_owner"))):
    from datetime import datetime, timedelta
    role = data.get('role') or ''
    if not role:
        raise HTTPException(status_code=400, detail="Не указана роль")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    code = str(uuid.uuid4())[:8].upper()
    expires_in_days = int(data.get('expiresInDays') or 14)
    expires_at = datetime.now() + timedelta(days=expires_in_days)
    cur.execute(
        "INSERT INTO invite_codes (code, role, supplier_id, preset_name, preset_category, created_by, expires_at) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (code, role, data.get('supplierId'), data.get('presetName'),
         data.get('presetCategory'), data.get('createdBy'), expires_at))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.delete("/invite-codes/{id}")
def delete_invite_code(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "system_owner"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM invite_codes WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/companies")
def list_companies(_current_user: dict = Depends(get_current_user)):
    """Список компаний-клиентов системы. Сейчас одна (СтройКа).
       В будущем — multi-tenancy: поставщик видит заявки от разных компаний."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, name, short_name as \"shortName\", inn, logo_url as \"logoUrl\", plan, active FROM companies WHERE active=TRUE ORDER BY id")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

# === SaaS: Кабинет системы (только для system_owner) ===

@app.get("/system/companies")
def system_companies_list(_current_user: dict = Depends(require_roles("system_owner"))):
    """Полный список компаний с биллингом — только для владельца платформы."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT id, name, short_name, inn, contact_name, contact_phone, contact_email,
                          plan, trial_until, plan_expires_at, monthly_fee, payment_status,
                          suspended_at, suspended_reason, max_projects, max_users,
                          last_active_at, notes, active, created_at
                   FROM companies ORDER BY id""")
    rows = [dict(r) for r in cur.fetchall()]
    # Считаем подсчёты по каждой компании
    for c in rows:
        cur.execute("SELECT COUNT(*) FROM users WHERE company_id=%s", (c['id'],))
        c['users_count'] = cur.fetchone()['count']
        cur.execute("SELECT COUNT(*) FROM projects WHERE company_id=%s", (c['id'],))
        c['projects_count'] = cur.fetchone()['count']
        cur.execute("SELECT COALESCE(SUM(amount),0) as t FROM company_payments WHERE company_id=%s AND status='paid'", (c['id'],))
        c['total_paid'] = float(cur.fetchone()['t'] or 0)
    conn.close()
    return rows

@app.post("/system/companies")
def system_create_company(data: dict, _current_user: dict = Depends(require_roles("system_owner"))):
    """Создание новой компании-клиента + инвайт-код её директору."""
    from datetime import datetime, timedelta
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    plan = data.get('plan') or 'demo'
    trial_days = int(data.get('trialDays') or 30)
    trial_until = (datetime.now() + timedelta(days=trial_days)).date() if plan == 'demo' else None
    cur.execute("""INSERT INTO companies (name, short_name, inn, kpp, contact_name, contact_phone,
                                          contact_email, plan, trial_until, monthly_fee,
                                          payment_status, max_projects, max_users, active, notes)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (data.get('name'), data.get('shortName'), data.get('inn'), data.get('kpp'),
         data.get('contactName'), data.get('contactPhone'), data.get('contactEmail'),
         plan, trial_until, float(data.get('monthlyFee') or 0),
         'trial' if plan == 'demo' else 'active',
         int(data.get('maxProjects') or 0) or None, int(data.get('maxUsers') or 0) or None,
         True, data.get('notes')))
    new_id = cur.fetchone()['id']
    # Создаём инвайт-код для директора этой новой компании
    invite_code = str(uuid.uuid4())[:8].upper()
    expires = datetime.now() + timedelta(days=30)
    cur.execute("INSERT INTO invite_codes (code, role, preset_name, expires_at, created_by) VALUES (%s,%s,%s,%s,%s)",
        (invite_code, 'директор', data.get('name'), expires, data.get('createdBy') or 'system_owner'))
    conn.close()
    return {"id": new_id, "inviteCode": invite_code, "trialUntil": str(trial_until) if trial_until else None}

@app.put("/system/companies/{id}")
def system_update_company(id: int, data: dict, _current_user: dict = Depends(require_roles("system_owner"))):
    """Обновление компании: смена тарифа, продление триала, заморозка."""
    from datetime import datetime
    conn = get_db()
    cur = conn.cursor()
    sets, vals = [], []
    fields_map = [
        ('plan','plan'),('trialUntil','trial_until'),('planExpiresAt','plan_expires_at'),
        ('monthlyFee','monthly_fee'),('paymentStatus','payment_status'),
        ('maxProjects','max_projects'),('maxUsers','max_users'),
        ('contactName','contact_name'),('contactPhone','contact_phone'),
        ('contactEmail','contact_email'),('notes','notes'),('active','active'),
    ]
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s"); vals.append(data[js_key])
    if data.get('action') == 'suspend':
        sets.append("suspended_at=%s"); vals.append(datetime.now())
        sets.append("suspended_reason=%s"); vals.append(data.get('reason') or '')
        sets.append("active=%s"); vals.append(False)
    elif data.get('action') == 'resume':
        sets.append("suspended_at=NULL")
        sets.append("active=%s"); vals.append(True)
    if not sets:
        conn.close()
        return {"ok": False, "error": "no fields"}
    vals.append(id)
    cur.execute("UPDATE companies SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.close()
    return {"ok": True}

@app.get("/system/dashboard")
def system_dashboard(_current_user: dict = Depends(require_roles("system_owner"))):
    """Сводка для главной страницы кабинета системы."""
    from datetime import datetime, date
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT COUNT(*) as c FROM companies WHERE active=TRUE")
    active = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM companies WHERE plan='demo' AND active=TRUE")
    in_demo = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM companies WHERE suspended_at IS NOT NULL")
    suspended = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM companies WHERE payment_status='overdue'")
    overdue = cur.fetchone()['c']
    today = date.today()
    cur.execute("SELECT COALESCE(SUM(amount),0) as t FROM company_payments WHERE status='paid' AND date_trunc('month', payment_date)=date_trunc('month', %s::date)", (today,))
    month_revenue = float(cur.fetchone()['t'] or 0)
    cur.execute("SELECT COALESCE(SUM(amount),0) as t FROM company_payments WHERE status='paid' AND date_trunc('year', payment_date)=date_trunc('year', %s::date)", (today,))
    year_revenue = float(cur.fetchone()['t'] or 0)
    cur.execute("SELECT COUNT(*) as c FROM demo_requests WHERE status='Новая'")
    new_demos = cur.fetchone()['c']
    conn.close()
    return {
        "activeCompanies": active, "inDemo": in_demo, "suspended": suspended,
        "overdue": overdue, "monthRevenue": month_revenue, "yearRevenue": year_revenue,
        "newDemoRequests": new_demos
    }

@app.get("/system/payments")
def system_payments_list(_current_user: dict = Depends(require_roles("system_owner"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT p.*, c.name as company_name FROM company_payments p
                   LEFT JOIN companies c ON c.id=p.company_id
                   ORDER BY p.payment_date DESC LIMIT 200""")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.post("/system/payments")
def system_create_payment(data: dict, _current_user: dict = Depends(require_roles("system_owner"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""INSERT INTO company_payments (company_id, amount, payment_date, method,
                                                  invoice_number, status, period_start, period_end,
                                                  notes, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (data.get('companyId'), float(data.get('amount') or 0), data.get('paymentDate') or None,
         data.get('method'), data.get('invoiceNumber'), data.get('status') or 'paid',
         data.get('periodStart') or None, data.get('periodEnd') or None,
         data.get('notes'), data.get('createdBy')))
    new_id = cur.fetchone()['id']
    # Обновляем компанию: продлеваем plan_expires_at
    if data.get('periodEnd') and data.get('companyId'):
        cur.execute("UPDATE companies SET plan_expires_at=%s, payment_status='active' WHERE id=%s",
            (data['periodEnd'], data['companyId']))
    conn.close()
    return {"id": new_id, "ok": True}

@app.get("/demo-requests")
def list_demo_requests(_current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "system_owner"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM demo_requests ORDER BY created_at DESC LIMIT 200")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

@app.post("/demo-request")
def create_demo_request(data: dict):
    """Публичный endpoint — приходит с лендинга / формы на сайте."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""INSERT INTO demo_requests (company_name, contact_name, phone, email,
                                                employees_count, projects_count, comment, source)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (data.get('companyName'), data.get('contactName'), data.get('phone'),
         data.get('email'), data.get('employeesCount'), data.get('projectsCount'),
         data.get('comment'), data.get('source') or 'landing'))
    new_id = cur.fetchone()['id']
    conn.close()
    return {"id": new_id, "ok": True, "message": "Заявка принята, с вами свяжутся в течение рабочего дня"}

@app.put("/demo-requests/{id}")
def update_demo_request(id: int, data: dict, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "system_owner"))):
    from datetime import datetime
    conn = get_db()
    cur = conn.cursor()
    sets, vals = [], []
    for k, c in [('status','status'),('notes','notes'),('assignedCompanyId','assigned_company_id')]:
        if k in data:
            sets.append(c + "=%s"); vals.append(data[k])
    if data.get('status') in ('Обработана','Отклонена'):
        sets.append("processed_at=%s"); vals.append(datetime.now())
    if not sets:
        conn.close()
        return {"ok": False}
    vals.append(id)
    cur.execute("UPDATE demo_requests SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.close()
    return {"ok": True}

@app.get("/invite-codes/{code}/info")
def invite_code_info(code: str):
    """Возвращает данные приглашения для подсветки формы регистрации."""
    from datetime import datetime
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM invite_codes WHERE code=%s", (code.upper().strip(),))
    row = cur.fetchone()
    conn.close()
    if not row:
        return {"valid": False, "error": "Код не найден"}
    if row.get('used'):
        return {"valid": False, "error": "Код уже использован"}
    if row.get('expires_at') and row['expires_at'] < datetime.now():
        return {"valid": False, "error": "Срок действия ссылки истёк"}
    return {
        "valid": True,
        "role": row['role'],
        "presetName": row.get('preset_name') or '',
        "presetCategory": row.get('preset_category') or '',
        "supplierId": row.get('supplier_id'),
    }

@app.get("/suppliers")
def get_suppliers(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if supplier_id:
            cur.execute("SELECT * FROM suppliers WHERE id=%s ORDER BY name", (supplier_id,))
        else:
            cur.close(); conn.close()
            return []
    elif role in SUPPLY_ROLES or role in WAREHOUSE_ROLES or role in FINANCE_ROLES:
        cur.execute("SELECT * FROM suppliers ORDER BY name")
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/suppliers")
def create_supplier(s: SupplierModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO suppliers (name,phone,email,specialization,category,rating,status) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (s.name,s.phone,s.email,s.specialization,s.category,s.rating,s.status))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/suppliers/{id}")
def update_supplier(id: int, s: SupplierModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE suppliers SET name=%s,phone=%s,email=%s,specialization=%s,category=%s,rating=%s,status=%s WHERE id=%s",
                (s.name,s.phone,s.email,s.specialization,s.category,s.rating,s.status,id))
    conn.close()
    return {"ok": True}

@app.delete("/suppliers/{id}")
def delete_supplier(id: int, _current_user: dict = Depends(require_roles("директор", "зам_директора", "снабженец", "кладовщик"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM suppliers WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

SUPPLY_SELECT = ("SELECT id,material_name as \"materialName\",quantity,unit,project,"
                 "created_by as \"createdBy\",date,status,notes,"
                 "selected_suppliers as \"selectedSuppliers\","
                 "requested_by_role as \"requestedByRole\","
                 "requested_by_id as \"requestedById\","
                 "urgency,category,"
                 "prorab_id as \"prorabId\",prorab_name as \"prorabName\","
                 "prorab_confirmed_at as \"prorabConfirmedAt\","
                 "director_id as \"directorId\",director_name as \"directorName\","
                 "director_approved_at as \"directorApprovedAt\","
                 "reject_reason as \"rejectReason\","
                 "items_json as \"itemsJson\","
                 "created_at as \"createdAt\" "
                 "FROM supply_requests")

@app.get("/supply-requests")
def get_supply_requests(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if role in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.execute(SUPPLY_SELECT + " ORDER BY id DESC")
    elif role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if not supplier_id:
            cur.close(); conn.close()
            return []
        cur.execute(SUPPLY_SELECT + " WHERE selected_suppliers::text LIKE %s ORDER BY id DESC", ('%"'+str(supplier_id)+'"%',))
    else:
        projects = user_project_names(current_user)
        clauses = ["requested_by_id=%s", "created_by=%s"]
        params = [current_user.get("id"), current_user.get("name") or ""]
        if projects:
            clauses.append("project = ANY(%s)")
            params.append(projects)
        cur.execute(SUPPLY_SELECT + " WHERE (" + " OR ".join(clauses) + ") ORDER BY id DESC", params)
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/supply-requests")
def create_supply_request(r: SupplyRequestModel, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    import json as _json
    from datetime import datetime
    role = (r.requestedByRole or "").strip()
    if role in ("директор", "зам_директора"):
        initial_status = "Утверждена"
    elif role == "прораб":
        initial_status = "Подтверждена прорабом"
    else:
        initial_status = "Новая"
    now = datetime.now()
    prorab_id = r.requestedById if role == "прораб" else None
    prorab_name = r.createdBy if role == "прораб" else None
    prorab_at = now if role == "прораб" else None
    director_id = r.requestedById if role in ("директор", "зам_директора") else None
    director_name = r.createdBy if role in ("директор", "зам_директора") else None
    director_at = now if role in ("директор", "зам_директора") else None
    # Нормализация items: если items пустой но есть materialName — упаковываем как single-item
    items = [it for it in (r.items or []) if (it or {}).get("materialName") and float((it or {}).get("quantity") or 0) > 0]
    if not items and r.materialName:
        items = [{"materialName": r.materialName, "quantity": r.quantity, "unit": r.unit}]
    if not items:
        raise HTTPException(status_code=400, detail="Заявка должна содержать хотя бы одну позицию")
    # material_name / quantity / unit заполняем агрегатом для совместимости со старым UI
    # Если позиций больше одной — пишем «N позиций» в material_name; quantity = сумма всех
    if len(items) == 1:
        agg_name = items[0]["materialName"]
        agg_qty = float(items[0].get("quantity") or 0)
        agg_unit = items[0].get("unit") or r.unit
    else:
        agg_name = items[0]["materialName"] + " и ещё " + str(len(items)-1) + " поз."
        agg_qty = float(len(items))  # количество позиций
        agg_unit = "поз."
    items_json = _json.dumps(items, ensure_ascii=False)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "INSERT INTO supply_requests "
        "(material_name,quantity,unit,project,created_by,date,notes,selected_suppliers,"
        "status,requested_by_role,requested_by_id,urgency,category,"
        "prorab_id,prorab_name,prorab_confirmed_at,"
        "director_id,director_name,director_approved_at,items_json) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (agg_name, agg_qty, agg_unit, r.project, r.createdBy, r.date, r.notes,
         r.selectedSuppliers, initial_status, role, r.requestedById, r.urgency, r.category,
         prorab_id, prorab_name, prorab_at,
         director_id, director_name, director_at, items_json))
    new_id = cur.fetchone()['id']
    cur.execute(SUPPLY_SELECT + " WHERE id=%s", (new_id,))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/supply-requests/{id}")
def update_supply_request(id: int, data: dict, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    from datetime import datetime
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    action = data.get('action')
    now = datetime.now()
    if action == 'confirm_prorab':
        cur.execute(
            "UPDATE supply_requests SET status=%s, prorab_id=%s, prorab_name=%s, prorab_confirmed_at=%s WHERE id=%s",
            ('Подтверждена прорабом', data.get('userId'), data.get('userName'), now, id))
    elif action == 'approve_director':
        cur.execute(
            "UPDATE supply_requests SET status=%s, director_id=%s, director_name=%s, director_approved_at=%s WHERE id=%s",
            ('Утверждена', data.get('userId'), data.get('userName'), now, id))
    elif action == 'reject':
        cur.execute(
            "UPDATE supply_requests SET status=%s, reject_reason=%s WHERE id=%s",
            ('Отклонена', data.get('rejectReason') or data.get('reason') or '', id))
    elif action == 'cancel':
        cur.execute("UPDATE supply_requests SET status=%s WHERE id=%s", ('Отменена', id))
    elif 'status' in data:
        cur.execute("UPDATE supply_requests SET status=%s WHERE id=%s", (data['status'], id))
    cur.execute(SUPPLY_SELECT + " WHERE id=%s", (id,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else {"ok": True}

@app.delete("/supply-requests/{id}")
def delete_supply_request(id: int, rollback_received: bool = False, _current_user: dict = Depends(require_roles("директор", "зам_директора", "снабженец", "кладовщик", "прораб"))):
    from datetime import datetime
    conn = get_db()
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("SELECT id, project, material_name FROM supply_requests WHERE id=%s FOR UPDATE", (id,))
        req = cur.fetchone()
        if not req:
            conn.rollback()
            raise HTTPException(status_code=404, detail="Заявка не найдена")
        if req.get("project"):
            require_project_access(_current_user, req.get("project"))

        cur.execute("SELECT * FROM supply_deliveries WHERE request_id=%s FOR UPDATE", (id,))
        deliveries = cur.fetchall()
        received_deliveries = [d for d in deliveries if float(d.get("received_quantity") or 0) > 0]
        if received_deliveries and not rollback_received:
            conn.rollback()
            raise HTTPException(status_code=400, detail="По заявке уже есть принятая поставка. Для удаления с откатом склада используйте rollback_received=true.")

        restored = 0
        if rollback_received:
            for d in received_deliveries:
                qty = float(d.get("received_quantity") or 0)
                name = d.get("material_name") or ""
                project = d.get("project") or ""
                unit = d.get("unit") or "шт"
                if not name or not project or qty <= 0:
                    continue
                cur.execute("SELECT id, quantity FROM materials WHERE name=%s AND project=%s FOR UPDATE", (name, project))
                mat = cur.fetchone()
                if not mat:
                    conn.rollback()
                    raise HTTPException(status_code=400, detail="Нельзя откатить поставку: материал «"+name+"» не найден на складе объекта.")
                stock_qty = float(mat.get("quantity") or 0)
                if stock_qty + 0.000001 < qty:
                    conn.rollback()
                    raise HTTPException(status_code=400, detail="Нельзя откатить поставку «"+name+"»: на складе осталось "+str(stock_qty)+" "+unit+", а принять было "+str(qty)+" "+unit+". Материал уже списывали.")
                remaining_qty = stock_qty - qty
                if remaining_qty <= 0.000001:
                    cur.execute("DELETE FROM materials WHERE id=%s", (mat.get("id"),))
                else:
                    cur.execute("UPDATE materials SET quantity=%s WHERE id=%s", (remaining_qty, mat.get("id")))
                cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                            (name, "откат поставки (удаление заявки)", qty, datetime.now().date().isoformat(), project, _current_user.get("name") or "", datetime.now().strftime("%d.%m.%Y, %H:%M")))
                restored += 1

        delivery_ids = [d.get("id") for d in deliveries if d.get("id")]
        if delivery_ids:
            cur.execute("DELETE FROM material_inspection_journal WHERE delivery_id = ANY(%s)", (delivery_ids,))
            cur.execute("DELETE FROM cable_journal WHERE delivery_id = ANY(%s)", (delivery_ids,))
            cur.execute("DELETE FROM supply_history WHERE delivery_id = ANY(%s)", (delivery_ids,))
            cur.execute("DELETE FROM supply_claims WHERE delivery_id = ANY(%s)", (delivery_ids,))
        cur.execute("DELETE FROM supply_claims WHERE request_id=%s", (id,))
        cur.execute("DELETE FROM supplier_invoices WHERE request_id=%s", (id,))
        cur.execute("DELETE FROM supply_deliveries WHERE request_id=%s", (id,))
        cur.execute("DELETE FROM supplier_offers WHERE request_id=%s", (id,))
        cur.execute("DELETE FROM supply_requests WHERE id=%s", (id,))
        conn.commit()
        return {"ok": True, "deliveriesDeleted": len(deliveries), "materialsRolledBack": restored}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.get("/supply-requests/{id}/stock-check")
def supply_request_stock_check(id: int, current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    """Возвращает остатки похожих материалов на складе и бюджет проекта.
       Используется UI чтобы показать «есть X из Y, закупить Z»."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT material_name, quantity, unit, project FROM supply_requests WHERE id=%s", (id,))
        req = cur.fetchone()
        if not req:
            return {"error": "Заявка не найдена"}
        name = (req['material_name'] or '').strip()
        project = (req['project'] or '').strip()
        if project:
            require_project_access(current_user, project)
        needed = float(req['quantity'] or 0)
        # Поиск похожих материалов на основном складе (ILIKE по части имени)
        stock_matches = []
        total_available = 0.0
        if name:
            # Берём первое значащее слово (>=4 символа) для гибкого поиска
            tokens = [t for t in name.split() if len(t) >= 4]
            search = tokens[0] if tokens else name
            cur.execute(
                "SELECT id, name, quantity, unit, price, category FROM warehouse_main WHERE name ILIKE %s ORDER BY quantity DESC LIMIT 10",
                ('%' + search + '%',))
            for row in cur.fetchall():
                stock_matches.append({
                    "id": row['id'], "name": row['name'],
                    "quantity": float(row['quantity'] or 0),
                    "unit": row['unit'], "price": float(row['price'] or 0),
                    "category": row['category']
                })
                total_available += float(row['quantity'] or 0)
            # Также ищем на складе объекта (materials)
            if project:
                try:
                    cur.execute(
                        "SELECT id, name, quantity, unit, price, category FROM materials WHERE project=%s AND name ILIKE %s LIMIT 10",
                        (project, '%' + search + '%'))
                    for row in cur.fetchall():
                        stock_matches.append({
                            "id": row['id'], "name": '[На объекте] ' + (row['name'] or ''),
                            "quantity": float(row['quantity'] or 0),
                            "unit": row['unit'], "price": float(row['price'] or 0),
                            "category": row['category']
                        })
                        total_available += float(row['quantity'] or 0)
                except Exception:
                    pass
        # Бюджет проекта
        project_budget = 0.0
        project_approved_cost = 0.0
        project_pending_cost = 0.0
        if project:
            try:
                cur.execute("SELECT budget FROM projects WHERE name=%s", (project,))
                pr = cur.fetchone()
                project_budget = float(pr['budget'] or 0) if pr else 0.0
            except Exception:
                project_budget = 0.0
            # Считаем оценочную стоимость уже утверждённых/ожидающих заявок
            try:
                cur.execute(
                    "SELECT material_name, quantity, status FROM supply_requests WHERE project=%s",
                    (project,))
                requests_rows = cur.fetchall()
                for row in requests_rows:
                    mname = (row['material_name'] or '').strip()
                    if not mname:
                        continue
                    first_word = mname.split()[0]
                    try:
                        cur.execute(
                            "SELECT AVG(price) AS avg_price FROM warehouse_main WHERE name ILIKE %s",
                            ('%' + first_word + '%',))
                        avg_row = cur.fetchone()
                        avg_price = float(avg_row['avg_price'] or 0) if avg_row else 0.0
                    except Exception:
                        avg_price = 0.0
                    est_cost = float(row['quantity'] or 0) * avg_price
                    if row['status'] == 'Утверждена':
                        project_approved_cost += est_cost
                    elif row['status'] in ('Новая', 'Подтверждена прорабом'):
                        project_pending_cost += est_cost
            except Exception:
                pass
        shortage = max(0.0, needed - total_available)
        budget_risk = 0.0
        if project_budget > 0:
            budget_risk = (project_approved_cost + project_pending_cost) / project_budget * 100
        return {
            "requestId": id,
            "materialName": name,
            "needed": needed,
            "unit": req['unit'],
            "totalAvailable": total_available,
            "shortage": shortage,
            "stockMatches": stock_matches,
            "project": project,
            "projectBudget": project_budget,
            "projectApprovedCost": project_approved_cost,
            "projectPendingCost": project_pending_cost,
            "budgetRiskPercent": budget_risk,
        }
    except Exception as e:
        import traceback
        print("STOCK-CHECK ERROR id=" + str(id) + ": " + str(e))
        print(traceback.format_exc())
        return {"error": "Ошибка сервера: " + str(e)[:200]}
    finally:
        if conn:
            try: conn.close()
            except Exception: pass

OFFERS_SELECT = ("SELECT id, request_id as \"requestId\", supplier_id as \"supplierId\","
                 "price_per_unit as \"pricePerUnit\", total_price as \"totalPrice\","
                 "delivery_days as \"deliveryDays\", notes, status,"
                 "delivery_status as \"deliveryStatus\","
                 "payment_terms as \"paymentTerms\", vat_included as \"vatIncluded\","
                 "pdf_url as \"pdfUrl\", valid_until as \"validUntil\","
                 "supplier_message as \"supplierMessage\","
                 "requested_at as \"requestedAt\", responded_at as \"respondedAt\","
                 "ai_recommended as \"aiRecommended\","
                 "items_kp_json as \"itemsKpJson\" "
                 "FROM supplier_offers")

@app.get("/supplier-offers")
def get_supplier_offers(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if role in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.execute(OFFERS_SELECT + " ORDER BY id DESC")
    elif role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if not supplier_id:
            cur.close(); conn.close()
            return []
        cur.execute(OFFERS_SELECT + " WHERE supplier_id=%s ORDER BY id DESC", (supplier_id,))
    else:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute(OFFERS_SELECT + " WHERE request_id IN (SELECT id FROM supply_requests WHERE project = ANY(%s)) ORDER BY id DESC", (projects,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/supplier-offers")
def create_supplier_offer(o: SupplierOfferModel, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO supplier_offers (request_id,supplier_id,price_per_unit,total_price,delivery_days,notes) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
                (o.requestId,o.supplierId,o.pricePerUnit,o.totalPrice,o.deliveryDays,o.notes))
    new_id = cur.fetchone()['id']
    cur.execute(OFFERS_SELECT + " WHERE id=%s", (new_id,))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/supplier-offers/{id}")
def update_supplier_offer(id: int, data: dict, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    from datetime import datetime
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    action = data.get('action')
    if action == 'respond':
        # Поставщик отвечает на КП: цена, срок, условия, НДС, PDF, комментарий
        import json as _json
        items_kp = data.get('itemsKp') or []
        # Если пришёл массив постатейного КП — считаем итог автоматически
        items_kp_json = None
        if items_kp and isinstance(items_kp, list):
            # Нормализация: каждый item должен иметь pricePerUnit и quantity
            normalized = []
            calc_total = 0.0
            for it in items_kp:
                if not isinstance(it, dict): continue
                p = float(it.get('pricePerUnit') or 0)
                q = float(it.get('quantity') or 0)
                line_total = p * q
                normalized.append({
                    'materialName': it.get('materialName',''),
                    'quantity': q,
                    'unit': it.get('unit','шт'),
                    'pricePerUnit': p,
                    'totalPrice': line_total,
                    'deliveryDays': int(it.get('deliveryDays') or 0) if it.get('deliveryDays') else None,
                    'notes': it.get('notes','')
                })
                calc_total += line_total
            items_kp_json = _json.dumps(normalized, ensure_ascii=False)
            # Для совместимости: pricePerUnit = средневзвешенная, totalPrice = сумма
            total = float(data.get('totalPrice') or calc_total)
            ppu = float(data.get('pricePerUnit') or (calc_total / max(1, len(normalized))))
        else:
            # Старый путь — одна цена за единицу
            ppu = float(data.get('pricePerUnit') or 0)
            qty_for_total = float(data.get('quantity') or 0)
            total = float(data.get('totalPrice') or (ppu * qty_for_total))
        cur.execute(
            "UPDATE supplier_offers SET status=%s, price_per_unit=%s, total_price=%s, delivery_days=%s, "
            "payment_terms=%s, vat_included=%s, pdf_url=%s, valid_until=%s, supplier_message=%s, "
            "items_kp_json=COALESCE(%s, items_kp_json), responded_at=%s WHERE id=%s",
            ('Получено', ppu, total, int(data.get('deliveryDays') or 0),
             data.get('paymentTerms') or 'Постоплата',
             bool(data.get('vatIncluded', True)),
             data.get('pdfUrl') or None,
             data.get('validUntil') or None,
             data.get('supplierMessage') or '',
             items_kp_json,
             datetime.now(), id))
    elif action == 'select':
        # Директор выбрал это КП
        cur.execute("UPDATE supplier_offers SET status=%s WHERE id=%s", ('Утверждено', id))
        # Остальные КП по этой заявке — отклонены
        cur.execute("SELECT request_id FROM supplier_offers WHERE id=%s", (id,))
        r = cur.fetchone()
        if r and r['request_id']:
            cur.execute("UPDATE supplier_offers SET status=%s WHERE request_id=%s AND id<>%s AND status<>%s",
                ('Отклонено', r['request_id'], id, 'Отклонено'))
    elif action == 'reject':
        cur.execute("UPDATE supplier_offers SET status=%s WHERE id=%s", ('Отклонено', id))
    else:
        if 'status' in data:
            cur.execute("UPDATE supplier_offers SET status=%s WHERE id=%s", (data['status'], id))
        if 'deliveryStatus' in data:
            cur.execute("UPDATE supplier_offers SET delivery_status=%s WHERE id=%s", (data['deliveryStatus'], id))
    cur.execute(OFFERS_SELECT + " WHERE id=%s", (id,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else {"ok": True}

@app.post("/supply-requests/{id}/request-kp")
def request_kp_from_suppliers(id: int, data: dict, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    """Директор отправляет запрос КП нескольким поставщикам.
       data: {supplierIds: [1,2,3], aiRecommendedIds: [1,2]}"""
    supplier_ids = data.get('supplierIds') or []
    ai_ids = set(data.get('aiRecommendedIds') or [])
    if not supplier_ids:
        return {"error": "Не выбраны поставщики"}
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    # Получаем количество из заявки для preview total
    cur.execute("SELECT quantity FROM supply_requests WHERE id=%s", (id,))
    req = cur.fetchone()
    qty = float(req['quantity']) if req else 0
    created = []
    for sid in supplier_ids:
        # Не создаём дубль если уже есть offer
        cur.execute("SELECT id FROM supplier_offers WHERE request_id=%s AND supplier_id=%s", (id, sid))
        if cur.fetchone():
            continue
        cur.execute(
            "INSERT INTO supplier_offers (request_id, supplier_id, status, ai_recommended, requested_at) "
            "VALUES (%s, %s, %s, %s, NOW()) RETURNING id",
            (id, sid, 'Ожидает ответа', sid in ai_ids))
        created.append(cur.fetchone()['id'])
    # Обновляем статус заявки
    cur.execute("UPDATE supply_requests SET status=%s WHERE id=%s AND status='Утверждена'",
        ('КП запрошены', id))
    conn.close()
    return {"ok": True, "created": len(created), "ids": created}

@app.get("/supply-requests/{id}/suggest-suppliers")
def suggest_suppliers_for_request(id: int, current_user: dict = Depends(require_roles(*SUPPLY_INTERNAL_ROLES))):
    """Возвращает список поставщиков подходящих под категорию материала, ранжированных
       по рейтингу + истории успешных поставок. AI-комментарий опциональный."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT material_name, category, project FROM supply_requests WHERE id=%s", (id,))
    req = cur.fetchone()
    if not req:
        conn.close()
        return {"error": "not found"}
    if req.get('project'):
        require_project_access(current_user, req.get('project') or "")
    category = (req['category'] or '').strip()
    material_name = (req['material_name'] or '').lower()
    # Фильтр по категории если задана. Иначе берём всех активных
    if category:
        cur.execute(
            "SELECT id, name, category, specialization, rating, phone, email, status FROM suppliers "
            "WHERE (status IS NULL OR status='Активный' OR status='') AND category=%s ORDER BY rating DESC NULLS LAST",
            (category,))
    else:
        cur.execute(
            "SELECT id, name, category, specialization, rating, phone, email, status FROM suppliers "
            "WHERE (status IS NULL OR status='Активный' OR status='') ORDER BY rating DESC NULLS LAST")
    suppliers = [dict(r) for r in cur.fetchall()]
    # Считаем историю успешных поставок (status='Доставлено') по каждому поставщику
    cur.execute("SELECT supplier_id, COUNT(*) as deliveries FROM supply_history WHERE status='Доставлено' GROUP BY supplier_id")
    deliveries_map = {r['supplier_id']: r['deliveries'] for r in cur.fetchall()}
    # Считаем уже отправленные КП этой заявке
    cur.execute("SELECT supplier_id FROM supplier_offers WHERE request_id=%s", (id,))
    already_requested = {r['supplier_id'] for r in cur.fetchall()}
    conn.close()
    for s in suppliers:
        s['deliveriesCount'] = int(deliveries_map.get(s['id'], 0))
        s['alreadyRequested'] = s['id'] in already_requested
        # Простая логика: AI рекомендует если category совпадает + (rating>=4 OR deliveries>=1)
        score = 0
        if s['rating']: score += float(s['rating'])
        if s['deliveriesCount']: score += min(s['deliveriesCount'], 5)
        if material_name and (s.get('specialization') or '').lower() and any(w in material_name for w in (s['specialization'] or '').lower().split()):
            score += 2
        s['_score'] = score
        s['aiRecommend'] = score >= 5  # порог рекомендации
    suppliers.sort(key=lambda x: -x['_score'])
    return {
        "requestId": id,
        "materialName": req['material_name'],
        "category": category,
        "suppliers": suppliers[:15],  # топ 15
        "aiRecommendedCount": sum(1 for s in suppliers if s.get('aiRecommend')),
    }

@app.get("/supply-requests/{id}/compare-kp")
def compare_kp_for_request(id: int, current_user: dict = Depends(require_roles(*SUPPLY_INTERNAL_ROLES))):
    """Сравнивает все полученные КП по заявке и возвращает AI-вердикт.
       Учитывает цену, срок поставки, условия оплаты, НДС, рейтинг поставщика."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT material_name, quantity, unit, project FROM supply_requests WHERE id=%s", (id,))
    req = cur.fetchone()
    if not req:
        conn.close()
        return {"error": "Заявка не найдена"}
    if req.get('project'):
        require_project_access(current_user, req.get('project') or "")
    cur.execute(
        "SELECT o.id, o.supplier_id, o.price_per_unit, o.total_price, o.delivery_days, "
        "o.payment_terms, o.vat_included, o.valid_until, o.supplier_message, o.status, "
        "s.name as supplier_name, s.rating "
        "FROM supplier_offers o LEFT JOIN suppliers s ON s.id=o.supplier_id "
        "WHERE o.request_id=%s AND o.status IN ('Получено','Утверждено')",
        (id,))
    offers = [dict(r) for r in cur.fetchall()]
    conn.close()
    if len(offers) < 2:
        return {
            "error": "Нужно минимум 2 полученных КП для сравнения",
            "offersCount": len(offers),
        }
    # Готовим короткую сводку для AI
    offers_summary = []
    for o in offers:
        offers_summary.append({
            "offerId": o['id'],
            "supplier": o['supplier_name'] or 'Поставщик #'+str(o['supplier_id']),
            "rating": float(o['rating'] or 0),
            "pricePerUnit": float(o['price_per_unit'] or 0),
            "totalPrice": float(o['total_price'] or 0),
            "deliveryDays": int(o['delivery_days'] or 0),
            "paymentTerms": o['payment_terms'] or '',
            "vatIncluded": bool(o['vat_included']),
            "validUntil": str(o['valid_until']) if o['valid_until'] else None,
            "message": o['supplier_message'] or '',
        })
    # Считаем простой score:
    # - ниже цена = лучше
    # - меньше срок = лучше
    # - предоплата = чуть хуже (риск)
    # - выше рейтинг = лучше
    prices = [o['pricePerUnit'] for o in offers_summary if o['pricePerUnit']>0] or [1]
    days = [o['deliveryDays'] for o in offers_summary if o['deliveryDays']>0] or [1]
    min_price = min(prices); max_price = max(prices)
    min_days = min(days); max_days = max(days)
    def _terms_risk(t):
        t = (t or '').lower()
        if 'предоплат' in t and '100' in t: return 0.0  # риск для покупателя
        if '50' in t: return 0.5
        if 'постоплат' in t or 'отсрочк' in t: return 1.0
        return 0.5
    for o in offers_summary:
        sc_price = 1.0 - (o['pricePerUnit']-min_price)/(max_price-min_price) if max_price>min_price else 1.0
        sc_days = 1.0 - (o['deliveryDays']-min_days)/(max_days-min_days) if max_days>min_days else 1.0
        sc_terms = _terms_risk(o['paymentTerms'])
        sc_rating = o['rating']/5.0 if o['rating'] else 0.5
        # веса: цена 40%, срок 20%, условия 20%, рейтинг 20%
        o['score'] = round((sc_price*0.4 + sc_days*0.2 + sc_terms*0.2 + sc_rating*0.2) * 100, 1)
    offers_summary.sort(key=lambda x: -x['score'])
    best = offers_summary[0]
    # AI вердикт текстом
    ai_text = None
    try:
        import openai as oa
        prompt = (
            f"Сравни {len(offers_summary)} коммерческих предложений на материал "
            f"«{req['material_name']}» (нужно {req['quantity']} {req['unit']}).\n"
            f"Предложения:\n"
        )
        for o in offers_summary:
            prompt += (
                f"- {o['supplier']} (рейтинг {o['rating']}/5): "
                f"{o['pricePerUnit']} ₽/ед, итого {o['totalPrice']} ₽, "
                f"срок {o['deliveryDays']} дн, оплата «{o['paymentTerms']}», "
                f"{'с НДС' if o['vatIncluded'] else 'без НДС'}"
                + (f", «{o['message']}»" if o['message'] else "")
                + "\n"
            )
        prompt += (
            "\nДай короткий вывод (2-3 предложения по-русски): кого выбрать и почему. "
            "Учитывай не только цену но и срок, условия оплаты, рейтинг поставщика. "
            "Не используй markdown, не используй ```."
        )
        client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
        r = client.responses.create(
            model="gpt://"+YANDEX_FOLDER_ID+"/yandexgpt-5.1/latest",
            temperature=0.2,
            instructions="Ты помощник директора строительной компании. Сравниваешь коммерческие предложения и даёшь короткий вывод.",
            input=prompt,
            max_output_tokens=400,
        )
        ai_text = (r.output_text or '').strip()
    except Exception as e:
        print("compare-kp AI error:", e)
        ai_text = None
    return {
        "requestId": id,
        "materialName": req['material_name'],
        "quantity": float(req['quantity'] or 0),
        "unit": req['unit'],
        "offersCount": len(offers_summary),
        "bestOfferId": best['offerId'],
        "bestSupplier": best['supplier'],
        "ranking": offers_summary,
        "aiText": ai_text,
    }

@app.post("/supplier-offers/{id}/create-invoice")
def create_invoice_from_offer(id: int, data: dict, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    """Поставщик выставляет счёт по выигранному КП.
       Автоматически создаёт supplier_invoice."""
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT o.id, o.supplier_id, o.request_id, o.total_price, o.payment_terms, o.vat_included, "
        "s.name as supplier_name, r.project as project_name, r.material_name "
        "FROM supplier_offers o "
        "LEFT JOIN suppliers s ON s.id=o.supplier_id "
        "LEFT JOIN supply_requests r ON r.id=o.request_id "
        "WHERE o.id=%s AND o.status=%s",
        (id, 'Утверждено'))
    offer = cur.fetchone()
    if not offer:
        conn.close()
        return {"error": "Утверждённое КП не найдено"}
    invoice_number = data.get('invoiceNumber') or ''
    invoice_date = data.get('invoiceDate') or None
    amount = float(data.get('amount') or offer['total_price'] or 0)
    vat_amount = float(data.get('vatAmount') or 0)
    file_url = data.get('fileUrl') or data.get('photoUrl') or ''
    description = data.get('description') or ('Материал: '+(offer['material_name'] or ''))
    cur.execute(
        "INSERT INTO supplier_invoices "
        "(supplier_id, supplier_name, project_name, invoice_number, invoice_date, "
        "amount, vat_amount, description, file_url, status, offer_id, request_id, "
        "payment_terms, material_name) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (offer['supplier_id'], offer['supplier_name'], offer['project_name'],
         invoice_number, invoice_date, amount, vat_amount, description, file_url,
         'На утверждении', offer['id'], offer['request_id'],
         offer['payment_terms'], offer['material_name']))
    new_id = cur.fetchone()['id']
    conn.close()
    return {"ok": True, "id": new_id}

DELIVERY_SELECT = """
    SELECT d.id, d.offer_id as "offerId", d.request_id as "requestId",
           d.supplier_id as "supplierId", d.supplier_name as "supplierName",
           d.project, d.material_name as "materialName",
           d.planned_quantity as "plannedQuantity",
           d.shipped_quantity as "shippedQuantity",
           d.received_quantity as "receivedQuantity",
           d.unit, d.price_per_unit as "pricePerUnit",
           d.total_price as "totalPrice", d.status,
           d.waybill_number as "waybillNumber",
           d.waybill_date as "waybillDate",
           d.vehicle_number as "vehicleNumber",
           d.driver_name as "driverName",
           d.document_url as "documentUrl", d.photo_url as "photoUrl",
           d.shipped_at as "shippedAt", d.received_at as "receivedAt",
           d.received_by as "receivedBy",
           d.quality_status as "qualityStatus",
           d.quality_notes as "qualityNotes",
           d.shortage_quantity as "shortageQuantity",
           d.ai_check_result as "aiCheckResult",
           d.claim_id as "claimId", d.created_at as "createdAt"
    FROM supply_deliveries d
"""

CLAIM_SELECT = """
    SELECT id, delivery_id as "deliveryId", request_id as "requestId",
           offer_id as "offerId", supplier_id as "supplierId",
           project, material_name as "materialName",
           claim_type as "claimType", description,
           expected_quantity as "expectedQuantity",
           received_quantity as "receivedQuantity",
           shortage_quantity as "shortageQuantity",
           photo_url as "photoUrl", status, created_by as "createdBy",
           created_at as "createdAt", resolved_at as "resolvedAt", resolution
    FROM supply_claims
"""

def _float_or_zero(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0

def _detect_cable_info(name):
    import re as _re
    raw = (name or "").strip()
    up = raw.upper().replace("Ё", "Е")
    compact = _re.sub(r"[\s\"'«»()\\/._-]", "", up)
    start_prefixes = [
        "ВВГ", "АВВГ", "ВББШВ", "ПВВ", "ПВС", "СИП", "КВВГ", "КГ",
        "ТППЭП", "ТПВ", "КВПЭФ", "NYM", "NYY", "ПУНП",
        "UTP", "FTP", "SFTP", "FUTP", "UUTP", "SFUTP",
        "КПС", "КПСЭ", "КПСВВ", "КСВВ", "КСПВ", "КСПЭВ", "ППГ",
        "JYSTY", "JEH", "JHSTH"
    ]
    indicators = start_prefixes + ["КАБЕЛ", "ПРОВОД", "FRLS", "FRHF", "ОПС"]
    is_cable = any(compact.startswith(p) for p in start_prefixes) or any(p in compact for p in indicators)
    cable_type = ""
    if any(p in compact for p in ["КПС", "КПСЭ", "КПСВВ", "ОПС", "ПОЖАР"]):
        cable_type = "Пожарная сигнализация"
    elif any(p in compact for p in ["UTP", "FTP", "SFTP", "FUTP", "UUTP", "SFUTP"]):
        cable_type = "СКС / интернет"
    elif any(p in compact for p in ["КСВВ", "КСПВ", "КСПЭВ", "КВПЭФ", "ТППЭП", "ТПВ", "JYSTY", "JEH", "JHSTH"]):
        cable_type = "Слаботочка / сигнализация"
    elif is_cable:
        cable_type = "Силовой кабель"
    cores = None
    section = None
    m3 = _re.search(r"(\d+)\s*[х×x*]\s*(\d+)\s*[х×x*]\s*(\d+(?:[.,]\d+)?)", raw, _re.IGNORECASE)
    if m3:
        cores = int(m3.group(1)) * int(m3.group(2))
        section = float(m3.group(3).replace(",", "."))
    else:
        m2 = _re.search(r"(\d+)\s*[х×x*]\s*(\d+(?:[.,]\d+)?)", raw, _re.IGNORECASE)
        if m2:
            cores = int(m2.group(1))
            section = float(m2.group(2).replace(",", "."))
    return {"isCable": is_cable, "cableType": cable_type, "cores": cores, "section": section}

def _add_project_material(cur, name, unit, qty, price, project):
    if not name or not project or qty <= 0:
        return
    cur.execute("SELECT id FROM materials WHERE name=%s AND project=%s LIMIT 1", (name, project))
    existing = cur.fetchone()
    if existing:
        cur.execute("UPDATE materials SET quantity=COALESCE(quantity,0)+%s, unit=%s, price=%s WHERE id=%s",
                    (qty, unit or "", price or 0, existing['id'] if isinstance(existing, dict) else existing[0]))
    else:
        cur.execute("INSERT INTO materials (name, unit, quantity, price, min_quantity, project, category) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (name, unit or "шт", qty, price or 0, 0, project, "Закупка"))
    cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (name, "приход (поставка)", qty, __import__("datetime").date.today().isoformat(), project, "Снабжение", __import__("datetime").datetime.now().strftime("%d.%m.%Y, %H:%M")))

def _create_delivery_quality_records(cur, delivery):
    name = (delivery.get('material_name') or '').strip()
    if not name:
        return
    delivery_id = delivery.get('id')
    project = delivery.get('project') or ''
    supplier = delivery.get('supplier_name') or ''
    qty = _float_or_zero(delivery.get('received_quantity'))
    unit = delivery.get('unit') or 'шт'
    received_at = delivery.get('received_at') or __import__("datetime").date.today()
    has_inspection = False
    if delivery_id:
        try:
            cur.execute("SELECT id FROM material_inspection_journal WHERE delivery_id=%s LIMIT 1", (delivery_id,))
            has_inspection = bool(cur.fetchone())
        except Exception as e:
            print("DELIVERY INSPECTION CHECK ERROR:", str(e))
    if not has_inspection:
        try:
            cur.execute("""INSERT INTO material_inspection_journal
                           (project_name, delivery_id, material_name, unit, quantity, supplier,
                            received_at, visual_inspection_result, remarks, inspected)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (project, delivery_id, name, unit, qty, supplier, received_at,
                         delivery.get('quality_status') or 'Принято',
                         delivery.get('quality_notes') or '', True))
        except Exception as e:
            print("DELIVERY INSPECTION INSERT ERROR:", str(e))
            try:
                cur.execute("""INSERT INTO material_inspection_journal
                               (project_name, material_name, unit, quantity, supplier,
                                received_at, visual_inspection_result, remarks, inspected)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (project, name, unit, qty, supplier, received_at,
                             delivery.get('quality_status') or 'Принято',
                             delivery.get('quality_notes') or '', True))
            except Exception as legacy_error:
                print("DELIVERY INSPECTION LEGACY INSERT ERROR:", str(legacy_error))
    cable_info = _detect_cable_info(name)
    if cable_info["isCable"]:
        has_cable = False
        if delivery_id:
            try:
                cur.execute("SELECT id FROM cable_journal WHERE delivery_id=%s LIMIT 1", (delivery_id,))
                has_cable = bool(cur.fetchone())
            except Exception as e:
                print("DELIVERY CABLE CHECK ERROR:", str(e))
        if not has_cable:
            try:
                cur.execute("""SELECT id FROM cable_journal
                               WHERE project_name=%s AND cable_brand=%s AND COALESCE(length_received,0)=%s
                               LIMIT 1""", (project, name, qty))
                has_cable = bool(cur.fetchone())
            except Exception as e:
                print("DELIVERY CABLE LEGACY CHECK ERROR:", str(e))
        if not has_cable:
            try:
                cur.execute("""INSERT INTO cable_journal
                               (project_name, delivery_id, cable_brand, cable_type, cross_section, cores_count,
                                length_received, supplier, received_at)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (project, delivery_id, name, cable_info["cableType"],
                             cable_info["section"], cable_info["cores"], qty, supplier, received_at))
            except Exception as e:
                print("DELIVERY CABLE INSERT ERROR:", str(e))
                try:
                    cur.execute("""INSERT INTO cable_journal
                                   (project_name, cable_brand, cross_section, cores_count,
                                    length_received, supplier, received_at)
                                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                                (project, name, cable_info["section"], cable_info["cores"],
                                 qty, supplier, received_at))
                except Exception as legacy_error:
                    print("DELIVERY CABLE LEGACY INSERT ERROR:", str(legacy_error))

def _create_supply_delivery_history(cur, delivery, status=None, received_qty=None, confirmed_by=None):
    delivery_id = delivery.get('id')
    request_id = delivery.get('request_id')
    received_qty = _float_or_zero(received_qty if received_qty is not None else delivery.get('received_quantity'))
    status = status or delivery.get('status') or 'Принято'
    confirmed_by = confirmed_by if confirmed_by is not None else (delivery.get('received_by') or '')
    try:
        cur.execute("ALTER TABLE supply_history ADD COLUMN IF NOT EXISTS request_id INT")
        cur.execute("ALTER TABLE supply_history ADD COLUMN IF NOT EXISTS delivery_id INT")
    except Exception as e:
        print("SUPPLY HISTORY MIGRATION ERROR:", str(e))
    if delivery_id:
        try:
            cur.execute("SELECT id FROM supply_history WHERE delivery_id=%s LIMIT 1", (delivery_id,))
            existing = cur.fetchone()
            if existing:
                cur.execute("""UPDATE supply_history SET quantity=%s, status=%s, confirmed_by=%s
                               WHERE id=%s""",
                            (received_qty, status, confirmed_by,
                             existing['id'] if isinstance(existing, dict) else existing[0]))
                return
        except Exception as e:
            print("SUPPLY HISTORY CHECK ERROR:", str(e))
    try:
        cur.execute("""INSERT INTO supply_history
                       (supplier_id, material_name, quantity, unit, price_per_unit, total_price,
                        project, date, status, confirmed_by, request_id, delivery_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (delivery.get('supplier_id'), delivery.get('material_name'), received_qty, delivery.get('unit'),
                     _float_or_zero(delivery.get('price_per_unit')),
                     _float_or_zero(delivery.get('price_per_unit')) * received_qty,
                     delivery.get('project'), (delivery.get('received_at') or __import__("datetime").date.today()).date().isoformat()
                     if hasattr(delivery.get('received_at'), 'date') else (delivery.get('received_at') or __import__("datetime").date.today().isoformat()),
                     status, confirmed_by, request_id, delivery_id))
    except Exception as e:
        print("SUPPLY HISTORY LINKED INSERT ERROR:", str(e))
        try:
            cur.execute("""INSERT INTO supply_history
                           (supplier_id, material_name, quantity, unit, price_per_unit, total_price,
                            project, date, status, confirmed_by)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                        (delivery.get('supplier_id'), delivery.get('material_name'), received_qty, delivery.get('unit'),
                         _float_or_zero(delivery.get('price_per_unit')),
                         _float_or_zero(delivery.get('price_per_unit')) * received_qty,
                         delivery.get('project'), __import__("datetime").date.today().isoformat(),
                         status, confirmed_by))
        except Exception as legacy_error:
            print("SUPPLY HISTORY LEGACY INSERT ERROR:", str(legacy_error))

def _ensure_cable_journal_row(cur, *, project, cable_brand, qty, supplier="", received_at=None, delivery_id=None, invoice_id=None):
    cable_info = _detect_cable_info(cable_brand)
    if not cable_info["isCable"]:
        return False
    exists = False
    if delivery_id:
        try:
            cur.execute("SELECT id FROM cable_journal WHERE delivery_id=%s LIMIT 1", (delivery_id,))
            exists = bool(cur.fetchone())
        except Exception as e:
            print("CABLE BACKFILL DELIVERY CHECK ERROR:", str(e))
    if not exists and invoice_id:
        try:
            cur.execute("SELECT id FROM cable_journal WHERE invoice_id=%s AND cable_brand=%s LIMIT 1", (invoice_id, cable_brand))
            exists = bool(cur.fetchone())
        except Exception as e:
            print("CABLE BACKFILL INVOICE CHECK ERROR:", str(e))
    if not exists:
        try:
            cur.execute("""SELECT id FROM cable_journal
                           WHERE project_name=%s AND cable_brand=%s AND COALESCE(length_received,0)=%s
                           LIMIT 1""", (project or "", cable_brand or "", _float_or_zero(qty)))
            exists = bool(cur.fetchone())
        except Exception as e:
            print("CABLE BACKFILL LEGACY CHECK ERROR:", str(e))
    if exists:
        return False
    try:
        cur.execute("""INSERT INTO cable_journal
                       (project_name, delivery_id, invoice_id, cable_brand, cable_type, cross_section, cores_count,
                        length_received, supplier, received_at)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (project or "", delivery_id, invoice_id, cable_brand or "", cable_info["cableType"],
                     cable_info["section"], cable_info["cores"], _float_or_zero(qty), supplier or "", received_at))
        return True
    except Exception as e:
        print("CABLE BACKFILL INSERT ERROR:", str(e))
        try:
            cur.execute("""INSERT INTO cable_journal
                           (project_name, cable_brand, cross_section, cores_count,
                            length_received, supplier, received_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (project or "", cable_brand or "", cable_info["section"], cable_info["cores"],
                         _float_or_zero(qty), supplier or "", received_at))
            return True
        except Exception as legacy_error:
            print("CABLE BACKFILL LEGACY INSERT ERROR:", str(legacy_error))
    return False

def _backfill_cable_journal(cur, project_names=None):
    import json as _json
    repaired = 0
    project_names = list(project_names or [])
    delivery_where = "WHERE COALESCE(received_quantity,0)>0 AND status IN ('Принято','Проблема')"
    delivery_params = []
    if project_names:
        delivery_where += " AND project = ANY(%s)"
        delivery_params.append(project_names)
    cur.execute("""SELECT id, project, material_name, received_quantity, supplier_name, received_at
                   FROM supply_deliveries """ + delivery_where, delivery_params)
    for d in cur.fetchall():
        delivery_id, project, name, qty, supplier, received_at = d
        if _ensure_cable_journal_row(cur, project=project, cable_brand=name, qty=qty,
                                     supplier=supplier, received_at=received_at, delivery_id=delivery_id):
            repaired += 1
    invoice_where = ""
    invoice_params = []
    if project_names:
        invoice_where = " WHERE project = ANY(%s) OR location = ANY(%s)"
        invoice_params = [project_names, project_names]
    cur.execute("SELECT id, project, location, supplier_name, date, items FROM warehouse_invoices" + invoice_where, invoice_params)
    for inv in cur.fetchall():
        invoice_id, project, location, supplier, date_value, items_json = inv
        target_project = project or (location if location and location != "Основной склад" else "")
        if project_names and target_project not in project_names:
            continue
        try:
            items = _json.loads(items_json) if items_json else []
        except Exception:
            items = []
        for item in items:
            if not isinstance(item, dict):
                continue
            name = (item.get("name") or item.get("materialName") or "").strip()
            qty = _float_or_zero(item.get("quantity"))
            if _ensure_cable_journal_row(cur, project=target_project, cable_brand=name, qty=qty,
                                         supplier=supplier, received_at=date_value, invoice_id=invoice_id):
                repaired += 1
    material_where = "WHERE COALESCE(quantity,0)>0"
    material_params = []
    if project_names:
        material_where += " AND project = ANY(%s)"
        material_params.append(project_names)
    cur.execute("SELECT name, quantity, project FROM materials " + material_where, material_params)
    for material in cur.fetchall():
        name, qty, project = material
        if _ensure_cable_journal_row(cur, project=project, cable_brand=name, qty=qty):
            repaired += 1
    return repaired

@app.get("/supply-deliveries")
def list_supply_deliveries(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if role in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.execute(DELIVERY_SELECT + " ORDER BY d.id DESC")
    elif role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if not supplier_id:
            cur.close(); conn.close()
            return []
        cur.execute(DELIVERY_SELECT + " WHERE d.supplier_id=%s ORDER BY d.id DESC", (supplier_id,))
    else:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute(DELIVERY_SELECT + " WHERE d.project = ANY(%s) ORDER BY d.id DESC", (projects,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]

@app.get("/supply-claims")
def list_supply_claims(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if role in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.execute(CLAIM_SELECT + " ORDER BY id DESC")
    elif role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if not supplier_id:
            cur.close(); conn.close()
            return []
        cur.execute(CLAIM_SELECT + " WHERE supplier_id=%s ORDER BY id DESC", (supplier_id,))
    else:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute(CLAIM_SELECT + " WHERE project = ANY(%s) ORDER BY id DESC", (projects,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]

@app.post("/supplier-offers/{id}/ship")
def ship_supplier_offer(id: int, data: dict, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    from datetime import datetime
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT o.id, o.request_id, o.supplier_id, o.price_per_unit, o.total_price,
               o.payment_terms, s.name as supplier_name,
               r.project, r.material_name, r.quantity, r.unit
        FROM supplier_offers o
        LEFT JOIN suppliers s ON s.id=o.supplier_id
        LEFT JOIN supply_requests r ON r.id=o.request_id
        WHERE o.id=%s AND o.status=%s
    """, (id, 'Утверждено'))
    offer = cur.fetchone()
    if not offer:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Утверждённое КП не найдено")
    cur.execute("SELECT id, status, paid_amount, amount FROM supplier_invoices WHERE offer_id=%s ORDER BY id DESC LIMIT 1", (id,))
    inv = cur.fetchone()
    terms = (offer.get('payment_terms') or '').lower()
    need_payment = ('предоплат' in terms) or ('50/50' in terms) or ('50' in terms and 'постоплат' not in terms)
    if need_payment:
        paid = _float_or_zero(inv['paid_amount']) if inv else 0
        amount = _float_or_zero(inv['amount']) if inv else _float_or_zero(offer['total_price'])
        required = amount if '100' in terms or 'предоплат' in terms else amount * 0.5
        if not inv:
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Сначала поставщик должен выставить счёт, а бухгалтерия — оплатить по условиям КП")
        if paid + 0.01 < required:
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail=f"По условиям «{offer.get('payment_terms') or ''}» перед отгрузкой нужно оплатить минимум {round(required, 2)} ₽. Сейчас оплачено {round(paid, 2)} ₽")
    shipped_qty = _float_or_zero(data.get('shippedQuantity') or offer['quantity'])
    cur.execute("SELECT id FROM supply_deliveries WHERE offer_id=%s LIMIT 1", (id,))
    existing = cur.fetchone()
    vals = (
        offer['request_id'], offer['supplier_id'], offer['supplier_name'] or '',
        offer['project'] or '', offer['material_name'] or '',
        _float_or_zero(offer['quantity']), shipped_qty, offer['unit'] or '',
        _float_or_zero(offer['price_per_unit']), _float_or_zero(offer['total_price']),
        data.get('waybillNumber') or '', data.get('waybillDate') or None,
        data.get('vehicleNumber') or '', data.get('driverName') or '',
        data.get('documentUrl') or '', data.get('photoUrl') or '', datetime.now()
    )
    if existing:
        cur.execute("SELECT status FROM supply_deliveries WHERE id=%s", (existing['id'],))
        existing_status = cur.fetchone()
        if existing_status and existing_status['status'] in ('Принято', 'Проблема'):
            cur.close(); conn.close()
            raise HTTPException(status_code=400, detail="Поставка уже принята. Повторная отгрузка запрещена — создайте новую заявку/КП для допоставки.")
        cur.execute("""UPDATE supply_deliveries SET
                       request_id=%s, supplier_id=%s, supplier_name=%s, project=%s,
                       material_name=%s, planned_quantity=%s, shipped_quantity=%s, unit=%s,
                       price_per_unit=%s, total_price=%s, waybill_number=%s, waybill_date=%s,
                       vehicle_number=%s, driver_name=%s, document_url=%s, photo_url=%s,
                       shipped_at=%s, status='В пути'
                       WHERE id=%s RETURNING id""", vals + (existing['id'],))
        delivery_id = cur.fetchone()['id']
    else:
        cur.execute("""INSERT INTO supply_deliveries
                       (offer_id, request_id, supplier_id, supplier_name, project,
                        material_name, planned_quantity, shipped_quantity, unit,
                        price_per_unit, total_price, waybill_number, waybill_date,
                        vehicle_number, driver_name, document_url, photo_url, shipped_at, status)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                       RETURNING id""",
                    (id,) + vals + ('В пути',))
        delivery_id = cur.fetchone()['id']
    cur.execute("UPDATE supplier_offers SET delivery_status=%s WHERE id=%s", ('В пути', id))
    cur.execute("UPDATE supply_requests SET status=%s WHERE id=%s", ('В пути', offer['request_id']))
    cur.execute(DELIVERY_SELECT + " WHERE d.id=%s", (delivery_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return dict(row)

@app.put("/supply-deliveries/{id}/receive")
def receive_supply_delivery(id: int, data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    from datetime import datetime
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM supply_deliveries WHERE id=%s", (id,))
    delivery = cur.fetchone()
    if not delivery:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Поставка не найдена")
    if delivery['status'] in ('Принято', 'Проблема') or delivery.get('received_at'):
        try:
            _create_delivery_quality_records(cur, delivery)
        except Exception as e:
            print("DELIVERY QUALITY RECOVERY ERROR:", str(e))
        try:
            _create_supply_delivery_history(cur, delivery)
        except Exception as e:
            print("DELIVERY HISTORY RECOVERY ERROR:", str(e))
        try:
            cur.execute(DELIVERY_SELECT + " WHERE d.id=%s", (id,))
            row = cur.fetchone()
        except Exception as e:
            print("DELIVERY RECOVERY SELECT ERROR:", str(e))
            row = None
        cur.close(); conn.close()
        return {
            "ok": True,
            "delivery": dict(row) if row else {"id": id},
            "claimId": delivery.get('claim_id'),
            "alreadyReceived": True
        }
    received_qty = _float_or_zero(data.get('receivedQuantity'))
    planned_qty = _float_or_zero(delivery['planned_quantity'])
    shipped_qty = _float_or_zero(delivery['shipped_quantity']) or planned_qty
    quality_status = data.get('qualityStatus') or 'Принято'
    shortage = max(0.0, shipped_qty - received_qty)
    problem = shortage > 0 or quality_status in ('Брак', 'Несоответствие', 'Недостача', 'Частично')
    status = 'Проблема' if problem else 'Принято'
    received_at = datetime.now()
    cur.execute("""UPDATE supply_deliveries SET received_quantity=%s, received_at=%s,
                   received_by=%s, quality_status=%s, quality_notes=%s,
                   shortage_quantity=%s, photo_url=COALESCE(NULLIF(%s,''), photo_url),
                   status=%s WHERE id=%s""",
                (received_qty, received_at, data.get('receivedBy') or '',
                 quality_status, data.get('qualityNotes') or '',
                 shortage, data.get('photoUrl') or '', status, id))
    claim_id = None
    if received_qty > 0 and quality_status not in ('Брак',):
        _add_project_material(cur, delivery['material_name'], delivery['unit'], received_qty,
                              _float_or_zero(delivery['price_per_unit']), delivery['project'])
    cur.execute("SELECT * FROM supply_deliveries WHERE id=%s", (id,))
    updated = cur.fetchone()
    _create_delivery_quality_records(cur, updated)
    if problem:
        claim_type = 'Недостача' if shortage > 0 else quality_status
        description = data.get('claimDescription') or (
            f"По поставке «{delivery['material_name']}» ожидалось {shipped_qty:g} {delivery['unit']}, "
            f"принято {received_qty:g} {delivery['unit']}. "
            f"Статус качества: {quality_status}. {data.get('qualityNotes') or ''}"
        )
        cur.execute("""INSERT INTO supply_claims
                       (delivery_id, request_id, offer_id, supplier_id, project, material_name,
                        claim_type, description, expected_quantity, received_quantity,
                        shortage_quantity, photo_url, created_by)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (id, delivery['request_id'], delivery['offer_id'], delivery['supplier_id'],
                     delivery['project'], delivery['material_name'], claim_type, description,
                     shipped_qty, received_qty, shortage, data.get('photoUrl') or '',
                     data.get('receivedBy') or ''))
        claim_id = cur.fetchone()['id']
        cur.execute("UPDATE supply_deliveries SET claim_id=%s WHERE id=%s", (claim_id, id))
    cur.execute("UPDATE supplier_offers SET delivery_status=%s WHERE id=%s", (status, delivery['offer_id']))
    cur.execute("UPDATE supply_requests SET status=%s WHERE id=%s",
                ('Поставлено' if status == 'Принято' else 'Проблема поставки', delivery['request_id']))
    _create_supply_delivery_history(cur, updated, status, received_qty, data.get('receivedBy') or '')
    cur.execute(DELIVERY_SELECT + " WHERE d.id=%s", (id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"ok": True, "delivery": dict(row), "claimId": claim_id}

@app.post("/supply-deliveries/{id}/ai-check")
def ai_check_supply_delivery(id: int, data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM supply_deliveries WHERE id=%s", (id,))
    d = cur.fetchone()
    if not d:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Поставка не найдена")
    parsed_items = data.get('parsedItems') or []
    doc_text = data.get('documentText') or ''
    expected = {
        "material": d['material_name'],
        "quantity": _float_or_zero(d['planned_quantity']),
        "unit": d['unit'] or '',
        "supplier": d['supplier_name'] or '',
        "project": d['project'] or '',
    }
    result_text = ""
    try:
        if parsed_items:
            best = None
            exp_name = (expected['material'] or '').lower()
            for it in parsed_items:
                nm = (it.get('name') or it.get('materialName') or '').lower()
                if nm and (nm in exp_name or exp_name in nm or any(w in nm for w in exp_name.split()[:2])):
                    best = it
                    break
            if best:
                qty = _float_or_zero(best.get('quantity'))
                diff = expected['quantity'] - qty
                result_text = (
                    f"В документе найдена похожая позиция: {best.get('name') or best.get('materialName')} "
                    f"{qty:g} {best.get('unit') or expected['unit']}. "
                    + ("Количество совпадает." if abs(diff) < 0.0001 else f"Расхождение: {diff:g} {expected['unit']}.")
                )
            else:
                result_text = "В распознанном документе не нашёл позицию, похожую на заявку. Проверьте накладную вручную."
        elif doc_text:
            import openai as oa
            prompt = (
                "Сверь поставку с текстом накладной. Ожидалось: "
                f"{expected['material']} — {expected['quantity']} {expected['unit']}, "
                f"поставщик {expected['supplier']}, объект {expected['project']}.\n"
                "Текст/заметки документа:\n" + doc_text[:4000] +
                "\nОтветь коротко: совпадает ли материал и количество, есть ли риск недопоставки или пересорта."
            )
            client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
            r = client.responses.create(
                model="gpt://"+YANDEX_FOLDER_ID+"/yandexgpt-5.1/latest",
                temperature=0.1,
                instructions="Ты помощник кладовщика на стройке. Проверяешь накладную перед приёмкой.",
                input=prompt,
                max_output_tokens=500,
            )
            result_text = (r.output_text or '').strip()
        else:
            result_text = "Загрузите фото накладной или вставьте текст/позиции документа для AI-сверки."
    except Exception as e:
        print("DELIVERY AI CHECK ERROR:", e)
        result_text = "AI-сверка не сработала, проверьте поставку вручную."
    cur.execute("UPDATE supply_deliveries SET ai_check_result=%s WHERE id=%s", (result_text, id))
    cur.close(); conn.close()
    return {"ok": True, "result": result_text, "expected": expected}

@app.put("/supply-claims/{id}")
def update_supply_claim(id: int, data: dict, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    fields_map = [('status','status'),('resolution','resolution'),('resolvedAt','resolved_at')]
    sets, vals = [], []
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            vals.append(data[js_key] or None if js_key == 'resolvedAt' else data[js_key])
    if data.get('status') in ('Закрыта', 'Решена') and 'resolvedAt' not in data:
        sets.append("resolved_at=NOW()")
    if not sets:
        cur.close(); conn.close()
        return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE supply_claims SET " + ", ".join(sets) + " WHERE id=%s", vals)
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/supply-history")
def get_supply_history(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if role in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.execute("SELECT id,supplier_id as \"supplierId\",material_name as \"materialName\",quantity,unit,price_per_unit as \"pricePerUnit\",total_price as \"totalPrice\",project,date,status,confirmed_by as \"confirmedBy\" FROM supply_history ORDER BY id DESC")
    elif role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if not supplier_id:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,supplier_id as \"supplierId\",material_name as \"materialName\",quantity,unit,price_per_unit as \"pricePerUnit\",total_price as \"totalPrice\",project,date,status,confirmed_by as \"confirmedBy\" FROM supply_history WHERE supplier_id=%s ORDER BY id DESC", (supplier_id,))
    else:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,supplier_id as \"supplierId\",material_name as \"materialName\",quantity,unit,price_per_unit as \"pricePerUnit\",total_price as \"totalPrice\",project,date,status,confirmed_by as \"confirmedBy\" FROM supply_history WHERE project = ANY(%s) ORDER BY id DESC", (projects,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/supply-history")
def create_supply_history(d: SupplyHistoryModel, _current_user: dict = Depends(require_roles("директор", "зам_директора", "снабженец", "кладовщик", "прораб", "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO supply_history (supplier_id,material_name,quantity,unit,price_per_unit,total_price,project,date,status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (d.supplierId,d.materialName,d.quantity,d.unit,d.pricePerUnit,d.totalPrice,d.project,d.date,d.status))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/supply-history/{id}")
def update_supply_history(id: int, data: dict, _current_user: dict = Depends(require_roles("директор", "зам_директора", "снабженец", "кладовщик", "прораб", "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor()
    status = data.get('status','')
    confirmed_by = data.get('confirmedBy','')
    cur.execute("UPDATE supply_history SET status=%s,confirmed_by=%s WHERE id=%s", (status,confirmed_by,id))
    conn.close()
    return {"ok": True}

@app.get("/work-journal")
def get_work_journal(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    select_sql = """SELECT id,master_id as "masterId",master_name as "masterName",project,description,unit,quantity,
                          price_per_unit as "pricePerUnit",total,date,status,comment,
                          photo_url as "photoUrl",confirmed_by as "confirmedBy",confirmed_at as "confirmedAt",
                          materials_used as "materialsUsed",
                          estimate_id as "estimateId", section_name as "sectionName",
                          responsible_itr as "responsibleItr", weather,
                          time_start as "timeStart", time_end as "timeEnd",
                          hidden_work as "hiddenWork", quality_status as "qualityStatus",
                          normatives, project_docs as "projectDocs",
                          ai_filled as "aiFilled",
                          unexpected_work_id as "unexpectedWorkId"
                   FROM work_journal"""
    role = current_user.get("role")
    if can_see_all_company_data(current_user) or role in ("прораб", "стройконтроль", "технадзор"):
        projects = user_project_names(current_user)
        if role in ("прораб", "стройконтроль", "технадзор") and projects:
            cur.execute(select_sql + " WHERE project = ANY(%s) ORDER BY id DESC", (projects,))
        else:
            cur.execute(select_sql + " ORDER BY id DESC")
    elif role in ("мастер", "субподрядчик"):
        cur.execute(select_sql + " WHERE master_id=%s OR master_name=%s ORDER BY id DESC", (current_user.get("id"), current_user.get("name") or ""))
    elif role == "заказчик":
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute(select_sql + " WHERE project = ANY(%s) AND status='Подтверждено' ORDER BY id DESC", (projects,))
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

def _parse_materials_used(raw):
    import json as _json
    if not raw:
        return []
    if isinstance(raw, list):
        return raw
    try:
        parsed = _json.loads(raw)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []

def _personal_material_balance(cur, project: str, person_id, person_name: str, material_name: str):
    key = (material_name or "").strip().lower()
    if not key:
        return {"issued": 0, "used": 0, "available": 0}
    cur.execute("""SELECT COALESCE(SUM(quantity),0)
                   FROM material_transfers
                   WHERE project_name=%s AND to_person=%s AND signed=TRUE
                     AND LOWER(TRIM(material_name))=LOWER(TRIM(%s))""",
                (project, person_name or "", material_name or ""))
    issued_row = cur.fetchone()
    issued = float((next(iter(issued_row.values())) if isinstance(issued_row, dict) else ((issued_row or [0])[0])) or 0)
    if person_id:
        cur.execute("""SELECT materials_used FROM work_journal
                       WHERE project=%s
                         AND COALESCE(status,'') <> 'Отклонено'
                         AND (master_id=%s OR master_name=%s)""",
                    (project, person_id, person_name or ""))
    else:
        cur.execute("""SELECT materials_used FROM work_journal
                       WHERE project=%s
                         AND COALESCE(status,'') <> 'Отклонено'
                         AND master_name=%s""",
                    (project, person_name or ""))
    used = 0
    for row in cur.fetchall() or []:
        raw = row.get("materials_used") if isinstance(row, dict) else row[0]
        for m in _parse_materials_used(raw):
            if (m.get("name") or "").strip().lower() == key:
                try:
                    used += float(m.get("quantity") or 0)
                except Exception:
                    pass
    return {"issued": issued, "used": used, "available": issued - used}

def _apply_material_work_writeoff(cur, project: str, material: dict, actor: dict, date_value: str, fallback_master_name: str = ""):
    name = material.get("name") or ""
    qty = float(material.get("quantity") or 0)
    unit = material.get("unit") or ""
    role = actor.get("role") or ""
    actor_id = actor.get("id")
    actor_name = actor.get("name") or fallback_master_name or ""
    if not name or qty <= 0:
        return
    if role in ("мастер", "субподрядчик"):
        balance = _personal_material_balance(cur, project, actor_id, actor_name, name)
        if balance["issued"] <= 0:
            raise HTTPException(status_code=400, detail="Материал «"+name+"» не выдан мастеру «"+actor_name+"» или получение не подтверждено")
        if balance["available"] < qty:
            raise HTTPException(status_code=400, detail="У мастера «"+actor_name+"» доступно "+str(round(balance["available"], 3))+" "+unit+" «"+name+"», запрошено "+str(qty)+". Лишний материал нужно выдать или вернуть/уточнить списание")
        cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_to,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (name, "расход (работа мастера)", qty, date_value or None, project, actor_name, actor_name, __import__("datetime").datetime.now().strftime("%d.%m.%Y, %H:%M")))
        return

    cur.execute("SELECT id, quantity FROM materials WHERE name=%s AND project=%s FOR UPDATE", (name, project))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="Материал «"+name+"» не найден на складе объекта «"+project+"»")
    mat_id = row.get("id") if isinstance(row, dict) else row[0]
    stock_qty = float((row.get("quantity") if isinstance(row, dict) else row[1]) or 0)
    if stock_qty < qty:
        raise HTTPException(status_code=400, detail="На складе «"+project+"» только "+str(stock_qty)+" "+unit+" «"+name+"», запрошено "+str(qty))
    cur.execute("UPDATE materials SET quantity=quantity-%s WHERE id=%s", (qty, mat_id))
    cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (name, "расход (работа)", qty, date_value or None, project, actor_name, __import__("datetime").datetime.now().strftime("%d.%m.%Y, %H:%M")))

@app.post("/work-journal")
def create_work_journal(w: WorkJournalModel, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    require_project_access(_current_user, w.project)
    used = [m for m in (w.materialsUsed or []) if m.get("name") and float(m.get("quantity") or 0) > 0]

    conn = get_db()
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        for m in used:
            _apply_material_work_writeoff(cur, w.project, m, _current_user, w.date, w.masterName)

        import json as _json
        materials_json = _json.dumps(used, ensure_ascii=False) if used else None
        cur.execute("""INSERT INTO work_journal
                       (master_id,master_name,project,description,unit,quantity,price_per_unit,total,date,
                        comment,photo_url,materials_used,
                        estimate_id,section_name,responsible_itr,weather,time_start,time_end,
                        hidden_work,quality_status,normatives,project_docs)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
                    (w.masterId,w.masterName,w.project,w.description,w.unit,w.quantity,w.pricePerUnit,w.total,w.date,
                     w.comment,w.photoUrl,materials_json,
                     w.estimateId,w.sectionName,w.responsibleItr,w.weather,w.timeStart,w.timeEnd,
                     w.hiddenWork,w.qualityStatus,w.normatives,w.projectDocs))
        row = cur.fetchone()
        conn.commit()
        return dict(row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.put("/work-journal/{id}")
def update_work_journal(id: int, data: dict, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "work_journal", id, _current_user, "project")
    # Динамически обновляем только переданные поля. Старая логика (status/confirmedBy/...) продолжает работать.
    fields_map = [
        ('status', 'status'),
        ('confirmedBy', 'confirmed_by'),
        ('confirmedAt', 'confirmed_at'),
        ('comment', 'comment'),
        ('responsibleItr', 'responsible_itr'),
        ('weather', 'weather'),
        ('timeStart', 'time_start'),
        ('timeEnd', 'time_end'),
        ('qualityStatus', 'quality_status'),
        ('normatives', 'normatives'),
        ('projectDocs', 'project_docs'),
        ('sectionName', 'section_name'),
        ('hiddenWork', 'hidden_work'),
    ]
    sets, vals = [], []
    ai_resetting_fields = {'responsibleItr','weather','timeStart','timeEnd','qualityStatus','normatives','projectDocs'}
    reset_ai = False
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            vals.append(data[js_key])
            if js_key in ai_resetting_fields:
                reset_ai = True
    if reset_ai:
        sets.append("ai_filled=FALSE")
    if not sets:
        cur.close(); conn.close()
        return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE work_journal SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/work-journal/{id}")
def delete_work_journal(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер"))):
    import json as _json
    from datetime import datetime
    conn = get_db()
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        require_row_project_access(cur, "work_journal", id, _current_user, "project")
        cur.execute("SELECT project, master_id, master_name, date, materials_used FROM work_journal WHERE id=%s FOR UPDATE", (id,))
        work = cur.fetchone()
        if not work:
            conn.rollback()
            cur.close(); conn.close()
            raise HTTPException(status_code=404, detail="Запись журнала не найдена")
        used = []
        raw_used = work.get("materials_used")
        if raw_used:
            try:
                parsed = _json.loads(raw_used) if isinstance(raw_used, str) else raw_used
                if isinstance(parsed, list):
                    used = [m for m in parsed if isinstance(m, dict) and m.get("name") and float(m.get("quantity") or 0) > 0]
            except Exception:
                used = []
        for m in used:
            name = m.get("name") or ""
            qty = float(m.get("quantity") or 0)
            unit = m.get("unit") or "шт"
            project_name = work.get("project") or ""
            master_name = work.get("master_name") or ""
            master_id = work.get("master_id")
            personal_balance = _personal_material_balance(cur, project_name, master_id, master_name, name)
            if personal_balance["issued"] > 0:
                cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_to,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                            (name, "отмена списания мастера", qty, str(work.get("date") or ""), project_name, master_name, _current_user.get("name") or "", datetime.now().strftime("%d.%m.%Y, %H:%M")))
                continue
            cur.execute("SELECT id FROM materials WHERE name=%s AND project=%s FOR UPDATE", (name, project_name))
            mat = cur.fetchone()
            if mat:
                cur.execute("UPDATE materials SET quantity=COALESCE(quantity,0)+%s, unit=COALESCE(NULLIF(unit,''),%s) WHERE id=%s", (qty, unit, mat["id"]))
            else:
                cur.execute("INSERT INTO materials (name, unit, quantity, price, min_quantity, project, category) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                            (name, unit, qty, 0, 0, project_name, "Возврат"))
            cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                        (name, "возврат (удаление работы)", qty, str(work.get("date") or ""), project_name, master_name, datetime.now().strftime("%d.%m.%Y, %H:%M")))
        cur.execute("DELETE FROM piecework WHERE work_journal_id=%s", (id,))
        cur.execute("DELETE FROM work_journal WHERE id=%s", (id,))
        conn.commit()
        return {"ok": True, "materialsRestored": len(used)}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.post("/work-journal/{id}/ai-prefill")
def ai_prefill_work_journal(id: int, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    import openai as oa, json as j, re
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "work_journal", id, _current_user, "project")
    cur.execute("""SELECT description, section_name, unit, quantity, materials_used, project, hidden_work
                   FROM work_journal WHERE id=%s""", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="запись не найдена")
    description, section_name, unit, quantity, materials_used, project, hidden_work = row
    cur.close()

    user_text = (
        "Описание работы: " + (description or "—") + "\n"
        "Раздел сметы: " + (section_name or "—") + "\n"
        "Объект: " + (project or "—") + "\n"
        "Объём: " + str(quantity or 0) + " " + (unit or "") + "\n"
        "Скрытые работы: " + ("да" if hidden_work else "нет") + "\n"
        "Использованные материалы: " + (materials_used or "—") + "\n\n"
        "Верни СТРОГО JSON с тремя полями (без markdown, без тройных кавычек):\n"
        '{"normatives": "...", "projectDocs": "...", "qualityNote": "..."}\n'
        "Где:\n"
        "- normatives: перечень применимых СНиП/СП/ГОСТ через запятую (только реально применимые к виду работ)\n"
        "- projectDocs: типовые проектные документы (разделы рабочего проекта, листы)\n"
        "- qualityNote: краткая формулировка о соответствии работ нормативной/проектной документации, "
        "официальный канцелярский русский, 1-2 предложения."
    )
    instructions = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, никаких тройных кавычек, никакого текста до или после JSON. Только сам JSON."
    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)

    def _call(model_id):
        try:
            r = client.responses.create(
                model="gpt://" + YANDEX_FOLDER_ID + "/" + model_id,
                temperature=0.1, instructions=instructions, input=user_text, max_output_tokens=2000,
            )
            return (r.output_text or ""), None
        except Exception as e:
            return "", str(e)

    answer, err = _call("qwen3.6-35b-a3b/latest")
    if not (answer or "").strip():
        print("AI-PREFILL work_journal primary empty, fallback. err=" + str(err))
        answer, err = _call("yandexgpt-5.1/latest")
    if not (answer or "").strip():
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул пустой ответ: " + str(err))

    text = answer.strip()
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        text = m.group(0)
    try:
        parsed = j.loads(text)
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул невалидный JSON: " + str(e)[:200])

    normatives = (parsed.get("normatives") or "").strip()
    project_docs = (parsed.get("projectDocs") or "").strip()
    quality_note = (parsed.get("qualityNote") or "").strip()

    cur = conn.cursor()
    # qualityNote дописываем в comment (если comment пустой), нормативы и проектные доки — в свои поля
    cur.execute("""UPDATE work_journal
                   SET normatives=%s, project_docs=%s,
                       comment = CASE WHEN COALESCE(comment,'')='' THEN %s ELSE comment END,
                       ai_filled=TRUE
                   WHERE id=%s""",
                (normatives, project_docs, quality_note, id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "normatives": normatives, "projectDocs": project_docs, "qualityNote": quality_note, "aiFilled": True}

# Ключевые слова скрытых работ (резерв, если ИИ недоступен) — СНиП 12-01-2004
HIDDEN_WORK_KEYWORDS = [
    "арматур","армиров","бетонир","фундамент","основани","гидроизол","пароизол","теплоизол",
    "утеплен","стяжк","засыпк","обратн","грунт","свая","сваи","ростверк","монолит","опалубк",
    "закладн","заземлен","молниезащит","кабел","проводк","электропроводк","трубопровод","разводк",
    "канализац","водопровод","дренаж","вентиляц","воздуховод","штроб","закрыт","скрыт",
    "мембран","рулонн","праймер","битум","сеткa","сетк","каркас","анкер","дюбел","примыкан",
]

def _detect_hidden_by_keywords(name):
    n = (name or "").lower()
    return any(k in n for k in HIDDEN_WORK_KEYWORDS)

@app.post("/estimates/{id}/ai-detect-hidden")
def ai_detect_hidden_works(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    import openai as oa, json as j, re
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "estimates", id, _current_user, "project_name")
    cur.execute("SELECT sections_json FROM estimates WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="смета не найдена")
    try:
        sections = j.loads(row[0]) if row[0] else []
    except Exception:
        sections = []
    if not isinstance(sections, list):
        sections = []
    # Собираем уникальные названия работ
    names = []
    for sec in sections:
        for it in (sec.get("items") or []):
            nm = (it.get("name") or "").strip()
            if nm and nm not in names:
                names.append(nm)
    if not names:
        cur.close(); conn.close()
        return {"ok": True, "count": 0, "sections": sections, "method": "empty"}

    hidden_set = set()
    method = "keywords"
    # Пытаемся через ИИ; при любой ошибке — откат на ключевые слова
    if YANDEX_API_KEY and YANDEX_FOLDER_ID:
        user_text = (
            "Ниже список наименований строительных работ из сметы. Определи, какие из них являются "
            "СКРЫТЫМИ работами, требующими оформления Акта освидетельствования скрытых работ (АОСР) по "
            "СНиП 12-01-2004 — это работы, скрываемые последующими конструкциями (земляные, фундаменты, "
            "армирование, бетонирование, гидро-/паро-/теплоизоляция, стяжки, скрытые инженерные сети — "
            "кабели/трубопроводы в стенах и полах, заземление, закладные и т.п.). Отделочные и монтажные "
            "видимые работы НЕ являются скрытыми.\n\n"
            "Список работ (JSON-массив):\n" + j.dumps(names, ensure_ascii=False) + "\n\n"
            'Верни СТРОГО JSON без markdown: {"hidden": ["точное название работы из списка", ...]}'
        )
        instructions = "Ты отвечаешь СТРОГО валидным JSON без markdown и пояснений. Только JSON."
        try:
            client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
            r = client.responses.create(
                model="gpt://" + YANDEX_FOLDER_ID + "/yandexgpt-5.1/latest",
                temperature=0.1, instructions=instructions, input=user_text, max_output_tokens=2000,
            )
            text = (r.output_text or "").strip()
            m = re.search(r"\{.*\}", text, re.DOTALL)
            if m:
                parsed = j.loads(m.group(0))
                arr = parsed.get("hidden") or []
                if isinstance(arr, list):
                    hidden_set = set((s or "").strip() for s in arr if isinstance(s, str))
                    method = "ai"
        except Exception as e:
            print("AI-DETECT-HIDDEN error, fallback to keywords:", str(e))

    if not hidden_set:
        hidden_set = set(nm for nm in names if _detect_hidden_by_keywords(nm))
        method = "keywords"

    # Проставляем hiddenWork=true найденным (ручные отметки не снимаем)
    count = 0
    for sec in sections:
        for it in (sec.get("items") or []):
            nm = (it.get("name") or "").strip()
            if nm in hidden_set and not it.get("hiddenWork"):
                it["hiddenWork"] = True
                count += 1
    cur.execute("UPDATE estimates SET sections_json=%s WHERE id=%s", (j.dumps(sections, ensure_ascii=False), id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "count": count, "detected": len(hidden_set), "sections": sections, "method": method}

@app.get("/master-profiles")
def get_master_profiles(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if current_user.get("role") in FINANCE_ROLES or current_user.get("role") in ("прораб", "главный_инженер"):
        cur.execute("SELECT id,user_id as \"userId\",full_name as \"fullName\",passport,inn,contract_type as \"contractType\",bank_account as \"bankAccount\",bank_name as \"bankName\",phone,specialization,ogrnip,profile_completed as \"profileCompleted\" FROM master_profiles")
    elif current_user.get("role") in ("мастер", "субподрядчик"):
        cur.execute("SELECT id,user_id as \"userId\",full_name as \"fullName\",passport,inn,contract_type as \"contractType\",bank_account as \"bankAccount\",bank_name as \"bankName\",phone,specialization,ogrnip,profile_completed as \"profileCompleted\" FROM master_profiles WHERE user_id=%s", (current_user.get("id"),))
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.get("/master-profile/{user_id}")
def get_master_profile(user_id: int, current_user: dict = Depends(get_current_user)):
    if current_user.get("id") != user_id and current_user.get("role") not in FINANCE_ROLES and current_user.get("role") not in ("прораб", "главный_инженер"):
        raise HTTPException(status_code=403, detail="Нет доступа к профилю")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id,user_id as \"userId\",full_name as \"fullName\",passport,inn,contract_type as \"contractType\",bank_account as \"bankAccount\",bank_name as \"bankName\",phone,specialization,ogrnip,profile_completed as \"profileCompleted\" FROM master_profiles WHERE user_id=%s", (user_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return {"userId": user_id, "fullName": "", "profileCompleted": False}
    return dict(row)

@app.post("/master-profile")
def create_master_profile(p: MasterProfileModel, current_user: dict = Depends(get_current_user)):
    if current_user.get("id") != p.userId and current_user.get("role") not in FINANCE_ROLES:
        raise HTTPException(status_code=403, detail="Можно редактировать только свой профиль")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        INSERT INTO master_profiles (user_id,full_name,passport,inn,contract_type,bank_account,bank_name,phone,specialization,ogrnip,profile_completed)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
        ON CONFLICT (user_id) DO UPDATE SET
            full_name=EXCLUDED.full_name,passport=EXCLUDED.passport,inn=EXCLUDED.inn,
            contract_type=EXCLUDED.contract_type,bank_account=EXCLUDED.bank_account,
            bank_name=EXCLUDED.bank_name,phone=EXCLUDED.phone,
            specialization=EXCLUDED.specialization,ogrnip=EXCLUDED.ogrnip,profile_completed=TRUE
        RETURNING id,user_id as "userId",full_name as "fullName",passport,inn,
            contract_type as "contractType",bank_account as "bankAccount",
            bank_name as "bankName",phone,specialization,ogrnip,profile_completed as "profileCompleted"
    """, (p.userId,p.fullName,p.passport,p.inn,p.contractType,p.bankAccount,p.bankName,p.phone,p.specialization,p.ogrnip))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.get("/contracts")
def get_contracts(_current_user: dict = Depends(require_roles(*CONTRACT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    allowed_projects = visible_project_names(_current_user)
    if allowed_projects is not None and not allowed_projects:
        conn.close()
        return []
    where, params = [], []
    if allowed_projects is not None:
        where.append("project = ANY(%s)")
        params.append(allowed_projects)
    if _current_user.get("role") in ("мастер", "субподрядчик"):
        where.append("(master_id=%s OR master_name=%s)")
        params.extend([_current_user.get("id"), _current_user.get("name") or ""])
    q = "SELECT id,master_id as \"masterId\",master_name as \"masterName\",contract_type as \"contractType\",contract_number as \"contractNumber\",project,start_date as \"startDate\",end_date as \"endDate\" FROM contracts"
    if where:
        q += " WHERE " + " AND ".join(where)
    q += " ORDER BY id DESC"
    cur.execute(q, tuple(params))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/contracts")
def create_contract(c: ContractModel, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    require_project_access(_current_user, c.project)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO contracts (master_id,master_name,contract_type,contract_number,project,start_date,end_date) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (c.masterId,c.masterName,c.contractType,c.contractNumber,c.project,c.startDate,c.endDate))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.delete("/contracts/{id}")
def delete_contract(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM contracts WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/interim-acts")
def get_interim_acts(_current_user: dict = Depends(require_roles(*CONTRACT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    allowed_projects = visible_project_names(_current_user)
    if allowed_projects is not None and not allowed_projects:
        conn.close()
        return []
    where, params = [], []
    if allowed_projects is not None:
        where.append("project = ANY(%s)")
        params.append(allowed_projects)
    if _current_user.get("role") in ("мастер", "субподрядчик"):
        where.append("(master_id=%s OR master_name=%s)")
        params.extend([_current_user.get("id"), _current_user.get("name") or ""])
    q = "SELECT id,master_id as \"masterId\",master_name as \"masterName\",project,period_start as \"periodStart\",period_end as \"periodEnd\",total_amount as \"totalAmount\",paid_amount as \"paidAmount\",contract_id as \"contractId\",status,scan_url as \"scanUrl\" FROM interim_acts"
    if where:
        q += " WHERE " + " AND ".join(where)
    q += " ORDER BY id DESC"
    cur.execute(q, tuple(params))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/interim-acts")
def create_interim_act(a: InterimActModel, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    require_project_access(_current_user, a.project)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO interim_acts (master_id,master_name,project,period_start,period_end,total_amount,paid_amount,contract_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (a.masterId,a.masterName,a.project,a.periodStart,a.periodEnd,a.totalAmount,a.paidAmount,a.contractId))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/interim-acts/{id}")
def update_interim_act(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "interim_acts", id, _current_user, "project")
    if 'status' in data:
        cur.execute("UPDATE interim_acts SET status=%s WHERE id=%s", (data['status'],id))
    if 'paidAmount' in data:
        cur.execute("UPDATE interim_acts SET paid_amount=%s WHERE id=%s", (data['paidAmount'],id))
    if 'scanUrl' in data:
        cur.execute("UPDATE interim_acts SET scan_url=%s WHERE id=%s", (data['scanUrl'],id))
    conn.close()
    return {"ok": True}

@app.delete("/interim-acts/{id}")
def delete_interim_act(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM interim_acts WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/timesheet/{staff_id}")
def get_timesheet(staff_id: int, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES, "прораб", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT day FROM timesheet WHERE staff_id=%s", (staff_id,))
    rows = cur.fetchall()
    conn.close()
    return {"days": [r['day'] for r in rows]}

@app.post("/timesheet")
def toggle_timesheet(data: TimesheetModel, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES, "прораб", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id FROM timesheet WHERE staff_id=%s AND day=%s", (data.staffId,data.day))
    existing = cur.fetchone()
    if existing:
        cur.execute("DELETE FROM timesheet WHERE staff_id=%s AND day=%s", (data.staffId,data.day))
    else:
        cur.execute("INSERT INTO timesheet (staff_id,day) VALUES (%s,%s)", (data.staffId,data.day))
    conn.close()
    return {"ok": True}

@app.get("/timesheet")
def get_timesheet_all(_current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES, "прораб", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT staff_id, day FROM timesheet")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"staffId": r[0], "day": r[1]} for r in rows]

@app.get("/rooms")
def get_rooms(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    allowed_projects = visible_project_names(current_user)
    select_sql = """SELECT id,project,name,
                           floor_area as "floorArea",
                           wall_area as "wallArea",
                           ceiling_area as "ceilingArea",
                           height,
                           ceiling_type as "ceilingType",
                           wall_material as "wallMaterial",
                           floor_material as "floorMaterial",
                           windows,doors,notes,floor,liter,room_type as "roomType"
                    FROM rooms"""
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(select_sql + " WHERE project = ANY(%s) ORDER BY id", (allowed_projects,))
    else:
        cur.execute(select_sql + " ORDER BY id")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/rooms")
def create_room(r: RoomModel, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    require_project_access(_current_user, r.project)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""INSERT INTO rooms (project,name,floor_area,wall_area,ceiling_area,height,ceiling_type,wall_material,floor_material,windows,doors,notes,floor,liter,room_type)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                   RETURNING id,project,name,floor_area as "floorArea",wall_area as "wallArea",ceiling_area as "ceilingArea",height,ceiling_type as "ceilingType",wall_material as "wallMaterial",floor_material as "floorMaterial",windows,doors,notes,floor,liter,room_type as "roomType" """,
                (r.project,r.name,r.floorArea,r.wallArea,r.ceilingArea,r.height,r.ceilingType,r.wallMaterial,r.floorMaterial,r.windows,r.doors,r.notes,r.floor,r.liter,r.roomType))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/rooms/{id}")
def update_room(id: int, r: RoomModel, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "rooms", id, _current_user, "project")
    require_project_access(_current_user, r.project)
    cur.execute("""UPDATE rooms SET floor=%s,liter=%s,room_type=%s, project=%s,name=%s,
                   floor_area=%s,wall_area=%s,ceiling_area=%s,height=%s,
                   ceiling_type=%s,wall_material=%s,floor_material=%s,
                   windows=%s,doors=%s,notes=%s WHERE id=%s""",
                (r.floor,r.liter,r.roomType,r.project,r.name,r.floorArea,r.wallArea,r.ceilingArea,r.height,r.ceilingType,r.wallMaterial,r.floorMaterial,r.windows,r.doors,r.notes,id))
    conn.close()
    return {"ok": True}

@app.delete("/rooms/{id}")
def delete_room(id: int, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "rooms", id, _current_user, "project")
    cur.execute("DELETE FROM room_windows WHERE room_id=%s", (id,))
    cur.execute("DELETE FROM room_doors WHERE room_id=%s", (id,))
    cur.execute("DELETE FROM room_works WHERE room_id=%s", (id,))
    cur.execute("DELETE FROM rooms WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/room-works")
def get_room_works(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    allowed_projects = visible_project_names(current_user)
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,room_id as \"roomId\",project,room_name as \"roomName\",master_id as \"masterId\",master_name as \"masterName\",description,surface,unit,quantity,price_per_unit as \"pricePerUnit\",total,date,status,photo_url as \"photoUrl\",confirmed_by as \"confirmedBy\" FROM room_works WHERE project = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute("SELECT id,room_id as \"roomId\",project,room_name as \"roomName\",master_id as \"masterId\",master_name as \"masterName\",description,surface,unit,quantity,price_per_unit as \"pricePerUnit\",total,date,status,photo_url as \"photoUrl\",confirmed_by as \"confirmedBy\" FROM room_works ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/room-works")
def create_room_work(w: RoomWorkModel, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    require_project_access(_current_user, w.project)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO room_works (room_id,project,room_name,master_id,master_name,description,surface,unit,quantity,price_per_unit,total,date,photo_url) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (w.roomId,w.project,w.roomName,w.masterId,w.masterName,w.description,w.surface,w.unit,w.quantity,w.pricePerUnit,w.total,w.date,w.photoUrl))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/room-works/{id}")
def update_room_work(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES, "технадзор", "стройконтроль"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "room_works", id, _current_user, "project")
    if 'status' in data:
        cur.execute("UPDATE room_works SET status=%s,confirmed_by=%s WHERE id=%s",
                    (data['status'],data.get('confirmedBy',''),id))
    conn.close()
    return {"ok": True}

AI_FINDING_SELECT = """
SELECT id,
       project_name as "projectName",
       finding_type as "findingType",
       category,
       severity,
       title,
       description,
       source,
       linked_entity_type as "linkedEntityType",
       linked_entity_id as "linkedEntityId",
       suggested_action as "suggestedAction",
       assigned_role as "assignedRole",
       assigned_to as "assignedTo",
       status,
       dedupe_key as "dedupeKey",
       created_by as "createdBy",
       created_by_id as "createdById",
       created_at as "createdAt",
       updated_at as "updatedAt"
FROM ai_findings
"""

AI_TASK_SELECT = """
SELECT id,
       finding_id as "findingId",
       project_name as "projectName",
       title,
       description,
       assigned_role as "assignedRole",
       assigned_to as "assignedTo",
       status,
       due_date as "dueDate",
       accepted_by as "acceptedBy",
       accepted_at as "acceptedAt",
       closed_by as "closedBy",
       closed_at as "closedAt",
       action_label as "actionLabel",
       action_payload as "actionPayload",
       created_at as "createdAt",
       updated_at as "updatedAt"
FROM ai_tasks
"""

def _insert_ai_task(cur, data: dict, current_user: dict):
    project_name = data.get("projectName") or data.get("project_name") or ""
    require_project_access(current_user, project_name)
    cur.execute("""
        INSERT INTO ai_tasks (
            finding_id, project_name, title, description, assigned_role, assigned_to,
            status, due_date, action_label, action_payload, created_at, updated_at
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,NULLIF(%s,'')::date,%s,%s,NOW(),NOW())
        RETURNING id
    """, (
        data.get("findingId") or data.get("finding_id"),
        project_name,
        data.get("title") or "",
        data.get("description") or "",
        data.get("assignedRole") or data.get("assigned_role") or "",
        data.get("assignedTo") or data.get("assigned_to") or "",
        data.get("status") or "Новое",
        data.get("dueDate") or data.get("due_date") or "",
        data.get("actionLabel") or data.get("action_label") or "Принять к исполнению",
        data.get("actionPayload") or data.get("action_payload") or "",
    ))
    row = cur.fetchone()
    return row["id"] if isinstance(row, dict) else row[0]

def _ensure_ai_task_for_finding(cur, finding_id: int, finding: dict, current_user: dict):
    cur.execute("SELECT id FROM ai_tasks WHERE finding_id=%s AND status NOT IN ('Закрыто','Отклонено') LIMIT 1", (finding_id,))
    if cur.fetchone():
        return None
    return _insert_ai_task(cur, {
        "findingId": finding_id,
        "projectName": finding.get("projectName"),
        "title": finding.get("title"),
        "description": finding.get("suggestedAction") or finding.get("description") or "",
        "assignedRole": finding.get("assignedRole") or "прораб",
        "assignedTo": finding.get("assignedTo") or "",
        "status": "Новое",
        "actionLabel": "Принять к исполнению",
        "actionPayload": json.dumps({
            "linkedEntityType": finding.get("linkedEntityType") or "",
            "linkedEntityId": finding.get("linkedEntityId") or "",
            "dedupeKey": finding.get("dedupeKey") or "",
        }, ensure_ascii=False),
    }, current_user)

def _upsert_ai_finding(cur, finding: dict, current_user: dict):
    project_name = finding.get("projectName") or ""
    require_project_access(current_user, project_name)
    dedupe_key = finding.get("dedupeKey") or ""
    if dedupe_key:
        cur.execute("""
            SELECT id FROM ai_findings
            WHERE project_name=%s AND dedupe_key=%s AND status NOT IN ('Закрыто','Отклонено')
            LIMIT 1
        """, (project_name, dedupe_key))
        row = cur.fetchone()
        if row:
            finding_id = row["id"] if isinstance(row, dict) else row[0]
            cur.execute("""
                UPDATE ai_findings
                SET finding_type=%s, category=%s, severity=%s, title=%s, description=%s,
                    source=%s, linked_entity_type=%s, linked_entity_id=%s,
                    suggested_action=%s, assigned_role=%s, assigned_to=%s,
                    updated_at=NOW()
                WHERE id=%s
            """, (
                finding.get("findingType") or "rule",
                finding.get("category") or "Общее",
                finding.get("severity") or "Проверить",
                finding.get("title") or "",
                finding.get("description") or "",
                finding.get("source") or "system_rules",
                finding.get("linkedEntityType") or "",
                finding.get("linkedEntityId") or "",
                finding.get("suggestedAction") or "",
                finding.get("assignedRole") or "",
                finding.get("assignedTo") or "",
                finding_id,
            ))
            _ensure_ai_task_for_finding(cur, finding_id, finding, current_user)
            return finding_id, False
    cur.execute("""
        INSERT INTO ai_findings (
            project_name, finding_type, category, severity, title, description, source,
            linked_entity_type, linked_entity_id, suggested_action, assigned_role,
            assigned_to, status, dedupe_key, created_by, created_by_id, created_at, updated_at
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
        RETURNING id
    """, (
        project_name,
        finding.get("findingType") or "rule",
        finding.get("category") or "Общее",
        finding.get("severity") or "Проверить",
        finding.get("title") or "",
        finding.get("description") or "",
        finding.get("source") or "system_rules",
        finding.get("linkedEntityType") or "",
        finding.get("linkedEntityId") or "",
        finding.get("suggestedAction") or "",
        finding.get("assignedRole") or "",
        finding.get("assignedTo") or "",
        finding.get("status") or "Новое",
        dedupe_key,
        current_user.get("name") or "",
        current_user.get("id"),
    ))
    row = cur.fetchone()
    finding_id = row["id"] if isinstance(row, dict) else row[0]
    _ensure_ai_task_for_finding(cur, finding_id, finding, current_user)
    return finding_id, True

@app.get("/ai-findings")
def list_ai_findings(project_name: str = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute(AI_FINDING_SELECT + " WHERE project_name=%s ORDER BY updated_at DESC, id DESC", (project_name,))
    else:
        allowed_projects = visible_project_names(current_user)
        if allowed_projects is not None:
            if not allowed_projects:
                cur.close(); conn.close()
                return []
            cur.execute(AI_FINDING_SELECT + " WHERE project_name = ANY(%s) ORDER BY updated_at DESC, id DESC", (allowed_projects,))
        else:
            cur.execute(AI_FINDING_SELECT + " ORDER BY updated_at DESC, id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]

@app.post("/ai-findings")
def create_ai_finding(data: AiFindingModel, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES, "мастер", "субподрядчик"))):
    payload = data.dict()
    if not payload.get("projectName"):
        raise HTTPException(status_code=400, detail="projectName required")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    finding_id, _ = _upsert_ai_finding(cur, payload, current_user)
    cur.execute(AI_FINDING_SELECT + " WHERE id=%s", (finding_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return dict(row)

@app.put("/ai-findings/{id}")
def update_ai_finding(id: int, data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    require_row_project_access(cur, "ai_findings", id, current_user, "project_name")
    editable = {
        "findingType": "finding_type",
        "category": "category",
        "severity": "severity",
        "title": "title",
        "description": "description",
        "source": "source",
        "linkedEntityType": "linked_entity_type",
        "linkedEntityId": "linked_entity_id",
        "suggestedAction": "suggested_action",
        "assignedRole": "assigned_role",
        "assignedTo": "assigned_to",
        "status": "status",
        "dedupeKey": "dedupe_key",
    }
    sets, vals = [], []
    for js_key, db_col in editable.items():
        if js_key in data:
            sets.append(db_col + "=%s")
            vals.append(data[js_key])
    if sets:
        vals.append(id)
        cur.execute("UPDATE ai_findings SET " + ", ".join(sets) + ", updated_at=NOW() WHERE id=%s", vals)
        if data.get("status") in ("Закрыто", "Отклонено"):
            cur.execute("UPDATE ai_tasks SET status=%s, closed_by=%s, closed_at=NOW(), updated_at=NOW() WHERE finding_id=%s AND status NOT IN ('Закрыто','Отклонено')",
                        (data.get("status"), current_user.get("name") or "", id))
    cur.execute(AI_FINDING_SELECT + " WHERE id=%s", (id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return dict(row)

@app.post("/ai-findings/generate")
def generate_ai_findings(data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    project_name = data.get("projectName") or data.get("project_name") or data.get("project") or ""
    if not project_name:
        raise HTTPException(status_code=400, detail="projectName required")
    require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT id, project, name, floor_area, wall_area, ceiling_area, height, windows, doors, notes, floor, liter, room_type
                   FROM rooms WHERE project=%s ORDER BY id""", (project_name,))
    rooms_rows = [dict(r) for r in cur.fetchall()]
    room_ids = [r["id"] for r in rooms_rows]
    windows_by_room, doors_by_room = {}, {}
    if room_ids:
        cur.execute("""SELECT w.id,w.room_id,w.name,w.width,w.height,w.reveal_depth
                       FROM room_windows w WHERE w.room_id = ANY(%s) ORDER BY w.id""", (room_ids,))
        for w in cur.fetchall():
            windows_by_room.setdefault(w["room_id"], []).append(dict(w))
        cur.execute("""SELECT d.id,d.room_id,d.name,d.width,d.height,d.reveal_depth
                       FROM room_doors d WHERE d.room_id = ANY(%s) ORDER BY d.id""", (room_ids,))
        for d in cur.fetchall():
            doors_by_room.setdefault(d["room_id"], []).append(dict(d))
    findings = []
    for room in rooms_rows:
        room_title = (room.get("name") or f"Помещение {room.get('id')}")
        room_prefix = f"{room_title}"
        for field, label, action in [
            ("floor_area", "площади пола", "Заполнить площадь пола по проекту или фактическому обмеру."),
            ("wall_area", "площади стен", "Заполнить площадь стен. Окна и двери потом вычтутся из чистой площади."),
            ("ceiling_area", "площади потолка", "Заполнить площадь потолка по проекту или фактическому обмеру."),
            ("height", "высоты помещения", "Заполнить высоту помещения. Она нужна для проверки стен, откосов и расчёта материалов."),
        ]:
            if float(room.get(field) or 0) <= 0:
                findings.append({
                    "projectName": project_name,
                    "findingType": "rule",
                    "category": "Помещения",
                    "severity": "Не хватает данных",
                    "title": f"Нет {label}: {room_prefix}",
                    "description": f"В помещении «{room_title}» не заполнена {label}. Из-за этого система не сможет корректно сверять смету, факт работ и списание материалов.",
                    "source": "system_rules",
                    "linkedEntityType": "room",
                    "linkedEntityId": str(room.get("id")),
                    "suggestedAction": action,
                    "assignedRole": "прораб",
                    "dedupeKey": f"room:{room.get('id')}:{field}",
                })
        for w in windows_by_room.get(room["id"], []):
            win_title = w.get("name") or f"Окно {w.get('id')}"
            if float(w.get("width") or 0) <= 0 or float(w.get("height") or 0) <= 0:
                findings.append({
                    "projectName": project_name,
                    "findingType": "rule",
                    "category": "Помещения",
                    "severity": "Не хватает данных",
                    "title": f"Нет размеров окна: {room_title} / {win_title}",
                    "description": "Без ширины и высоты окна нельзя правильно вычесть проём из площади стен.",
                    "source": "system_rules",
                    "linkedEntityType": "room_window",
                    "linkedEntityId": str(w.get("id")),
                    "suggestedAction": "Заполнить ширину и высоту окна.",
                    "assignedRole": "прораб",
                    "dedupeKey": f"window:{w.get('id')}:size",
                })
            if float(w.get("reveal_depth") or 0) <= 0:
                findings.append({
                    "projectName": project_name,
                    "findingType": "rule",
                    "category": "Помещения",
                    "severity": "Проверить",
                    "title": f"Нет глубины оконного откоса: {room_title} / {win_title}",
                    "description": "Откосы считаются отдельным объёмом. Без глубины нельзя посчитать штукатурку/отделку откосов.",
                    "source": "system_rules",
                    "linkedEntityType": "room_window",
                    "linkedEntityId": str(w.get("id")),
                    "suggestedAction": "Указать глубину оконного откоса.",
                    "assignedRole": "прораб",
                    "dedupeKey": f"window:{w.get('id')}:reveal",
                })
        for d in doors_by_room.get(room["id"], []):
            door_title = d.get("name") or f"Дверь {d.get('id')}"
            if float(d.get("width") or 0) <= 0 or float(d.get("height") or 0) <= 0:
                findings.append({
                    "projectName": project_name,
                    "findingType": "rule",
                    "category": "Помещения",
                    "severity": "Не хватает данных",
                    "title": f"Нет размеров двери: {room_title} / {door_title}",
                    "description": "Без ширины и высоты двери нельзя правильно вычесть проём из площади стен.",
                    "source": "system_rules",
                    "linkedEntityType": "room_door",
                    "linkedEntityId": str(d.get("id")),
                    "suggestedAction": "Заполнить ширину и высоту двери.",
                    "assignedRole": "прораб",
                    "dedupeKey": f"door:{d.get('id')}:size",
                })
            if float(d.get("reveal_depth") or 0) <= 0:
                findings.append({
                    "projectName": project_name,
                    "findingType": "rule",
                    "category": "Помещения",
                    "severity": "Проверить",
                    "title": f"Нет глубины дверного откоса: {room_title} / {door_title}",
                    "description": "Дверные откосы считаются отдельно от стен. Без глубины система не сможет проверить объём.",
                    "source": "system_rules",
                    "linkedEntityType": "room_door",
                    "linkedEntityId": str(d.get("id")),
                    "suggestedAction": "Указать глубину дверного откоса.",
                    "assignedRole": "прораб",
                    "dedupeKey": f"door:{d.get('id')}:reveal",
                })
    cur.execute("""SELECT id, description, unit, quantity, date, master_name
                   FROM work_journal
                   WHERE project=%s AND COALESCE(status,'') <> 'Отклонено'
                   ORDER BY id DESC LIMIT 80""", (project_name,))
    journal_rows = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT description, date FROM room_works WHERE project=%s", (project_name,))
    bound_keys = set((str(r.get("description") or "").strip().lower(), str(r.get("date") or "")) for r in cur.fetchall())
    for wj in journal_rows:
        unit = (wj.get("unit") or "").lower()
        if unit not in ("м2", "м", "шт", "м3"):
            continue
        desc = (wj.get("description") or "").strip()
        if not desc:
            continue
        key = (desc.lower(), str(wj.get("date") or ""))
        if key in bound_keys:
            continue
        findings.append({
            "projectName": project_name,
            "findingType": "rule",
            "category": "ЖПР",
            "severity": "Проверить",
            "title": f"Работа без помещения: {desc[:90]}",
            "description": "Запись журнала работ не привязана к помещению. Из-за этого сложнее сверять факт с обмерами, сметой и материалами.",
            "source": "system_rules",
            "linkedEntityType": "work_journal",
            "linkedEntityId": str(wj.get("id")),
            "suggestedAction": "Уточнить помещение и поверхность для этой работы.",
            "assignedRole": "прораб",
            "dedupeKey": f"work_journal:{wj.get('id')}:room_binding",
        })
    created = 0
    updated = 0
    for finding in findings:
        _, is_created = _upsert_ai_finding(cur, finding, current_user)
        if is_created:
            created += 1
        else:
            updated += 1
    cur.close(); conn.close()
    return {"ok": True, "created": created, "updated": updated, "total": len(findings)}

@app.get("/ai-tasks")
def list_ai_tasks(project_name: str = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute(AI_TASK_SELECT + " WHERE project_name=%s ORDER BY updated_at DESC, id DESC", (project_name,))
    else:
        allowed_projects = visible_project_names(current_user)
        if allowed_projects is not None:
            if not allowed_projects:
                cur.close(); conn.close()
                return []
            cur.execute(AI_TASK_SELECT + " WHERE project_name = ANY(%s) ORDER BY updated_at DESC, id DESC", (allowed_projects,))
        else:
            cur.execute(AI_TASK_SELECT + " ORDER BY updated_at DESC, id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [dict(r) for r in rows]

@app.post("/ai-tasks")
def create_ai_task(data: AiTaskModel, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES, "мастер", "субподрядчик"))):
    payload = data.dict()
    if not payload.get("projectName"):
        raise HTTPException(status_code=400, detail="projectName required")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    task_id = _insert_ai_task(cur, payload, current_user)
    cur.execute(AI_TASK_SELECT + " WHERE id=%s", (task_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return dict(row)

@app.put("/ai-tasks/{id}")
def update_ai_task(id: int, data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    require_row_project_access(cur, "ai_tasks", id, current_user, "project_name")
    editable = {
        "title": "title",
        "description": "description",
        "assignedRole": "assigned_role",
        "assignedTo": "assigned_to",
        "status": "status",
        "dueDate": "due_date",
        "actionLabel": "action_label",
        "actionPayload": "action_payload",
    }
    sets, vals = [], []
    for js_key, db_col in editable.items():
        if js_key in data:
            if js_key == "dueDate":
                sets.append(db_col + "=NULLIF(%s,'')::date")
            else:
                sets.append(db_col + "=%s")
            vals.append(data[js_key] or "")
    if data.get("status") in ("Принято к исполнению", "В работе"):
        sets.append("accepted_by=%s")
        vals.append(current_user.get("name") or "")
        sets.append("accepted_at=COALESCE(accepted_at,NOW())")
    if data.get("status") in ("Закрыто", "Отклонено", "Исправлено"):
        sets.append("closed_by=%s")
        vals.append(current_user.get("name") or "")
        sets.append("closed_at=NOW()")
    if sets:
        vals.append(id)
        cur.execute("UPDATE ai_tasks SET " + ", ".join(sets) + ", updated_at=NOW() WHERE id=%s", vals)
    cur.execute(AI_TASK_SELECT + " WHERE id=%s", (id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    return dict(row)

@app.get("/tools")
def get_tools(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if can_see_all_company_data(current_user) or role in ("кладовщик", "снабженец"):
        cur.execute("SELECT id,name,inventory_number as \"inventoryNumber\",cost,status,location,project,master_id as \"masterId\",master_name as \"masterName\",issue_type as \"issueType\",photo_url as \"photoUrl\",notes FROM tools ORDER BY name")
    elif role == "прораб":
        allowed_projects = user_project_names(current_user)
        cur.execute("""SELECT id,name,inventory_number as "inventoryNumber",cost,status,location,project,master_id as "masterId",master_name as "masterName",issue_type as "issueType",photo_url as "photoUrl",notes
                       FROM tools
                       WHERE COALESCE(project,'')='' OR project = ANY(%s) OR location = ANY(%s)
                       ORDER BY name""", (allowed_projects, allowed_projects))
    elif role in ("мастер", "субподрядчик"):
        cur.execute("""SELECT id,name,inventory_number as "inventoryNumber",cost,status,location,project,master_id as "masterId",master_name as "masterName",issue_type as "issueType",photo_url as "photoUrl",notes
                       FROM tools
                       WHERE master_id=%s OR master_name=%s
                       ORDER BY name""", (current_user.get("id"), current_user.get("name") or ""))
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/tools")
def create_tool(t: ToolModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    if _current_user.get("role") == "прораб" and t.project:
        require_project_access(_current_user, t.project)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO tools (name,inventory_number,cost,status,location,project,master_id,master_name,issue_type,photo_url,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (t.name,t.inventoryNumber,t.cost,t.status,t.location,t.project,t.masterId,t.masterName,t.issueType,t.photoUrl,t.notes))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/tools/{id}")
def update_tool(id: int, t: ToolModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_tool_access(cur, id, _current_user)
    if _current_user.get("role") == "прораб" and t.project:
        require_project_access(_current_user, t.project)
    cur.execute("UPDATE tools SET name=%s,inventory_number=%s,cost=%s,status=%s,location=%s,project=%s,master_id=%s,master_name=%s,issue_type=%s,photo_url=%s,notes=%s WHERE id=%s",
                (t.name,t.inventoryNumber,t.cost,t.status,t.location,t.project,t.masterId,t.masterName,t.issueType,t.photoUrl,t.notes,id))
    conn.close()
    return {"ok": True}

@app.delete("/tools/{id}")
def delete_tool(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_tool_access(cur, id, _current_user)
    cur.execute("DELETE FROM tool_history WHERE tool_id=%s", (id,))
    cur.execute("DELETE FROM tools WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/tool-history")
def get_tool_history(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    role = current_user.get("role")
    if can_see_all_company_data(current_user) or role in ("кладовщик", "снабженец"):
        cur.execute("SELECT id,tool_id as \"toolId\",tool_name as \"toolName\",action,from_location as \"fromLocation\",to_location as \"toLocation\",master_name as \"masterName\",project,issue_type as \"issueType\",condition,date,created_by as \"createdBy\" FROM tool_history ORDER BY id DESC")
    elif role == "прораб":
        allowed_projects = user_project_names(current_user)
        cur.execute("""SELECT id,tool_id as "toolId",tool_name as "toolName",action,from_location as "fromLocation",to_location as "toLocation",master_name as "masterName",project,issue_type as "issueType",condition,date,created_by as "createdBy"
                       FROM tool_history
                       WHERE COALESCE(project,'')='' OR project = ANY(%s)
                       ORDER BY id DESC""", (allowed_projects,))
    elif role in ("мастер", "субподрядчик"):
        cur.execute("""SELECT id,tool_id as "toolId",tool_name as "toolName",action,from_location as "fromLocation",to_location as "toLocation",master_name as "masterName",project,issue_type as "issueType",condition,date,created_by as "createdBy"
                       FROM tool_history
                       WHERE master_name=%s
                       ORDER BY id DESC""", (current_user.get("name") or "",))
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/tool-history")
def create_tool_history(h: ToolHistoryModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    require_tool_access(cur, h.toolId, _current_user)
    if _current_user.get("role") == "прораб" and h.project:
        require_project_access(_current_user, h.project)
    cur.execute("INSERT INTO tool_history (tool_id,tool_name,action,from_location,to_location,master_name,project,issue_type,condition,date,created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (h.toolId,h.toolName,h.action,h.fromLocation,h.toLocation,h.masterName,h.project,h.issueType,h.condition,h.date,h.createdBy))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.get("/inventory")
def get_inventory(current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер", "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if can_see_all_company_data(current_user) or current_user.get("role") in ("кладовщик", "снабженец"):
        cur.execute("SELECT * FROM inventory ORDER BY id DESC")
    else:
        allowed_projects = user_project_names(current_user)
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT * FROM inventory WHERE project = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/inventory")
def create_inventory(inv: InventoryModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    if _current_user.get("role") == "прораб":
        require_project_access(_current_user, inv.project)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("INSERT INTO inventory (project,date,created_by,notes) VALUES (%s,%s,%s,%s) RETURNING *",
                (inv.project,inv.date,inv.createdBy,inv.notes))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.put("/inventory/{id}")
def update_inventory(id: int, data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_inventory_access(cur, id, _current_user)
    if 'status' in data:
        cur.execute("UPDATE inventory SET status=%s WHERE id=%s", (data['status'],id))
    conn.close()
    return {"ok": True}

@app.delete("/inventory/{id}")
def delete_inventory(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_inventory_access(cur, id, _current_user)
    cur.execute("DELETE FROM inventory_items WHERE inventory_id=%s", (id,))
    cur.execute("DELETE FROM inventory WHERE id=%s", (id,))
    conn.close()
    return {"ok": True}

@app.get("/inventory/{id}/items")
def get_inventory_items(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер", "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    require_inventory_access(cur, id, _current_user)
    cur.execute("SELECT * FROM inventory_items WHERE inventory_id=%s", (id,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/inventory-items")
def create_inventory_item(item: InventoryItemModel, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    require_inventory_access(cur, item.inventoryId, _current_user)
    cur.execute("INSERT INTO inventory_items (inventory_id,material_name,unit,expected,actual,difference,notes) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (item.inventoryId,item.materialName,item.unit,item.expected,item.actual,item.difference,item.notes))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.post("/inventory/{id}/items")
def create_inventory_item_for_inventory(id: int, data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    require_inventory_access(cur, id, _current_user)
    cur.execute("INSERT INTO inventory_items (inventory_id,material_name,unit,expected,actual,difference,notes) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                (id,data.get("materialName",""),data.get("unit",""),float(data.get("expected") or 0),float(data.get("actual") or 0),float(data.get("difference") or 0),data.get("notes","")))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.get("/pd-consents")
def get_pd_consents(_current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id,user_id as \"userId\",signed_at as \"signedAt\",scan_url as \"scanUrl\",uploaded_by as \"uploadedBy\" FROM pd_consents")
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/pd-consents")
def create_pd_consent(p: PdConsentModel, _current_user: dict = Depends(get_current_user)):
    if _current_user.get("role") not in STAFF_MANAGE_ROLES and int(p.userId) != int(_current_user.get("id") or 0):
        raise HTTPException(status_code=403, detail="Можно подписать только своё согласие")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        INSERT INTO pd_consents (user_id,signed_at,scan_url,uploaded_by)
        VALUES (%s,%s,%s,%s)
        ON CONFLICT (user_id) DO UPDATE SET
            signed_at=EXCLUDED.signed_at,scan_url=EXCLUDED.scan_url,uploaded_by=EXCLUDED.uploaded_by
        RETURNING id,user_id as "userId",signed_at as "signedAt",scan_url as "scanUrl",uploaded_by as "uploadedBy"
    """, (p.userId,p.signedAt,p.scanUrl,p.uploadedBy))
    row = cur.fetchone()
    conn.close()
    return dict(row)

@app.delete("/pd-consents/{user_id}")
def delete_pd_consent(user_id: int, _current_user: dict = Depends(require_roles(*STAFF_MANAGE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM pd_consents WHERE user_id=%s", (user_id,))
    conn.close()
    return {"ok": True}

@app.post("/upload-photo")
async def upload_photo(file: UploadFile = File(...), _current_user: dict = Depends(get_current_user)):
    ext = os.path.splitext(file.filename)[1]
    filename = str(uuid.uuid4()) + ext
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"url": "/uploads/" + filename}

@app.get("/room-windows")
def get_room_windows(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(current_user)
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("""SELECT w.id,w.room_id,w.name,w.width,w.height,w.window_type,w.reveal_depth,w.reveal_material,w.order_num
                       FROM room_windows w
                       JOIN rooms r ON r.id = w.room_id
                       WHERE r.project = ANY(%s)
                       ORDER BY w.id""", (allowed_projects,))
    else:
        cur.execute("SELECT id,room_id,name,width,height,window_type,reveal_depth,reveal_material,order_num FROM room_windows ORDER BY id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"room_id":r[1],"name":r[2],"width":r[3],"height":r[4],"window_type":r[5],"reveal_depth":r[6],"reveal_material":r[7],"order_num":r[8]} for r in rows]

@app.post("/room-windows")
def create_room_window(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    room_id = data.get('roomId') or data.get('room_id')
    require_room_access(cur, room_id, _current_user)
    cur.execute("INSERT INTO room_windows (room_id,name,width,height,window_type,reveal_depth,reveal_material,order_num) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (room_id,data.get('name',''),float(data.get('width',0)),float(data.get('height',0)),data.get('windowType') or data.get('window_type','ПВХ'),float(data.get('revealDepth') or data.get('reveal_depth') or 0),data.get('revealMaterial') or data.get('reveal_material','Штукатурка'),int(data.get('orderNum') or data.get('order_num') or 0)))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return row

@app.put("/room-windows/{id}")
def update_room_window(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_room_child_access(cur, "room_windows", id, _current_user)
    room_id = data.get('roomId') or data.get('room_id')
    if room_id:
        require_room_access(cur, room_id, _current_user)
    cur.execute("""UPDATE room_windows
                   SET room_id=%s,name=%s,width=%s,height=%s,window_type=%s,reveal_depth=%s,reveal_material=%s,order_num=%s
                   WHERE id=%s""",
        (room_id,data.get('name',''),float(data.get('width',0)),float(data.get('height',0)),data.get('windowType') or data.get('window_type','ПВХ'),float(data.get('revealDepth') or data.get('reveal_depth') or 0),data.get('revealMaterial') or data.get('reveal_material','Штукатурка'),int(data.get('orderNum') or data.get('order_num') or 0),id))
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/room-windows/{id}")
def delete_room_window(id: int, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_room_child_access(cur, "room_windows", id, _current_user)
    cur.execute("DELETE FROM room_windows WHERE id=%s", (id,))
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/room-doors")
def get_room_doors(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(current_user)
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("""SELECT d.id,d.room_id,d.name,d.width,d.height,d.door_type,d.door_purpose,d.reveal_depth,d.reveal_material,d.order_num
                       FROM room_doors d
                       JOIN rooms r ON r.id = d.room_id
                       WHERE r.project = ANY(%s)
                       ORDER BY d.id""", (allowed_projects,))
    else:
        cur.execute("SELECT id,room_id,name,width,height,door_type,door_purpose,reveal_depth,reveal_material,order_num FROM room_doors ORDER BY id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"room_id":r[1],"name":r[2],"width":r[3],"height":r[4],"door_type":r[5],"door_purpose":r[6],"reveal_depth":r[7],"reveal_material":r[8],"order_num":r[9]} for r in rows]

@app.post("/room-doors")
def create_room_door(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    room_id = data.get('roomId') or data.get('room_id')
    require_room_access(cur, room_id, _current_user)
    cur.execute("INSERT INTO room_doors (room_id,name,width,height,door_type,door_purpose,reveal_depth,reveal_material,order_num) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *",
        (room_id,data.get('name',''),float(data.get('width',0)),float(data.get('height',0)),data.get('doorType') or data.get('door_type','Деревянная'),data.get('doorPurpose') or data.get('door_purpose','Межкомнатная'),float(data.get('revealDepth') or data.get('reveal_depth') or 0),data.get('revealMaterial') or data.get('reveal_material','Штукатурка'),int(data.get('orderNum') or data.get('order_num') or 0)))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return row

@app.put("/room-doors/{id}")
def update_room_door(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_room_child_access(cur, "room_doors", id, _current_user)
    room_id = data.get('roomId') or data.get('room_id')
    if room_id:
        require_room_access(cur, room_id, _current_user)
    cur.execute("""UPDATE room_doors
                   SET room_id=%s,name=%s,width=%s,height=%s,door_type=%s,door_purpose=%s,reveal_depth=%s,reveal_material=%s,order_num=%s
                   WHERE id=%s""",
        (room_id,data.get('name',''),float(data.get('width',0)),float(data.get('height',0)),data.get('doorType') or data.get('door_type','Деревянная'),data.get('doorPurpose') or data.get('door_purpose','Межкомнатная'),float(data.get('revealDepth') or data.get('reveal_depth') or 0),data.get('revealMaterial') or data.get('reveal_material','Штукатурка'),int(data.get('orderNum') or data.get('order_num') or 0),id))
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/room-doors/{id}")
def delete_room_door(id: int, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_room_child_access(cur, "room_doors", id, _current_user)
    cur.execute("DELETE FROM room_doors WHERE id=%s", (id,))
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/messages")
def get_messages(_current_user: dict = Depends(get_current_user)):
    import json as _j
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,chat_type,project_id,author_id,author_name,author_role,text,photo_url,created_at,read_by FROM messages ORDER BY created_at ASC LIMIT 200")
    rows = cur.fetchall()
    cur.close(); conn.close()
    out = []
    for r in rows:
        try: rb = r[9] if isinstance(r[9], list) else (_j.loads(r[9]) if r[9] else [])
        except: rb = []
        out.append({"id":r[0],"chat_type":r[1],"project_id":r[2],"author_id":r[3],"author_name":r[4],"author_role":r[5],"text":r[6],"photo_url":r[7],"created_at":str(r[8]),"readBy":rb})
    return out

@app.post("/messages")
def create_message(data: dict, current_user: dict = Depends(get_current_user)):
    import json as _j
    conn = get_db()
    cur = conn.cursor()
    # Автор сразу попадает в read_by
    author_id = current_user.get("id")
    read_by = _j.dumps([author_id] if author_id else [])
    cur.execute("INSERT INTO messages (chat_type,project_id,author_id,author_name,author_role,text,photo_url,read_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s::jsonb) RETURNING *",
        (data.get('chatType','company'),data.get('projectId'),author_id,current_user.get("name",""),current_user.get("role",""),data.get('text',''),data.get('photoUrl',''),read_by))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return row

@app.post("/messages/mark-read")
def mark_messages_read(data: dict, current_user: dict = Depends(get_current_user)):
    """Помечает все сообщения как прочитанные данным пользователем.
       data: {userId, chatType?, projectId?} — фильтры опциональны."""
    user_id = current_user.get("id")
    chat_type = data.get('chatType')
    project_id = data.get('projectId')
    conn = get_db()
    cur = conn.cursor()
    where = ["NOT (read_by @> %s::jsonb)"]
    params = [str(user_id) if isinstance(user_id, int) else __import__('json').dumps([user_id])]
    # подправим: read_by это массив id, ищем где user_id НЕ входит
    import json as _j
    params[0] = _j.dumps([user_id])
    if chat_type:
        where.append("chat_type=%s"); params.append(chat_type)
    if project_id is not None:
        where.append("project_id=%s"); params.append(project_id)
    sql = "UPDATE messages SET read_by = read_by || %s::jsonb WHERE " + " AND ".join(where)
    params.insert(0, _j.dumps([user_id]))
    cur.execute(sql, params)
    updated = cur.rowcount
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "updated": updated}

import urllib.request
import json as json_lib

@app.post("/ai-chat")
def ai_chat(data: dict, current_user: dict = Depends(get_current_user)):
    from openai import OpenAI
    FOLDER_ID = YANDEX_FOLDER_ID
    API_KEY = YANDEX_API_KEY
    messages = data.get("messages", [])
    json_only = data.get("jsonOnly", False)
    skip_context = data.get("skipContext", False) or json_only

    if json_only:
        context = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, никаких ```, никакого текста до или после JSON. Только сам JSON."
    elif skip_context:
        context = ""
    else:
        conn = get_db()
        cur = conn.cursor()
        allowed_projects = visible_project_names(current_user)
        if allowed_projects is None:
            cur.execute("SELECT name, status FROM projects")
        elif allowed_projects:
            cur.execute("SELECT name, status FROM projects WHERE name = ANY(%s)", (allowed_projects,))
        else:
            cur.execute("SELECT name, status FROM projects WHERE FALSE")
        projects = cur.fetchall()
        if current_user.get("role") in WAREHOUSE_ROLES or current_user.get("role") in FINANCE_ROLES:
            cur.execute("SELECT name, quantity, unit FROM warehouse_main")
            materials = cur.fetchall()
        else:
            materials = []
        if can_see_all_company_data(current_user):
            cur.execute("SELECT name, role FROM users WHERE role NOT IN ('заказчик','поставщик')")
            staff = cur.fetchall()
        else:
            staff = [(current_user.get("name",""), current_user.get("role",""))]
        cur.close()
        cur2 = conn.cursor()
        if allowed_projects is None:
            cur2.execute("SELECT name, quantity, unit, project FROM materials WHERE quantity > 0 ORDER BY project")
        elif allowed_projects:
            cur2.execute("SELECT name, quantity, unit, project FROM materials WHERE quantity > 0 AND project = ANY(%s) ORDER BY project", (allowed_projects,))
        else:
            cur2.execute("SELECT name, quantity, unit, project FROM materials WHERE FALSE")
        obj_materials = cur2.fetchall()
        if allowed_projects is None:
            cur2.execute("SELECT project_name, brigade_name, status FROM brigade_contracts")
        elif allowed_projects:
            cur2.execute("SELECT project_name, brigade_name, status FROM brigade_contracts WHERE project_name = ANY(%s)", (allowed_projects,))
        else:
            cur2.execute("SELECT project_name, brigade_name, status FROM brigade_contracts WHERE FALSE")
        brigades = cur2.fetchall()
        if current_user.get("role") in FINANCE_ROLES:
            cur2.execute("SELECT project_name, amount, note FROM project_payments ORDER BY id DESC LIMIT 10")
            payments = cur2.fetchall()
        else:
            payments = []
        cur2.close()
        context = "Ты ИИ помощник строительной компании СтройКа. Отвечай кратко на русском.\n"
        context += "ПРОЕКТЫ ("+str(len(projects))+"): "+", ".join([p[0]+" ("+p[1]+")" for p in projects])+"\n"
        context += "СОТРУДНИКИ: "+", ".join([s[0]+" - "+s[1] for s in staff[:15]])+"\n"
        context += "СКЛАД ОСНОВНОЙ: "+", ".join([m[0]+": "+str(m[1])+" "+m[2] for m in materials[:20]])+"\n"
        context += "МАТЕРИАЛЫ НА ОБЪЕКТАХ: "+", ".join([m[0]+": "+str(m[1])+" "+m[2]+" ("+str(m[3])+")" for m in obj_materials[:20]])+"\n"
        context += "НАРЯДЫ: "+", ".join([b[0]+": "+b[1]+" - "+b[2] for b in brigades[:10]])+"\n"
        context += "ОПЛАТЫ: "+", ".join([pay[0]+": "+str(int(pay[1]))+" руб - "+str(pay[2]) for pay in payments[:10]])+"\n"
    import openai as oa
    client = oa.OpenAI(api_key=API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=FOLDER_ID)
    user_text = messages[-1].get("content","") if messages else ""

    def _call(model_id, max_tokens):
        try:
            r = client.responses.create(
                model="gpt://"+FOLDER_ID+"/"+model_id,
                temperature=0.1 if json_only else 0.2,
                instructions=context,
                input=user_text,
                max_output_tokens=max_tokens,
            )
            return (r.output_text or ""), None
        except Exception as e:
            return "", str(e)

    primary_model = "qwen3.6-35b-a3b/latest" if json_only else "yandexgpt-5.1/latest"
    primary_tokens = 4000 if json_only else 2000
    answer, err = _call(primary_model, primary_tokens)
    if not (answer or "").strip():
        print("AI PRIMARY EMPTY model=" + primary_model + " err=" + str(err))
        fallback_model = "yandexgpt-5.1/latest" if json_only else "qwen3.6-35b-a3b/latest"
        print("AI FALLBACK trying " + fallback_model)
        answer, err = _call(fallback_model, primary_tokens)
        if not (answer or "").strip():
            print("AI FALLBACK ALSO EMPTY err=" + str(err))
            answer = "Ошибка: ИИ вернул пустой ответ. Попробуйте ещё раз или сократите запрос."
    print("AI ANSWER LEN:", len(answer or ""))
    print("AI ANSWER HEAD:", (answer or "")[:200])
    return {"response": answer}


@app.get("/warehouses")
def get_warehouses(_current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер", "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,name,city,address,notes FROM warehouses ORDER BY id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"name":r[1],"city":r[2],"address":r[3],"notes":r[4]} for r in rows]

@app.post("/warehouses")
def create_warehouse(data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO warehouses (name,city,address,notes) VALUES (%s,%s,%s,%s) RETURNING id,name,city,address,notes",
        (data.get("name",""),data.get("city",""),data.get("address",""),data.get("notes","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"name":row[1],"city":row[2],"address":row[3],"notes":row[4]}

@app.put("/warehouses/{id}")
def update_warehouse(id: int, data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE warehouses SET name=%s,city=%s,address=%s,notes=%s WHERE id=%s RETURNING id,name,city,address,notes",
        (data.get("name",""),data.get("city",""),data.get("address",""),data.get("notes",""),id))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"name":row[1],"city":row[2],"address":row[3],"notes":row[4]}

@app.delete("/warehouses/{id}")
def delete_warehouse(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES, "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM warehouses WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/company-requisites")
def get_company_requisites(_current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,full_name,short_name,inn,kpp,ogrn,legal_address,actual_address,phone,email,director_name,director_position,basis,bank_name,bik,rs,ks FROM company_requisites ORDER BY id LIMIT 1")
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row: return {}
    return {"id":row[0],"fullName":row[1],"shortName":row[2],"inn":row[3],"kpp":row[4],"ogrn":row[5],"legalAddress":row[6],"actualAddress":row[7],"phone":row[8],"email":row[9],"directorName":row[10],"directorPosition":row[11],"basis":row[12],"bankName":row[13],"bik":row[14],"rs":row[15],"ks":row[16]}

@app.post("/company-requisites")
def save_company_requisites(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM company_requisites")
    cur.execute("INSERT INTO company_requisites (full_name,short_name,inn,kpp,ogrn,legal_address,actual_address,phone,email,director_name,director_position,basis,bank_name,bik,rs,ks) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("fullName",""),data.get("shortName",""),data.get("inn",""),data.get("kpp",""),data.get("ogrn",""),data.get("legalAddress",""),data.get("actualAddress",""),data.get("phone",""),data.get("email",""),data.get("directorName",""),data.get("directorPosition",""),data.get("basis",""),data.get("bankName",""),data.get("bik",""),data.get("rs",""),data.get("ks","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.get("/company-documents")
def get_company_documents(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in FINANCE_ROLES:
        return []
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,company_id,name,doc_type,file_url,expires_at,uploaded_by FROM company_documents ORDER BY id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"companyId":r[1],"name":r[2],"docType":r[3],"fileUrl":r[4],"expiresAt":r[5],"uploadedBy":r[6]} for r in rows]

@app.post("/company-documents")
def create_company_document(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO company_documents (company_id,name,doc_type,file_url,expires_at,uploaded_by) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("companyId"),data.get("name",""),data.get("docType",""),data.get("fileUrl",""),data.get("expiresAt",""),data.get("uploadedBy","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.delete("/company-documents/{id}")
def delete_company_document(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM company_documents WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/project-stages")
def get_project_stages(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(current_user)
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,project_id,project_name,name,status,start_date,end_date,progress,responsible,notes,order_num FROM project_stages WHERE project_name = ANY(%s) ORDER BY order_num,id", (allowed_projects,))
    else:
        cur.execute("SELECT id,project_id,project_name,name,status,start_date,end_date,progress,responsible,notes,order_num FROM project_stages ORDER BY order_num,id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectId":r[1],"projectName":r[2],"name":r[3],"status":r[4],"startDate":r[5],"endDate":r[6],"progress":r[7],"responsible":r[8],"notes":r[9],"orderNum":r[10]} for r in rows]

@app.post("/project-stages")
def create_project_stage(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    project_name = project_name_from_payload(cur, data)
    require_project_access(_current_user, project_name)
    cur.execute("INSERT INTO project_stages (project_id,project_name,name,status,start_date,end_date,progress,responsible,notes,order_num) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("projectId"),project_name,data.get("name",""),data.get("status","Не начат"),data.get("startDate",""),data.get("endDate",""),int(data.get("progress",0)),data.get("responsible",""),data.get("notes",""),int(data.get("orderNum",0))))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.put("/project-stages/{id}")
def update_project_stage(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_stages", id, _current_user, "project_name")
    cur.execute("UPDATE project_stages SET name=%s,status=%s,start_date=%s,end_date=%s,progress=%s,responsible=%s,notes=%s WHERE id=%s",
        (data.get("name",""),data.get("status",""),data.get("startDate",""),data.get("endDate",""),int(data.get("progress",0)),data.get("responsible",""),data.get("notes",""),id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.delete("/project-stages/{id}")
def delete_project_stage(id: int, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_stages", id, _current_user, "project_name")
    cur.execute("DELETE FROM project_stages WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/project-checklists")
def get_project_checklists(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(current_user)
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,project_id,project_name,name,template,status,created_by,created_at FROM project_checklists WHERE project_name = ANY(%s) ORDER BY id", (allowed_projects,))
    else:
        cur.execute("SELECT id,project_id,project_name,name,template,status,created_by,created_at FROM project_checklists ORDER BY id")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectId":r[1],"projectName":r[2],"name":r[3],"template":r[4],"status":r[5],"createdBy":r[6],"createdAt":str(r[7])} for r in rows]

@app.post("/project-checklists")
def create_project_checklist(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    project_name = project_name_from_payload(cur, data)
    require_project_access(_current_user, project_name)
    cur.execute("INSERT INTO project_checklists (project_id,project_name,name,template,status,created_by,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("projectId"),project_name,data.get("name",""),data.get("template",""),data.get("status","В работе"),data.get("createdBy",""),data.get("createdAt","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.delete("/project-checklists/{id}")
def delete_project_checklist(id: int, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_checklists", id, _current_user, "project_name")
    cur.execute("DELETE FROM checklist_items WHERE checklist_id=%s",(id,))
    cur.execute("DELETE FROM project_checklists WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/checklist-items/{checklist_id}")
def get_checklist_items(checklist_id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_checklist_access(cur, checklist_id, _current_user)
    cur.execute("SELECT id,checklist_id,name,checked,checked_by,checked_at,order_num FROM checklist_items WHERE checklist_id=%s ORDER BY order_num,id",(checklist_id,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"checklistId":r[1],"name":r[2],"checked":r[3],"checkedBy":r[4],"checkedAt":r[5],"orderNum":r[6]} for r in rows]

@app.post("/checklist-items")
def create_checklist_item(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_checklist_access(cur, int(data.get("checklistId") or 0), _current_user)
    cur.execute("INSERT INTO checklist_items (checklist_id,name,checked,checked_by,checked_at,order_num) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("checklistId"),data.get("name",""),data.get("checked",False),data.get("checkedBy",""),data.get("checkedAt",""),int(data.get("orderNum",0))))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.put("/checklist-items/{id}")
def update_checklist_item(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES, "технадзор", "стройконтроль"))):
    conn = get_db()
    cur = conn.cursor()
    require_checklist_item_access(cur, id, _current_user)
    cur.execute("UPDATE checklist_items SET checked=%s,checked_by=%s,checked_at=%s WHERE id=%s",
        (data.get("checked",False),data.get("checkedBy",""),data.get("checkedAt",""),id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/prescriptions")
def get_prescriptions(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(current_user)
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,project_name,number,issued_by,issued_by_role,violation,deadline,responsible,status,photo_url,fix_photo_url,fix_notes FROM prescriptions WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute("SELECT id,project_name,number,issued_by,issued_by_role,violation,deadline,responsible,status,photo_url,fix_photo_url,fix_notes FROM prescriptions ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"number":r[2],"issuedBy":r[3],"issuedByRole":r[4],"violation":r[5],"deadline":r[6],"responsible":r[7],"status":r[8],"photoUrl":r[9],"fixPhotoUrl":r[10],"fixNotes":r[11]} for r in rows]

@app.post("/prescriptions")
def create_prescription(data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES, "заказчик"))):
    project_name = data.get("projectName", "")
    require_project_access(current_user, project_name)
    issued_by = data.get("issuedBy","")
    issued_by_role = data.get("issuedByRole","")
    if current_user.get("role") == "заказчик":
        issued_by = current_user.get("name") or issued_by
        issued_by_role = "Заказчик"
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO prescriptions (project_name,number,issued_by,issued_by_role,violation,deadline,responsible,status,photo_url) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (project_name,data.get("number",""),issued_by,issued_by_role,data.get("violation",""),data.get("deadline",""),data.get("responsible",""),data.get("status","Открыто"),data.get("photoUrl","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.put("/prescriptions/{id}")
def update_prescription(id: int, data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "prescriptions", id, current_user, "project_name")
    role = current_user.get("role")
    new_status = data.get("status","")
    if role in ("мастер", "субподрядчик", "кладовщик", "снабженец") and new_status not in ("На проверке", ""):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Можно только отправить предписание на проверку")
    if role == "заказчик":
        cur.execute("SELECT issued_by, issued_by_role FROM prescriptions WHERE id=%s", (id,))
        row = cur.fetchone()
        if not row or (row[0] != current_user.get("name") and row[1] != "Заказчик"):
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к изменению предписания")
    cur.execute("UPDATE prescriptions SET status=%s,fix_photo_url=%s,fix_notes=%s WHERE id=%s",
        (data.get("status",""),data.get("fixPhotoUrl",""),data.get("fixNotes",""),id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.delete("/prescriptions/{id}")
def delete_prescription(id: int, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "prescriptions", id, current_user, "project_name")
    cur.execute("DELETE FROM prescriptions WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/project-chat/{project_name}")
def get_project_chat(project_name: str, current_user: dict = Depends(get_current_user)):
    require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,project_name,author_id,author_name,author_role,text,photo_url,created_at FROM project_chat WHERE project_name=%s ORDER BY created_at ASC LIMIT 200",(project_name,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"authorId":r[2],"authorName":r[3],"authorRole":r[4],"text":r[5],"photoUrl":r[6],"createdAt":str(r[7])} for r in rows]

@app.post("/project-chat")
def create_project_chat(data: dict, current_user: dict = Depends(get_current_user)):
    project_name = data.get("projectName", "")
    require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO project_chat (project_name,author_id,author_name,author_role,text,photo_url) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id",
        (project_name,current_user.get("id"),current_user.get("name",""),current_user.get("role",""),data.get("text",""),data.get("photoUrl","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

ESTIMATE_CHANGE_APPROVED_STATUSES = ("Утверждено", "Утверждено отдельной допработой")
ESTIMATE_CHANGE_CUSTOMER_STATUSES = ("Ожидает согласования", "Утверждено", "Утверждено отдельной допработой", "Включено в новую смету")

@app.get("/unexpected-works")
def get_unexpected_works(current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(current_user)
    role = current_user.get("role")
    cols = """id,project_name,description,unit,quantity,price,total,added_by,added_by_role,status,
              approved_by,approved_at,notes,photo_url,change_type,estimate_id,section_name,
              estimate_item_name,base_quantity,new_required_quantity,delta_quantity,
              included_in_estimate_id,reason"""
    if allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        if role == "заказчик":
            cur.execute(f"SELECT {cols} FROM unexpected_works WHERE project_name = ANY(%s) AND status = ANY(%s) ORDER BY id DESC", (allowed_projects, list(ESTIMATE_CHANGE_CUSTOMER_STATUSES)))
        else:
            cur.execute(f"SELECT {cols} FROM unexpected_works WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute(f"SELECT {cols} FROM unexpected_works ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"description":r[2],"unit":r[3],
             "quantity":float(r[4] or 0),"price":float(r[5] or 0),"total":float(r[6] or 0),
             "addedBy":r[7],"addedByRole":r[8],"status":r[9],"approvedBy":r[10],
             "approvedAt":r[11],"notes":r[12],"photoUrl":r[13],
             "changeType":r[14] or "Работа вне сметы","estimateId":r[15],
             "sectionName":r[16] or "","estimateItemName":r[17] or "",
             "baseQuantity":float(r[18] or 0),"newRequiredQuantity":float(r[19] or 0),
             "deltaQuantity":float(r[20] or 0),"includedInEstimateId":r[21],
             "reason":r[22] or ""} for r in rows]

@app.post("/unexpected-works")
def create_unexpected_work(data: dict, current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    require_project_access(current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO unexpected_works
                   (project_name,description,unit,quantity,price,total,added_by,added_by_role,status,notes,photo_url,
                    change_type,estimate_id,section_name,estimate_item_name,base_quantity,new_required_quantity,
                    delta_quantity,included_in_estimate_id,reason)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (data.get("projectName",""),data.get("description",""),data.get("unit","шт"),
         float(data.get("quantity",0)),float(data.get("price",0)),float(data.get("total",0)),
         data.get("addedBy",""),data.get("addedByRole",""),data.get("status","Ожидает согласования"),
         data.get("notes",""),data.get("photoUrl",""),
         data.get("changeType") or "Работа вне сметы", data.get("estimateId") or None,
         data.get("sectionName",""), data.get("estimateItemName",""),
         float(data.get("baseQuantity") or 0), float(data.get("newRequiredQuantity") or 0),
         float(data.get("deltaQuantity") or 0), data.get("includedInEstimateId") or None,
         data.get("reason","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.put("/unexpected-works/{id}")
def update_unexpected_work(id: int, data: dict, current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES, *LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    new_status = data.get("status","")
    price = float(data.get("price",0))
    total = float(data.get("total",0))
    # Считываем текущее состояние и описательные поля до апдейта
    cur.execute("SELECT status, project_name, description, unit, quantity, added_by, change_type FROM unexpected_works WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Запись не найдена")
    require_project_access(current_user, row[1] or "")
    old_status = row[0] if row else ""
    cur.execute("""UPDATE unexpected_works SET
                   status=%s, price=%s, total=%s, approved_by=%s, approved_at=%s,
                   included_in_estimate_id=COALESCE(%s,included_in_estimate_id),
                   reason=COALESCE(%s,reason)
                   WHERE id=%s""",
        (new_status, price, total, data.get("approvedBy",""), data.get("approvedAt",""),
         data.get("includedInEstimateId") or None, data.get("reason") if "reason" in data else None, id))
    # Если изменение стало утверждённой отдельной допработой и записи в журнале ещё нет — авто-создаём.
    # Статус «Включено в новую смету» в журнал не пишет, чтобы не задвоить объём.
    auto_journal_id = None
    if row and new_status in ESTIMATE_CHANGE_APPROVED_STATUSES and old_status not in ESTIMATE_CHANGE_APPROVED_STATUSES and (row[6] or "") != "Исключение объёма":
        proj, desc, unit, qty, added_by, change_type = row[1] or "", row[2] or "", row[3] or "шт", float(row[4] or 0), row[5] or "", row[6] or "Работа вне сметы"
        cur.execute("SELECT id FROM work_journal WHERE unexpected_work_id=%s LIMIT 1", (id,))
        existing = cur.fetchone()
        if not existing and desc:
            from datetime import date as _date
            today = _date.today().isoformat()
            try:
                cur.execute("""INSERT INTO work_journal
                               (master_id, master_name, project, description, unit, quantity, price_per_unit, total, date, status, comment,
                                unexpected_work_id)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                            (None, added_by or "(изменение к смете)", proj, desc, unit, qty, price, total, today,
                             "На проверке",
                             "Авто-запись по утверждённому изменению к смете №"+str(id)+" ("+change_type+")",
                             id))
                auto_journal_id = cur.fetchone()[0]
            except Exception as e:
                print("UNEXPECTED→JOURNAL ERROR:", str(e))
    conn.commit()
    cur.close(); conn.close()
    if row and new_status != old_status:
        log_audit(user_name=data.get("approvedBy") or "—", user_role="—",
                  action="status_change", entity_type="unexpected_work", entity_id=id,
                  description="Статус: "+(old_status or "—")+" → "+new_status+", сумма: "+str(total)+" ₽",
                  project_name=row[1] or "")
    return {"ok": True, "journalId": auto_journal_id}

@app.delete("/unexpected-works/{id}")
def delete_unexpected_work(id: int, current_user: dict = Depends(require_roles(*PROJECT_WRITE_ROLES, *LEADERSHIP_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "unexpected_works", id, current_user)
    cur.execute("DELETE FROM unexpected_works WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/parse-smeta")
async def parse_smeta(file: UploadFile = File(...)):
    import tempfile, os
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed"}
    
    contents = await file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(contents)
        tmp_path = tmp.name
    
    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active
        results = []
        current_section = "Без раздела"
        data_start_row = 1
        file_type = "unknown"
        
        for i, row in enumerate(ws.iter_rows(max_row=60, values_only=True)):
            vals = [str(v).strip() for v in row if v is not None]
            row_text = " ".join(vals).lower()
            if "обоснование" in row_text and "наименование" in row_text and ("работ" in row_text or "затрат" in row_text):
                data_start_row = i + 2
                file_type = "lsr"
                break
            elif "наименование" in row_text and "ед" in row_text and "кол" in row_text and "обоснование" not in row_text:
                data_start_row = i + 2
                file_type = "defect"
                break
            elif "обоснование" in row_text and "наименование" in row_text and "общее" in row_text:
                data_start_row = i + 2
                file_type = "vedomost"
                break
        
        work_prefixes = ["ГЭСН", "ФЕР", "ТЕР", "ГЭСНм", "ФЕРм", "ТЕРм", "ГЭСНи", "ФЕРи", "ГЭСНр", "ФЕРр", "ТЕРр"]
        material_prefixes = ["ФСБЦ", "ФССЦ", "ТЦ_", "КАЦ", "МАТ"]
        
        for i, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True)):
            try:
                first_val = str(row[0]).strip() if row[0] is not None else ""
                name_col = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                obosn = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                
                if "Раздел" in first_val or "РАЗДЕЛ" in first_val:
                    current_section = first_val
                    continue
                
                if file_type == "lsr":
                    if "Всего по позиции" in name_col:
                        if results and results[-1].get("total", 0) == 0:
                            try:
                                v = round(float(row[13]), 2) if len(row) > 13 and row[13] and isinstance(row[13], (int,float)) else 0
                                if v > 0:
                                    results[-1]["total"] = v
                                    if results[-1].get("type") == "work":
                                        results[-1]["totalWork"] = v
                                    else:
                                        results[-1]["totalMaterial"] = v
                            except:
                                pass
                        continue
                    
                    num = row[0]
                    if not num:
                        continue
                    try:
                        float(str(num).strip())
                    except:
                        continue
                    
                    if not name_col or len(name_col) < 5:
                        continue
                    if any(x in name_col for x in ["Объем=", "Итого", "ФОТ", "Всего", "Вспомогательные ненормируемые"]):
                        continue
                    if "Пр/" in obosn or "648/" in obosn:
                        continue
                    
                    if any(obosn.startswith(x) for x in work_prefixes):
                        item_type = "work"
                    elif any(obosn.startswith(x) for x in material_prefixes):
                        item_type = "material"
                    else:
                        item_type = "work"
                    
                    unit = str(row[7]).strip() if len(row) > 7 and row[7] else "шт"
                    qty = float(row[8]) if len(row) > 8 and isinstance(row[8], (int,float)) else 0

                    work_total = 0
                    mat_total = 0
                    try:
                        if item_type == "work" and len(row) > 13 and row[13] and isinstance(row[13], (int,float)):
                            work_total = round(float(row[13]), 2)
                    except:
                        pass
                    try:
                        if item_type == "material" and len(row) > 15 and row[15] and isinstance(row[15], (int,float)):
                            mat_total = round(float(row[15]), 2)
                    except:
                        pass

                    results.append({
                        "section": current_section,
                        "name": name_col,
                        "unit": unit,
                        "quantity": round(qty, 4),
                        "total": mat_total if item_type == "material" else work_total,
                        "totalWork": work_total,
                        "totalMaterial": mat_total,
                        "type": item_type
                    })
                
                elif file_type == "defect":
                    if all(v is None or (isinstance(v, (int,float)) and v < 10) for v in row[:5]):
                        continue
                    name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    unit = str(row[2]).strip() if len(row) > 2 and row[2] else "шт"
                    try:
                        qty = float(row[3]) if len(row) > 3 and row[3] and isinstance(row[3], (int,float)) else 0
                    except:
                        qty = 0
                    if name and len(name) > 5 and name not in ["Наименование", "2"] and not all(c.isdigit() or c == " " for c in name):
                        results.append({"section": current_section, "name": name, "unit": unit, "quantity": qty, "total": 0, "type": "work"})
                
                elif file_type == "vedomost":
                    name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    unit = str(row[3]).strip() if len(row) > 3 and row[3] else "шт"
                    try:
                        qty = float(row[4]) if len(row) > 4 and row[4] and isinstance(row[4], (int,float)) else 0
                        total = float(row[6]) if len(row) > 6 and row[6] and isinstance(row[6], (int,float)) else 0
                    except:
                        qty = total = 0
                    if name and len(name) > 5 and name not in ["Наименование", "3"]:
                        results.append({"section": current_section, "name": name, "unit": unit, "quantity": round(qty,4), "total": round(total,2), "type": "material"})
            except:
                continue
        
        os.unlink(tmp_path)
        return {"items": results, "count": len(results)}
    except Exception as e:
        try:
            os.unlink(tmp_path)
        except:
            pass
        return {"error": str(e)}


@app.get("/estimates")
def get_estimates(current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(current_user)
    role = current_user.get("role")
    base_cols = """e.id,e.project_id,e.project_name,e.name,e.version,e.sections_json,
                   COALESCE(e.smeta_type,'Заказчик'),COALESCE(e.work_package,'Основная'),COALESCE(e.is_template,FALSE),e.status,e.created_at,
                   (SELECT COUNT(*) FROM estimate_versions ev WHERE ev.estimate_id=e.id) as version_count,
                   (SELECT MAX(ev.created_at) FROM estimate_versions ev WHERE ev.estimate_id=e.id) as latest_version_at"""
    if role == "заказчик":
        if not allowed_projects:
            cur.execute(f"SELECT {base_cols} FROM estimates e WHERE FALSE")
        else:
            cur.execute(f"SELECT {base_cols} FROM estimates e WHERE e.project_name = ANY(%s) AND e.status='Активная' ORDER BY e.id DESC", (allowed_projects,))
    elif allowed_projects is not None and not allowed_projects:
        cur.execute(f"SELECT {base_cols} FROM estimates e WHERE COALESCE(e.is_template,FALSE)=TRUE ORDER BY e.id DESC")
    elif allowed_projects is None:
        cur.execute(f"SELECT {base_cols} FROM estimates e ORDER BY e.id DESC")
    else:
        cur.execute(f"SELECT {base_cols} FROM estimates e WHERE e.project_name = ANY(%s) OR COALESCE(e.is_template,FALSE)=TRUE ORDER BY e.id DESC", (allowed_projects,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    import json as j
    result = []
    for r in rows:
        try:
            sections = j.loads(r[5]) if r[5] else []
        except:
            sections = []
        result.append({"id":r[0],"projectId":r[1],"projectName":r[2],"name":r[3],"version":r[4],"sections":sections,"smetaType":r[6] or "Заказчик","workPackage":r[7] or "Основная","isTemplate":bool(r[8]),"status":r[9] or "Черновик","createdAt":str(r[10]) if r[10] else "","versionCount":int(r[11] or 0),"latestVersionAt":str(r[12]) if r[12] else ""})
    return result

@app.post("/estimates")
def create_estimate(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    import json as j
    require_project_access(_current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    status = data.get("status") or "Активная"
    smeta_type = data.get("smetaType") or "Заказчик"
    work_package = data.get("workPackage") or data.get("work_package") or "Основная"
    project_name = data.get("projectName","")
    if status == "Активная" and project_name:
        cur.execute("""UPDATE estimates SET status='Архив'
                       WHERE project_name=%s
                         AND COALESCE(smeta_type,'Заказчик')=%s
                         AND COALESCE(work_package,'Основная')=%s""",
                    (project_name, smeta_type, work_package))
    cur.execute("INSERT INTO estimates (project_id,project_name,name,version,sections_json,smeta_type,work_package,status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("projectId"),project_name,data.get("name",""),data.get("version","1.0"),j.dumps(data.get("sections",[]),ensure_ascii=False),smeta_type,work_package,status))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.put("/estimates/{id}")
def update_estimate(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик", "мастер", "субподрядчик"))):
    import json as j
    from datetime import date as _date
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT sections_json, version, project_name, COALESCE(smeta_type,'Заказчик'), status, COALESCE(work_package,'Основная') FROM estimates WHERE id=%s", (id,))
    prev = cur.fetchone()
    if not prev:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    project_name = (prev[2] or "") if prev else ""
    require_project_access(_current_user, project_name)

    old_sections = []
    if prev and prev[0]:
        try:
            old_sections = j.loads(prev[0])
            total = 0
            for s in old_sections:
                for it in (s.get("items") or []):
                    if it.get("isImported"):
                        total += float(it.get("priceWork") or 0)
                    else:
                        total += float(it.get("quantity") or 0) * (float(it.get("priceWork") or 0) + float(it.get("priceMaterial") or 0))
            cur.execute("INSERT INTO estimate_versions (estimate_id, version_label, sections_json, total, comment, created_by) VALUES (%s,%s,%s,%s,%s,%s)",
                (id, prev[1] or "", prev[0], total, data.get("versionComment",""), data.get("updatedBy","")))
        except Exception:
            pass

    new_sections = data.get("sections", []) or []
    for s in new_sections:
        for it in (s.get("items") or []):
            try:
                qty = float(it.get("quantity") or 0)
                done = float(it.get("doneQuantity") or 0)
            except Exception:
                continue
            if qty > 0 and done > qty:
                it["doneQuantity"] = qty
            elif done < 0:
                it["doneQuantity"] = 0
    new_status = data.get("status") or prev[4] or "Черновик"
    new_smeta_type = data.get("smetaType") or prev[3] or "Заказчик"
    new_work_package = data.get("workPackage") or data.get("work_package") or prev[5] or "Основная"
    if new_status == "Активная" and project_name:
        cur.execute("""UPDATE estimates SET status='Архив'
                       WHERE id<>%s
                         AND project_name=%s
                         AND COALESCE(smeta_type,'Заказчик')=%s
                         AND COALESCE(work_package,'Основная')=%s""",
                    (id, project_name, new_smeta_type, new_work_package))
    cur.execute("UPDATE estimates SET name=%s,version=%s,sections_json=%s,smeta_type=%s,work_package=%s,status=%s WHERE id=%s",
        (data.get("name",""),data.get("version","1.0"),j.dumps(new_sections,ensure_ascii=False),new_smeta_type,new_work_package,new_status,id))

    # Build lookup of old done qty by (section name, item name)
    old_done = {}
    for s in old_sections:
        for it in (s.get("items") or []):
            key = (s.get("name",""), it.get("name",""))
            old_done[key] = float(it.get("doneQuantity") or 0)

    today = _date.today().isoformat()
    journal_added = 0
    acts_added = 0
    auto_journal_status = "На проверке" if _current_user.get("role") in ("мастер", "субподрядчик") else "Подтверждено"
    auto_confirmed_by = "" if auto_journal_status != "Подтверждено" else (_current_user.get("name") or "")
    auto_confirmed_at = today if auto_journal_status == "Подтверждено" else None
    work_journal_materials = data.get("_workJournalMaterials") or {}
    def _materials_for_delta(section_idx, item_idx, section_name, item_name):
        keys = [
            str(id) + ":" + str(section_idx) + ":" + str(item_idx),
            str(section_idx) + ":" + str(item_idx),
            str(section_name or "") + "|" + str(item_name or ""),
        ]
        raw = []
        for k in keys:
            if k in work_journal_materials:
                raw = work_journal_materials.get(k) or []
                break
        used = []
        for m in raw if isinstance(raw, list) else []:
            try:
                qty = float(m.get("quantity") or 0)
            except Exception:
                qty = 0
            name = (m.get("name") or "").strip()
            if name and qty > 0:
                used.append({"name": name, "quantity": qty, "unit": m.get("unit") or "шт"})
        return used

    for section_idx, s in enumerate(new_sections):
        for item_idx, it in enumerate(s.get("items") or []):
            new_done = float(it.get("doneQuantity") or 0)
            key = (s.get("name",""), it.get("name",""))
            old_q = old_done.get(key, 0)
            delta = new_done - old_q
            if delta <= 0:
                continue
            brigade = (it.get("brigadeName") or "").strip()
            unit = it.get("unit") or "шт"
            price = float(it.get("priceWork") or 0) + float(it.get("priceMaterial") or 0)
            used_materials = _materials_for_delta(section_idx, item_idx, s.get("name",""), it.get("name",""))
            try:
                for m in used_materials:
                    _apply_material_work_writeoff(cur, project_name, m, _current_user, today, brigade)
                materials_json = j.dumps(used_materials, ensure_ascii=False) if used_materials else None
                cur.execute("""INSERT INTO work_journal
                               (master_id, master_name, project, description, unit, quantity, price_per_unit, total, date, status, comment,
                                materials_used, estimate_id, section_name, hidden_work, confirmed_by, confirmed_at)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (_current_user.get("id"), _current_user.get("name") or brigade or "(из сметы)", project_name, it.get("name",""), unit, delta, price, round(delta*price,2), today,
                     auto_journal_status,
                     "Авто-запись при изменении doneQuantity по позиции сметы №"+str(id),
                     materials_json, id, s.get("name",""), bool(it.get("hiddenWork")), auto_confirmed_by, auto_confirmed_at))
                journal_added += 1
            except HTTPException:
                raise
            except Exception as e:
                print("AUTO-JOURNAL ERROR:", str(e))

            # If item flagged as hidden work — create draft АОСР (Акт освидетельствования скрытых работ)
            if it.get("hiddenWork"):
                try:
                    # Не плодим дубли: один АОСР на позицию сметы (estimate_id + наименование работы)
                    cur.execute("SELECT COUNT(*) FROM hidden_works_acts WHERE estimate_id=%s AND work_name=%s",
                                (id, it.get("name","")))
                    already = (cur.fetchone()[0] or 0) > 0
                    if not already:
                        cur.execute("SELECT COUNT(*) FROM hidden_works_acts WHERE project_name=%s", (project_name,))
                        next_num = (cur.fetchone()[0] or 0) + 1
                        act_number = "АОСР-"+str(next_num)+"/"+str(id)
                        cur.execute("""INSERT INTO hidden_works_acts
                                       (project_name, estimate_id, act_number, work_name, section_name, brigade,
                                        quantity, unit, price_per_unit, total, work_date, status)
                                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (project_name, id, act_number, it.get("name",""), s.get("name",""),
                             brigade or "", delta, unit, price, round(delta*price,2), today, "Черновик"))
                        acts_added += 1
                except Exception as e:
                    print("AUTO-ACT ERROR:", str(e))

    # Смета — единый источник «Сделано»: синхронизируем выполнение в позиции бригады.
    # Позиция сметы привязана к бригаде через brigadeName (колонка «Кому»),
    # к позиции наряда — по совпадению наименования. Объём капается планом наряда.
    brigade_synced = 0
    try:
        cur.execute("SELECT id, brigade_name FROM brigade_contracts WHERE project_name=%s", (project_name,))
        bc_rows = cur.fetchall() or []
        bc_by_name = {}
        for bc_id, bname in bc_rows:
            bc_by_name.setdefault((bname or "").strip().lower(), []).append(bc_id)
        if bc_by_name:
            for s in new_sections:
                for it in (s.get("items") or []):
                    bn = (it.get("brigadeName") or "").strip().lower()
                    if not bn or bn not in bc_by_name:
                        continue
                    done = float(it.get("doneQuantity") or 0)
                    iname = (it.get("name") or "").strip().lower()
                    if not iname:
                        continue
                    qty = float(it.get("quantity") or 0)
                    for bc_id in bc_by_name[bn]:
                        cur.execute(
                            "UPDATE brigade_contract_items "
                            "SET quantity = CASE WHEN COALESCE(quantity,0)>0 THEN quantity ELSE %s END, "
                            "done_quantity = LEAST(%s, CASE WHEN COALESCE(quantity,0)>0 THEN quantity ELSE %s END) "
                            "WHERE contract_id=%s AND LOWER(TRIM(description))=%s",
                            (qty, done, qty, bc_id, iname))
                        brigade_synced += cur.rowcount or 0
    except Exception as e:
        print("BRIGADE-SYNC ERROR:", str(e))

    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "journalEntries": journal_added, "hiddenWorkActs": acts_added, "brigadeItemsSynced": brigade_synced}

@app.put("/estimates/{id}/status")
def update_estimate_status(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    status = data.get("status") or "Черновик"
    if status not in ("Активная", "Архив", "Черновик"):
        raise HTTPException(status_code=400, detail="Недопустимый статус сметы")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT project_name, COALESCE(smeta_type,'Заказчик'), COALESCE(work_package,'Основная') FROM estimates WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    project_name, smeta_type, work_package = row[0] or "", row[1] or "Заказчик", row[2] or "Основная"
    require_project_access(_current_user, project_name)
    if status == "Активная" and project_name:
        cur.execute("""UPDATE estimates SET status='Архив'
                       WHERE id<>%s
                         AND project_name=%s
                         AND COALESCE(smeta_type,'Заказчик')=%s
                         AND COALESCE(work_package,'Основная')=%s""",
                    (id, project_name, smeta_type, work_package))
    cur.execute("UPDATE estimates SET status=%s WHERE id=%s", (status, id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "status": status}

@app.post("/estimates/{id}/include-changes")
def include_estimate_changes(id: int, data: dict, current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    """Создать новую активную версию сметы с утверждёнными изменениями без задвоения КС."""
    import copy
    import json as j
    import re
    import time
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""SELECT project_id, project_name, name, version, sections_json,
                          COALESCE(smeta_type,'Заказчик'), COALESCE(work_package,'Основная'), status
                   FROM estimates WHERE id=%s""", (id,))
    est = cur.fetchone()
    if not est:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    project_id, project_name, name, version, sections_json, smeta_type, work_package, status = est
    require_project_access(current_user, project_name or "")
    try:
        sections = j.loads(sections_json) if sections_json else []
    except Exception:
        sections = []
    sections = copy.deepcopy(sections)

    change_ids = data.get("changeIds") or data.get("change_ids") or []
    if change_ids:
        change_ids = [int(x) for x in change_ids if str(x).strip().isdigit()]
    cols = """id, description, unit, quantity, price, total, change_type, estimate_id,
              section_name, estimate_item_name, base_quantity, new_required_quantity,
              delta_quantity, reason"""
    if change_ids:
        cur.execute(f"""SELECT {cols} FROM unexpected_works
                       WHERE project_name=%s
                         AND status = ANY(%s)
                         AND included_in_estimate_id IS NULL
                         AND id = ANY(%s)
                       ORDER BY id""", (project_name, list(ESTIMATE_CHANGE_APPROVED_STATUSES), change_ids))
    else:
        cur.execute(f"""SELECT {cols} FROM unexpected_works
                       WHERE project_name=%s
                         AND status = ANY(%s)
                         AND included_in_estimate_id IS NULL
                         AND (estimate_id=%s OR estimate_id IS NULL)
                       ORDER BY id""", (project_name, list(ESTIMATE_CHANGE_APPROVED_STATUSES), id))
    changes = cur.fetchall() or []
    if not changes:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Нет утверждённых изменений для включения")

    def _num(v):
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    def _norm(v):
        return re.sub(r"[^a-zа-я0-9]+", " ", str(v or "").lower().replace("ё", "е")).strip()

    def _next_version(v):
        raw = str(v or "1.0").strip()
        parts = raw.split(".")
        try:
            parts[-1] = str(int(parts[-1]) + 1)
            return ".".join(parts)
        except Exception:
            return raw + "+1"

    def _item_total(it):
        qty = _num(it.get("quantity"))
        work = qty * _num(it.get("priceWork"))
        mat = _num(it.get("priceMaterial")) if it.get("isImported") else qty * _num(it.get("priceMaterial"))
        return work + mat

    def _item_unit_price(it):
        qty = _num(it.get("quantity"))
        return (_item_total(it) / qty) if qty > 0 else (_num(it.get("priceWork")) + _num(it.get("priceMaterial")))

    def _find_item(section_name, item_name):
        sec_key, item_key = _norm(section_name), _norm(item_name)
        fallback = None
        for s in sections:
            s_match = _norm(s.get("name")) == sec_key if sec_key else True
            for it in (s.get("items") or []):
                if _norm(it.get("name")) == item_key:
                    if s_match:
                        return s, it
                    fallback = fallback or (s, it)
        return fallback

    def _changes_section():
        for s in sections:
            if _norm(s.get("name")) == _norm("Изменения к смете"):
                return s
        s = {"id": int(time.time() * 1000), "name": "Изменения к смете", "items": []}
        sections.append(s)
        return s

    not_applied = []
    applied_ids = []
    for idx, r in enumerate(changes):
        ch_id, desc, unit, qty, price, total, change_type, estimate_id, section_name, estimate_item_name, base_qty, new_qty, delta_qty, reason = r
        qty = _num(delta_qty) or _num(qty)
        price = _num(price)
        change_type = change_type or "Работа вне сметы"
        target = _find_item(section_name, estimate_item_name or desc) if estimate_item_name or desc else None
        if change_type in ("Дополнительный объём к строке сметы", "Исключение объёма") and target:
            _s, it = target
            old_qty = _num(it.get("quantity"))
            required = _num(new_qty)
            if required <= 0:
                required = old_qty + qty if change_type == "Дополнительный объём к строке сметы" else max(0, old_qty - qty)
            it["quantity"] = required
            if _num(it.get("doneQuantity")) > required:
                it["doneQuantity"] = required
            if price > 0:
                if _num(it.get("priceWork")) > 0 or not _num(it.get("priceMaterial")):
                    it["priceWork"] = price
                else:
                    it["priceMaterial"] = price
            it["sourceUnexpectedWorkId"] = ch_id
            it["changeType"] = change_type
            applied_ids.append(ch_id)
            continue
        if change_type == "Исключение объёма":
            not_applied.append(desc or estimate_item_name or str(ch_id))
            continue
        section = _changes_section()
        item_id = int(time.time() * 1000) + idx + 1
        section.setdefault("items", []).append({
            "id": item_id,
            "itemType": "work",
            "name": desc or estimate_item_name or "Изменение к смете",
            "unit": unit or "шт",
            "quantity": qty,
            "priceWork": price,
            "priceMaterial": 0,
            "measurementBasis": "manual",
            "sourceUnexpectedWorkId": ch_id,
            "changeType": change_type,
            "reason": reason or "",
        })
        applied_ids.append(ch_id)

    if not_applied:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Не найдены строки для исключения объёма: " + ", ".join(not_applied[:5]))
    if not applied_ids:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Не удалось применить изменения")

    new_version = data.get("version") or _next_version(version)
    new_name = data.get("name") or (name or "Смета") + " — ред. " + str(new_version)
    conn.autocommit = False
    try:
        cur.execute("""UPDATE estimates SET status='Архив'
                       WHERE project_name=%s
                         AND COALESCE(smeta_type,'Заказчик')=%s
                         AND COALESCE(work_package,'Основная')=%s""",
                    (project_name, smeta_type, work_package))
        cur.execute("""INSERT INTO estimates
                       (project_id, project_name, name, version, sections_json, smeta_type, work_package, status)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,'Активная') RETURNING id, created_at""",
                    (project_id, project_name, new_name, new_version, j.dumps(sections, ensure_ascii=False), smeta_type, work_package))
        new_id, created_at = cur.fetchone()
        cur.execute("""UPDATE unexpected_works
                       SET status='Включено в новую смету',
                           included_in_estimate_id=%s,
                           approved_by=COALESCE(NULLIF(approved_by,''), %s),
                           approved_at=COALESCE(approved_at, TO_CHAR(CURRENT_DATE,'YYYY-MM-DD'))
                       WHERE id = ANY(%s)""", (new_id, current_user.get("name") or "", applied_ids))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    cur.close(); conn.close()
    return {"ok": True, "id": new_id, "includedChangeIds": applied_ids,
            "estimate": {"id": new_id, "projectId": project_id, "projectName": project_name,
                         "name": new_name, "version": new_version, "sections": sections,
                         "smetaType": smeta_type, "workPackage": work_package,
                         "status": "Активная", "createdAt": str(created_at) if created_at else ""}}

@app.put("/estimates/{id}/toggle-template")
def toggle_estimate_template(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "сметчик", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "estimates", id, _current_user, "project_name")
    cur.execute("UPDATE estimates SET is_template = NOT COALESCE(is_template,FALSE) WHERE id=%s RETURNING is_template", (id,))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"ok": True, "isTemplate": bool(row[0]) if row else False}

@app.get("/estimates/{id}/versions")
def get_estimate_versions(id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "estimates", id, _current_user, "project_name")
    cur.execute("SELECT id, version_label, total, comment, created_by, created_at FROM estimate_versions WHERE estimate_id=%s ORDER BY id DESC", (id,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"versionLabel":r[1] or "","total":float(r[2] or 0),"comment":r[3] or "","createdBy":r[4] or "","createdAt":str(r[5])} for r in rows]

@app.get("/estimate-version/{version_id}")
def get_estimate_version_detail(version_id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    import json as j
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""SELECT ev.id, ev.estimate_id, ev.version_label, ev.sections_json, ev.total, ev.comment, ev.created_by, ev.created_at, e.project_name
                   FROM estimate_versions ev
                   JOIN estimates e ON e.id = ev.estimate_id
                   WHERE ev.id=%s""", (version_id,))
    r = cur.fetchone()
    if not r:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="version not found")
    require_project_access(_current_user, r[8] or "")
    cur.close(); conn.close()
    try:
        sections = j.loads(r[3]) if r[3] else []
    except:
        sections = []
    return {"id":r[0],"estimateId":r[1],"versionLabel":r[2] or "","sections":sections,"total":float(r[4] or 0),"comment":r[5] or "","createdBy":r[6] or "","createdAt":str(r[7])}

@app.delete("/estimates/{id}")
def delete_estimate(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "сметчик", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "estimates", id, _current_user, "project_name")
    cur.execute("DELETE FROM estimate_versions WHERE estimate_id=%s", (id,))
    cur.execute("DELETE FROM estimates WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/brigade-contracts")
def get_brigade_contracts(project_name: str = None, _current_user: dict = Depends(require_roles(*CONTRACT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    # plan_amount = сумма договора из позиций; done_amount = выполненное к оплате; paid_amount = зафиксированные оплаты
    base = ("SELECT bc.id,bc.project_id,bc.project_name,bc.brigade_name,bc.contractor_type,bc.contractor_id,"
            "bc.total_amount,bc.status,bc.signed_at,bc.notes,bc.created_at,bc.pricelist_id,"
            "COALESCE((SELECT SUM(COALESCE(bci.quantity,0)*COALESCE(bci.price_brigade,0)) FROM brigade_contract_items bci WHERE bci.contract_id=bc.id),0) AS plan_amount,"
            "COALESCE((SELECT SUM(CASE WHEN COALESCE(bci.quantity,0)>0 THEN GREATEST(0, LEAST(COALESCE(bci.done_quantity,0), COALESCE(bci.quantity,0))) * COALESCE(bci.price_brigade,0) ELSE 0 END) FROM brigade_contract_items bci WHERE bci.contract_id=bc.id),0) AS done_amount,"
            "COALESCE((SELECT SUM(COALESCE(bp.amount,0)) FROM brigade_payments bp WHERE bp.contract_id=bc.id),0) AS paid_amount,"
            "bc.act_scan_url "
            "FROM brigade_contracts bc")
    allowed_projects = visible_project_names(_current_user)
    where, params = [], []
    if project_name:
        if allowed_projects is not None and project_name not in allowed_projects:
            cur.close(); conn.close()
            return []
        where.append("bc.project_name=%s")
        params.append(project_name)
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        where.append("bc.project_name = ANY(%s)")
        params.append(allowed_projects)
    if _current_user.get("role") in ("мастер", "субподрядчик"):
        where.append("(bc.contractor_id=%s OR bc.brigade_name=%s)")
        params.extend([_current_user.get("id"), _current_user.get("name") or ""])
    q = base
    if where:
        q += " WHERE " + " AND ".join(where)
    q += " ORDER BY bc.id DESC"
    cur.execute(q, tuple(params))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectId":r[1],"projectName":r[2],"brigadeName":r[3],"contractorType":r[4],"contractorId":r[5],
             "totalAmount":float((r[12] if float(r[12] or 0) > 0 else r[6]) or 0),
             "status":r[7],"signedAt":str(r[8]) if r[8] else "","notes":r[9] or "","createdAt":str(r[10]),
             "pricelistId":r[11],"planAmount":float(r[12] or 0),"doneAmount":float(r[13] or 0),
             "paidAmount":float(r[14] or 0),"actScanUrl":r[15] or ""} for r in rows]

@app.get("/brigade-payments")
def get_brigade_payments(contract_id: int = None, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    if contract_id:
        cur.execute("SELECT id,contract_id,amount,paid_by,paid_date,note,created_at FROM brigade_payments WHERE contract_id=%s ORDER BY id DESC", (contract_id,))
    else:
        cur.execute("SELECT id,contract_id,amount,paid_by,paid_date,note,created_at FROM brigade_payments ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"contractId":r[1],"amount":float(r[2] or 0),"paidBy":r[3] or "","paidDate":str(r[4]) if r[4] else "","note":r[5] or "","createdAt":str(r[6])} for r in rows]

@app.post("/brigade-payments")
def create_brigade_payment(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COALESCE(act_scan_url,''), project_name, brigade_name FROM brigade_contracts WHERE id=%s", (data.get("contractId"),))
    contract = cur.fetchone()
    if not contract:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Договор бригады не найден")
    if not (contract[0] or "").strip():
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Оплата заблокирована: загрузите скан подписанного акта")
    amount = float(data.get("amount") or 0)
    paid_by = data.get("paidBy","")
    paid_date = data.get("paidDate") or None
    cur.execute("INSERT INTO brigade_payments (contract_id,amount,paid_by,paid_date,note) VALUES (%s,%s,%s,%s,%s) RETURNING id",
        (data.get("contractId"), amount, paid_by, paid_date, data.get("note","")))
    new_id = cur.fetchone()[0]
    project_payment_id = None
    project_name = contract[1] or ""
    brigade_name = contract[2] or ""
    if project_name and amount:
        payment_note = "Оплата бригаде " + brigade_name
        cur.execute("""SELECT id FROM project_payments
                       WHERE project_name=%s AND amount=%s AND COALESCE(note,'')=%s
                         AND date IS NOT DISTINCT FROM %s AND COALESCE(added_by,'')=%s
                       ORDER BY id DESC LIMIT 1""",
                    (project_name, amount, payment_note, paid_date, paid_by))
        existing = cur.fetchone()
        if existing:
            project_payment_id = existing[0]
        else:
            cur.execute("INSERT INTO project_payments (project_name,amount,note,date,added_by) VALUES (%s,%s,%s,%s,%s) RETURNING id",
                        (project_name, amount, payment_note, paid_date, paid_by))
            project_payment_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id, "projectPaymentId": project_payment_id}

@app.delete("/brigade-payments/{id}")
def delete_brigade_payment(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""SELECT bp.amount, bp.paid_by, bp.paid_date, bc.project_name, bc.brigade_name
                   FROM brigade_payments bp
                   LEFT JOIN brigade_contracts bc ON bc.id=bp.contract_id
                   WHERE bp.id=%s""", (id,))
    payment = cur.fetchone()
    cur.execute("DELETE FROM brigade_payments WHERE id=%s",(id,))
    if payment and payment[3]:
        cur.execute("""DELETE FROM project_payments
                       WHERE id = (
                         SELECT id FROM project_payments
                         WHERE project_name=%s AND amount=%s AND COALESCE(note,'')=%s
                           AND date IS NOT DISTINCT FROM %s AND COALESCE(added_by,'')=%s
                         ORDER BY id DESC LIMIT 1
                       )""",
                    (payment[3], payment[0], "Оплата бригаде " + (payment[4] or ""), payment[2], payment[1] or ""))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

# --- Реестр документов объекта (договоры, акты, доп.соглашения, журналы — со сканами) ---
@app.get("/project-documents")
def get_project_documents(project_name: str = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(_current_user)
    side_filter = None
    if _current_user.get("role") == "заказчик":
        side_filter = "customer"
    elif _current_user.get("role") in ("мастер", "субподрядчик"):
        side_filter = "contractor"
    side_sql = " AND side=%s" if side_filter else ""
    if project_name:
        if allowed_projects is not None and project_name not in allowed_projects:
            cur.close(); conn.close()
            return []
        params = [project_name]
        if side_filter:
            params.append(side_filter)
        cur.execute("SELECT id,project_name,side,doc_type,number,doc_date,counterparty,sign_status,scan_url,amount,notes,uploaded_by,created_at FROM project_documents WHERE project_name=%s" + side_sql + " ORDER BY id DESC", tuple(params))
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        params = [allowed_projects]
        if side_filter:
            params.append(side_filter)
        cur.execute("SELECT id,project_name,side,doc_type,number,doc_date,counterparty,sign_status,scan_url,amount,notes,uploaded_by,created_at FROM project_documents WHERE project_name = ANY(%s)" + side_sql + " ORDER BY id DESC", tuple(params))
    else:
        params = []
        if side_filter:
            params.append(side_filter)
        cur.execute("SELECT id,project_name,side,doc_type,number,doc_date,counterparty,sign_status,scan_url,amount,notes,uploaded_by,created_at FROM project_documents" + (" WHERE side=%s" if side_filter else "") + " ORDER BY id DESC", tuple(params))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"side":r[2],"docType":r[3] or "","number":r[4] or "","docDate":str(r[5]) if r[5] else "","counterparty":r[6] or "","signStatus":r[7] or "","scanUrl":r[8] or "","amount":float(r[9] or 0),"notes":r[10] or "","uploadedBy":r[11] or "","createdAt":str(r[12])} for r in rows]

@app.post("/project-documents")
def create_project_document(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    require_project_access(_current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO project_documents (project_name,side,doc_type,number,doc_date,counterparty,sign_status,scan_url,amount,notes,uploaded_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("projectName",""), data.get("side","customer"), data.get("docType",""), data.get("number",""),
         data.get("docDate") or None, data.get("counterparty",""), data.get("signStatus","Не подписан"),
         data.get("scanUrl",""), data.get("amount") or 0, data.get("notes",""), data.get("uploadedBy","")))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id}

@app.put("/project-documents/{id}")
def update_project_document(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_documents", id, _current_user)
    fields = {"side":"side","docType":"doc_type","number":"number","docDate":"doc_date","counterparty":"counterparty","signStatus":"sign_status","scanUrl":"scan_url","amount":"amount","notes":"notes"}
    sets, vals = [], []
    for k, col in fields.items():
        if k in data:
            sets.append(col+"=%s")
            vals.append(data.get(k) if data.get(k) != "" or col not in ("doc_date",) else None)
    if sets:
        vals.append(id)
        cur.execute("UPDATE project_documents SET "+",".join(sets)+" WHERE id=%s", tuple(vals))
        conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/project-documents/{id}")
def delete_project_document(id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_documents", id, _current_user)
    cur.execute("DELETE FROM project_documents WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

# --- Проект / обмеры объекта: исходники объёмов для помещений и сметы ---
@app.get("/project-measurements")
def get_project_measurements(project_name: str = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(_current_user)
    select_sql = """SELECT id,project_name,source_type,doc_type,title,file_url,status,rooms_created,notes,
                           uploaded_by,reviewed_by,reviewed_at,created_at
                    FROM project_measurements"""
    if project_name:
        if allowed_projects is not None and project_name not in allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(select_sql + " WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(select_sql + " WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute(select_sql + " ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{
        "id": r[0],
        "projectName": r[1],
        "sourceType": r[2] or "Фактический обмер",
        "docType": r[3] or "Обмер",
        "title": r[4] or "",
        "fileUrl": r[5] or "",
        "status": r[6] or "Черновик",
        "roomsCreated": int(r[7] or 0),
        "notes": r[8] or "",
        "uploadedBy": r[9] or "",
        "reviewedBy": r[10] or "",
        "reviewedAt": str(r[11]) if r[11] else "",
        "createdAt": str(r[12]) if r[12] else "",
    } for r in rows]

@app.post("/project-measurements")
def create_project_measurement(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    project_name = data.get("projectName", "")
    require_project_access(_current_user, project_name)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO project_measurements
                   (project_name,source_type,doc_type,title,file_url,status,rooms_created,notes,uploaded_by)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (project_name, data.get("sourceType","Фактический обмер"), data.get("docType","Обмер"),
         data.get("title",""), data.get("fileUrl",""), data.get("status","Черновик"),
         int(data.get("roomsCreated") or 0), data.get("notes",""), data.get("uploadedBy","")))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id}

@app.put("/project-measurements/{id}")
def update_project_measurement(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_measurements", id, _current_user)
    fields = {
        "sourceType": "source_type",
        "docType": "doc_type",
        "title": "title",
        "fileUrl": "file_url",
        "status": "status",
        "roomsCreated": "rooms_created",
        "notes": "notes",
    }
    sets, vals = [], []
    for k, col in fields.items():
        if k in data:
            sets.append(col + "=%s")
            vals.append(int(data.get(k) or 0) if k == "roomsCreated" else data.get(k, ""))
    if data.get("status") == "Принято":
        sets.append("reviewed_by=%s")
        vals.append(_current_user.get("name") or "")
        sets.append("reviewed_at=NOW()")
    if sets:
        vals.append(id)
        cur.execute("UPDATE project_measurements SET " + ",".join(sets) + " WHERE id=%s", tuple(vals))
        conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/project-measurements/{id}")
def delete_project_measurement(id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_measurements", id, _current_user)
    cur.execute("DELETE FROM measurement_room_drafts WHERE measurement_id=%s", (id,))
    cur.execute("DELETE FROM project_measurements WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

MEASUREMENT_ROOM_DRAFT_SELECT = """SELECT id,measurement_id,project_name,name,floor,liter,room_type,
                                          floor_area,wall_area,ceiling_area,height,windows,doors,
                                          notes,status,created_by,accepted_room_id,created_at
                                   FROM measurement_room_drafts"""

def _room_draft_dict(r):
    return {
        "id": r[0],
        "measurementId": r[1],
        "projectName": r[2] or "",
        "name": r[3] or "",
        "floor": int(r[4] or 1),
        "liter": r[5] or "",
        "roomType": r[6] or "Комната",
        "floorArea": float(r[7] or 0),
        "wallArea": float(r[8] or 0),
        "ceilingArea": float(r[9] or 0),
        "height": float(r[10] or 0),
        "windows": int(r[11] or 0),
        "doors": int(r[12] or 0),
        "notes": r[13] or "",
        "status": r[14] or "Черновик ИИ",
        "createdBy": r[15] or "",
        "acceptedRoomId": r[16],
        "createdAt": str(r[17]) if r[17] else "",
    }

def _clean_ai_json(text: str):
    clean = (text or "").replace("```json", "").replace("```", "").strip()
    if not clean:
        return None
    try:
        return json.loads(clean)
    except Exception:
        start = clean.find("{")
        end = clean.rfind("}")
        if start >= 0 and end > start:
            try:
                return json.loads(clean[start:end+1])
            except Exception:
                return None
    return None

def _fallback_room_drafts_from_text(text: str):
    import re
    rooms = []
    for line in (text or "").splitlines():
        src = line.strip()
        if len(src) < 4:
            continue
        nums = [float(x.replace(",", ".")) for x in re.findall(r"(\d+(?:[,.]\d+)?)\s*(?:м2|м²|кв\.?\s*м|м\b)", src.lower())]
        if not nums:
            continue
        name = re.sub(r"\d+(?:[,.]\d+)?\s*(?:м2|м²|кв\.?\s*м|м\b)", "", src, flags=re.I).strip(" -:;,.")
        if not name:
            name = "Помещение " + str(len(rooms) + 1)
        floor_area = nums[0]
        height = nums[1] if len(nums) > 1 and nums[1] < 6 else 0
        rooms.append({
            "name": name[:120],
            "floor": 1,
            "liter": "",
            "roomType": "Комната",
            "floorArea": floor_area,
            "wallArea": 0,
            "ceilingArea": floor_area,
            "height": height,
            "windows": 0,
            "doors": 0,
            "notes": "Черновик из строки обмера: " + src[:300],
        })
    return rooms[:80]

def _draft_rooms_with_ai(measurement: dict):
    title = measurement.get("title") or ""
    notes = measurement.get("notes") or ""
    source_type = measurement.get("source_type") or ""
    doc_type = measurement.get("doc_type") or ""
    prompt = (
        "Извлеки помещения из строительного проекта или обмера. Верни только JSON без markdown. "
        "Формат: {\"rooms\":[{\"name\":\"\",\"floor\":1,\"liter\":\"\",\"roomType\":\"Комната\","
        "\"floorArea\":0,\"wallArea\":0,\"ceilingArea\":0,\"height\":0,\"windows\":0,\"doors\":0,\"notes\":\"\"}]}. "
        "Если точного значения нет — ставь 0 или пустую строку. Не выдумывай площади. "
        "Оконные и дверные проёмы отдельно не считай в wallArea, только количество если явно есть.\n\n"
        "Тип источника: " + source_type + "\n"
        "Тип документа: " + doc_type + "\n"
        "Название: " + title + "\n"
        "Текст/примечания:\n" + notes
    )
    if not (YANDEX_API_KEY and YANDEX_FOLDER_ID):
        return _fallback_room_drafts_from_text(title + "\n" + notes), "fallback"
    import openai as oa
    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    image_payload = None
    file_url = measurement.get("file_url") or ""
    if file_url:
        local_path = file_url.lstrip("/")
        if local_path.startswith("uploads/") and os.path.exists(local_path):
            ext = os.path.splitext(local_path)[1].lower()
            if ext in (".jpg", ".jpeg", ".png", ".webp"):
                with open(local_path, "rb") as f:
                    image_payload = base64.b64encode(f.read()).decode("utf-8")
    try:
        if image_payload:
            response = client.responses.create(
                model=f"gpt://{YANDEX_FOLDER_ID}/qwen3.6-35b-a3b/latest",
                temperature=0.1,
                instructions="Ты извлекаешь помещения из обмеров и проектов. Верни только валидный JSON.",
                input=[{"role":"user","content":[
                    {"type":"input_image","image_url":"data:image/jpeg;base64,"+image_payload},
                    {"type":"input_text","text":prompt}
                ]}],
                max_output_tokens=2500
            )
        else:
            response = client.responses.create(
                model=f"gpt://{YANDEX_FOLDER_ID}/qwen3.6-35b-a3b/latest",
                temperature=0.1,
                instructions="Ты извлекаешь помещения из обмеров и проектов. Верни только валидный JSON.",
                input=prompt,
                max_output_tokens=2500
            )
        parsed = _clean_ai_json(response.output_text or "")
        rooms = parsed.get("rooms", []) if isinstance(parsed, dict) else []
        if isinstance(rooms, list) and rooms:
            return rooms[:100], "ai"
    except Exception as e:
        print("MEASUREMENT AI DRAFT ERROR:", str(e))
    return _fallback_room_drafts_from_text(title + "\n" + notes), "fallback"

@app.get("/measurement-room-drafts")
def get_measurement_room_drafts(project_name: str = None, measurement_id: int = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(_current_user)
    where, params = [], []
    if project_name:
        require_project_access(_current_user, project_name)
        where.append("project_name=%s")
        params.append(project_name)
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        where.append("project_name = ANY(%s)")
        params.append(allowed_projects)
    if measurement_id:
        where.append("measurement_id=%s")
        params.append(measurement_id)
    q = MEASUREMENT_ROOM_DRAFT_SELECT
    if where:
        q += " WHERE " + " AND ".join(where)
    q += " ORDER BY id DESC"
    cur.execute(q, tuple(params))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [_room_draft_dict(r) for r in rows]

@app.post("/project-measurements/{id}/ai-draft-rooms")
def generate_measurement_room_drafts(id: int, data: dict = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    data = data or {}
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM project_measurements WHERE id=%s", (id,))
    measurement = cur.fetchone()
    if not measurement:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Источник обмера не найден")
    require_project_access(_current_user, measurement.get("project_name") or "")
    rooms, source = _draft_rooms_with_ai(dict(measurement))
    if data.get("replaceExisting"):
        cur.execute("DELETE FROM measurement_room_drafts WHERE measurement_id=%s AND status='Черновик ИИ'", (id,))
    created = 0
    for room in rooms:
        name = (room.get("name") or "").strip()
        if not name:
            continue
        cur.execute("""INSERT INTO measurement_room_drafts
                       (measurement_id,project_name,name,floor,liter,room_type,floor_area,wall_area,ceiling_area,height,windows,doors,notes,status,created_by)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (id, measurement.get("project_name") or "", name[:255], int(room.get("floor") or 1),
             str(room.get("liter") or ""), str(room.get("roomType") or room.get("room_type") or "Комната"),
             float(room.get("floorArea") or room.get("floor_area") or 0),
             float(room.get("wallArea") or room.get("wall_area") or 0),
             float(room.get("ceilingArea") or room.get("ceiling_area") or room.get("floorArea") or room.get("floor_area") or 0),
             float(room.get("height") or 0), int(room.get("windows") or 0), int(room.get("doors") or 0),
             str(room.get("notes") or ""), "Черновик ИИ", _current_user.get("name") or ""))
        created += 1
    if created:
        cur.execute("UPDATE project_measurements SET status=%s WHERE id=%s", ("На проверке", id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "created": created, "source": source}

@app.put("/measurement-room-drafts/{id}")
def update_measurement_room_draft(id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "measurement_room_drafts", id, _current_user, "project_name")
    fields = {
        "name": "name", "floor": "floor", "liter": "liter", "roomType": "room_type",
        "floorArea": "floor_area", "wallArea": "wall_area", "ceilingArea": "ceiling_area",
        "height": "height", "windows": "windows", "doors": "doors", "notes": "notes", "status": "status",
    }
    sets, vals = [], []
    for k, col in fields.items():
        if k in data:
            sets.append(col + "=%s")
            vals.append(data.get(k))
    if sets:
        vals.append(id)
        cur.execute("UPDATE measurement_room_drafts SET " + ",".join(sets) + " WHERE id=%s", tuple(vals))
        conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/measurement-room-drafts/{id}/accept")
def accept_measurement_room_draft(id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "measurement_room_drafts", id, _current_user, "project_name")
    cur.execute("""SELECT measurement_id,project_name,name,floor,liter,room_type,floor_area,wall_area,ceiling_area,height,windows,doors,notes,status,accepted_room_id
                   FROM measurement_room_drafts WHERE id=%s""", (id,))
    d = cur.fetchone()
    if not d:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Черновик не найден")
    if d[14]:
        cur.close(); conn.close()
        return {"ok": True, "roomId": d[14], "alreadyAccepted": True}
    cur.execute("""INSERT INTO rooms (project,name,floor_area,wall_area,ceiling_area,height,ceiling_type,wall_material,floor_material,windows,doors,notes,floor,liter,room_type)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (d[1], d[2], float(d[6] or 0), float(d[7] or 0), float(d[8] or 0), float(d[9] or 0),
         "Простой", "Штукатурка", "Стяжка", int(d[10] or 0), int(d[11] or 0),
         ((d[12] or "") + "\nИсточник: черновик обмера №" + str(id)).strip(), int(d[3] or 1), d[4] or "", d[5] or "Комната"))
    room_id = cur.fetchone()[0]
    cur.execute("UPDATE measurement_room_drafts SET status='Принято', accepted_room_id=%s WHERE id=%s", (room_id, id))
    if d[0]:
        cur.execute("""UPDATE project_measurements
                       SET rooms_created=(SELECT COUNT(*) FROM measurement_room_drafts WHERE measurement_id=%s AND accepted_room_id IS NOT NULL)
                       WHERE id=%s""", (d[0], d[0]))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "roomId": room_id}

@app.delete("/measurement-room-drafts/{id}")
def delete_measurement_room_draft(id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "measurement_room_drafts", id, _current_user, "project_name")
    cur.execute("DELETE FROM measurement_room_drafts WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

# --- Переписка по объекту (письма, уведомления, претензии) ---
@app.get("/project-letters")
def get_project_letters(project_name: str = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(_current_user)
    side_filter = None
    if _current_user.get("role") == "заказчик":
        side_filter = "customer"
    elif _current_user.get("role") in ("мастер", "субподрядчик"):
        side_filter = "contractor"
    side_sql = " AND side=%s" if side_filter else ""
    if project_name:
        if allowed_projects is not None and project_name not in allowed_projects:
            cur.close(); conn.close()
            return []
        params = [project_name]
        if side_filter:
            params.append(side_filter)
        cur.execute("SELECT id,project_name,side,direction,subject,body,counterparty,letter_date,file_url,author,created_at FROM project_letters WHERE project_name=%s" + side_sql + " ORDER BY id DESC", tuple(params))
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        params = [allowed_projects]
        if side_filter:
            params.append(side_filter)
        cur.execute("SELECT id,project_name,side,direction,subject,body,counterparty,letter_date,file_url,author,created_at FROM project_letters WHERE project_name = ANY(%s)" + side_sql + " ORDER BY id DESC", tuple(params))
    else:
        params = []
        if side_filter:
            params.append(side_filter)
        cur.execute("SELECT id,project_name,side,direction,subject,body,counterparty,letter_date,file_url,author,created_at FROM project_letters" + (" WHERE side=%s" if side_filter else "") + " ORDER BY id DESC", tuple(params))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"side":r[2],"direction":r[3] or "","subject":r[4] or "","body":r[5] or "","counterparty":r[6] or "","letterDate":str(r[7]) if r[7] else "","fileUrl":r[8] or "","author":r[9] or "","createdAt":str(r[10])} for r in rows]

@app.post("/project-letters")
def create_project_letter(data: dict, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    require_project_access(_current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO project_letters (project_name,side,direction,subject,body,counterparty,letter_date,file_url,author) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("projectName",""), data.get("side","customer"), data.get("direction","outgoing"), data.get("subject",""),
         data.get("body",""), data.get("counterparty",""), data.get("letterDate") or None, data.get("fileUrl",""), data.get("author","")))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id}

@app.delete("/project-letters/{id}")
def delete_project_letter(id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "project_letters", id, _current_user)
    cur.execute("DELETE FROM project_letters WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/salary-payments")
def get_salary_payments(_current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,staff_id,staff_name,month,amount,paid_by,paid_date,note,created_at FROM salary_payments ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"staffId":r[1],"staffName":r[2] or "","month":r[3] or "","amount":float(r[4] or 0),"paidBy":r[5] or "","paidDate":r[6] or "","note":r[7] or "","createdAt":str(r[8])} for r in rows]

@app.post("/salary-payments")
def create_salary_payment(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO salary_payments (staff_id,staff_name,month,amount,paid_by,paid_date,note) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("staffId"), data.get("staffName",""), data.get("month",""), data.get("amount") or 0, data.get("paidBy",""), data.get("paidDate") or "", data.get("note","")))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id}

@app.delete("/salary-payments/{id}")
def delete_salary_payment(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM salary_payments WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/crm-leads")
def get_crm_leads(_current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "менеджер_crm"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,name,phone,email,source,budget,notes,stage,created_by,created_at FROM crm_leads ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"name":r[1] or "","phone":r[2] or "","email":r[3] or "","source":r[4] or "","budget":float(r[5] or 0),"notes":r[6] or "","stage":r[7] or "Новый","createdBy":r[8] or "","createdAt":r[9] or ""} for r in rows]

@app.post("/crm-leads")
def create_crm_lead(data: dict, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "менеджер_crm"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO crm_leads (name,phone,email,source,budget,notes,stage,created_by,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("name",""), data.get("phone",""), data.get("email",""), data.get("source",""), data.get("budget") or 0, data.get("notes",""), data.get("stage","Новый"), data.get("createdBy",""), data.get("createdAt","")))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": new_id}

@app.put("/crm-leads/{id}")
def update_crm_lead(id: int, data: dict, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "менеджер_crm"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE crm_leads SET name=%s,phone=%s,email=%s,source=%s,budget=%s,notes=%s,stage=%s WHERE id=%s",
        (data.get("name",""), data.get("phone",""), data.get("email",""), data.get("source",""), data.get("budget") or 0, data.get("notes",""), data.get("stage","Новый"), id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/crm-leads/{id}")
def delete_crm_lead(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "менеджер_crm"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM crm_leads WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/brigade-contracts")
def create_brigade_contract(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    require_project_access(_current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    pricelist_id = data.get("pricelistId") or None
    cur.execute("INSERT INTO brigade_contracts (project_id,project_name,brigade_name,contractor_type,contractor_id,total_amount,status,notes,pricelist_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("projectId") or None,data.get("projectName",""),data.get("brigadeName",""),data.get("contractorType","Бригада"),data.get("contractorId") or None,data.get("totalAmount",0),data.get("status","Черновик"),data.get("notes",""),pricelist_id))
    conn.commit()
    row = cur.fetchone()
    new_id = row[0]
    inserted = 0
    if pricelist_id:
        try:
            pl_id_int = int(pricelist_id)
            cur.execute("SELECT coefficient FROM pricelists WHERE id=%s", (pl_id_int,))
            cr = cur.fetchone()
            coef = float(cr[0] or 1.0) if cr else 1.0
            cur.execute("SELECT name, unit, price, category FROM pricelist_items WHERE pricelist_id=%s AND (item_type IS NULL OR item_type='work')", (pl_id_int,))
            for it in cur.fetchall():
                price = float(it[2] or 0)
                cur.execute("INSERT INTO brigade_contract_items (contract_id, estimate_section, description, unit, quantity, price_smeta, price_brigade, done_quantity) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (new_id, it[3] or "", it[0], it[1] or "шт", 0, price, round(price * coef, 2), 0))
                inserted += 1
            recalc_brigade_contract_total(cur, new_id)
            conn.commit()
        except Exception as e:
            print("AUTO-LOAD FROM PRICELIST ERROR:", str(e))
    cur.close(); conn.close()
    return {"id": new_id, "ok": True, "itemsLoaded": inserted}

@app.post("/brigade-contracts/{contract_id}/load-from-pricelist")
def load_brigade_items_from_pricelist(contract_id: int, with_materials: bool = False, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    import json as _json
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "brigade_contracts", contract_id, _current_user)
    cur.execute("SELECT pricelist_id, project_name FROM brigade_contracts WHERE id=%s", (contract_id,))
    r = cur.fetchone()
    if not r or not r[0]:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="К наряду не привязан прайс-лист")
    pl_id = int(r[0])
    project_name = r[1] or ""

    cur.execute("SELECT coefficient FROM pricelists WHERE id=%s", (pl_id,))
    cr = cur.fetchone()
    coef = float(cr[0] or 1.0) if cr else 1.0

    # Build estimate name -> qty lookup for this project
    estimate_lookup = {}
    if project_name:
        cur.execute("""SELECT sections_json FROM estimates
                       WHERE project_name=%s
                         AND COALESCE(smeta_type,'Заказчик')='Заказчик'
                         AND status='Активная'
                       ORDER BY COALESCE(work_package,'Основная'), id DESC""", (project_name,))
        est_rows = cur.fetchall()
        if not est_rows:
            cur.execute("SELECT sections_json FROM estimates WHERE project_name=%s ORDER BY id DESC LIMIT 1", (project_name,))
            est_rows = cur.fetchall()
        for est_row in est_rows:
            if est_row and est_row[0]:
                try:
                    sections = _json.loads(est_row[0])
                    for s in sections:
                        for it in (s.get("items") or []):
                            nm = (it.get("name") or "").strip().lower()
                            if nm:
                                estimate_lookup[nm] = estimate_lookup.get(nm, 0) + float(it.get("quantity") or 0)
                except Exception:
                    pass

    cur.execute("SELECT description FROM brigade_contract_items WHERE contract_id=%s", (contract_id,))
    existing_names = {row[0] for row in cur.fetchall()}
    if with_materials:
        cur.execute("SELECT name, unit, price, category, item_type FROM pricelist_items WHERE pricelist_id=%s", (pl_id,))
    else:
        cur.execute("SELECT name, unit, price, category, item_type FROM pricelist_items WHERE pricelist_id=%s AND (item_type IS NULL OR item_type='work')", (pl_id,))
    rows = cur.fetchall()
    inserted = 0
    matched = 0
    for it in rows:
        if it[0] in existing_names:
            continue
        price = float(it[2] or 0)
        nm_low = (it[0] or "").strip().lower()
        # Try exact match first
        qty = estimate_lookup.get(nm_low, 0)
        # Try substring match if exact fails
        if not qty and estimate_lookup:
            for est_nm, est_qty in estimate_lookup.items():
                if est_nm and (nm_low in est_nm or est_nm in nm_low):
                    qty = est_qty
                    break
        if qty:
            matched += 1
        cur.execute("INSERT INTO brigade_contract_items (contract_id, estimate_section, description, unit, quantity, price_smeta, price_brigade, done_quantity) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (contract_id, it[3] or "", it[0], it[1] or "шт", qty, price, round(price * coef, 2), 0))
        inserted += 1
    recalc_brigade_contract_total(cur, contract_id)
    conn.commit(); cur.close(); conn.close()
    return {"ok": True, "itemsLoaded": inserted, "matchedFromEstimate": matched}

@app.put("/brigade-contracts/{id}")
def update_brigade_contract(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "brigade_contracts", id, _current_user)
    cur.execute("UPDATE brigade_contracts SET brigade_name=%s,contractor_type=%s,total_amount=%s,status=%s,signed_at=%s,notes=%s,pricelist_id=%s,act_scan_url=COALESCE(%s,act_scan_url) WHERE id=%s",
        (data.get("brigadeName",""),data.get("contractorType","Бригада"),data.get("totalAmount",0),data.get("status","Черновик"),data.get("signedAt") or None,data.get("notes",""),data.get("pricelistId") or None,data.get("actScanUrl"),id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.delete("/brigade-contracts/{id}")
def delete_brigade_contract(id: int, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "brigade_contracts", id, _current_user)
    cur.execute("""DELETE FROM project_payments pp
                   USING brigade_payments bp, brigade_contracts bc
                   WHERE bp.contract_id=bc.id
                     AND bc.id=%s
                     AND pp.project_name=bc.project_name
                     AND pp.amount=bp.amount
                     AND COALESCE(pp.note,'')='Оплата бригаде ' || COALESCE(bc.brigade_name,'')
                     AND pp.date IS NOT DISTINCT FROM bp.paid_date
                     AND COALESCE(pp.added_by,'')=COALESCE(bp.paid_by,'')""", (id,))
    cur.execute("DELETE FROM brigade_payments WHERE contract_id=%s", (id,))
    cur.execute("DELETE FROM brigade_contract_items WHERE contract_id=%s", (id,))
    cur.execute("DELETE FROM brigade_contracts WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/brigade-contract-items-all")
def list_all_brigade_contract_items(project_name: str = None, _current_user: dict = Depends(require_roles(*CONTRACT_ROLES))):
    """Все позиции нарядов сразу — для подсчёта прогресса по бюджету."""
    conn = get_db()
    cur = conn.cursor()
    allowed_projects = visible_project_names(_current_user)
    where, params = [], []
    if project_name:
        if allowed_projects is not None and project_name not in allowed_projects:
            cur.close(); conn.close()
            return []
        where.append("bc.project_name=%s")
        params.append(project_name)
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        where.append("bc.project_name = ANY(%s)")
        params.append(allowed_projects)
    if _current_user.get("role") in ("мастер", "субподрядчик"):
        where.append("(bc.contractor_id=%s OR bc.brigade_name=%s)")
        params.extend([_current_user.get("id"), _current_user.get("name") or ""])
    q = """SELECT bci.id, bci.contract_id, bci.description, bci.unit, bci.quantity,
                  bci.price_smeta, bci.price_brigade, bci.done_quantity, bci.estimate_section,
                  bc.project_name
           FROM brigade_contract_items bci
           JOIN brigade_contracts bc ON bc.id = bci.contract_id"""
    if where:
        q += " WHERE " + " AND ".join(where)
    q += " ORDER BY bci.id DESC"
    cur.execute(q, tuple(params))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"contractId":r[1],"name":r[2] or "","unit":r[3] or "",
             "quantity":float(r[4] or 0),"priceSmeta":float(r[5] or 0),
             "priceBrigade":float(r[6] or 0),
             "doneQuantity":max(0, min(float(r[7] or 0), float(r[4] or 0))) if float(r[4] or 0) > 0 else 0,
             "rawDoneQuantity":float(r[7] or 0),
             "hasInvalidDoneQuantity":(float(r[4] or 0) <= 0 and float(r[7] or 0) > 0) or float(r[7] or 0) < 0 or (float(r[4] or 0) > 0 and float(r[7] or 0) > float(r[4] or 0)),
             "estimateSection":r[8] or "","projectName":r[9] or ""} for r in rows]

@app.post("/estimates/{estimate_id}/distribute")
def distribute_estimate_to_brigades(estimate_id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    import json as _json
    assignments = data.get("assignments") or []
    default_coef = float(data.get("defaultCoefficient") or 0.6)
    if not assignments:
        raise HTTPException(status_code=400, detail="Нет распределений")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT name, project_id, project_name FROM estimates WHERE id=%s", (estimate_id,))
    est = cur.fetchone()
    if not est:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    estimate_name, project_id, project_name = est[0] or "", est[1], est[2] or ""
    require_project_access(_current_user, project_name)

    # group assignments by brigade
    brigades_map = {}
    for a in assignments:
        bname = (a.get("brigadeName") or "").strip()
        if not bname:
            continue
        key = bname.lower()
        if key not in brigades_map:
            brigades_map[key] = {
                "brigadeName": bname,
                "contractorType": a.get("contractorType") or "Своя бригада",
                "contractorId": a.get("contractorId"),
                "pricelistId": a.get("pricelistId"),
                "items": [],
            }
        brigades_map[key]["items"].append(a)

    # Load pricelists once for coefficient lookup
    cur.execute("SELECT id, coefficient FROM pricelists")
    pl_coef = {r[0]: float(r[1] or 1.0) for r in cur.fetchall()}

    created = []
    for bdata in brigades_map.values():
        # find coefficient
        pl_id = bdata.get("pricelistId")
        try:
            pl_id = int(pl_id) if pl_id else None
        except Exception:
            pl_id = None
        coef = pl_coef.get(pl_id, default_coef) if pl_id else default_coef
        # create brigade_contract
        cur.execute("""INSERT INTO brigade_contracts (project_id, project_name, brigade_name, contractor_type, contractor_id, total_amount, status, notes, pricelist_id)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
            (project_id or None, project_name, bdata["brigadeName"], bdata["contractorType"], bdata.get("contractorId") or None, 0, "Черновик", "Создан из сметы: " + estimate_name, pl_id))
        contract_id = cur.fetchone()[0]
        total = 0
        for it in bdata["items"]:
            qty = float(it.get("quantity") or 0)
            price_smeta = float(it.get("priceSmeta") or it.get("priceWork") or 0)
            price_brigade = round(price_smeta * coef, 2)
            cur.execute("""INSERT INTO brigade_contract_items (contract_id, estimate_section, description, unit, quantity, price_smeta, price_brigade, done_quantity)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                (contract_id, it.get("section",""), it.get("name",""), it.get("unit","шт"), qty, price_smeta, price_brigade, 0))
            total += qty * price_brigade
        cur.execute("UPDATE brigade_contracts SET total_amount=%s WHERE id=%s", (round(total, 2), contract_id))
        created.append({"id": contract_id, "brigadeName": bdata["brigadeName"], "totalAmount": round(total, 2), "itemsCount": len(bdata["items"])})

    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "createdContracts": created}

@app.post("/estimates/{estimate_id}/ai-distribute-suggest")
def ai_suggest_distribution(estimate_id: int, data: dict, _current_user: dict = Depends(require_roles(*CONTRACT_ROLES))):
    import openai as oa
    import json as _json
    brigade_names = data.get("brigadeNames") or []
    if not brigade_names:
        raise HTTPException(status_code=400, detail="Передайте список бригад (brigadeNames)")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT name, sections_json, project_name FROM estimates WHERE id=%s", (estimate_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    require_project_access(_current_user, row[2] or "")
    try:
        sections = _json.loads(row[1]) if row[1] else []
    except Exception:
        sections = []
    cur.close(); conn.close()

    # Flatten items
    items = []
    for s in sections:
        for it in (s.get("items") or []):
            items.append({"section": s.get("name",""), "name": it.get("name",""), "unit": it.get("unit",""), "quantity": it.get("quantity",0)})

    if not items:
        return {"ok": True, "assignments": []}

    instructions = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, ```, только сам JSON."
    prompt = ("Распредели позиции сметы между бригадами по специализации.\n\nБРИГАДЫ:\n" +
              "\n".join("- " + b for b in brigade_names) +
              "\n\nПОЗИЦИИ:\n" +
              "\n".join(str(i+1) + ". [" + it["section"] + "] " + it["name"] for i, it in enumerate(items)) +
              "\n\nОТВЕТЬ JSON формате:\n{\"assignments\": [{\"index\": номер_позиции_с_1, \"brigadeName\": \"имя бригады из списка выше\"}]}\n" +
              "Привязывай позицию к бригаде по специализации. Если непонятно — выбирай 'Общестроительные' или первую общестроительную бригаду из списка.")

    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    raw = ""
    for model_id in ("qwen3.6-35b-a3b/latest", "yandexgpt-5.1/latest"):
        try:
            r = client.responses.create(
                model="gpt://" + YANDEX_FOLDER_ID + "/" + model_id,
                temperature=0.1,
                instructions=instructions,
                input=prompt,
                max_output_tokens=4000,
            )
            raw = (r.output_text or "").strip()
            if raw:
                break
        except Exception as e:
            print("AI-DISTRIBUTE ERROR:", str(e))

    if not raw:
        raise HTTPException(status_code=500, detail="ИИ не ответил")

    import re as _re
    clean = _re.sub(r"^```(?:json)?\s*", "", raw).strip()
    clean = _re.sub(r"\s*```\s*$", "", clean).strip()
    s_idx = clean.find("{"); e_idx = clean.rfind("}")
    parsed = None
    if s_idx >= 0 and e_idx > s_idx:
        try:
            parsed = _json.loads(clean[s_idx:e_idx+1])
        except Exception:
            parsed = None
    if not parsed or not isinstance(parsed.get("assignments"), list):
        raise HTTPException(status_code=500, detail="ИИ вернул не JSON")

    # Map index -> brigade
    result = []
    for a in parsed["assignments"]:
        idx = int(a.get("index", 0)) - 1
        if 0 <= idx < len(items):
            result.append({"itemIndex": idx, "brigadeName": str(a.get("brigadeName") or "")})
    return {"ok": True, "assignments": result, "items": items}

@app.get("/brigade-contract-items/{contract_id}")
def get_brigade_contract_items(contract_id: int, _current_user: dict = Depends(require_roles(*CONTRACT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "brigade_contracts", contract_id, _current_user)
    require_worker_brigade_contract_access(cur, contract_id, _current_user)
    cur.execute("SELECT id,contract_id,estimate_section,description,unit,quantity,price_smeta,price_brigade,done_quantity FROM brigade_contract_items WHERE contract_id=%s ORDER BY id", (contract_id,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    def _status(q, done):
        try:
            q = float(q or 0); done = float(done or 0)
        except Exception:
            return "Не начато"
        if done < 0:
            return "Ошибка объёма"
        if q <= 0 and done > 0:
            return "Нет плана"
        if q > 0 and done > q:
            return "Сверх плана"
        if q > 0 and done >= q:
            return "Выполнено"
        if done > 0:
            return "В работе"
        return "Не начато"
    return [{"id":r[0],"contractId":r[1],"estimateSection":r[2],"name":r[3],"unit":r[4],
             "quantity":float(r[5] or 0),"priceSmeta":float(r[6] or 0),"priceBrigade":float(r[7] or 0),
             "doneQuantity":max(0, min(float(r[8] or 0), float(r[5] or 0))) if float(r[5] or 0) > 0 else 0,
             "rawDoneQuantity":float(r[8] or 0),
             "hasInvalidDoneQuantity":(float(r[5] or 0) <= 0 and float(r[8] or 0) > 0) or float(r[8] or 0) < 0 or (float(r[5] or 0) > 0 and float(r[8] or 0) > float(r[5] or 0)),
             "status":_status(r[5], r[8])} for r in rows]

@app.post("/brigade-contract-items")
def create_brigade_contract_item(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "brigade_contracts", data.get("contractId"), _current_user)
    cur.execute("INSERT INTO brigade_contract_items (contract_id,estimate_section,description,unit,quantity,price_smeta,price_brigade,done_quantity) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("contractId"),data.get("estimateSection",""),data.get("name","") or data.get("description",""),data.get("unit",""),data.get("quantity",0),data.get("priceSmeta",0),data.get("priceBrigade",0),data.get("doneQuantity",0)))
    row = cur.fetchone()
    recalc_brigade_contract_total(cur, data.get("contractId"))
    conn.commit()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.put("/brigade-contract-items/{id}")
def update_brigade_contract_item(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    require_brigade_item_access(cur, id, _current_user)
    quantity = float(data.get("quantity", 0) or 0)
    done_quantity = float(data.get("doneQuantity", 0) or 0)
    done_quantity = max(0, min(done_quantity, quantity)) if quantity > 0 else 0
    cur.execute("UPDATE brigade_contract_items SET quantity=%s,price_brigade=%s,price_smeta=%s,done_quantity=%s WHERE id=%s RETURNING contract_id",
        (quantity,data.get("priceBrigade",0),data.get("priceSmeta",0),done_quantity,id))
    row = cur.fetchone()
    if row:
        recalc_brigade_contract_total(cur, row[0])
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.delete("/brigade-contract-items/{id}")
def delete_brigade_contract_item(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    require_brigade_item_access(cur, id, _current_user)
    cur.execute("DELETE FROM brigade_contract_items WHERE id=%s RETURNING contract_id",(id,))
    row = cur.fetchone()
    if row:
        recalc_brigade_contract_total(cur, row[0])
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/brigade-acts")
def get_brigade_acts(contract_id: int = None, current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    if contract_id:
        require_row_project_access(cur, "brigade_contracts", contract_id, current_user, "project_name")
        cur.execute("SELECT id,contract_id,project_name,brigade_name,period_from,period_to,total_amount,status,created_at FROM brigade_acts WHERE contract_id=%s ORDER BY id DESC", (contract_id,))
    elif visible_project_names(current_user) is not None:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,contract_id,project_name,brigade_name,period_from,period_to,total_amount,status,created_at FROM brigade_acts WHERE project_name = ANY(%s) ORDER BY id DESC", (projects,))
    else:
        cur.execute("SELECT id,contract_id,project_name,brigade_name,period_from,period_to,total_amount,status,created_at FROM brigade_acts ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"contractId":r[1],"projectName":r[2],"brigadeName":r[3],"periodFrom":str(r[4]) if r[4] else "","periodTo":str(r[5]) if r[5] else "","totalAmount":float(r[6] or 0),"status":r[7],"createdAt":str(r[8])} for r in rows]

@app.post("/brigade-acts")
def create_brigade_act(data: dict, current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    require_project_access(current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO brigade_acts (contract_id,project_name,brigade_name,period_from,period_to,total_amount,status) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("contractId"),data.get("projectName",""),data.get("brigadeName",""),data.get("periodFrom") or None,data.get("periodTo") or None,data.get("totalAmount",0),data.get("status","Черновик")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.get("/material-transfers")
def get_material_transfers(project_name: str = None, current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute("SELECT id,project_name,from_location,to_person,to_person_role,material_name,quantity,unit,transfer_date,signed,signed_at,notes,created_by,created_at FROM material_transfers WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif visible_project_names(current_user) is not None:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,project_name,from_location,to_person,to_person_role,material_name,quantity,unit,transfer_date,signed,signed_at,notes,created_by,created_at FROM material_transfers WHERE project_name = ANY(%s) ORDER BY id DESC", (projects,))
    else:
        cur.execute("SELECT id,project_name,from_location,to_person,to_person_role,material_name,quantity,unit,transfer_date,signed,signed_at,notes,created_by,created_at FROM material_transfers ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"fromLocation":r[2],"toPerson":r[3],"toPersonRole":r[4],"materialName":r[5],"quantity":float(r[6] or 0),"unit":r[7],"transferDate":str(r[8]) if r[8] else "","signed":r[9],"signedAt":str(r[10]) if r[10] else "","notes":r[11] or "","createdBy":r[12] or "","createdAt":str(r[13])} for r in rows]

@app.post("/material-transfers")
def create_material_transfer(data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    from_location = data.get("fromLocation", "Основной склад")
    material_name = data.get("materialName", "")
    qty = float(data.get("quantity", 0) or 0)
    if not material_name or qty <= 0:
        raise HTTPException(status_code=400, detail="Укажите материал и количество больше 0")

    conn = get_db()
    conn.autocommit = False
    cur = conn.cursor()
    try:
        if from_location == "Основной склад":
            cur.execute("SELECT id, quantity FROM warehouse_main WHERE name=%s FOR UPDATE", (material_name,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=400, detail="Материал «"+material_name+"» не найден на основном складе")
            stock_id, stock_qty = row[0], float(row[1] or 0)
            if stock_qty < qty:
                raise HTTPException(status_code=400, detail="На основном складе только "+str(stock_qty)+", запрошено "+str(qty))
            cur.execute("UPDATE warehouse_main SET quantity=quantity-%s WHERE id=%s", (qty, stock_id))
        else:
            cur.execute("SELECT id, quantity FROM materials WHERE name=%s AND project=%s FOR UPDATE", (material_name, from_location))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=400, detail="Материал «"+material_name+"» не найден на складе объекта «"+from_location+"»")
            stock_id, stock_qty = row[0], float(row[1] or 0)
            if stock_qty < qty:
                raise HTTPException(status_code=400, detail="На складе «"+from_location+"» только "+str(stock_qty)+", запрошено "+str(qty))
            cur.execute("UPDATE materials SET quantity=quantity-%s WHERE id=%s", (qty, stock_id))

        cur.execute("INSERT INTO material_transfers (project_name,from_location,to_person,to_person_role,material_name,quantity,unit,transfer_date,notes,created_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (data.get("projectName",""), from_location, data.get("toPerson",""), data.get("toPersonRole",""), material_name, qty, data.get("unit","шт"), data.get("transferDate") or None, data.get("notes",""), data.get("createdBy","")))
        new_id = cur.fetchone()[0]

        cur.execute("INSERT INTO warehouse_history (material,type,quantity,date,project,issued_by,date_time) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (material_name, "расход", qty, data.get("transferDate") or None, from_location, data.get("createdBy",""), __import__("datetime").datetime.now().strftime("%d.%m.%Y, %H:%M")))

        conn.commit()
        return {"id": new_id, "ok": True}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

@app.put("/material-transfers/{id}/sign")
def sign_material_transfer(id: int, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE material_transfers SET signed=TRUE,signed_at=NOW() WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.delete("/material-transfers/{id}")
def delete_material_transfer(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM material_transfers WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/supplier-catalog")
def get_supplier_catalog(supplier_id: int = None, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    role = current_user.get("role")
    if role == "поставщик":
        own_supplier_id = current_supplier_id(cur, current_user)
        if not own_supplier_id:
            cur.close(); conn.close()
            return []
        if supplier_id and supplier_id != own_supplier_id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к каталогу этого поставщика")
        cur.execute("SELECT id,supplier_id,supplier_name,material_name,unit,price,min_quantity,delivery_days,in_stock,notes FROM supplier_catalog WHERE supplier_id=%s ORDER BY material_name", (own_supplier_id,))
    elif role in SUPPLY_ROLES or role in WAREHOUSE_ROLES or role in FINANCE_ROLES:
        if supplier_id:
            cur.execute("SELECT id,supplier_id,supplier_name,material_name,unit,price,min_quantity,delivery_days,in_stock,notes FROM supplier_catalog WHERE supplier_id=%s ORDER BY material_name", (supplier_id,))
        else:
            cur.execute("SELECT id,supplier_id,supplier_name,material_name,unit,price,min_quantity,delivery_days,in_stock,notes FROM supplier_catalog ORDER BY material_name")
    else:
        cur.close(); conn.close()
        return []
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"supplierId":r[1],"supplierName":r[2],"materialName":r[3],"unit":r[4],"price":float(r[5] or 0),"minQuantity":float(r[6] or 1),"deliveryDays":r[7] or 3,"inStock":r[8],"notes":r[9] or ""} for r in rows]

@app.post("/supplier-catalog")
def create_supplier_catalog(data: dict, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    supplier_id = int(data.get("supplierId") or 0)
    role = current_user.get("role")
    if role == "поставщик":
        own_supplier_id = current_supplier_id(cur, current_user)
        if not own_supplier_id or supplier_id != own_supplier_id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к каталогу этого поставщика")
    elif role not in ("директор", "зам_директора", "снабженец", "кладовщик"):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    cur.execute("INSERT INTO supplier_catalog (supplier_id,supplier_name,material_name,unit,price,min_quantity,delivery_days,in_stock,notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (supplier_id,data.get("supplierName",""),data.get("materialName",""),data.get("unit","шт"),data.get("price",0),data.get("minQuantity",1),data.get("deliveryDays",3),data.get("inStock",True),data.get("notes","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.put("/supplier-catalog/{id}")
def update_supplier_catalog(id: int, data: dict, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    role = current_user.get("role")
    if role == "поставщик":
        own_supplier_id = current_supplier_id(cur, current_user)
        cur.execute("SELECT supplier_id FROM supplier_catalog WHERE id=%s", (id,))
        row = cur.fetchone()
        if not row or row[0] != own_supplier_id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к каталогу этого поставщика")
    elif role not in ("директор", "зам_директора", "снабженец", "кладовщик"):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    cur.execute("UPDATE supplier_catalog SET price=%s,in_stock=%s,delivery_days=%s,notes=%s WHERE id=%s",
        (data.get("price",0),data.get("inStock",True),data.get("deliveryDays",3),data.get("notes",""),id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.delete("/supplier-catalog/{id}")
def delete_supplier_catalog(id: int, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    role = current_user.get("role")
    if role == "поставщик":
        own_supplier_id = current_supplier_id(cur, current_user)
        cur.execute("SELECT supplier_id FROM supplier_catalog WHERE id=%s", (id,))
        row = cur.fetchone()
        if not row or row[0] != own_supplier_id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к каталогу этого поставщика")
    elif role not in ("директор", "зам_директора", "снабженец", "кладовщик"):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    cur.execute("DELETE FROM supplier_catalog WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

# === Сн.5: история цен на материал ===
@app.get("/material-price-history")
def material_price_history(material: str = "", _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    """Прошлые цены по материалу: из истории поставок + текущие предложения каталога.
    Помогает снабженцу понять адекватную цену при создании заявки."""
    material = (material or "").strip()
    if not material:
        return {"history": [], "catalog": [], "stats": None}
    conn = get_db()
    cur = conn.cursor()
    like = "%" + material + "%"
    # реально закупленные позиции (история поставок)
    cur.execute(
        "SELECT h.material_name, h.price_per_unit, h.unit, h.quantity, h.project, h.date, "
        "COALESCE(s.name, h.confirmed_by, '') AS supplier_name "
        "FROM supply_history h LEFT JOIN suppliers s ON s.id = h.supplier_id "
        "WHERE h.material_name ILIKE %s AND COALESCE(h.price_per_unit,0) > 0 "
        "ORDER BY h.id DESC LIMIT 15", (like,))
    history = [{"materialName": r[0], "price": float(r[1] or 0), "unit": r[2] or "",
               "quantity": float(r[3] or 0), "project": r[4] or "", "date": r[5] or "",
               "supplierName": r[6] or ""} for r in cur.fetchall()]
    # актуальные предложения из каталогов поставщиков
    cur.execute(
        "SELECT material_name, supplier_name, price, unit, delivery_days, in_stock "
        "FROM supplier_catalog WHERE material_name ILIKE %s AND COALESCE(price,0) > 0 "
        "ORDER BY price ASC LIMIT 15", (like,))
    catalog = [{"materialName": r[0], "supplierName": r[1] or "", "price": float(r[2] or 0),
                "unit": r[3] or "", "deliveryDays": r[4] or 3, "inStock": r[5]} for r in cur.fetchall()]
    cur.close(); conn.close()
    prices = [h["price"] for h in history] + [c["price"] for c in catalog]
    stats = None
    if prices:
        stats = {"min": round(min(prices), 2), "max": round(max(prices), 2),
                 "avg": round(sum(prices) / len(prices), 2), "count": len(prices)}
    return {"history": history, "catalog": catalog, "stats": stats}

# === Сн.5: шаблоны заявок на материалы (общие на компанию) ===
@app.get("/supply-request-templates")
def get_supply_request_templates(_current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id,name,category,items_json,created_by,created_by_id,created_at "
                "FROM supply_request_templates ORDER BY name")
    rows = cur.fetchall()
    cur.close(); conn.close()
    import json as _json
    out = []
    for r in rows:
        try:
            items = _json.loads(r[3]) if r[3] else []
        except Exception:
            items = []
        out.append({"id": r[0], "name": r[1] or "", "category": r[2] or "", "items": items,
                    "createdBy": r[4] or "", "createdById": r[5], "createdAt": str(r[6]) if r[6] else ""})
    return out

@app.post("/supply-request-templates")
def create_supply_request_template(data: dict, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    import json as _json
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Укажите название шаблона")
    items = [it for it in (data.get("items") or [])
             if (it or {}).get("materialName") and float((it or {}).get("quantity") or 0) > 0]
    if not items:
        raise HTTPException(status_code=400, detail="Шаблон должен содержать хотя бы одну позицию")
    items = [{"materialName": it["materialName"], "quantity": float(it.get("quantity") or 0),
              "unit": it.get("unit") or "шт"} for it in items]
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO supply_request_templates (name,category,items_json,created_by,created_by_id) "
                "VALUES (%s,%s,%s,%s,%s) RETURNING id",
                (name, data.get("category", ""), _json.dumps(items, ensure_ascii=False),
                 data.get("createdBy", ""), data.get("createdById")))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@app.delete("/supply-request-templates/{id}")
def delete_supply_request_template(id: int, _current_user: dict = Depends(require_roles(*SUPPLY_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM supply_request_templates WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.put("/suppliers/{id}/requisites")
def update_supplier_requisites(id: int, data: dict, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    role = current_user.get("role")
    if role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if supplier_id != id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к этому поставщику")
    elif role not in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    # Расширенный апдейт реквизитов: все поля опциональные
    cur.execute("""UPDATE suppliers SET
        inn=COALESCE(%s, inn), kpp=COALESCE(%s, kpp), ogrn=COALESCE(%s, ogrn),
        legal_address=COALESCE(%s, legal_address),
        actual_address=COALESCE(%s, actual_address),
        bank=COALESCE(%s, bank), bik=COALESCE(%s, bik),
        account=COALESCE(%s, account), kor_account=COALESCE(%s, kor_account),
        director_name=COALESCE(%s, director_name),
        director_position=COALESCE(%s, director_position),
        contract_url=COALESCE(%s, contract_url),
        contract_number=COALESCE(%s, contract_number),
        contract_date=COALESCE(%s, contract_date),
        license_url=COALESCE(%s, license_url),
        price_url=COALESCE(%s, price_url),
        website=COALESCE(%s, website),
        notes=COALESCE(%s, notes),
        phone=COALESCE(%s, phone), email=COALESCE(%s, email),
        category=COALESCE(%s, category), specialization=COALESCE(%s, specialization)
        WHERE id=%s""",
        (data.get("inn"), data.get("kpp"), data.get("ogrn"),
         data.get("legalAddress") or data.get("address"),
         data.get("actualAddress"),
         data.get("bank"), data.get("bik"),
         data.get("account"), data.get("korAccount"),
         data.get("directorName"), data.get("directorPosition"),
         data.get("contractUrl"), data.get("contractNumber"), data.get("contractDate") or None,
         data.get("licenseUrl"), data.get("priceUrl"),
         data.get("website"), data.get("notes"),
         data.get("phone"), data.get("email"),
         data.get("category"), data.get("specialization"),
         id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/supplier-documents")
def list_supplier_documents(supplier_id: int = None, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    where = ""
    params = []
    role = current_user.get("role")
    if role == "поставщик":
        own_supplier_id = current_supplier_id(cur, current_user)
        if not own_supplier_id:
            cur.close(); conn.close()
            return []
        if supplier_id and supplier_id != own_supplier_id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к документам этого поставщика")
        where = " WHERE supplier_id=%s"
        params = [own_supplier_id]
    elif role in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        if supplier_id:
            where = " WHERE supplier_id=%s"
            params = [supplier_id]
    else:
        cur.close(); conn.close()
        return []
    cur.execute("SELECT id, supplier_id, doc_type, title, file_url, status, signed_at, expires_at, notes, uploaded_by, created_at FROM supplier_documents" + where + " ORDER BY created_at DESC", params)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"supplierId":r[1],"docType":r[2] or "","title":r[3] or "",
             "fileUrl":r[4] or "","status":r[5] or "","signedAt":str(r[6]) if r[6] else "",
             "expiresAt":str(r[7]) if r[7] else "","notes":r[8] or "",
             "uploadedBy":r[9] or "","createdAt":str(r[10])} for r in rows]

@app.post("/supplier-documents")
def create_supplier_document(data: dict, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    role = current_user.get("role")
    supplier_id = int(data.get('supplierId') or 0)
    if role == "поставщик":
        own_supplier_id = current_supplier_id(cur, current_user)
        if not own_supplier_id or supplier_id != own_supplier_id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к документам этого поставщика")
    elif role not in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    cur.execute(
        "INSERT INTO supplier_documents (supplier_id, doc_type, title, file_url, status, signed_at, expires_at, notes, uploaded_by) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (supplier_id, data.get('docType') or 'Другое', data.get('title') or '',
         data.get('fileUrl') or '', data.get('status') or 'Загружен',
         data.get('signedAt') or None, data.get('expiresAt') or None,
         data.get('notes') or '', data.get('uploadedBy') or ''))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()
    return {"id": new_id, "ok": True}

@app.delete("/supplier-documents/{id}")
def delete_supplier_document(id: int, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    role = current_user.get("role")
    if role == "поставщик":
        own_supplier_id = current_supplier_id(cur, current_user)
        cur.execute("SELECT supplier_id FROM supplier_documents WHERE id=%s", (id,))
        row = cur.fetchone()
        if not row or row[0] != own_supplier_id:
            cur.close(); conn.close()
            raise HTTPException(status_code=403, detail="Нет доступа к документам этого поставщика")
    elif role not in ("директор", "зам_директора", "снабженец", "кладовщик", "бухгалтер"):
        cur.close(); conn.close()
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    cur.execute("DELETE FROM supplier_documents WHERE id=%s", (id,))
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/warehouse-invoices")
def get_warehouse_invoices(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in WAREHOUSE_ROLES and not can_see_all_company_data(current_user):
        return []
    conn = get_db()
    cur = conn.cursor()
    if current_user.get("role") == "прораб":
        allowed_projects = user_project_names(current_user)
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute("SELECT id,number,date,supplier_id,supplier_name,accepted_by,location,project,vat,items,total_base,total_vat,total_with_vat,status,added_by,photo_url FROM warehouse_invoices WHERE project = ANY(%s) OR location = ANY(%s) ORDER BY id DESC", (allowed_projects, allowed_projects))
    else:
        cur.execute("SELECT id,number,date,supplier_id,supplier_name,accepted_by,location,project,vat,items,total_base,total_vat,total_with_vat,status,added_by,photo_url FROM warehouse_invoices ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    import json as j
    result = []
    for r in rows:
        try: items = j.loads(r[9]) if r[9] else []
        except: items = []
        result.append({"id":r[0],"number":r[1],"date":str(r[2]) if r[2] else "","supplierId":r[3],"supplierName":r[4] or "","acceptedBy":r[5] or "","location":r[6] or "","project":r[7] or "","vat":r[8] or "Без НДС","items":items,"totalBase":float(r[10] or 0),"totalVat":float(r[11] or 0),"totalWithVat":float(r[12] or 0),"status":r[13] or "Принята","addedBy":r[14] or "","photoUrl":r[15] or ""})
    return result

@app.post("/warehouse-invoices")
def create_warehouse_invoice(data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    import json as j
    target_project = data.get("project") or (data.get("location") if data.get("location") != "Основной склад" else "")
    if _current_user.get("role") == "прораб" and target_project:
        require_project_access(_current_user, target_project)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO warehouse_invoices (number,date,supplier_id,supplier_name,accepted_by,location,project,vat,items,total_base,total_vat,total_with_vat,status,added_by,photo_url) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("number",""),data.get("date") or None,data.get("supplierId") or None,data.get("supplierName",""),data.get("acceptedBy",""),data.get("location",""),data.get("project",""),data.get("vat","Без НДС"),j.dumps(data.get("items",[]),ensure_ascii=False),data.get("totalBase",0),data.get("totalVat",0),data.get("totalWithVat",0),data.get("status","Принята"),data.get("addedBy",""),data.get("photoUrl","")))
    invoice_id = cur.fetchone()[0]
    # Авто-создание записей в журнале входного контроля материалов (по СП 48.13330)
    items_list = data.get("items", []) or []
    proj = data.get("project","")
    sup = data.get("supplierName","")
    rcv_date = data.get("date") or None
    inspections_added = 0
    cables_added = 0
    for it in items_list:
        name = (it.get("name") or "").strip()
        if not name:
            continue
        qty = float(it.get("quantity",0) or 0)
        unit = it.get("unit","шт")
        try:
            cur.execute("""INSERT INTO material_inspection_journal
                           (project_name, invoice_id, material_name, unit, quantity, supplier, received_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (proj, invoice_id, name, unit, qty, sup, rcv_date))
            inspections_added += 1
        except Exception as e:
            print("INSPECTION INSERT ERROR:", str(e))
        # Автоопределение: это кабель?
        cable_info = _detect_cable_info(name)
        if cable_info["isCable"]:
            try:
                cur.execute("""INSERT INTO cable_journal
                               (project_name, invoice_id, cable_brand, cable_type, cross_section, cores_count,
                                length_received, supplier, received_at)
                               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (proj, invoice_id, name, cable_info["cableType"], cable_info["section"],
                             cable_info["cores"], qty, sup, rcv_date))
                cables_added += 1
            except Exception as e:
                print("CABLE INSERT ERROR:", str(e))
    conn.commit()
    cur.close(); conn.close()
    return {"id": invoice_id, "ok": True, "inspectionsAdded": inspections_added, "cablesAdded": cables_added}

@app.delete("/warehouse-invoices/{id}")
def delete_warehouse_invoice(id: int, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    if _current_user.get("role") == "прораб":
        cur.execute("SELECT COALESCE(project,''), COALESCE(location,'') FROM warehouse_invoices WHERE id=%s", (id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            raise HTTPException(status_code=404, detail="Накладная не найдена")
        target_project = row[0] or (row[1] if row[1] != "Основной склад" else "")
        if target_project:
            require_project_access(_current_user, target_project)
    cur.execute("DELETE FROM material_inspection_journal WHERE invoice_id=%s", (id,))
    cur.execute("DELETE FROM cable_journal WHERE invoice_id=%s", (id,))
    cur.execute("DELETE FROM warehouse_invoices WHERE id=%s",(id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

def _norm_list(value):
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    if not value:
        return []
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return []
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [str(v).strip() for v in parsed if str(v).strip()]
        except Exception:
            pass
        return [v.strip() for v in raw.replace("\n", ",").split(",") if v.strip()]
    return []

def _material_norm_row(row):
    return {
        "id": row["id"],
        "ruleKey": row["rule_key"] or "",
        "name": row["name"] or "",
        "work": _norm_list(row["work_keywords"]),
        "blockWork": _norm_list(row["block_work_keywords"]),
        "material": _norm_list(row["material_keywords"]),
        "workUnit": row["work_unit"] or "",
        "materialUnit": row["material_unit"] or "",
        "qtyPerUnit": float(row["qty_per_unit"] or 0),
        "thicknessBaseMm": float(row["thickness_base_mm"]) if row["thickness_base_mm"] is not None else None,
        "defaultThicknessMm": float(row["default_thickness_mm"]) if row["default_thickness_mm"] is not None else None,
        "label": row["label"] or "",
        "active": bool(row["active"]),
        "updatedBy": row["updated_by"] or "",
        "updatedAt": str(row["updated_at"]) if row["updated_at"] else "",
    }

def _material_norm_values(data: MaterialNormModel, user_name: str):
    rule_key = (data.ruleKey or "").strip() or ("custom_" + uuid.uuid4().hex[:8])
    return (
        rule_key,
        data.name.strip() or rule_key,
        json.dumps(_norm_list(data.work), ensure_ascii=False),
        json.dumps(_norm_list(data.blockWork), ensure_ascii=False),
        json.dumps(_norm_list(data.material), ensure_ascii=False),
        data.workUnit or "м2",
        data.materialUnit or "кг",
        data.qtyPerUnit or 0,
        data.thicknessBaseMm,
        data.defaultThicknessMm,
        data.label.strip(),
        bool(data.active),
        user_name or "",
    )

def _material_norm_override_row(row):
    return {
        "id": row["id"],
        "baseNormId": row["base_norm_id"],
        "projectName": row["project_name"] or "",
        "estimateId": row["estimate_id"],
        "sectionName": row["section_name"] or "",
        "workName": row["work_name"] or "",
        "materialName": row["material_name"] or "",
        "work": _norm_list(row["work_keywords"]),
        "blockWork": _norm_list(row["block_work_keywords"]),
        "material": _norm_list(row["material_keywords"]),
        "workUnit": row["work_unit"] or "",
        "materialUnit": row["material_unit"] or "",
        "qtyPerUnit": float(row["qty_per_unit"] or 0),
        "thicknessBaseMm": float(row["thickness_base_mm"]) if row["thickness_base_mm"] is not None else None,
        "defaultThicknessMm": float(row["default_thickness_mm"]) if row["default_thickness_mm"] is not None else None,
        "label": row["label"] or "",
        "reason": row["reason"] or "",
        "active": bool(row["active"]),
        "updatedBy": row["updated_by"] or "",
        "updatedAt": str(row["updated_at"]) if row["updated_at"] else "",
    }

def _material_norm_override_values(data: MaterialNormOverrideModel, user_name: str):
    return (
        data.baseNormId,
        (data.projectName or "").strip(),
        data.estimateId,
        (data.sectionName or "").strip(),
        (data.workName or "").strip(),
        (data.materialName or "").strip(),
        json.dumps(_norm_list(data.work), ensure_ascii=False),
        json.dumps(_norm_list(data.blockWork), ensure_ascii=False),
        json.dumps(_norm_list(data.material), ensure_ascii=False),
        data.workUnit or "м2",
        data.materialUnit or "кг",
        data.qtyPerUnit or 0,
        data.thicknessBaseMm,
        data.defaultThicknessMm,
        (data.label or "").strip(),
        (data.reason or "").strip(),
        bool(data.active),
        user_name or "",
    )

@app.get("/material-norms")
def list_material_norms(_current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT id, rule_key, name, work_keywords, block_work_keywords, material_keywords,
                          work_unit, material_unit, qty_per_unit, thickness_base_mm,
                          default_thickness_mm, label, active, updated_by, updated_at
                   FROM material_norms
                   WHERE active=TRUE
                   ORDER BY name, id""")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [_material_norm_row(r) for r in rows]

@app.post("/material-norms")
def create_material_norm(data: MaterialNormModel, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    vals = _material_norm_values(data, current_user.get("name", ""))
    try:
        cur.execute("""
            INSERT INTO material_norms (
                rule_key, name, work_keywords, block_work_keywords, material_keywords,
                work_unit, material_unit, qty_per_unit, thickness_base_mm,
                default_thickness_mm, label, active, updated_by, updated_at
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            RETURNING id, rule_key, name, work_keywords, block_work_keywords, material_keywords,
                      work_unit, material_unit, qty_per_unit, thickness_base_mm,
                      default_thickness_mm, label, active, updated_by, updated_at
        """, vals)
    except psycopg2.errors.UniqueViolation:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Код нормы уже используется")
    row = cur.fetchone()
    cur.close(); conn.close()
    return _material_norm_row(row)

@app.get("/material-norms/overrides")
def list_material_norm_overrides(project_name: str = None, estimate_id: int = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = ["active=TRUE"]
    params = []
    if project_name:
        require_project_access(current_user, project_name)
        where.append("project_name=%s")
        params.append(project_name)
    else:
        allowed = visible_project_names(current_user)
        if allowed is not None:
            if not allowed:
                cur.close(); conn.close()
                return []
            where.append("project_name = ANY(%s)")
            params.append(allowed)
    if estimate_id:
        where.append("estimate_id=%s")
        params.append(estimate_id)
    cur.execute(f"""SELECT * FROM material_norm_overrides
                    WHERE {' AND '.join(where)}
                    ORDER BY updated_at DESC, id DESC""", tuple(params))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [_material_norm_override_row(r) for r in rows]

@app.post("/material-norms/overrides")
def create_material_norm_override(data: MaterialNormOverrideModel, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    project_name = (data.projectName or "").strip()
    if not project_name:
        raise HTTPException(status_code=400, detail="Укажите объект для поправки нормы")
    require_project_access(current_user, project_name)
    if (data.qtyPerUnit or 0) <= 0:
        raise HTTPException(status_code=400, detail="Расход должен быть больше 0")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    vals = _material_norm_override_values(data, current_user.get("name", ""))
    cur.execute("""
        INSERT INTO material_norm_overrides (
            base_norm_id, project_name, estimate_id, section_name, work_name, material_name,
            work_keywords, block_work_keywords, material_keywords, work_unit, material_unit,
            qty_per_unit, thickness_base_mm, default_thickness_mm, label, reason, active,
            updated_by, updated_at, created_at
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
        RETURNING *
    """, vals)
    row = cur.fetchone()
    cur.close(); conn.close()
    return _material_norm_override_row(row)

@app.put("/material-norms/overrides/{id}")
def update_material_norm_override(id: int, data: MaterialNormOverrideModel, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    project_name = (data.projectName or "").strip()
    if not project_name:
        raise HTTPException(status_code=400, detail="Укажите объект для поправки нормы")
    require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    vals = _material_norm_override_values(data, current_user.get("name", ""))
    cur.execute("""
        UPDATE material_norm_overrides
        SET base_norm_id=%s, project_name=%s, estimate_id=%s, section_name=%s,
            work_name=%s, material_name=%s, work_keywords=%s, block_work_keywords=%s,
            material_keywords=%s, work_unit=%s, material_unit=%s, qty_per_unit=%s,
            thickness_base_mm=%s, default_thickness_mm=%s, label=%s, reason=%s,
            active=%s, updated_by=%s, updated_at=NOW()
        WHERE id=%s
        RETURNING *
    """, vals + (id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Поправка нормы не найдена")
    return _material_norm_override_row(row)

@app.delete("/material-norms/overrides/{id}")
def delete_material_norm_override(id: int, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT project_name FROM material_norm_overrides WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Поправка нормы не найдена")
    require_project_access(current_user, row["project_name"] or "")
    cur.execute("UPDATE material_norm_overrides SET active=FALSE, updated_by=%s, updated_at=NOW() WHERE id=%s", (current_user.get("name", ""), id))
    cur.close(); conn.close()
    return {"ok": True}

@app.put("/material-norms/{id}")
def update_material_norm(id: int, data: MaterialNormModel, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    vals = _material_norm_values(data, current_user.get("name", ""))
    cur.execute("""
        UPDATE material_norms
        SET rule_key=%s, name=%s, work_keywords=%s, block_work_keywords=%s,
            material_keywords=%s, work_unit=%s, material_unit=%s, qty_per_unit=%s,
            thickness_base_mm=%s, default_thickness_mm=%s, label=%s, active=%s,
            updated_by=%s, updated_at=NOW()
        WHERE id=%s
        RETURNING id, rule_key, name, work_keywords, block_work_keywords, material_keywords,
                  work_unit, material_unit, qty_per_unit, thickness_base_mm,
                  default_thickness_mm, label, active, updated_by, updated_at
    """, vals + (id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Норма не найдена")
    return _material_norm_row(row)

@app.delete("/material-norms/{id}")
def delete_material_norm(id: int, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE material_norms SET active=FALSE, updated_by=%s, updated_at=NOW() WHERE id=%s", (current_user.get("name", ""), id))
    affected = cur.rowcount
    cur.close(); conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Норма не найдена")
    return {"ok": True}

def _safe_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default

def _norm_key_text(value: str) -> str:
    import re
    return re.sub(r"\s+", " ", re.sub(r"[.,;:()«»\"'`/\\]+", " ", str(value or "").lower())).strip()

def _norm_keywords_from_text(value: str, limit: int = 5) -> list[str]:
    stop = {"работ", "работа", "монтаж", "установка", "устройство", "демонтаж", "разбор", "прочее", "раздел", "материал", "для", "при", "под"}
    words = []
    for token in _norm_key_text(value).split():
        if len(token) < 4 or token in stop:
            continue
        if token not in words:
            words.append(token)
    return words[:limit]

def _estimate_item_type_backend(item: dict, section_name: str = "") -> str:
    raw = str(item.get("itemType") or item.get("type") or item.get("kind") or "").lower()
    text = _norm_key_text((item.get("name") or "") + " " + section_name)
    if raw in ("work", "работа", "works", "работы"):
        return "work"
    if raw in ("material", "материал", "materials", "материалы") or "материал" in raw:
        return "material"
    if raw in ("equipment", "оборудование", "delivery", "доставка", "other", "прочее"):
        return "other"
    material_markers = ("смесь", "штукатурка", "шпатлевка", "шпаклевка", "клей", "краска", "грунтовка", "кабель", "провод", "гофра", "лист гкл", "профиль", "саморез", "кирпич", "бетон")
    work_markers = ("монтаж", "установка", "устройство", "демонтаж", "разбор", "кладка", "окраска", "штукатур", "стяжка", "облицовка", "прокладка")
    if any(m in text for m in material_markers) and not any(w in text for w in work_markers):
        return "material"
    return "work"

def _norm_base_unit(value: str) -> str:
    import re
    text = _norm_key_text(str(value or "").replace("²", "2").replace("³", "3"))
    text = re.sub(r"^\d{2,}\s*", "", text).strip()
    compact = text.replace(" ", "")
    if compact in ("м", "мп", "пм", "погм", "погметр", "метр", "метра", "метров"):
        return "м"
    if compact in ("м2", "квм", "квадратныйметр", "квадратногометра", "квадратныхметров"):
        return "м2"
    if compact in ("м3", "кубм", "кубическийметр", "кубическогометра", "кубическихметров"):
        return "м3"
    if compact in ("шт", "штук", "штука", "штуки"):
        return "шт"
    if compact in ("кг", "килограмм", "килограмма", "килограммов"):
        return "кг"
    if compact in ("л", "литр", "литра", "литров"):
        return "л"
    if compact in ("т", "тонна", "тонны", "тонн"):
        return "т"
    if compact in ("мешок", "мешка", "мешков"):
        return "мешок"
    return text

def _work_no_material_norm_reason(work_name: str = "", section_name: str = "") -> str:
    text = _norm_key_text((work_name or "") + " " + (section_name or ""))
    if any(w in text for w in ("демонтаж", "разбор", "разборка", "снятие", "отбивка")):
        return "Демонтажная/разборочная работа — материал по норме не требуется."
    if any(w in text for w in ("очистка", "обеспыливание", "пробивка", "погрузка", "перевозка", "затаривание")):
        return "Подготовительная или транспортная операция — материал по норме не требуется."
    if "без стоимости оборудования" in text or "ранее демонтирован" in text:
        return "Установка ранее демонтированного оборудования — новый материал по норме не требуется."
    return ""

def _material_no_norm_reason(material_name: str = "") -> str:
    text = _norm_key_text(material_name or "")
    if any(w in text for w in ("сверло", "бур", "диск отрез", "круг отрез", "коронка алмаз", "оснастк", "инструмент")):
        return "Расходный инструмент/оснастка — не участвует в подборе нормы строительного материала."
    return ""

def _norm_rule_specific_enough(rule: dict, work_text: str) -> bool:
    text = _norm_key_text(work_text or "")
    words = text.split()
    has_cable_word = any(
        w.startswith("кабел") or w.startswith("гофр") or (w.startswith("провод") and not w.startswith("трубопровод"))
        for w in words
    )
    rule_key = _norm_key_text(rule.get("rule_key") or "")
    material_text = _norm_key_text(" ".join(_norm_list(rule.get("material_keywords"))))
    is_cable_rule = (
        "cable" in rule_key or
        any(w in material_text for w in ("кабел", "провод", "utp", "ftp", "гофр", "кабель канал"))
    )
    if is_cable_rule and "проклад" in text and not has_cable_word:
        return False
    if "pipe_insulation" in rule_key and not any(w in text for w in ("изоляц", "теплоизоляц", "утепл", "энергофлекс")):
        return False
    return True

def _norm_material_unit_compatible(rule: dict, material_name: str = "", material_unit: str = "") -> bool:
    rule_unit = _norm_base_unit(rule.get("material_unit") or "")
    unit = _norm_base_unit(material_unit or "")
    if not rule_unit or not unit or rule_unit == unit:
        return True
    material_text = _norm_key_text(material_name or "")
    pair = {rule_unit, unit}
    if pair == {"кг", "мешок"} and any(w in material_text for w in ("цемент", "штукатур", "шпаклев", "шпатлев", "клей", "пескобетон", "стяжк", "смесь")):
        return True
    if pair == {"л", "кг"} and any(w in material_text for w in ("краск", "эмал", "грунтов", "лак")):
        return True
    if pair == {"т", "кг"}:
        return True
    if pair == {"м", "кг"} and any(w in material_text for w in ("арматур", "арм")):
        return True
    return False

def _norm_material_family_compatible(rule: dict, material_name: str = "") -> bool:
    rule_key = _norm_key_text(rule.get("rule_key") or "")
    rule_text = _norm_key_text(" ".join([rule_key, rule.get("name") or "", " ".join(_norm_list(rule.get("material_keywords")))]))
    material_text = _norm_key_text(material_name or "")
    def has(words):
        return any(w in material_text for w in words)
    fitting_words = ("муфта", "угольник", "угол", "тройник", "переход", "кран", "фитинг", "отвод", "вентиль", "клапан")
    pipe_words = ("труба", "трубы", "трубн", "ppr", "pprc", "ppr c", "ppr-c", "полипропилен")
    pipe_clamp_words = ("хомут", "кронштейн", "клипс", "клипса", "скоба", "держатель", "опора", "подвес", "консоль")
    insulation_words = ("изоляц", "теплоизоляц", "утепл", "утеплител", "минераловат", "минвата", "вата", "изовер", "технониколь", "пенополиэтилен", "энергофлекс")
    fastener_words = ("дюбел", "дюбель", "саморез", "шуруп", "анкер", "гвозд", "болт")
    if "pipe_pp" in rule_key:
        return has(pipe_words) and not has(fitting_words) and not has(pipe_clamp_words) and not has(fastener_words)
    if "pipe_fittings" in rule_key:
        return has(fitting_words)
    if "pipe_clamps" in rule_key:
        return has(pipe_clamp_words)
    if "pipe_insulation" in rule_key:
        return has(insulation_words) and not has(fastener_words)
    if "thermal_insulation_board" in rule_key:
        return has(insulation_words) and not has(fastener_words)
    if "radiator_mount" in rule_key:
        return has(("кронштейн", "консоль", "крепление радиатор", "крепеж радиатор", "комплект крепления"))
    if "изоляция труб" in rule_text:
        return has(insulation_words) and not has(fastener_words)
    if "трубопровод" in rule_text and has(fastener_words) and not has(pipe_clamp_words):
        return False
    return True

def _norm_rule_matches(rule: dict, work_name: str, section_name: str, material_name: str, work_unit: str = "", material_unit: str = "") -> bool:
    work_text = _norm_key_text((work_name or "") + " " + (section_name or ""))
    mat_text = _norm_key_text(material_name or "")
    work_words = _norm_list(rule.get("work_keywords"))
    block_words = _norm_list(rule.get("block_work_keywords"))
    material_words = _norm_list(rule.get("material_keywords"))
    if _work_no_material_norm_reason(work_name, section_name):
        return False
    if work_unit and rule.get("work_unit") and _norm_base_unit(work_unit) != _norm_base_unit(rule.get("work_unit")):
        return False
    if not _norm_rule_specific_enough(rule, work_text):
        return False
    if not _norm_material_unit_compatible(rule, material_name, material_unit):
        return False
    if not _norm_material_family_compatible(rule, material_name):
        return False
    return (
        any(_norm_key_text(w) and _norm_key_text(w) in work_text for w in work_words) and
        not any(_norm_key_text(w) and _norm_key_text(w) in work_text for w in block_words) and
        any(_norm_key_text(w) and _norm_key_text(w) in mat_text for w in material_words)
    )

def _estimate_norm_candidate_work(material_name: str, works: list[dict], section_name: str = "", material_unit: str = "") -> dict:
    material_text = _norm_key_text(material_name)
    if not material_text or not works:
        return {}
    pairs = [
        (("штукатур", "ротбанд", "гипсов"), ("штукатур",)),
        (("грунтов",), ("грунтов",)),
        (("краск", "эмал"), ("окраск", "покраск")),
        (("шпаклев", "шпатлев"), ("шпаклев", "шпатлев")),
        (("клей", "плиточ", "затир"), ("плитк", "керамогранит", "облицов")),
        (("кабель", "провод", "ввг", "nym", "utp", "ftp", "кпс", "кпкв"), ("кабел", "провод", "проклад")),
        (("гофр", "кабель канал", "короб"), ("гофр", "труб", "кабель", "короб")),
        (("коробка ответв", "распаечн"), ("коробка ответв", "распаечн")),
        (("розет", "выключател", "штепсель"), ("розет", "выключател", "штепсель")),
        (("щит", "шкаф", "бокс распредел"), ("щит", "шкаф", "распределительн")),
        (("труб", "полипропилен", "муфта", "угольник", "хомут"), ("трубопровод", "водоснаб", "отоплен")),
        (("радиатор", "кронштейн", "клапан"), ("радиатор", "отоплен")),
        (("светильник", "табло", "транспарант"), ("светильник", "табло", "транспарант")),
        (("фанер",), ("фанер", "пол")),
        (("osb", "древесн", "плиты"), ("основан", "фанер", "пол")),
        (("линолеум", "плинтус"), ("линолеум", "плинтус")),
        (("двер", "блок двер"), ("двер", "блок")),
        (("окон", "подокон"), ("окон", "подокон")),
        (("бетон",), ("бетон",)),
        (("кирпич",), ("кладк", "кирпич")),
        (("сетка", "маяч"), ("штукатур",)),
        (("дюбель", "саморез"), ("креп", "гипсокарт", "гкл", "профиль", "теплоизоляц")),
        (("сверло", "алмаз"), ("пробив", "отверст")),
        (("изоляц", "пенополиэтилен"), ("изоляц", "теплоизоляц")),
    ]
    stop = {"материал", "изделия", "разные", "прочие", "наружный", "диаметр", "номинальный", "давление", "размер"}
    material_tokens = {t for t in material_text.split() if len(t) >= 5 and t not in stop}
    material_base_unit = _norm_base_unit(material_unit)
    best = (0, {})
    for work in works:
        work_text = _norm_key_text(work.get("name") or "")
        if _work_no_material_norm_reason(work.get("name"), section_name):
            continue
        work_words = work_text.split()
        has_cable_word = any(
            w.startswith("кабел") or w.startswith("гофр") or (w.startswith("провод") and not w.startswith("трубопровод"))
            for w in work_words
        )
        if any(w in material_text for w in ("кабел", "провод", "utp", "ftp", "кпс", "кпкв")) and not has_cable_word:
            continue
        semantic_score = 0
        for material_words, work_words in pairs:
            if any(w in material_text for w in material_words) and any(w in work_text for w in work_words):
                semantic_score += 5
        work_tokens = {t for t in work_text.split() if len(t) >= 5 and t not in stop}
        semantic_score += min(3, len(material_tokens & work_tokens))
        if semantic_score <= 0:
            continue
        score = semantic_score
        work_base_unit = _norm_base_unit(work.get("unit"))
        if material_base_unit and work_base_unit:
            if material_base_unit == work_base_unit:
                score += 4
            elif material_base_unit in ("шт", "компл") and work_base_unit in ("шт", "соединений", "точк"):
                score += 2
        if score > best[0]:
            best = (score, work)
    return best[1] if best[0] >= 2 else {}

def _material_norm_suggestion_row(row):
    return {
        "id": row["id"],
        "projectName": row["project_name"] or "",
        "suggestionType": row["suggestion_type"] or "",
        "status": row["status"] or "Новая",
        "severity": row["severity"] or "Проверить",
        "workName": row["work_name"] or "",
        "sectionName": row["section_name"] or "",
        "workUnit": row["work_unit"] or "",
        "materialName": row["material_name"] or "",
        "materialUnit": row["material_unit"] or "",
        "currentNormId": row["current_norm_id"],
        "currentQtyPerUnit": _safe_float(row["current_qty_per_unit"]),
        "suggestedQtyPerUnit": _safe_float(row["suggested_qty_per_unit"]),
        "sampleCount": int(row["sample_count"] or 0),
        "actualQty": _safe_float(row["actual_qty"]),
        "normQty": _safe_float(row["norm_qty"]),
        "confidence": _safe_float(row["confidence"]),
        "reason": row["reason"] or "",
        "aiSummary": row["ai_summary"] or "",
        "work": _norm_list(row["work_keywords"]),
        "material": _norm_list(row["material_keywords"]),
        "blockWork": _norm_list(row["block_work_keywords"]),
        "label": row["label"] or "",
        "source": row["source"] or "rules",
        "dedupeKey": row["dedupe_key"] or "",
        "createdBy": row["created_by"] or "",
        "appliedNormId": row["applied_norm_id"],
        "createdAt": str(row["created_at"]) if row["created_at"] else "",
        "updatedAt": str(row["updated_at"]) if row["updated_at"] else "",
    }

def _load_active_norm_rules(cur, project_name: str = "", estimate_id: int = None):
    cur.execute("""SELECT id, rule_key, name, work_keywords, block_work_keywords, material_keywords,
                          work_unit, material_unit, qty_per_unit, thickness_base_mm,
                          default_thickness_mm, label
                   FROM material_norms WHERE active=TRUE ORDER BY id""")
    base_rules = [dict(r) for r in cur.fetchall()]
    if not project_name:
        return base_rules
    params = [project_name]
    estimate_clause = ""
    if estimate_id:
        estimate_clause = "AND (estimate_id IS NULL OR estimate_id=%s)"
        params.append(estimate_id)
    cur.execute(f"""SELECT * FROM material_norm_overrides
                    WHERE active=TRUE AND project_name=%s {estimate_clause}
                    ORDER BY CASE WHEN estimate_id IS NULL THEN 1 ELSE 0 END, id DESC""", tuple(params))
    overrides = []
    for row in cur.fetchall():
        r = dict(row)
        overrides.append({
            "id": r.get("base_norm_id"),
            "override_id": r.get("id"),
            "base_norm_id": r.get("base_norm_id"),
            "rule_key": "override_" + str(r.get("id")),
            "name": r.get("material_name") or r.get("label") or "Поправка нормы",
            "work_keywords": r.get("work_keywords"),
            "block_work_keywords": r.get("block_work_keywords"),
            "material_keywords": r.get("material_keywords"),
            "work_unit": r.get("work_unit"),
            "material_unit": r.get("material_unit"),
            "qty_per_unit": r.get("qty_per_unit"),
            "thickness_base_mm": r.get("thickness_base_mm"),
            "default_thickness_mm": r.get("default_thickness_mm"),
            "label": r.get("label") or ("Поправка объекта: " + str(r.get("qty_per_unit") or 0)),
            "project_name": r.get("project_name"),
            "estimate_id": r.get("estimate_id"),
            "reason": r.get("reason"),
        })
    return overrides + base_rules

def _upsert_material_norm_suggestion(cur, suggestion: dict, current_user: dict):
    project_name = suggestion.get("projectName") or ""
    if project_name:
        require_project_access(current_user, project_name)
    dedupe_key = suggestion.get("dedupeKey") or ""
    fields = (
        project_name,
        suggestion.get("suggestionType") or "without_norm",
        suggestion.get("severity") or "Проверить",
        suggestion.get("workName") or "",
        suggestion.get("sectionName") or "",
        suggestion.get("workUnit") or "",
        suggestion.get("materialName") or "",
        suggestion.get("materialUnit") or "",
        suggestion.get("currentNormId"),
        suggestion.get("currentQtyPerUnit") or 0,
        suggestion.get("suggestedQtyPerUnit") or 0,
        int(suggestion.get("sampleCount") or 0),
        suggestion.get("actualQty") or 0,
        suggestion.get("normQty") or 0,
        suggestion.get("confidence") or 0,
        suggestion.get("reason") or "",
        suggestion.get("aiSummary") or "",
        json.dumps(_norm_list(suggestion.get("work")), ensure_ascii=False),
        json.dumps(_norm_list(suggestion.get("material")), ensure_ascii=False),
        json.dumps(_norm_list(suggestion.get("blockWork")), ensure_ascii=False),
        suggestion.get("label") or "",
        suggestion.get("source") or "rules",
        dedupe_key,
        current_user.get("name") or "",
    )
    if dedupe_key:
        cur.execute("""SELECT id FROM material_norm_suggestions
                       WHERE dedupe_key=%s AND status NOT IN ('Принята','Отклонена')
                       LIMIT 1""", (dedupe_key,))
        row = cur.fetchone()
        if row:
            sid = row["id"] if isinstance(row, dict) else row[0]
            cur.execute("""
                UPDATE material_norm_suggestions
                SET project_name=%s, suggestion_type=%s, severity=%s, work_name=%s,
                    section_name=%s, work_unit=%s, material_name=%s, material_unit=%s,
                    current_norm_id=%s, current_qty_per_unit=%s, suggested_qty_per_unit=%s,
                    sample_count=%s, actual_qty=%s, norm_qty=%s, confidence=%s,
                    reason=%s, ai_summary=%s, work_keywords=%s, material_keywords=%s,
                    block_work_keywords=%s, label=%s, source=%s, dedupe_key=%s,
                    created_by=%s, updated_at=NOW()
                WHERE id=%s
            """, fields + (sid,))
            return sid, False
    cur.execute("""
        INSERT INTO material_norm_suggestions (
            project_name, suggestion_type, severity, work_name, section_name, work_unit,
            material_name, material_unit, current_norm_id, current_qty_per_unit,
            suggested_qty_per_unit, sample_count, actual_qty, norm_qty, confidence,
            reason, ai_summary, work_keywords, material_keywords, block_work_keywords,
            label, source, dedupe_key, created_by, created_at, updated_at
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
        RETURNING id
    """, fields)
    row = cur.fetchone()
    return (row["id"] if isinstance(row, dict) else row[0]), True

def _enhance_norm_suggestions_with_ai(suggestions: list[dict]) -> list[dict]:
    if not suggestions or not (YANDEX_API_KEY and YANDEX_FOLDER_ID):
        return suggestions
    try:
        import openai as oa
        import re as _re
        compact = [{
            "dedupeKey": s.get("dedupeKey"),
            "type": s.get("suggestionType"),
            "work": s.get("workName"),
            "section": s.get("sectionName"),
            "material": s.get("materialName"),
            "unit": s.get("materialUnit"),
            "suggestedQtyPerUnit": s.get("suggestedQtyPerUnit"),
            "reason": s.get("reason"),
        } for s in suggestions[:12]]
        instructions = "Ты отвечаешь строго валидным JSON без markdown. Не выдумывай объёмы, только уточняй ключевые слова, краткое объяснение и уверенность."
        prompt = (
            "Нужно улучшить предложения норм расхода материалов в строительной CRM.\n"
            "Верни JSON: {\"suggestions\":[{\"dedupeKey\":\"...\",\"workKeywords\":[\"...\"],\"materialKeywords\":[\"...\"],\"blockWorkKeywords\":[\"демонтаж\",\"разбор\"],\"label\":\"краткая норма\",\"reason\":\"почему предложено\",\"confidence\":0.0-0.95}]}\n\n"
            "ДАННЫЕ:\n" + json.dumps(compact, ensure_ascii=False)
        )
        client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
        raw = ""
        for model_id in ("qwen3.6-35b-a3b/latest", "yandexgpt-5.1/latest"):
            try:
                r = client.responses.create(
                    model="gpt://" + YANDEX_FOLDER_ID + "/" + model_id,
                    temperature=0.1,
                    instructions=instructions,
                    input=prompt,
                    max_output_tokens=3000,
                )
                raw = (r.output_text or "").strip()
                if raw:
                    break
            except Exception as e:
                print("MATERIAL NORM SUGGEST AI ERROR:", str(e))
        if not raw:
            return suggestions
        clean = _re.sub(r"^```(?:json)?\s*", "", raw).strip()
        clean = _re.sub(r"\s*```\s*$", "", clean).strip()
        s_idx = clean.find("{"); e_idx = clean.rfind("}")
        parsed = json.loads(clean[s_idx:e_idx+1]) if s_idx >= 0 and e_idx > s_idx else {}
        ai_by_key = {str(x.get("dedupeKey") or ""): x for x in (parsed.get("suggestions") or []) if isinstance(x, dict)}
        out = []
        for s in suggestions:
            ai = ai_by_key.get(str(s.get("dedupeKey") or ""))
            if ai:
                s = dict(s)
                s["work"] = _norm_list(ai.get("workKeywords")) or s.get("work") or []
                s["material"] = _norm_list(ai.get("materialKeywords")) or s.get("material") or []
                s["blockWork"] = _norm_list(ai.get("blockWorkKeywords")) or s.get("blockWork") or []
                s["label"] = str(ai.get("label") or s.get("label") or "")
                s["aiSummary"] = str(ai.get("reason") or s.get("aiSummary") or "")
                s["confidence"] = max(_safe_float(s.get("confidence")), min(0.95, _safe_float(ai.get("confidence"))))
                s["source"] = "rules+yandexgpt"
            out.append(s)
        return out
    except Exception as e:
        print("MATERIAL NORM SUGGEST AI PARSE ERROR:", str(e))
        return suggestions

def _generate_material_norm_suggestions(cur, current_user: dict, project_name: str = "", use_ai: bool = True) -> list[dict]:
    allowed_projects = visible_project_names(current_user)
    norm_rules = _load_active_norm_rules(cur, project_name)
    suggestions = {}
    where = "WHERE COALESCE(status,'') <> 'Отклонено'"
    params = []
    if project_name:
        where += " AND project=%s"
        params.append(project_name)
    elif allowed_projects is not None:
        if not allowed_projects:
            return []
        where += " AND project = ANY(%s)"
        params.append(allowed_projects)
    cur.execute(f"""SELECT id, project, description, section_name, unit, quantity, materials_used, date
                    FROM work_journal {where}
                    ORDER BY id DESC LIMIT 800""", tuple(params))
    for row in cur.fetchall():
        work = dict(row)
        try:
            used = json.loads(work.get("materials_used") or "[]")
        except Exception:
            used = []
        if not isinstance(used, list):
            used = []
        work_qty = _safe_float(work.get("quantity"))
        if work_qty <= 0:
            continue
        for mat in used:
            material_name = str((mat or {}).get("name") or "").strip()
            mat_qty = _safe_float((mat or {}).get("quantity"))
            if not material_name or mat_qty <= 0:
                continue
            if _material_no_norm_reason(material_name):
                continue
            mat_unit = str((mat or {}).get("unit") or "шт")
            norm_qty = _safe_float((mat or {}).get("normQuantity"))
            matching_rule = next((r for r in norm_rules if _norm_rule_matches(r, work.get("description"), work.get("section_name"), material_name, work.get("unit"), mat_unit)), None)
            if norm_qty > 0 and mat_qty <= norm_qty * 1.15:
                continue
            suggestion_type = "over_norm" if norm_qty > 0 else "without_norm"
            dedupe = "|".join([
                suggestion_type,
                _norm_key_text(work.get("project")),
                _norm_key_text(work.get("description"))[:70],
                _norm_key_text(material_name)[:70],
                _norm_key_text(mat_unit),
            ])
            item = suggestions.setdefault(dedupe, {
                "projectName": work.get("project") or "",
                "suggestionType": suggestion_type,
                "severity": "Критично" if suggestion_type == "without_norm" else "Проверить",
                "workName": work.get("description") or "",
                "sectionName": work.get("section_name") or "",
                "workUnit": work.get("unit") or "",
                "materialName": material_name,
                "materialUnit": mat_unit,
                "currentNormId": matching_rule.get("id") if matching_rule and suggestion_type == "over_norm" else None,
                "currentQtyPerUnit": _safe_float(matching_rule.get("qty_per_unit")) if matching_rule else 0,
                "_workQty": 0.0,
                "actualQty": 0.0,
                "normQty": 0.0,
                "sampleCount": 0,
                "reason": "",
                "work": _norm_keywords_from_text((work.get("description") or "") + " " + (work.get("section_name") or "")),
                "material": _norm_keywords_from_text(material_name),
                "blockWork": ["демонтаж", "разбор"],
                "dedupeKey": dedupe,
            })
            item["_workQty"] += work_qty
            item["actualQty"] += mat_qty
            item["normQty"] += norm_qty
            item["sampleCount"] += 1
    for s in suggestions.values():
        work_qty = _safe_float(s.pop("_workQty", 0))
        s["suggestedQtyPerUnit"] = round(s["actualQty"] / work_qty, 4) if work_qty > 0 else 0
        if s["suggestionType"] == "without_norm":
            s["reason"] = "Материал списывали в ЖПР без привязанной нормы. Нужно завести норму или уточнить материал."
            s["confidence"] = min(0.9, 0.55 + s["sampleCount"] * 0.08)
        else:
            over = max(0, s["actualQty"] - s["normQty"])
            pct = round(over / s["normQty"] * 100) if s["normQty"] > 0 else 0
            s["reason"] = f"Фактическое списание выше нормы на {pct}%. Нужно проверить: ошибка списания, слой/условия работы или устаревшая норма."
            s["confidence"] = min(0.85, 0.5 + s["sampleCount"] * 0.06)
        s["label"] = f"{s['materialName']} {s['suggestedQtyPerUnit']} {s['materialUnit']} / {s['workUnit'] or 'ед.'}"

    estimate_where = "WHERE COALESCE(e.is_template,FALSE)=FALSE AND e.status='Активная' AND COALESCE(e.smeta_type,'Заказчик')='Заказчик'"
    estimate_params = []
    if project_name:
        estimate_where += " AND e.project_name=%s"
        estimate_params.append(project_name)
    elif allowed_projects is not None:
        if not allowed_projects:
            return list(suggestions.values())
        estimate_where += " AND e.project_name = ANY(%s)"
        estimate_params.append(allowed_projects)
    cur.execute(f"""SELECT e.id, e.project_name, e.sections_json FROM estimates e {estimate_where} ORDER BY e.id DESC LIMIT 120""", tuple(estimate_params))
    for est in cur.fetchall():
        est_norm_rules = _load_active_norm_rules(cur, est["project_name"] or "", est["id"])
        try:
            sections = json.loads(est["sections_json"] or "[]")
        except Exception:
            sections = []
        for section in sections:
            section_name = section.get("name") or ""
            items = section.get("items") or []
            works = [it for it in items if _estimate_item_type_backend(it, section_name) == "work"]
            mats = [it for it in items if _estimate_item_type_backend(it, section_name) == "material"]
            for mat in mats:
                material_name = str(mat.get("name") or "").strip()
                if not material_name:
                    continue
                if _material_no_norm_reason(material_name):
                    continue
                mat_qty = _safe_float(mat.get("quantity"))
                if mat_qty <= 0:
                    continue
                if any(_norm_rule_matches(r, w.get("name"), section_name, material_name, w.get("unit"), mat.get("unit") or "") for r in est_norm_rules for w in works):
                    continue
                work = _estimate_norm_candidate_work(material_name, works, section_name, mat.get("unit") or "")
                if not work:
                    continue
                work_qty = _safe_float(work.get("quantity"))
                suggested = round(mat_qty / work_qty, 4) if work_qty > 0 and mat_qty > 0 else 0
                dedupe = "|".join([
                    "estimate_material_without_norm",
                    _norm_key_text(est["project_name"]),
                    _norm_key_text(section_name)[:70],
                    _norm_key_text(material_name)[:70],
                ])
                suggestions.setdefault(dedupe, {
                    "projectName": est["project_name"] or "",
                    "suggestionType": "estimate_material_without_norm",
                    "severity": "Проверить",
                    "workName": work.get("name") or ("Раздел: " + section_name),
                    "sectionName": section_name,
                    "workUnit": work.get("unit") or "",
                    "materialName": material_name,
                    "materialUnit": mat.get("unit") or "",
                    "currentNormId": None,
                    "currentQtyPerUnit": 0,
                    "suggestedQtyPerUnit": suggested,
                    "sampleCount": 1,
                    "actualQty": mat_qty,
                    "normQty": 0,
                    "confidence": 0.45 if suggested else 0.35,
                    "reason": "В активной смете есть материал, но справочник норм не связывает его с работами этого раздела.",
                    "work": _norm_keywords_from_text((work.get("name") or "") + " " + section_name),
                    "material": _norm_keywords_from_text(material_name),
                    "blockWork": ["демонтаж", "разбор"],
                    "label": f"{material_name}" + (f" {suggested} {mat.get('unit') or ''} / {work.get('unit') or 'ед.'}" if suggested else ""),
                    "dedupeKey": dedupe,
                })
    result = list(suggestions.values())
    return _enhance_norm_suggestions_with_ai(result) if use_ai else result

@app.get("/material-norm-suggestions")
def list_material_norm_suggestions(project_name: str = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute("""SELECT * FROM material_norm_suggestions WHERE project_name=%s ORDER BY updated_at DESC, id DESC""", (project_name,))
    else:
        allowed_projects = visible_project_names(current_user)
        if allowed_projects is not None:
            if not allowed_projects:
                cur.close(); conn.close()
                return []
            cur.execute("""SELECT * FROM material_norm_suggestions WHERE project_name = ANY(%s) ORDER BY updated_at DESC, id DESC""", (allowed_projects,))
        else:
            cur.execute("""SELECT * FROM material_norm_suggestions ORDER BY updated_at DESC, id DESC""")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [_material_norm_suggestion_row(dict(r)) for r in rows]

@app.post("/material-norm-suggestions/generate")
def generate_material_norm_suggestions(data: dict = None, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    data = data or {}
    project_name = data.get("projectName") or data.get("project_name") or ""
    dry_run = bool(data.get("dryRun") or data.get("dry_run"))
    use_ai = False if dry_run else (data.get("useAi", data.get("use_ai", True)) is not False)
    if project_name:
        require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    suggestions = _generate_material_norm_suggestions(cur, current_user, project_name, use_ai=use_ai)
    if dry_run:
        for idx, suggestion in enumerate(suggestions, start=1):
            suggestion["id"] = "preview-" + str(idx)
            suggestion["status"] = "Предпросмотр"
            suggestion["source"] = suggestion.get("source") or "rules-preview"
        cur.close(); conn.close()
        return {"ok": True, "dryRun": True, "created": 0, "updated": 0, "findings": 0, "total": len(suggestions), "aiUsed": False, "suggestions": suggestions}
    created = updated = findings = 0
    for suggestion in suggestions:
        sid, is_created = _upsert_material_norm_suggestion(cur, suggestion, current_user)
        suggestion["id"] = sid
        if is_created:
            created += 1
        else:
            updated += 1
        if suggestion.get("projectName") and suggestion.get("suggestionType") in ("without_norm", "over_norm", "estimate_material_without_norm"):
            _, f_created = _upsert_ai_finding(cur, {
                "projectName": suggestion.get("projectName"),
                "findingType": "ai",
                "category": "Материалы",
                "severity": suggestion.get("severity") or "Проверить",
                "title": "Нужно уточнить норму: " + (suggestion.get("materialName") or ""),
                "description": (suggestion.get("aiSummary") or suggestion.get("reason") or "") + "\nРабота: " + (suggestion.get("workName") or ""),
                "source": suggestion.get("source") or "rules+yandexgpt",
                "linkedEntityType": "material_norm_suggestion",
                "linkedEntityId": str(sid),
                "suggestedAction": "Проверить предложение в Сметы → Нормы материалов и принять норму или отклонить.",
                "assignedRole": "сметчик",
                "dedupeKey": "material_norm_suggestion:" + (suggestion.get("dedupeKey") or str(sid)),
            }, current_user)
            if f_created:
                findings += 1
    cur.close(); conn.close()
    return {"ok": True, "created": created, "updated": updated, "findings": findings, "total": len(suggestions), "aiUsed": bool(YANDEX_API_KEY and YANDEX_FOLDER_ID)}

@app.put("/material-norm-suggestions/{id}")
def update_material_norm_suggestion(id: int, data: MaterialNormSuggestionUpdateModel, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM material_norm_suggestions WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Предложение не найдено")
    require_project_access(current_user, row["project_name"] or "")
    status = data.status or "Отклонена"
    cur.execute("UPDATE material_norm_suggestions SET status=%s, updated_at=NOW() WHERE id=%s RETURNING *", (status, id))
    row = cur.fetchone()
    cur.close(); conn.close()
    return _material_norm_suggestion_row(dict(row))

@app.post("/material-norm-suggestions/{id}/task")
def create_task_from_material_norm_suggestion(id: int, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM material_norm_suggestions WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Предложение не найдено")
    suggestion = _material_norm_suggestion_row(dict(row))
    require_project_access(current_user, suggestion["projectName"])
    finding_id, _ = _upsert_ai_finding(cur, {
        "projectName": suggestion["projectName"],
        "findingType": "ai",
        "category": "Материалы",
        "severity": suggestion["severity"],
        "title": "Уточнить норму: " + suggestion["materialName"],
        "description": (suggestion["aiSummary"] or suggestion["reason"]) + "\nРабота: " + suggestion["workName"],
        "source": suggestion["source"],
        "linkedEntityType": "material_norm_suggestion",
        "linkedEntityId": str(id),
        "suggestedAction": "Проверить предложенную норму и принять её в справочник или отклонить.",
        "assignedRole": "сметчик",
        "dedupeKey": "material_norm_suggestion:" + (suggestion["dedupeKey"] or str(id)),
    }, current_user)
    cur.close(); conn.close()
    return {"ok": True, "findingId": finding_id}

@app.post("/material-norm-suggestions/{id}/accept")
def accept_material_norm_suggestion(id: int, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM material_norm_suggestions WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Предложение не найдено")
    suggestion = _material_norm_suggestion_row(dict(row))
    require_project_access(current_user, suggestion["projectName"])
    qty = suggestion["suggestedQtyPerUnit"]
    if qty <= 0:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="В предложении нет расчётной нормы. Уточните вручную.")
    work_keywords = suggestion["work"] or _norm_keywords_from_text(suggestion["workName"] + " " + suggestion["sectionName"])
    material_keywords = suggestion["material"] or _norm_keywords_from_text(suggestion["materialName"])
    block_keywords = suggestion["blockWork"] or ["демонтаж", "разбор"]
    label = suggestion["label"] or (suggestion["materialName"] + " " + str(qty) + " " + suggestion["materialUnit"] + " / " + (suggestion["workUnit"] or "ед."))
    if suggestion["currentNormId"] and suggestion["suggestionType"] == "over_norm":
        cur.execute("""
            UPDATE material_norms
            SET name=%s, work_keywords=%s, block_work_keywords=%s, material_keywords=%s,
                work_unit=%s, material_unit=%s, qty_per_unit=%s, label=%s,
                updated_by=%s, updated_at=NOW()
            WHERE id=%s
            RETURNING id
        """, (
            suggestion["materialName"],
            json.dumps(work_keywords, ensure_ascii=False),
            json.dumps(block_keywords, ensure_ascii=False),
            json.dumps(material_keywords, ensure_ascii=False),
            suggestion["workUnit"] or "м2",
            suggestion["materialUnit"] or "шт",
            qty,
            label,
            current_user.get("name") or "",
            suggestion["currentNormId"],
        ))
        norm_row = cur.fetchone()
        norm_id = norm_row["id"] if norm_row else suggestion["currentNormId"]
    else:
        base_key = "ai_" + _norm_key_text(suggestion["materialName"] + "_" + suggestion["workName"]).replace(" ", "_")[:70]
        rule_key = base_key or ("ai_" + uuid.uuid4().hex[:8])
        norm_row = None
        for attempt in range(6):
            cur.execute("""
                INSERT INTO material_norms (
                    rule_key, name, work_keywords, block_work_keywords, material_keywords,
                    work_unit, material_unit, qty_per_unit, label, active, updated_by, updated_at
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s,NOW())
                ON CONFLICT (rule_key) DO NOTHING
                RETURNING id
            """, (
                rule_key,
                suggestion["materialName"],
                json.dumps(work_keywords, ensure_ascii=False),
                json.dumps(block_keywords, ensure_ascii=False),
                json.dumps(material_keywords, ensure_ascii=False),
                suggestion["workUnit"] or "м2",
                suggestion["materialUnit"] or "шт",
                qty,
                label,
                current_user.get("name") or "",
            ))
            norm_row = cur.fetchone()
            if norm_row:
                break
            rule_key = base_key + "_" + uuid.uuid4().hex[:4] if attempt < 5 else "ai_" + uuid.uuid4().hex[:10]
        norm_id = norm_row["id"] if norm_row else None
    cur.execute("UPDATE material_norm_suggestions SET status='Принята', applied_norm_id=%s, updated_at=NOW() WHERE id=%s RETURNING *", (norm_id, id))
    row = cur.fetchone()
    cur.execute("""UPDATE ai_findings
                   SET status='Закрыто', updated_at=NOW()
                   WHERE linked_entity_type='material_norm_suggestion' AND linked_entity_id=%s""", (str(id),))
    cur.close(); conn.close()
    return _material_norm_suggestion_row(dict(row))

@app.post("/material-norm-suggestions/{id}/accept-override")
def accept_material_norm_suggestion_as_override(id: int, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM material_norm_suggestions WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Предложение не найдено")
    suggestion = _material_norm_suggestion_row(dict(row))
    require_project_access(current_user, suggestion["projectName"])
    qty = suggestion["suggestedQtyPerUnit"]
    if qty <= 0:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="В предложении нет расчётной нормы. Уточните вручную.")
    work_keywords = suggestion["work"] or _norm_keywords_from_text(suggestion["workName"] + " " + suggestion["sectionName"])
    material_keywords = suggestion["material"] or _norm_keywords_from_text(suggestion["materialName"])
    block_keywords = suggestion["blockWork"] or ["демонтаж", "разбор"]
    label = suggestion["label"] or ("Поправка объекта: " + suggestion["materialName"] + " " + str(qty) + " " + suggestion["materialUnit"] + " / " + (suggestion["workUnit"] or "ед."))
    cur.execute("""
        INSERT INTO material_norm_overrides (
            base_norm_id, project_name, section_name, work_name, material_name,
            work_keywords, block_work_keywords, material_keywords, work_unit, material_unit,
            qty_per_unit, label, reason, active, updated_by, updated_at, created_at
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s,NOW(),NOW())
        RETURNING id
    """, (
        suggestion["currentNormId"],
        suggestion["projectName"],
        suggestion["sectionName"],
        suggestion["workName"],
        suggestion["materialName"],
        json.dumps(work_keywords, ensure_ascii=False),
        json.dumps(block_keywords, ensure_ascii=False),
        json.dumps(material_keywords, ensure_ascii=False),
        suggestion["workUnit"] or "м2",
        suggestion["materialUnit"] or "шт",
        qty,
        label,
        suggestion["reason"] or suggestion["aiSummary"] or "Принято как поправка объекта, базовая норма не изменена",
        current_user.get("name") or "",
    ))
    norm_row = cur.fetchone()
    override_id = norm_row["id"] if norm_row else None
    cur.execute("UPDATE material_norm_suggestions SET status='Принята', applied_norm_id=%s, updated_at=NOW() WHERE id=%s RETURNING *", (override_id, id))
    row = cur.fetchone()
    cur.close(); conn.close()
    return _material_norm_suggestion_row(dict(row))

@app.get("/material-inspection")
def list_material_inspections(project_name: str = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = """id, project_name, invoice_id, material_name, unit, quantity, supplier,
              batch_number, passport_number, certificate_number, test_protocol_number,
              visual_inspection_result, remarks, inspector_name,
              received_at, inspected_at, inspected, normatives, ai_filled, created_at"""
    allowed_projects = visible_project_names(_current_user)
    if project_name:
        if allowed_projects is not None and project_name not in allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM material_inspection_journal WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM material_inspection_journal WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute(f"SELECT {cols} FROM material_inspection_journal ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1] or "","invoiceId":r[2],"materialName":r[3] or "",
             "unit":r[4] or "","quantity":float(r[5] or 0),"supplier":r[6] or "",
             "batchNumber":r[7] or "","passportNumber":r[8] or "","certificateNumber":r[9] or "",
             "testProtocolNumber":r[10] or "","visualInspectionResult":r[11] or "",
             "remarks":r[12] or "","inspectorName":r[13] or "",
             "receivedAt":str(r[14]) if r[14] else "","inspectedAt":str(r[15]) if r[15] else "",
             "inspected":bool(r[16]),"normatives":r[17] or "",
             "aiFilled":bool(r[18]),"createdAt":str(r[19])} for r in rows]

@app.put("/material-inspection/{id}")
def update_material_inspection(id: int, data: dict, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "material_inspection_journal", id, _current_user)
    fields_map = [
        ('batchNumber', 'batch_number'),
        ('passportNumber', 'passport_number'),
        ('certificateNumber', 'certificate_number'),
        ('testProtocolNumber', 'test_protocol_number'),
        ('visualInspectionResult', 'visual_inspection_result'),
        ('remarks', 'remarks'),
        ('inspectorName', 'inspector_name'),
        ('inspectedAt', 'inspected_at'),
        ('inspected', 'inspected'),
        ('normatives', 'normatives'),
    ]
    sets, vals = [], []
    ai_resetting_fields = {'visualInspectionResult','remarks','passportNumber','certificateNumber','testProtocolNumber','normatives'}
    reset_ai = False
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            v = data[js_key]
            if js_key == 'inspectedAt' and not v:
                v = None
            vals.append(v)
            if js_key in ai_resetting_fields:
                reset_ai = True
    if reset_ai:
        sets.append("ai_filled=FALSE")
    if not sets:
        cur.close(); conn.close()
        return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE material_inspection_journal SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/material-inspection/{id}/ai-suggest")
def ai_suggest_material_inspection(id: int, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    import openai as oa, json as j, re
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "material_inspection_journal", id, _current_user)
    cur.execute("SELECT material_name, unit, quantity FROM material_inspection_journal WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="запись не найдена")
    material_name, unit, quantity = row[0] or "", row[1] or "", float(row[2] or 0)
    cur.close()

    user_text = (
        "Материал: " + material_name + "\n"
        "Единица: " + unit + "\n"
        "Количество: " + str(quantity) + "\n\n"
        "Верни СТРОГО JSON с двумя полями (без markdown, без тройных кавычек):\n"
        '{"normatives": "...", "requiredDocs": "..."}\n'
        "Где:\n"
        "- normatives: перечень применимых ГОСТ/СП/СНиП для входного контроля этого материала через запятую\n"
        "- requiredDocs: какие документы качества должны быть у поставщика "
        "(паспорт качества, сертификат соответствия, протокол испытаний, декларация). 1-2 предложения."
    )
    instructions = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, никаких тройных кавычек."
    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    def _call(model_id):
        try:
            r = client.responses.create(model="gpt://"+YANDEX_FOLDER_ID+"/"+model_id, temperature=0.1, instructions=instructions, input=user_text, max_output_tokens=1500)
            return (r.output_text or ""), None
        except Exception as e:
            return "", str(e)
    answer, err = _call("qwen3.6-35b-a3b/latest")
    if not (answer or "").strip():
        print("AI-SUGGEST inspection primary empty, fallback. err=" + str(err))
        answer, err = _call("yandexgpt-5.1/latest")
    if not (answer or "").strip():
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул пустой ответ: " + str(err))
    text = answer.strip()
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        text = m.group(0)
    try:
        parsed = j.loads(text)
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул невалидный JSON: " + str(e)[:200])
    normatives = (parsed.get("normatives") or "").strip()
    required_docs = (parsed.get("requiredDocs") or "").strip()
    cur = conn.cursor()
    cur.execute("""UPDATE material_inspection_journal
                   SET normatives=%s,
                       remarks = CASE WHEN COALESCE(remarks,'')='' THEN %s ELSE remarks END,
                       ai_filled=TRUE
                   WHERE id=%s""",
                (normatives, ("Требуемые документы: " + required_docs) if required_docs else "", id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "normatives": normatives, "requiredDocs": required_docs, "aiFilled": True}

@app.get("/cable-journal")
def list_cable_journal(project_name: str = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = """id, project_name, invoice_id, cable_brand, cross_section, cores_count,
              length_received, length_installed, drum_number, manufacturer, supplier,
              certificate_number, passport_number, insulation_before, insulation_after,
              installation_location, installation_method,
              installed_at, received_at, responsible_itr, normatives, ai_filled, created_at,
              cable_type"""
    allowed_projects = visible_project_names(_current_user)
    backfill_projects = None
    if project_name:
        backfill_projects = [project_name]
    elif allowed_projects is not None:
        backfill_projects = allowed_projects
    try:
        _backfill_cable_journal(cur, backfill_projects)
    except Exception as e:
        print("CABLE JOURNAL BACKFILL ERROR:", str(e))
    if project_name:
        if allowed_projects is not None and project_name not in allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM cable_journal WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif allowed_projects is not None:
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM cable_journal WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute(f"SELECT {cols} FROM cable_journal ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1] or "","invoiceId":r[2],
             "cableBrand":r[3] or "","crossSection":float(r[4] or 0),"coresCount":r[5],
             "lengthReceived":float(r[6] or 0),"lengthInstalled":float(r[7] or 0),
             "drumNumber":r[8] or "","manufacturer":r[9] or "","supplier":r[10] or "",
             "certificateNumber":r[11] or "","passportNumber":r[12] or "",
             "insulationBefore":float(r[13] or 0),"insulationAfter":float(r[14] or 0),
             "installationLocation":r[15] or "","installationMethod":r[16] or "",
             "installedAt":str(r[17]) if r[17] else "","receivedAt":str(r[18]) if r[18] else "",
             "responsibleItr":r[19] or "","normatives":r[20] or "",
             "aiFilled":bool(r[21]),"createdAt":str(r[22]),"cableType":r[23] or ""} for r in rows]

@app.put("/cable-journal/{id}")
def update_cable_journal(id: int, data: dict, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "cable_journal", id, _current_user)
    fields_map = [
        ('cableBrand','cable_brand'),
        ('cableType','cable_type'),
        ('crossSection','cross_section'),
        ('coresCount','cores_count'),
        ('lengthReceived','length_received'),
        ('lengthInstalled','length_installed'),
        ('drumNumber','drum_number'),
        ('manufacturer','manufacturer'),
        ('certificateNumber','certificate_number'),
        ('passportNumber','passport_number'),
        ('insulationBefore','insulation_before'),
        ('insulationAfter','insulation_after'),
        ('installationLocation','installation_location'),
        ('installationMethod','installation_method'),
        ('installedAt','installed_at'),
        ('responsibleItr','responsible_itr'),
        ('normatives','normatives'),
    ]
    sets, vals = [], []
    ai_resetting_fields = {'insulationBefore','insulationAfter','installationLocation','installationMethod','normatives','drumNumber'}
    reset_ai = False
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            v = data[js_key]
            if js_key == 'installedAt' and not v:
                v = None
            vals.append(v)
            if js_key in ai_resetting_fields:
                reset_ai = True
    if reset_ai:
        sets.append("ai_filled=FALSE")
    if not sets:
        cur.close(); conn.close()
        return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE cable_journal SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/cable-journal/{id}/ai-suggest")
def ai_suggest_cable_journal(id: int, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    import openai as oa, json as j, re
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "cable_journal", id, _current_user)
    cur.execute("SELECT cable_brand, cross_section, cores_count, length_received, cable_type FROM cable_journal WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="запись не найдена")
    brand, section, cores, length, cable_type = row[0] or "", float(row[1] or 0), row[2], float(row[3] or 0), row[4] or ""
    cur.close()

    user_text = (
        "Марка кабеля: " + brand + "\n"
        "Тип системы: " + (cable_type or "не определён") + "\n"
        "Сечение жилы (мм²): " + str(section) + "\n"
        "Количество жил: " + (str(cores) if cores else "—") + "\n"
        "Длина с барабана/бухты (м): " + str(length) + "\n\n"
        "Верни СТРОГО JSON с тремя полями (без markdown, без тройных кавычек):\n"
        '{"normatives": "...", "minInsulation": "...", "recommendations": "..."}\n'
        "Где:\n"
        "- normatives: применимые ГОСТ/СП/ПУЭ/СП для этой марки и типа системы: силовая электрика, СКС/интернет, слаботочка или пожарная сигнализация\n"
        "- minInsulation: минимальное сопротивление изоляции в МΩ по ПУЭ для этой марки (одно число или диапазон)\n"
        "- recommendations: 1-2 предложения по способу прокладки и испытаниям перед сдачей."
    )
    instructions = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, никаких тройных кавычек."
    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    def _call(model_id):
        try:
            r = client.responses.create(model="gpt://"+YANDEX_FOLDER_ID+"/"+model_id, temperature=0.1, instructions=instructions, input=user_text, max_output_tokens=1500)
            return (r.output_text or ""), None
        except Exception as e:
            return "", str(e)
    answer, err = _call("qwen3.6-35b-a3b/latest")
    if not (answer or "").strip():
        print("AI-SUGGEST cable primary empty, fallback. err=" + str(err))
        answer, err = _call("yandexgpt-5.1/latest")
    if not (answer or "").strip():
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул пустой ответ: " + str(err))
    text = answer.strip()
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        text = m.group(0)
    try:
        parsed = j.loads(text)
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул невалидный JSON: " + str(e)[:200])
    normatives = (parsed.get("normatives") or "").strip()
    min_insulation = (parsed.get("minInsulation") or "").strip()
    recommendations = (parsed.get("recommendations") or "").strip()
    full_normatives = normatives
    if min_insulation:
        full_normatives = "Мин. R изоляции по ПУЭ: " + min_insulation + " МΩ. " + normatives
    if recommendations:
        full_normatives = (full_normatives + "\n\nРекомендации: " + recommendations).strip()
    cur = conn.cursor()
    cur.execute("UPDATE cable_journal SET normatives=%s, ai_filled=TRUE WHERE id=%s", (full_normatives, id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "normatives": full_normatives, "minInsulation": min_insulation, "recommendations": recommendations, "aiFilled": True}

@app.get("/supervisor-acts")
def list_supervisor_acts(project_name: str = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = "id, project_name, act_number, act_type, description, findings, recommendations, issued_by, issued_by_role, date, photo_url, file_url, status, created_at"
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute(f"SELECT {cols} FROM supervisor_acts WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif visible_project_names(current_user) is not None:
        allowed_projects = visible_project_names(current_user)
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM supervisor_acts WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute(f"SELECT {cols} FROM supervisor_acts ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1] or "","actNumber":r[2] or "","actType":r[3] or "",
             "description":r[4] or "","findings":r[5] or "","recommendations":r[6] or "",
             "issuedBy":r[7] or "","issuedByRole":r[8] or "",
             "date":str(r[9]) if r[9] else "","photoUrl":r[10] or "","fileUrl":r[11] or "",
             "status":r[12] or "Открыт","createdAt":str(r[13])} for r in rows]

@app.post("/supervisor-acts")
def create_supervisor_act(data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    require_project_access(current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO supervisor_acts
                   (project_name, act_number, act_type, description, findings, recommendations,
                    issued_by, issued_by_role, date, photo_url, file_url, status)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (data.get("projectName",""), data.get("actNumber","") or ("САО-"+str(int(__import__("datetime").datetime.now().timestamp()))[-6:]),
                 data.get("actType","Осмотр"), data.get("description",""), data.get("findings",""),
                 data.get("recommendations",""),
                 data.get("issuedBy",""), data.get("issuedByRole","Технадзор"),
                 data.get("date") or None, data.get("photoUrl",""), data.get("fileUrl",""),
                 data.get("status","Открыт")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id": row[0], "ok": True}

@app.put("/supervisor-acts/{id}")
def update_supervisor_act(id: int, data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "supervisor_acts", id, current_user, "project_name")
    fields_map = [
        ('actType','act_type'),('description','description'),('findings','findings'),
        ('recommendations','recommendations'),('photoUrl','photo_url'),('fileUrl','file_url'),
        ('status','status'),
    ]
    sets, vals = [], []
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            vals.append(data[js_key])
    if not sets:
        cur.close(); conn.close()
        return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE supervisor_acts SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/supervisor-acts/{id}")
def delete_supervisor_act(id: int, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "supervisor_acts", id, current_user, "project_name")
    cur.execute("DELETE FROM supervisor_acts WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/tb-journal")
def list_tb_journal(project_name: str = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = "id, project_name, master_name, instructor, instruction_type, program, instruction_text, participants_json, photo_url, date, ai_filled, created_at"
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute(f"SELECT {cols} FROM tb_journal WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif visible_project_names(current_user) is not None:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM tb_journal WHERE project_name = ANY(%s) ORDER BY id DESC", (projects,))
    else:
        cur.execute(f"SELECT {cols} FROM tb_journal ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    import json as j
    out = []
    for r in rows:
        try: participants = j.loads(r[7]) if r[7] else []
        except: participants = []
        out.append({"id":r[0],"projectName":r[1] or "","masterName":r[2] or "",
             "instructor":r[3] or "","instructionType":r[4] or "",
             "program":r[5] or "","instructionText":r[6] or "",
             "participants":participants,"photoUrl":r[8] or "",
             "date":str(r[9]) if r[9] else "","aiFilled":bool(r[10]),
             "createdAt":str(r[11])})
    return out

@app.post("/tb-journal")
def create_tb_entry(data: dict, current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    import json as j
    require_project_access(current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    parts = j.dumps(data.get("participants") or [], ensure_ascii=False)
    cur.execute("""INSERT INTO tb_journal
                   (project_name, master_name, instructor, instruction_type, program, instruction_text,
                    participants_json, photo_url, date, ai_filled)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (data.get("projectName",""), data.get("masterName",""),
                 data.get("instructor",""), data.get("instructionType","Первичный инструктаж"),
                 data.get("program",""), data.get("instructionText",""),
                 parts, data.get("photoUrl",""), data.get("date") or None,
                 bool(data.get("aiFilled", False))))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id": row[0], "ok": True}

@app.delete("/tb-journal/{id}")
def delete_tb_entry(id: int, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "tb_journal", id, current_user, "project_name")
    cur.execute("DELETE FROM tb_journal WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/tb-journal/ai-generate")
def ai_generate_tb_instruction(data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    """AI генерирует текст инструктажа по ГОСТ 12.0.004-2015 для указанного типа работ."""
    import openai as oa
    instruction_type = data.get("instructionType", "Первичный инструктаж")
    work_context = data.get("workContext", "")
    project_name = data.get("projectName", "")
    if project_name:
        require_project_access(current_user, project_name)

    user_text = (
        "Тип инструктажа: " + instruction_type + "\n"
        "Объект: " + project_name + "\n"
        "Контекст работ: " + (work_context or "общие строительно-монтажные работы") + "\n\n"
        "Сгенерируй текст инструктажа по охране труда по ГОСТ 12.0.004-2015. "
        "Включи: 1) программу (3-5 пунктов), 2) основные требования техники безопасности под этот тип работ, "
        "3) последовательность действий при ЧП. "
        "Используй официальный канцелярский русский. Объём 8-15 строк."
    )
    instructions = "Ты эксперт по охране труда в строительстве. Отвечай прямым связным текстом, без markdown."
    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    def _call(model_id):
        try:
            r = client.responses.create(
                model="gpt://" + YANDEX_FOLDER_ID + "/" + model_id,
                temperature=0.2, instructions=instructions, input=user_text, max_output_tokens=2000,
            )
            return (r.output_text or ""), None
        except Exception as e:
            return "", str(e)
    answer, err = _call("yandexgpt-5.1/latest")
    if not (answer or "").strip():
        print("AI-TB primary empty, fallback. err=" + str(err))
        answer, err = _call("qwen3.6-35b-a3b/latest")
    if not (answer or "").strip():
        raise HTTPException(status_code=502, detail="AI вернул пустой ответ: " + str(err))
    return {"ok": True, "instructionText": answer.strip()}

@app.get("/audit-log")
def list_audit_log(limit: int = 200, _current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "бухгалтер"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT id, user_id, user_name, user_role, action, entity_type, entity_id, description, project_name, created_at FROM audit_log ORDER BY id DESC LIMIT %s", (limit,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"userId":r[1],"userName":r[2] or "","userRole":r[3] or "",
             "action":r[4] or "","entityType":r[5] or "","entityId":r[6],
             "description":r[7] or "","projectName":r[8] or "",
             "createdAt":str(r[9])} for r in rows]

@app.get("/inspection-orders")
def list_inspection_orders(project_name: str = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = "id, project_name, order_number, body, inspector, description, recommendations, deadline, status, photo_url, file_url, date, response, response_date, created_at"
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute(f"SELECT {cols} FROM inspection_orders WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif visible_project_names(current_user) is not None:
        allowed_projects = visible_project_names(current_user)
        if not allowed_projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM inspection_orders WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute(f"SELECT {cols} FROM inspection_orders ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1] or "","orderNumber":r[2] or "","body":r[3] or "",
             "inspector":r[4] or "","description":r[5] or "","recommendations":r[6] or "",
             "deadline":str(r[7]) if r[7] else "","status":r[8] or "Открыто",
             "photoUrl":r[9] or "","fileUrl":r[10] or "",
             "date":str(r[11]) if r[11] else "","response":r[12] or "",
             "responseDate":str(r[13]) if r[13] else "",
             "createdAt":str(r[14])} for r in rows]

@app.post("/inspection-orders")
def create_inspection_order(data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    require_project_access(current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO inspection_orders
                   (project_name, order_number, body, inspector, description, recommendations,
                    deadline, status, photo_url, file_url, date)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (data.get("projectName",""), data.get("orderNumber","") or ("ГСН-"+str(int(__import__("datetime").datetime.now().timestamp()))[-6:]),
                 data.get("body","ГСН"), data.get("inspector",""), data.get("description",""),
                 data.get("recommendations",""), data.get("deadline") or None,
                 data.get("status","Открыто"), data.get("photoUrl",""),
                 data.get("fileUrl",""), data.get("date") or None))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id": row[0], "ok": True}

@app.put("/inspection-orders/{id}")
def update_inspection_order(id: int, data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "inspection_orders", id, current_user, "project_name")
    fields_map = [('status','status'),('response','response'),('responseDate','response_date'),
                  ('recommendations','recommendations'),('photoUrl','photo_url'),('fileUrl','file_url')]
    sets, vals = [], []
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            v = data[js_key]
            if js_key == 'responseDate' and not v:
                v = None
            vals.append(v)
    if not sets:
        cur.close(); conn.close()
        return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE inspection_orders SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/expense-reports")
def list_expense_reports(employee_id: int = None, project_name: str = None, current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = "id, employee_id, employee_name, project_name, report_type, purpose, total_amount, issued_amount, spent_amount, balance, items_json, photo_url, date_from, date_to, status, approved_by, approved_at, created_at"
    where, params = [], []
    if employee_id: where.append("employee_id=%s"); params.append(employee_id)
    if project_name: where.append("project_name=%s"); params.append(project_name)
    q = f"SELECT {cols} FROM expense_reports"
    if where: q += " WHERE " + " AND ".join(where)
    q += " ORDER BY id DESC"
    cur.execute(q, params)
    rows = cur.fetchall()
    cur.close(); conn.close()
    import json as j
    out = []
    for r in rows:
        try: items = j.loads(r[10]) if r[10] else []
        except: items = []
        out.append({"id":r[0],"employeeId":r[1],"employeeName":r[2] or "",
             "projectName":r[3] or "","reportType":r[4] or "Авансовый отчёт",
             "purpose":r[5] or "","totalAmount":float(r[6] or 0),"issuedAmount":float(r[7] or 0),
             "spentAmount":float(r[8] or 0),"balance":float(r[9] or 0),"items":items,
             "photoUrl":r[11] or "","dateFrom":str(r[12]) if r[12] else "",
             "dateTo":str(r[13]) if r[13] else "","status":r[14] or "На утверждении",
             "approvedBy":r[15] or "","approvedAt":str(r[16]) if r[16] else "",
             "createdAt":str(r[17])})
    return out

@app.post("/expense-reports")
def create_expense_report(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    import json as j
    conn = get_db()
    cur = conn.cursor()
    items = j.dumps(data.get("items") or [], ensure_ascii=False)
    cur.execute("""INSERT INTO expense_reports
                   (employee_id, employee_name, project_name, report_type, purpose,
                    total_amount, issued_amount, spent_amount, balance, items_json, photo_url,
                    date_from, date_to, status)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (data.get("employeeId"), data.get("employeeName",""), data.get("projectName",""),
                 data.get("reportType","Авансовый отчёт"), data.get("purpose",""),
                 float(data.get("totalAmount",0)), float(data.get("issuedAmount",0)),
                 float(data.get("spentAmount",0)), float(data.get("balance",0)),
                 items, data.get("photoUrl",""), data.get("dateFrom") or None,
                 data.get("dateTo") or None, data.get("status","На утверждении")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id": row[0], "ok": True}

@app.put("/expense-reports/{id}")
def update_expense_report(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    fields_map = [('status','status'),('approvedBy','approved_by'),('approvedAt','approved_at'),
                  ('spentAmount','spent_amount'),('balance','balance'),('purpose','purpose')]
    sets, vals = [], []
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            v = data[js_key]
            if js_key == 'approvedAt' and not v: v = None
            vals.append(v)
    if not sets:
        cur.close(); conn.close(); return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE expense_reports SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/expense-reports/{id}")
def delete_expense_report(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM expense_reports WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/supplier-invoices")
def list_supplier_invoices(project_name: str = None, status: str = None, current_user: dict = Depends(get_current_user)):
    role = current_user.get("role")
    if role not in SUPPLIER_INVOICE_VIEW_ROLES:
        return []
    conn = get_db()
    cur = conn.cursor()
    cols = "id, supplier_id, supplier_name, project_name, invoice_number, invoice_date, amount, vat_amount, description, file_url, photo_url, status, approved_by, approved_at, paid_at, paid_by, paid_note, created_at, paid_amount, offer_id, request_id, payment_terms, material_name"
    where, params = [], []
    if project_name: where.append("project_name=%s"); params.append(project_name)
    if status: where.append("status=%s"); params.append(status)
    if role == "поставщик":
        supplier_id = current_supplier_id(cur, current_user)
        if not supplier_id:
            cur.close(); conn.close()
            return []
        where.append("supplier_id=%s"); params.append(supplier_id)
    elif role not in FINANCE_ROLES:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        where.append("project_name = ANY(%s)"); params.append(projects)
    q = f"SELECT {cols} FROM supplier_invoices"
    if where: q += " WHERE " + " AND ".join(where)
    q += " ORDER BY id DESC"
    cur.execute(q, params)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"supplierId":r[1],"supplierName":r[2] or "",
             "projectName":r[3] or "","invoiceNumber":r[4] or "",
             "invoiceDate":str(r[5]) if r[5] else "","amount":float(r[6] or 0),
             "totalAmount":float(r[6] or 0),
             "vatAmount":float(r[7] or 0),"description":r[8] or "",
             "fileUrl":r[9] or "","photoUrl":r[10] or "",
             "status":r[11] or "На утверждении","approvedBy":r[12] or "",
             "approvedAt":str(r[13]) if r[13] else "","paidAt":str(r[14]) if r[14] else "",
             "paidBy":r[15] or "","paidNote":r[16] or "",
             "createdAt":str(r[17]),"paidAmount":float(r[18] or 0),
             "offerId":r[19],"requestId":r[20],
             "paymentTerms":r[21] or "","materialName":r[22] or ""} for r in rows]

@app.post("/supplier-invoices")
def create_supplier_invoice(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "поставщик"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO supplier_invoices
                   (supplier_id, supplier_name, project_name, invoice_number, invoice_date,
                    amount, vat_amount, description, file_url, photo_url, status)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (data.get("supplierId"), data.get("supplierName",""), data.get("projectName",""),
                 data.get("invoiceNumber",""), data.get("invoiceDate") or None,
                 float(data.get("amount",0)), float(data.get("vatAmount",0)),
                 data.get("description",""), data.get("fileUrl",""), data.get("photoUrl",""),
                 data.get("status","На утверждении")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id": row[0], "ok": True}

@app.put("/supplier-invoices/{id}")
def update_supplier_invoice(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    fields_map = [('status','status'),('approvedBy','approved_by'),('approvedAt','approved_at'),
                  ('paidAt','paid_at'),('paidBy','paid_by'),('paidNote','paid_note'),
                  ('description','description'),('amount','amount'),('vatAmount','vat_amount'),
                  ('paidAmount','paid_amount')]
    sets, vals = [], []
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            v = data[js_key]
            if js_key in ('approvedAt','paidAt') and not v: v = None
            vals.append(v)
    if not sets:
        cur.close(); conn.close(); return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE supplier_invoices SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    if 'status' in data:
        log_audit(user_name=data.get("approvedBy") or data.get("paidBy") or "—", user_role="—",
                  action="status_change", entity_type="supplier_invoice", entity_id=id,
                  description="Новый статус: "+data.get('status',''),
                  project_name="")
    return {"ok": True}

@app.delete("/supplier-invoices/{id}")
def delete_supplier_invoice(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM supplier_invoices WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/warranty-defects")
def list_warranty_defects(project_name: str = None, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = "id, project_name, description, found_at, reported_by, reporter_phone, status, assigned_to, fix_notes, fixed_at, photo_url, severity, created_at"
    if project_name:
        require_project_access(current_user, project_name)
        cur.execute(f"SELECT {cols} FROM warranty_defects WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif visible_project_names(current_user) is not None:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM warranty_defects WHERE project_name = ANY(%s) ORDER BY id DESC", (projects,))
    else:
        cur.execute(f"SELECT {cols} FROM warranty_defects ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1] or "","description":r[2] or "",
             "foundAt":str(r[3]) if r[3] else "","reportedBy":r[4] or "",
             "reporterPhone":r[5] or "","status":r[6] or "Открыт","assignedTo":r[7] or "",
             "fixNotes":r[8] or "","fixedAt":str(r[9]) if r[9] else "",
             "photoUrl":r[10] or "","severity":r[11] or "",
             "createdAt":str(r[12])} for r in rows]

@app.post("/warranty-defects")
def create_warranty_defect(data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    require_project_access(current_user, data.get("projectName", ""))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO warranty_defects
                   (project_name, description, found_at, reported_by, reporter_phone,
                    status, assigned_to, photo_url, severity)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (data.get("projectName",""), data.get("description",""),
                 data.get("foundAt") or None, data.get("reportedBy",""),
                 data.get("reporterPhone",""), data.get("status","Открыт"),
                 data.get("assignedTo",""), data.get("photoUrl",""),
                 data.get("severity","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id": row[0], "ok": True}

@app.put("/warranty-defects/{id}")
def update_warranty_defect(id: int, data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "warranty_defects", id, current_user, "project_name")
    fields_map = [('status','status'),('assignedTo','assigned_to'),('fixNotes','fix_notes'),
                  ('fixedAt','fixed_at'),('severity','severity'),('photoUrl','photo_url')]
    sets, vals = [], []
    for js_key, db_col in fields_map:
        if js_key in data:
            sets.append(db_col + "=%s")
            v = data[js_key]
            if js_key == 'fixedAt' and not v: v = None
            vals.append(v)
    if not sets:
        cur.close(); conn.close(); return {"ok": True}
    vals.append(id)
    cur.execute("UPDATE warranty_defects SET " + ", ".join(sets) + " WHERE id=%s", vals)
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.delete("/warranty-defects/{id}")
def delete_warranty_defect(id: int, current_user: dict = Depends(require_roles(*LEADERSHIP_ROLES, "прораб", "главный_инженер"))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "warranty_defects", id, current_user, "project_name")
    cur.execute("DELETE FROM warranty_defects WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/unexpected-works/{id}/ai-estimate")
def ai_estimate_unexpected_work(id: int, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    """AI оценивает стоимость изменения к смете по аналогии со сметой и прайсами."""
    import openai as oa, json as j, re
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT description, unit, quantity, project_name FROM unexpected_works WHERE id=%s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="запись не найдена")
    desc, unit, qty, proj_name = row[0] or "", row[1] or "", float(row[2] or 0), row[3] or ""
    require_project_access(current_user, proj_name)
    # Подбираем похожие позиции из смет и прайсов как контекст
    cur.execute("""SELECT name, unit, price_per_unit FROM (
                       SELECT name, unit, price as price_per_unit FROM pricelist_items WHERE LOWER(name) LIKE %s
                   ) sub LIMIT 10""", ('%' + desc.lower().split()[0][:5] + '%',))
    similar = cur.fetchall()
    cur.close(); conn.close()

    similar_lines = [s[0]+' · '+(s[1] or '')+' · '+str(s[2] or 0)+' ₽/'+(s[1] or 'шт') for s in similar]
    user_text = (
        "Описание работы: " + desc + "\n"
        "Единица: " + unit + "\n"
        "Объём: " + str(qty) + "\n\n"
        "Похожие позиции из прайсов (для ориентира):\n" + ('\n'.join(similar_lines) if similar_lines else '(нет данных)') + "\n\n"
        "Верни СТРОГО JSON: {\"pricePerUnit\": число, \"justification\": \"строка\"}\n"
        "pricePerUnit — оценочная цена за единицу в рублях для строительных работ в России в 2026 году.\n"
        "justification — 1-2 строки обоснования (например: «аналог из прайса 850 ₽/м², увеличено на 15% за сложность ручной работы»)."
    )
    instructions = "Ты эксперт по строительной смете. Отвечай СТРОГО JSON без markdown."
    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    def _call(model_id):
        try:
            r = client.responses.create(model="gpt://"+YANDEX_FOLDER_ID+"/"+model_id, temperature=0.2, instructions=instructions, input=user_text, max_output_tokens=800)
            return (r.output_text or ""), None
        except Exception as e:
            return "", str(e)
    answer, err = _call("qwen3.6-35b-a3b/latest")
    if not (answer or "").strip():
        answer, err = _call("yandexgpt-5.1/latest")
    if not (answer or "").strip():
        raise HTTPException(status_code=502, detail="AI вернул пустой ответ: "+str(err))
    text = answer.strip()
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m: text = m.group(0)
    try:
        parsed = j.loads(text)
    except Exception as e:
        raise HTTPException(status_code=502, detail="AI вернул невалидный JSON")
    price = float(parsed.get("pricePerUnit") or 0)
    justification = (parsed.get("justification") or "").strip()
    estimated_total = round(price * qty, 2)
    return {"ok": True, "pricePerUnit": price, "estimatedTotal": estimated_total, "justification": justification, "similar": similar_lines}

@app.get("/unexpected-works/limit-check")
def check_unexpected_limit(project_name: str, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    """Проверка превышения контрольного % изменений к смете от бюджета проекта."""
    require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT budget FROM projects WHERE name=%s", (project_name,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="проект не найден")
    budget = float(row[0] or 0)
    cur.execute("SELECT COALESCE(SUM(total),0) FROM unexpected_works WHERE project_name=%s AND status = ANY(%s) AND COALESCE(change_type,'') <> %s AND included_in_estimate_id IS NULL", (project_name, list(ESTIMATE_CHANGE_APPROVED_STATUSES), "Исключение объёма"))
    approved_sum = float(cur.fetchone()[0] or 0)
    cur.execute("SELECT COALESCE(SUM(total),0) FROM unexpected_works WHERE project_name=%s AND status='Ожидает согласования' AND COALESCE(change_type,'') <> %s", (project_name, "Исключение объёма"))
    pending_sum = float(cur.fetchone()[0] or 0)
    cur.close(); conn.close()
    LIMIT_PCT = 10.0  # лимит 10% от бюджета без особого согласования
    percent = (approved_sum / budget * 100) if budget > 0 else 0
    over_limit = percent > LIMIT_PCT
    return {"projectName": project_name, "budget": budget, "approvedSum": approved_sum,
            "pendingSum": pending_sum, "percentOfBudget": round(percent, 2),
            "limitPct": LIMIT_PCT, "overLimit": over_limit,
            "warning": "Утверждённые изменения к смете превысили "+str(LIMIT_PCT)+"% от бюджета — стоит оформить доп.соглашение или новую редакцию сметы" if over_limit else None}

@app.delete("/inspection-orders/{id}")
def delete_inspection_order(id: int, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    require_row_project_access(cur, "inspection_orders", id, current_user, "project_name")
    cur.execute("DELETE FROM inspection_orders WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

def log_audit(user_name="", user_role="", action="", entity_type="", entity_id=None, description="", project_name=""):
    """Запись действия в audit_log. Тихо игнорирует ошибки чтобы не ломать основные операции."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""INSERT INTO audit_log (user_name, user_role, action, entity_type, entity_id, description, project_name)
                       VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                    (user_name, user_role, action, entity_type, entity_id, description, project_name))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print("AUDIT LOG ERROR:", str(e))

def save_doc_version(document_type, document_id, snapshot_json, changed_by="", change_reason=""):
    """Сохранить snapshot документа в document_versions. Возвращает label новой версии."""
    try:
        import json as _j
        from datetime import datetime as _dt
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM document_versions WHERE document_type=%s AND document_id=%s",
                    (document_type, document_id))
        count = cur.fetchone()[0]
        label = "v" + str(count+1) + "_" + _dt.now().strftime("%Y%m%d_%H%M%S")
        snap = snapshot_json if isinstance(snapshot_json, str) else _j.dumps(snapshot_json, ensure_ascii=False, default=str)
        cur.execute("""INSERT INTO document_versions (document_type, document_id, version_label, snapshot_json, changed_by, change_reason)
                       VALUES (%s,%s,%s,%s,%s,%s)""",
                    (document_type, document_id, label, snap, changed_by, change_reason))
        conn.commit()
        cur.close(); conn.close()
        return label
    except Exception as e:
        print("VERSION SAVE ERROR:", str(e))
        return None

@app.get("/document-versions")
def list_document_versions(document_type: str = None, document_id: int = None, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = "id, document_type, document_id, version_label, changed_by, change_reason, created_at"
    if document_type and document_id is not None:
        cur.execute(f"SELECT {cols} FROM document_versions WHERE document_type=%s AND document_id=%s ORDER BY created_at DESC",
                    (document_type, document_id))
    elif document_type:
        cur.execute(f"SELECT {cols} FROM document_versions WHERE document_type=%s ORDER BY created_at DESC LIMIT 200",
                    (document_type,))
    else:
        cur.execute(f"SELECT {cols} FROM document_versions ORDER BY created_at DESC LIMIT 200")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"documentType":r[1],"documentId":r[2],"versionLabel":r[3],
             "changedBy":r[4] or "","changeReason":r[5] or "","createdAt":str(r[6])} for r in rows]

@app.get("/document-versions/{vid}")
def get_document_version(vid: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    import json as _j
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""SELECT id, document_type, document_id, version_label, snapshot_json,
                          changed_by, change_reason, created_at
                   FROM document_versions WHERE id=%s""", (vid,))
    r = cur.fetchone()
    cur.close(); conn.close()
    if not r:
        return {"error": "not found"}
    try: snap = _j.loads(r[4]) if r[4] else {}
    except: snap = {}
    return {"id":r[0],"documentType":r[1],"documentId":r[2],"versionLabel":r[3],
            "snapshot":snap,"changedBy":r[5] or "","changeReason":r[6] or "",
            "createdAt":str(r[7])}

@app.post("/audit-log")
def create_audit_entry(data: dict, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""INSERT INTO audit_log
                   (user_id, user_name, user_role, action, entity_type, entity_id, description, project_name)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                (current_user.get("id"), current_user.get("name",""), current_user.get("role",""),
                 data.get("action",""), data.get("entityType",""), data.get("entityId"),
                 data.get("description",""), data.get("projectName","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id": row[0], "ok": True}

# Хранилище онлайн статусов
online_users = {}

@app.post("/online")
def update_online(data: dict, current_user: dict = Depends(get_current_user)):
    user_id = current_user.get("id")
    if user_id:
        online_users[str(user_id)] = {
            "userId": user_id,
            "userName": current_user.get("name",""),
            "userRole": current_user.get("role",""),
            "lastSeen": data.get("lastSeen",""),
            "page": data.get("page","")
        }
    return {"ok": True}

@app.get("/online")
def get_online(_current_user: dict = Depends(get_current_user)):
    import time
    now = time.time()
    # Возвращаем пользователей активных за последние 2 минуты
    return list(online_users.values())

@app.get("/project-payments")
def get_project_payments(project_name: str = "", current_user: dict = Depends(get_current_user)):
    role = current_user.get("role")
    if role not in FINANCE_ROLES and role != "заказчик":
        return []
    conn = get_db()
    cur = conn.cursor()
    customer_projects = user_project_names(current_user) if role == "заказчик" else None
    if role == "заказчик" and project_name and project_name not in customer_projects:
        cur.close(); conn.close()
        return []
    if role == "заказчик" and not customer_projects:
        cur.close(); conn.close()
        return []
    if role == "заказчик":
        visible_pay_projects = [project_name] if project_name else customer_projects
        cur.execute("""SELECT id,project_name,amount,note,date,added_by
                       FROM (
                           SELECT DISTINCT ON (project_name, amount, COALESCE(note,''), date, COALESCE(added_by,''))
                                  id,project_name,amount,note,date,added_by
                           FROM project_payments
                           WHERE project_name = ANY(%s) AND amount > 0
                           ORDER BY project_name, amount, COALESCE(note,''), date, COALESCE(added_by,''), id DESC
                       ) p
                       ORDER BY id DESC""", (visible_pay_projects,))
    elif project_name:
        cur.execute("""SELECT id,project_name,amount,note,date,added_by
                       FROM (
                           SELECT DISTINCT ON (project_name, amount, COALESCE(note,''), date, COALESCE(added_by,''))
                                  id,project_name,amount,note,date,added_by
                           FROM project_payments
                           WHERE project_name=%s
                           ORDER BY project_name, amount, COALESCE(note,''), date, COALESCE(added_by,''), id DESC
                       ) p
                       ORDER BY id DESC""", (project_name,))
    else:
        cur.execute("""SELECT id,project_name,amount,note,date,added_by
                       FROM (
                           SELECT DISTINCT ON (project_name, amount, COALESCE(note,''), date, COALESCE(added_by,''))
                                  id,project_name,amount,note,date,added_by
                           FROM project_payments
                           ORDER BY project_name, amount, COALESCE(note,''), date, COALESCE(added_by,''), id DESC
                       ) p
                       ORDER BY id DESC""")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"amount":float(r[2] or 0),"note":r[3] or "","date":str(r[4]) if r[4] else "","addedBy":r[5] or ""} for r in rows]

@app.post("/project-payments")
def create_project_payment(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    project_name = data.get("projectName","")
    amount = data.get("amount",0)
    note = data.get("note","")
    pay_date = data.get("date") or None
    added_by = data.get("addedBy") or data.get("paidBy") or ""
    if note:
        cur.execute("""SELECT id FROM project_payments
                       WHERE project_name=%s AND amount=%s AND COALESCE(note,'')=%s
                         AND date IS NOT DISTINCT FROM %s AND COALESCE(added_by,'')=%s
                       ORDER BY id DESC LIMIT 1""",
                    (project_name, amount, note, pay_date, added_by))
        existing = cur.fetchone()
        if existing:
            cur.close(); conn.close()
            return {"id": existing[0], "ok": True, "duplicate": True}
    cur.execute("INSERT INTO project_payments (project_name,amount,note,date,added_by) VALUES (%s,%s,%s,%s,%s) RETURNING id",
        (project_name, amount, note, pay_date, added_by))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.delete("/project-payments/{id}")
def delete_project_payment(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM project_payments WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

def send_vk_notification(vk_user_id: int, message: str):
    import requests
    if not VK_TOKEN:
        print("VK notify skipped: VK_TOKEN is not configured")
        return
    try:
        requests.post("https://api.vk.com/method/messages.send", params={
            "user_id": vk_user_id,
            "message": message,
            "random_id": 0,
            "access_token": VK_TOKEN,
            "v": "5.131"
        }, timeout=5)
    except Exception as e:
        print("VK error:", e)

@app.post("/vk-connect")
def vk_connect(data: dict, current_user: dict = Depends(get_current_user)):
    if data.get("email") and data.get("email") != current_user.get("email"):
        raise HTTPException(status_code=403, detail="Можно подключить только свой VK")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS vk_id INTEGER")
    cur.execute("UPDATE users SET vk_id=%s WHERE email=%s", (data.get("vkId"), data.get("email")))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/vk-notify")
def vk_notify(data: dict, _current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT vk_id FROM users WHERE id=%s", (data.get("userId"),))
    row = cur.fetchone()
    cur.close(); conn.close()
    if row and row[0]:
        send_vk_notification(row[0], data.get("message",""))
    return {"ok": True}

@app.get("/accountable-payments")
def get_accountable_payments(project_name: str = "", _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    if project_name:
        cur.execute("SELECT id,project_name,given_to,amount,payment_method,purpose,date,added_by,status,spent_amount FROM accountable_payments WHERE project_name=%s ORDER BY id DESC", (project_name,))
    else:
        cur.execute("SELECT id,project_name,given_to,amount,payment_method,purpose,date,added_by,status,spent_amount FROM accountable_payments ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"givenTo":r[2],"amount":float(r[3] or 0),"paymentMethod":r[4] or "Наличные","purpose":r[5] or "","date":str(r[6]) if r[6] else "","addedBy":r[7] or "","status":r[8] or "Открыт","spentAmount":float(r[9] or 0)} for r in rows]

@app.post("/accountable-payments")
def create_accountable_payment(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO accountable_payments (project_name,given_to,given_to_id,amount,payment_method,purpose,date,added_by) VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("projectName",""),data.get("givenTo",""),data.get("givenToId"),data.get("amount",0),data.get("paymentMethod","Наличные"),data.get("purpose",""),data.get("date") or None,data.get("addedBy","")))
    conn.commit()
    row = cur.fetchone()
    cur.close(); conn.close()
    return {"id":row[0],"ok":True}

@app.get("/accountable-expenses")
def get_accountable_expenses(payment_id: int = 0, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    if payment_id:
        cur.execute("SELECT id,payment_id,project_name,description,amount,photo_url,date,added_by FROM accountable_expenses WHERE payment_id=%s ORDER BY id DESC", (payment_id,))
    else:
        cur.execute("SELECT id,payment_id,project_name,description,amount,photo_url,date,added_by FROM accountable_expenses ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"paymentId":r[1],"projectName":r[2],"description":r[3],"amount":float(r[4] or 0),"photoUrl":r[5] or "","date":str(r[6]) if r[6] else "","addedBy":r[7] or ""} for r in rows]

@app.post("/accountable-expenses")
def create_accountable_expense(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO accountable_expenses (payment_id,project_name,description,amount,photo_url,date,added_by) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (data.get("paymentId"),data.get("projectName",""),data.get("description",""),data.get("amount",0),data.get("photoUrl",""),data.get("date") or None,data.get("addedBy","")))
    # Обновляем spent_amount
    cur.execute("UPDATE accountable_payments SET spent_amount=spent_amount+%s WHERE id=%s", (data.get("amount",0),data.get("paymentId")))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}
    return {"id":row[0],"ok":True}

@app.get("/own-expenses")
def get_own_expenses(project_name: str = "", employee_name: str = "", current_user: dict = Depends(require_roles(*OWN_EXPENSE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cols = "id,project_name,employee_name,description,amount,photo_url,date,status,approved_by,category,employee_id"
    where, params = [], []

    if project_name:
        require_project_access(current_user, project_name)
        where.append("project_name=%s")
        params.append(project_name)
    if employee_name:
        where.append("employee_name=%s")
        params.append(employee_name)

    role = current_user.get("role")
    if can_see_all_company_data(current_user):
        pass
    elif role in ("мастер", "субподрядчик"):
        where.append("(employee_id=%s OR employee_name=%s)")
        params.extend([current_user.get("id"), current_user.get("name") or ""])
    else:
        projects = user_project_names(current_user)
        if not projects:
            cur.close(); conn.close()
            return []
        where.append("project_name = ANY(%s)")
        params.append(projects)

    q = f"SELECT {cols} FROM own_expenses"
    if where:
        q += " WHERE " + " AND ".join(where)
    q += " ORDER BY id DESC"
    cur.execute(q, params)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"projectName":r[1],"employeeName":r[2],"description":r[3],"amount":float(r[4] or 0),"photoUrl":r[5] or "","date":str(r[6]) if r[6] else "","status":r[7] or "Ожидает","approvedBy":r[8] or "","category":r[9] or "other","employeeId":r[10]} for r in rows]

@app.post("/own-expenses")
def create_own_expense(data: dict, current_user: dict = Depends(require_roles(*OWN_EXPENSE_ROLES))):
    project_name = data.get("projectName", "")
    require_project_access(current_user, project_name)
    if current_user.get("role") in FINANCE_ROLES:
        employee_name = data.get("employeeName") or current_user.get("name") or ""
        employee_id = data.get("employeeId") or current_user.get("id")
    else:
        employee_name = current_user.get("name") or ""
        employee_id = current_user.get("id")
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO own_expenses (project_name,employee_name,employee_id,description,amount,photo_url,date,category) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
        (project_name,employee_name,employee_id,data.get("description",""),data.get("amount",0),data.get("photoUrl",""),data.get("date") or None,data.get("category","other")))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.put("/own-expenses/{id}")
def update_own_expense(id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE own_expenses SET status=%s,approved_by=%s WHERE id=%s",
        (data.get("status","Ожидает"),data.get("approvedBy",""),id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.delete("/own-expenses/{id}")
def delete_own_expense(id: int, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM own_expenses WHERE id=%s", (id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.get("/expenses")
def get_expenses(project: str = "", _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    if project:
        cur.execute("SELECT id,project,category,amount,note,date,added_by FROM expenses WHERE project=%s ORDER BY id DESC", (project,))
    else:
        cur.execute("SELECT id,project,category,amount,note,date,added_by FROM expenses ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id":r[0],"project":r[1],"category":r[2],"amount":float(r[3] or 0),"note":r[4] or "","date":str(r[5]) if r[5] else "","addedBy":r[6] or ""} for r in rows]

@app.post("/expenses")
def create_expense(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO expenses (project,category,amount,note,date,added_by) VALUES (%s,%s,%s,%s,%s,%s)",
        (data.get("project",""),data.get("category","other"),data.get("amount",0),data.get("note",""),data.get("date") or None,data.get("addedBy","")))
    conn.commit()
    cur.close(); conn.close()
    return {"ok":True}

@app.get("/project-ai-summary/{project_name}")
def get_project_ai_summary(project_name: str, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT payload_hash, summary, updated_at FROM project_ai_summary WHERE project_name=%s", (project_name,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return {"exists": False}
    return {"exists": True, "payloadHash": row[0], "summary": row[1] or "", "updatedAt": str(row[2])}

@app.post("/project-ai-summary")
def save_project_ai_summary(data: dict, current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    project_name = data.get("projectName", "")
    if not project_name:
        raise HTTPException(status_code=400, detail="projectName required")
    require_project_access(current_user, project_name)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO project_ai_summary (project_name, payload_hash, summary, updated_at)
        VALUES (%s, %s, %s, NOW())
        ON CONFLICT (project_name) DO UPDATE SET
            payload_hash = EXCLUDED.payload_hash,
            summary = EXCLUDED.summary,
            updated_at = NOW()
    """, (project_name, data.get("payloadHash", ""), data.get("summary", "")))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/ai-generate-estimate")
def ai_generate_estimate(data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    import openai as oa
    import json as _json
    description = (data.get("description") or "").strip()
    project_id = data.get("projectId")
    pricelist_id = data.get("pricelistId")
    name_hint = (data.get("name") or "Сгенерированная смета").strip()
    area = data.get("area")
    smeta_type = data.get("smetaType") or "Заказчик"
    status = data.get("status") or "Активная"
    work_package = data.get("workPackage") or data.get("work_package") or "Основная"
    if not description:
        raise HTTPException(status_code=400, detail="Опишите объект — поле обязательно")

    conn = get_db()
    cur = conn.cursor()

    pricelist_items = []
    project_name = ""
    if project_id:
        try:
            cur.execute("SELECT name, pricelist_id FROM projects WHERE id=%s", (int(project_id),))
            r = cur.fetchone()
            if r:
                project_name = r[0] or ""
                if not pricelist_id and r[1]:
                    pricelist_id = r[1]
        except Exception:
            pass
    if pricelist_id:
        cur.execute("SELECT id, name, unit, price, category FROM pricelist_items WHERE pricelist_id=%s ORDER BY category, name LIMIT 400", (int(pricelist_id),))
        pricelist_items = [{"id": r[0], "name": r[1], "unit": r[2], "price": float(r[3] or 0), "category": r[4] or ""} for r in cur.fetchall()]

    parts = []
    parts.append("Ты опытный сметчик. Составь смету для объекта на основе описания. ВЕРНИ СТРОГО ВАЛИДНЫЙ JSON без markdown.")
    parts.append("\nОПИСАНИЕ ОБЪЕКТА:\n" + description)
    if area:
        parts.append("Площадь: " + str(area) + " м²")
    if project_name:
        parts.append("Проект: " + project_name)
    if pricelist_items:
        parts.append("\nИспользуй ТОЛЬКО позиции из этого прайс-листа (поле name копируй точно):")
        for it in pricelist_items:
            parts.append("- ["+str(it["category"])+"] "+it["name"]+" — "+it["unit"]+" — "+str(int(it["price"]))+"₽")
    else:
        parts.append("\nПрайс-листа нет — используй типовые работы и материалы для подобных объектов с рыночными ценами по РФ 2026.")

    parts.append("""
ФОРМАТ ОТВЕТА — ровно такой JSON:
{
  "name": "название сметы одной строкой",
  "sections": [
    {"name": "Раздел 1. ...", "items": [
      {"name": "...", "unit": "м2|шт|кг|...", "quantity": число, "priceWork": число, "priceMaterial": число, "itemType": "work" или "material"}
    ]}
  ]
}

ПРАВИЛА:
1. Раздели смету на 4-8 разделов по этапам работ (демонтаж, основание, отделка, сантехника и т.д.).
2. В каждом разделе сначала идут работы (itemType="work" с priceWork>0, priceMaterial=0), затем материалы (itemType="material" с priceMaterial>0, priceWork=0).
3. Объёмы (quantity) прикинь реалистично по площади и описанию.
4. Если в прайсе нет нужной позиции — добавь её всё равно (только в крайнем случае) с рыночной ценой.
5. Числа БЕЗ пробелов и валюты — только цифры. quantity может быть дробным.
6. ТОЛЬКО валидный JSON. Никакого текста до или после.""")
    full_prompt = "\n".join(parts)

    instructions = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, ```, никакого текста до или после JSON. Только сам JSON."

    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    try:
        response = client.responses.create(
            model="gpt://" + YANDEX_FOLDER_ID + "/qwen3.6-35b-a3b/latest",
            temperature=0.2,
            instructions=instructions,
            input=full_prompt,
            max_output_tokens=6000,
        )
        raw = (response.output_text or "").strip()
    except Exception as e:
        cur.close(); conn.close()
        print("AI-GENERATE EXCEPTION:", str(e))
        raise HTTPException(status_code=500, detail="Ошибка ИИ: " + str(e))

    print("AI-GENERATE RAW LEN:", len(raw))
    print("AI-GENERATE RAW HEAD:", raw[:300])
    print("AI-GENERATE RAW TAIL:", raw[-300:] if len(raw) > 300 else "")

    import re as _re
    clean = raw.strip()
    clean = _re.sub(r"^```(?:json|JSON)?\s*", "", clean)
    clean = _re.sub(r"\s*```\s*$", "", clean)
    clean = clean.strip()

    s_idx = clean.find("{")
    e_idx = clean.rfind("}")
    parsed = None
    parse_error = None
    if s_idx >= 0 and e_idx > s_idx:
        candidate = clean[s_idx:e_idx + 1]
        try:
            parsed = _json.loads(candidate)
        except Exception as pe:
            parse_error = str(pe)
            # Trailing-junk fallback: cut from the last } backwards
            for cut in range(e_idx, s_idx, -1):
                if clean[cut] == "}":
                    try:
                        parsed = _json.loads(clean[s_idx:cut + 1])
                        parse_error = None
                        break
                    except Exception:
                        continue

    if not parsed or not isinstance(parsed.get("sections"), list):
        cur.close(); conn.close()
        print("AI-GENERATE PARSE FAILED. parse_error:", parse_error)
        raise HTTPException(status_code=500, detail="Не удалось распознать ответ ИИ как JSON. Попробуйте ещё раз или измените описание. Подробности в логах backend.")

    sections = []
    import time as _time
    for i, s in enumerate(parsed.get("sections") or []):
        items = []
        for j, it in enumerate(s.get("items") or []):
            items.append({
                "id": int(_time.time()*1000) + i*100 + j,
                "name": str(it.get("name") or ""),
                "unit": str(it.get("unit") or "шт"),
                "quantity": float(it.get("quantity") or 0),
                "priceWork": float(it.get("priceWork") or 0),
                "priceMaterial": float(it.get("priceMaterial") or 0),
                "itemType": str(it.get("itemType") or ("material" if float(it.get("priceMaterial") or 0) > 0 and float(it.get("priceWork") or 0) == 0 else "work")),
            })
        sections.append({"id": int(_time.time()*1000) + i, "name": str(s.get("name") or ("Раздел " + str(i+1))), "items": items})

    final_name = (parsed.get("name") or name_hint or "Сгенерированная смета")
    if status == "Активная" and project_name:
        cur.execute("""UPDATE estimates SET status='Архив'
                       WHERE project_name=%s
                         AND COALESCE(smeta_type,'Заказчик')=%s
                         AND COALESCE(work_package,'Основная')=%s""",
                    (project_name, smeta_type, work_package))
    cur.execute("""INSERT INTO estimates
                   (project_id, project_name, name, version, sections_json, smeta_type, work_package, status)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (int(project_id) if project_id else None, project_name, final_name, "1.0", _json.dumps(sections, ensure_ascii=False), smeta_type, work_package, status))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close(); conn.close()

    return {"ok": True, "id": new_id, "name": final_name, "projectId": project_id, "projectName": project_name, "sections": sections, "smetaType": smeta_type, "workPackage": work_package, "status": status}

@app.post("/pricelists/from-estimate")
def pricelist_from_estimate(data: dict, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    import json as _json
    estimate_id = data.get("estimateId")
    name = (data.get("name") or "").strip()
    for_who = (data.get("forWho") or "").strip()
    coefficient = float(data.get("coefficient") or 1.0)
    if not estimate_id:
        raise HTTPException(status_code=400, detail="Выберите смету")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT name, sections_json FROM estimates WHERE id=%s", (int(estimate_id),))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    estimate_name = row[0] or ""
    try:
        sections = _json.loads(row[1]) if row[1] else []
    except Exception:
        sections = []

    final_name = name or ("Прайс из сметы — " + estimate_name)
    cur.execute("INSERT INTO pricelists (name, description, for_who, coefficient) VALUES (%s, %s, %s, %s) RETURNING id",
                (final_name, "Создан из сметы: " + estimate_name, for_who, coefficient))
    pricelist_id = cur.fetchone()[0]

    inserted = 0
    seen = set()
    for s in sections:
        category = str(s.get("name") or "")
        for it in (s.get("items") or []):
            nm = str(it.get("name") or "").strip()
            if not nm:
                continue
            key = (nm.lower(), category.lower())
            if key in seen:
                continue
            seen.add(key)
            unit = str(it.get("unit") or "шт")
            item_type = str(it.get("itemType") or "")
            try:
                price_work = float(it.get("priceWork") or 0)
                price_material = float(it.get("priceMaterial") or 0)
                qty = float(it.get("quantity") or 0)
            except Exception:
                price_work = price_material = qty = 0
            is_imported = bool(it.get("isImported"))
            if item_type == "material" or (price_material > 0 and price_work == 0):
                kind = "material"
                price = price_material if not is_imported or qty == 0 else (price_material / qty if qty > 0 else price_material)
            else:
                kind = "work"
                price = price_work if not is_imported or qty == 0 else (price_work / qty if qty > 0 else price_work)
            cur.execute("INSERT INTO pricelist_items (pricelist_id, name, unit, price, category, specialization, item_type) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                        (pricelist_id, nm, unit, round(price, 2), category, for_who, kind))
            inserted += 1

    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "id": pricelist_id, "name": final_name, "itemsCount": inserted}

@app.post("/ai-generate-pricelist")
def ai_generate_pricelist(data: dict, _current_user: dict = Depends(require_roles(*PRICELIST_MANAGE_ROLES))):
    import openai as oa
    import json as _json
    description = (data.get("description") or "").strip()
    name_hint = (data.get("name") or "Прайс-лист (ИИ)").strip()
    for_who = (data.get("forWho") or "").strip()
    coefficient = float(data.get("coefficient") or 1.0)
    if not description:
        raise HTTPException(status_code=400, detail="Опишите для каких работ нужен прайс — поле обязательно")

    instructions = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, ```, никакого текста до или после JSON. Только сам JSON."

    parts = []
    parts.append("Ты опытный сметчик. Составь прайс-лист с реалистичными рыночными ценами по РФ 2026.")
    parts.append("\nОПИСАНИЕ:\n" + description)
    if for_who:
        parts.append("Для специализации: " + for_who)
    parts.append("""
ФОРМАТ ОТВЕТА — ровно такой JSON:
{
  "name": "название прайс-листа",
  "items": [
    {"name": "название позиции", "unit": "м2|шт|кг|м|м.п.|т|...", "price": число, "category": "категория"}
  ]
}

ПРАВИЛА:
1. 25-60 позиций, охватывающих описание.
2. Группируй позиции по категориям ("Демонтаж","Стены","Полы","Потолок","Сантехника","Электрика","Отделка","Прочее" и т.д.).
3. Цены — за единицу (за м², за шт, за кг). Без пробелов, без валюты, только цифры. Дробные значения допускаются.
4. Если работа сложная — указывай отдельные позиции по этапам.
5. ТОЛЬКО валидный JSON, никакого текста до/после.""")
    full_prompt = "\n".join(parts)

    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)

    def _call(model_id):
        try:
            r = client.responses.create(
                model="gpt://" + YANDEX_FOLDER_ID + "/" + model_id,
                temperature=0.2,
                instructions=instructions,
                input=full_prompt,
                max_output_tokens=5000,
            )
            return (r.output_text or "").strip(), None
        except Exception as e:
            return "", str(e)

    raw, err = _call("qwen3.6-35b-a3b/latest")
    if not raw.strip():
        print("AI-PRICELIST PRIMARY EMPTY, err=" + str(err))
        raw, err = _call("yandexgpt-5.1/latest")
    print("AI-PRICELIST RAW LEN:", len(raw))
    print("AI-PRICELIST RAW HEAD:", raw[:300])
    if not raw.strip():
        raise HTTPException(status_code=500, detail="ИИ вернул пустой ответ. Попробуйте ещё раз.")

    import re as _re
    clean = raw.strip()
    clean = _re.sub(r"^```(?:json|JSON)?\s*", "", clean)
    clean = _re.sub(r"\s*```\s*$", "", clean)
    clean = clean.strip()
    s_idx = clean.find("{")
    e_idx = clean.rfind("}")
    parsed = None
    if s_idx >= 0 and e_idx > s_idx:
        candidate = clean[s_idx:e_idx + 1]
        try:
            parsed = _json.loads(candidate)
        except Exception:
            for cut in range(e_idx, s_idx, -1):
                if clean[cut] == "}":
                    try:
                        parsed = _json.loads(clean[s_idx:cut + 1])
                        break
                    except Exception:
                        continue
    if not parsed or not isinstance(parsed.get("items"), list) or not parsed.get("items"):
        print("AI-PRICELIST PARSE FAILED")
        raise HTTPException(status_code=500, detail="Не удалось распознать ответ ИИ как JSON. Попробуйте ещё раз или измените описание.")

    final_name = parsed.get("name") or name_hint
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO pricelists (name, description, for_who, coefficient) VALUES (%s, %s, %s, %s) RETURNING id",
                (final_name, description[:500], for_who, coefficient))
    pricelist_id = cur.fetchone()[0]
    inserted = 0
    for it in parsed.get("items") or []:
        nm = str(it.get("name") or "").strip()
        if not nm:
            continue
        unit = str(it.get("unit") or "шт")
        try:
            price = float(it.get("price") or 0)
        except Exception:
            price = 0
        category = str(it.get("category") or "")
        cur.execute("INSERT INTO pricelist_items (pricelist_id, name, unit, price, category, specialization) VALUES (%s, %s, %s, %s, %s, %s)",
                    (pricelist_id, nm, unit, price, category, for_who))
        inserted += 1
    conn.commit()
    cur.close(); conn.close()

    return {"ok": True, "id": pricelist_id, "name": final_name, "itemsCount": inserted}

@app.get("/hidden-works-acts")
def list_hidden_works_acts(project_name: str = None, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cols = """id, project_name, estimate_id, act_number, work_name, section_name, brigade,
              quantity, unit, price_per_unit, total, work_date, status,
              signed_customer, signed_supervisor, signed_contractor, signed_subcontractor,
              signed_customer_at, signed_supervisor_at, signed_contractor_at, signed_subcontractor_at,
              conclusion, comments, materials_used, project_docs,
              photos, certificates, city, ai_filled,
              paid_status, paid_amount, paid_at, paid_by, paid_note,
              created_at"""
    allowed_projects = user_project_names(current_user)
    role = current_user.get("role")
    if role not in ("директор", "зам_директора", "бухгалтер", "главный_инженер", "сметчик", "прораб", "стройконтроль", "технадзор", "заказчик"):
        cur.close(); conn.close()
        return []
    if project_name:
        if allowed_projects and project_name not in allowed_projects and role not in ("директор", "зам_директора", "бухгалтер", "главный_инженер", "сметчик"):
            cur.close(); conn.close()
            return []
        cur.execute(f"SELECT {cols} FROM hidden_works_acts WHERE project_name=%s ORDER BY id DESC", (project_name,))
    elif allowed_projects and role not in ("директор", "зам_директора", "бухгалтер", "главный_инженер", "сметчик"):
        cur.execute(f"SELECT {cols} FROM hidden_works_acts WHERE project_name = ANY(%s) ORDER BY id DESC", (allowed_projects,))
    else:
        cur.execute(f"SELECT {cols} FROM hidden_works_acts ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close(); conn.close()
    result = []
    for r in rows:
        signed_customer = r[13] or ""
        signed_supervisor = r[14] or ""
        signed_contractor = r[15] or ""
        signed_subcontractor = r[16] or ""
        result.append({"id":r[0],"projectName":r[1],"estimateId":r[2],"actNumber":r[3],"workName":r[4],
             "sectionName":r[5],"brigade":r[6],"quantity":float(r[7] or 0),"unit":r[8],
             "pricePerUnit":float(r[9] or 0),"total":float(r[10] or 0),"workDate":str(r[11]) if r[11] else "",
             "status":hidden_work_effective_status(r[12], signed_customer, signed_supervisor, signed_contractor, signed_subcontractor),
             "signedCustomer":signed_customer,"signedSupervisor":signed_supervisor,
             "signedContractor":signed_contractor,"signedSubcontractor":signed_subcontractor,
             "signedCustomerAt":str(r[17]) if r[17] else "","signedSupervisorAt":str(r[18]) if r[18] else "",
             "signedContractorAt":str(r[19]) if r[19] else "","signedSubcontractorAt":str(r[20]) if r[20] else "",
             "conclusion":r[21] or "","comments":r[22] or "",
             "materialsUsed":r[23] or "","projectDocs":r[24] or "",
             "photos":r[25] or "","certificates":r[26] or "","city":r[27] or "",
             "aiFilled":bool(r[28]),
             "paidStatus":r[29] or "Не оплачен","paidAmount":float(r[30] or 0),
             "paidAt":str(r[31]) if r[31] else "","paidBy":r[32] or "","paidNote":r[33] or "",
             "createdAt":str(r[34])})
    return result

@app.put("/hidden-works-acts/{act_id}")
def update_hidden_works_act(act_id: int, data: dict, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    # Снапшот текущего состояния перед изменением (версионирование)
    try:
        cur.execute("SELECT row_to_json(t) FROM hidden_works_acts t WHERE id=%s", (act_id,))
        prev_row = cur.fetchone()
        if prev_row and prev_row[0]:
            save_doc_version("hidden_works_act", act_id, prev_row[0],
                             changed_by=data.get("_actor","—"),
                             change_reason="PUT /hidden-works-acts/"+str(act_id))
    except Exception as e:
        print("VERSION snapshot skipped:", str(e))
    # Подписи помогают контролю, но не являются обязательным стоп-краном.
    sc = data.get("signedCustomer","").strip()
    ss = data.get("signedSupervisor","").strip()
    sk = data.get("signedContractor","").strip()
    sb = data.get("signedSubcontractor","").strip()
    auto_status = hidden_work_effective_status(data.get("status","Черновик"), sc, ss, sk, sb)
    cur.execute("""UPDATE hidden_works_acts SET
                   status=%s,
                   signed_customer=%s, signed_supervisor=%s, signed_contractor=%s, signed_subcontractor=%s,
                   signed_customer_at=%s, signed_supervisor_at=%s, signed_contractor_at=%s, signed_subcontractor_at=%s,
                   conclusion=%s, comments=%s, project_docs=%s, materials_used=%s,
                   photos=%s, certificates=%s, city=%s,
                   ai_filled=FALSE
                   WHERE id=%s""",
        (auto_status,
         sc, ss, sk, sb,
         data.get("signedCustomerAt") or None, data.get("signedSupervisorAt") or None,
         data.get("signedContractorAt") or None, data.get("signedSubcontractorAt") or None,
         data.get("conclusion",""), data.get("comments",""),
         data.get("projectDocs",""), data.get("materialsUsed",""),
         data.get("photos",""), data.get("certificates",""), data.get("city",""),
         act_id))
    conn.commit(); cur.close(); conn.close()
    log_audit(user_name=data.get("_actor","система"), user_role="—",
              action="update", entity_type="hidden_works_act", entity_id=act_id,
              description="Изменён АОСР, статус: "+auto_status, project_name=data.get("_project",""))
    return {"ok": True, "status": auto_status}

@app.post("/hidden-works-acts/{act_id}/pay")
def pay_hidden_works_act(act_id: int, data: dict, _current_user: dict = Depends(require_roles(*FINANCE_ROLES))):
    """Отметить оплату в карточке АОСР без влияния на финансы объекта."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT project_name, act_number, total FROM hidden_works_acts WHERE id=%s", (act_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="акт не найден")
    proj, act_no, default_total = row[0] or "", row[1] or "", float(row[2] or 0)
    amount = float(data.get("amount") or default_total)
    paid_by = (data.get("paidBy") or "").strip()
    paid_note = (data.get("paidNote") or "Оплата по АОСР " + act_no).strip()
    paid_at = data.get("paidAt") or __import__("datetime").date.today().isoformat()
    cur.execute("""UPDATE hidden_works_acts
                   SET paid_status='Оплачен', paid_amount=%s, paid_at=%s, paid_by=%s, paid_note=%s
                   WHERE id=%s""",
                (amount, paid_at, paid_by, paid_note, act_id))
    conn.commit()
    cur.close(); conn.close()
    log_audit(user_name=paid_by or "—", user_role="—",
              action="pay", entity_type="hidden_works_act", entity_id=act_id,
              description="Отмечена оплата в карточке АОСР "+act_no+" на сумму "+str(amount)+" ₽",
              project_name=proj)
    return {"ok": True, "paidStatus": "Оплачен", "paidAmount": amount, "paidAt": paid_at, "financeDetached": True}

@app.delete("/hidden-works-acts/{act_id}")
def delete_hidden_works_act(act_id: int, _current_user: dict = Depends(require_roles(*PROJECT_DOCUMENT_WRITE_ROLES))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM hidden_works_acts WHERE id=%s", (act_id,))
    conn.commit(); cur.close(); conn.close()
    return {"ok": True}

@app.post("/hidden-works-acts/{act_id}/ai-prefill")
def ai_prefill_hidden_works_act(act_id: int, _current_user: dict = Depends(require_roles(*JOURNAL_WRITE_ROLES))):
    import openai as oa, json as j, re
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""SELECT work_name, section_name, brigade, materials_used, unit, quantity, project_name
                   FROM hidden_works_acts WHERE id=%s""", (act_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="act not found")
    work_name, section_name, brigade, materials_used, unit, quantity, project_name = row
    cur.close()

    user_text = (
        "Работа: " + (work_name or "—") + "\n"
        "Раздел сметы: " + (section_name or "—") + "\n"
        "Бригада: " + (brigade or "—") + "\n"
        "Объём: " + str(quantity or 0) + " " + (unit or "") + "\n"
        "Использованные материалы (если указаны): " + (materials_used or "—") + "\n\n"
        "Верни СТРОГО JSON-объект с тремя полями (без markdown, без тройных кавычек):\n"
        '{"conclusion": "...", "normatives": "...", "projectDocs": "..."}\n'
        "Где:\n"
        "- conclusion: формальная формулировка заключения комиссии о качестве работ "
        "со ссылкой на нормативы и проектные документы; разрешение на производство последующих работ; "
        "официальный канцелярский русский, 2-4 предложения.\n"
        "- normatives: перечень применимых СНиП/СП/ГОСТ через запятую (только реально применимые к этому виду работ).\n"
        "- projectDocs: перечень типовых проектных документов (разделы, листы), к которым относится эта работа."
    )

    instructions = "Ты отвечаешь СТРОГО валидным JSON. Никакого markdown, никаких тройных кавычек, никакого текста до или после JSON. Только сам JSON."
    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)

    def _call(model_id):
        try:
            r = client.responses.create(
                model="gpt://" + YANDEX_FOLDER_ID + "/" + model_id,
                temperature=0.1,
                instructions=instructions,
                input=user_text,
                max_output_tokens=2000,
            )
            return (r.output_text or ""), None
        except Exception as e:
            return "", str(e)

    answer, err = _call("qwen3.6-35b-a3b/latest")
    if not (answer or "").strip():
        print("AI-PREFILL primary empty, fallback. err=" + str(err))
        answer, err = _call("yandexgpt-5.1/latest")
    if not (answer or "").strip():
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул пустой ответ: " + str(err))

    text = answer.strip()
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if m:
        text = m.group(0)
    try:
        parsed = j.loads(text)
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=502, detail="AI вернул невалидный JSON: " + str(e)[:200])

    conclusion = (parsed.get("conclusion") or "").strip()
    normatives = (parsed.get("normatives") or "").strip()
    project_docs = (parsed.get("projectDocs") or "").strip()

    if normatives:
        full_conclusion = "Применимые нормативные документы: " + normatives + "\n\n" + conclusion
    else:
        full_conclusion = conclusion

    cur = conn.cursor()
    cur.execute("""UPDATE hidden_works_acts
                   SET conclusion=%s, project_docs=%s, ai_filled=TRUE
                   WHERE id=%s""",
                (full_conclusion, project_docs, act_id))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True, "conclusion": full_conclusion, "projectDocs": project_docs, "normatives": normatives, "aiFilled": True}

@app.get("/estimates/{estimate_id}/chat-history")
def get_estimate_chat(estimate_id: int, current_user: dict = Depends(get_current_user)):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT project_name FROM estimates WHERE id=%s", (estimate_id,))
    est = cur.fetchone()
    if not est:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    require_project_access(current_user, est[0] or "")
    cur.execute("SELECT id, role, content, created_at FROM estimate_chat_messages WHERE estimate_id=%s ORDER BY id ASC", (estimate_id,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return [{"id": r[0], "role": r[1], "content": r[2], "createdAt": str(r[3])} for r in rows]

@app.post("/estimate-chat")
def estimate_chat(data: dict, current_user: dict = Depends(get_current_user)):
    import openai as oa
    estimate_id = data.get("estimateId")
    user_message = (data.get("message") or "").strip()
    context = (data.get("context") or "").strip()
    history = data.get("history") or []
    if not estimate_id or not user_message:
        raise HTTPException(status_code=400, detail="estimateId and message required")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT project_name FROM estimates WHERE id=%s", (int(estimate_id),))
    est = cur.fetchone()
    if not est:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    require_project_access(current_user, est[0] or "")
    cur.execute("INSERT INTO estimate_chat_messages (estimate_id, role, content) VALUES (%s, %s, %s) RETURNING id, created_at",
                (estimate_id, "user", user_message))
    user_row = cur.fetchone()
    conn.commit()

    prompt_lines = []
    if context:
        prompt_lines.append("КОНТЕКСТ СМЕТЫ:\n" + context)
    if history:
        prompt_lines.append("\nПРЕДЫДУЩИЙ ДИАЛОГ:")
        for m in history[-20:]:
            r = m.get("role", "user")
            c = m.get("content", "")
            prompt_lines.append(("Пользователь: " if r == "user" else "Ассистент: ") + c)
    prompt_lines.append("\nНОВЫЙ ВОПРОС ПОЛЬЗОВАТЕЛЯ:\n" + user_message)
    prompt_lines.append("\nОтветь по-русски, используя факты из контекста сметы. Если для ответа недостаточно данных — скажи об этом. Если вопрос требует расчёта — приведи цифры явно.")
    full_prompt = "\n".join(prompt_lines)

    instructions = "Ты эксперт по строительным сметам. Помогаешь анализировать конкретную смету в формате диалога. Отвечаешь только по данной смете, не выдумываешь позиции. Используй конкретные числа из контекста."

    client = oa.OpenAI(api_key=YANDEX_API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=YANDEX_FOLDER_ID)
    try:
        response = client.responses.create(
            model="gpt://" + YANDEX_FOLDER_ID + "/yandexgpt-5.1/latest",
            temperature=0.3,
            instructions=instructions,
            input=full_prompt,
            max_output_tokens=1500,
        )
        answer = response.output_text or ""
    except Exception as e:
        answer = "Ошибка ИИ: " + str(e)

    cur.execute("INSERT INTO estimate_chat_messages (estimate_id, role, content) VALUES (%s, %s, %s) RETURNING id, created_at",
                (estimate_id, "assistant", answer))
    asst_row = cur.fetchone()
    conn.commit()
    cur.close(); conn.close()
    return {"response": answer, "userMessageId": user_row[0], "assistantMessageId": asst_row[0]}

@app.delete("/estimates/{estimate_id}/chat-history")
def clear_estimate_chat(estimate_id: int, current_user: dict = Depends(require_roles(*FINANCE_ROLES, "прораб", "главный_инженер", "сметчик"))):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT project_name FROM estimates WHERE id=%s", (estimate_id,))
    est = cur.fetchone()
    if not est:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Смета не найдена")
    require_project_access(current_user, est[0] or "")
    cur.execute("DELETE FROM estimate_chat_messages WHERE estimate_id=%s", (estimate_id,))
    conn.commit()
    cur.close(); conn.close()
    return {"ok": True}

@app.post("/scan-invoice")
def scan_invoice(data: dict, _current_user: dict = Depends(require_roles(*WAREHOUSE_ROLES))):
    import openai as oa
    FOLDER_ID = YANDEX_FOLDER_ID
    API_KEY = YANDEX_API_KEY
    base64_image = data.get("image", "")
    client = oa.OpenAI(api_key=API_KEY, base_url="https://ai.api.cloud.yandex.net/v1", project=FOLDER_ID)
    try:
        response = client.responses.create(
            model=f"gpt://{FOLDER_ID}/qwen3.6-35b-a3b/latest",
            temperature=0.1,
            instructions="Ты распознаёшь накладные. Верни только JSON без комментариев и без markdown.",
            input=[
                {
                    "role": "user",
                    "content": [
                        {"type": "input_image", "image_url": f"data:image/jpeg;base64,{base64_image}"},
                        {"type": "input_text", "text": "Распознай накладную. Верни JSON: {supplier:строка, items:[{name:строка,quantity:число,unit:строка,price:число}], total:число}. Только JSON без комментариев."}
                    ]
                }
            ],
            max_output_tokens=1000
        )
        answer = response.output_text
        import json
        clean = answer.replace("```json","").replace("```","").strip()
        parsed = json.loads(clean)
        print("SCAN OK:", parsed)
        return {"ok": True, "data": parsed}
    except Exception as e:
        print("SCAN ERROR:", str(e))
        return {"ok": False, "error": str(e)}
