"""Estimate Excel parser route.

Extracted verbatim from backend/main.py (Task 13.1, slice 40):
POST /parse-smeta — the single largest route in the monolith
(~1,070 lines): ЛСР/ГЭСН header detection, section tracking, unit
normalization and quantity parsing for imported Excel estimates.
"""

from fastapi import Depends, File, UploadFile


def register_smeta_parser_module(app, deps):
    get_current_user = deps["get_current_user"]

    @app.post("/parse-smeta")
    def parse_smeta(file: UploadFile = File(...), _current_user: dict = Depends(get_current_user)):
        import tempfile, os, re
        try:
            import openpyxl
        except ImportError:
            return {"error": "openpyxl not installed"}

        SMETA_PARSE_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
        SMETA_PARSE_MAX_BYTES = 15 * 1024 * 1024
        filename = os.path.basename(getattr(file, "filename", "") or "")
        ext = os.path.splitext(filename.lower())[1]
        if ext not in SMETA_PARSE_ALLOWED_EXTENSIONS:
            return {"error": "Поддерживаются только Excel-файлы .xlsx/.xlsm. Старый .xls сначала сохраните как .xlsx."}

        # A sync FastAPI handler runs in its worker thread. openpyxl is synchronous,
        # so this keeps a large estimate from blocking health checks and other users.
        contents = file.file.read()
        if len(contents) > SMETA_PARSE_MAX_BYTES:
            return {"error": "Файл сметы превышает лимит 15 МБ. Разделите смету или загрузите документом без импорта."}

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name
    
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            results = []
            current_section = "Без раздела"
            data_start_row = 1
            file_type = "unknown"
            header_rows = list(ws.iter_rows(max_row=80, values_only=True))
            document_header_text = " ".join(
                str(value)
                for row in header_rows[:15]
                for value in row
                if value is not None
            ).lower().replace("ё", "е")
            if "акт о приемке выполненных работ" in document_header_text:
                os.unlink(tmp_path)
                return {
                    "error": "Выбран акт о приемке выполненных работ, а не смета. "
                             "Не загружайте выполненные объемы как новый план. "
                             "Для пересчета сметы выберите исходный файл ЛСР/СК с откорректированными плановыми объемами."
                }
            lsr_header_row_idx = None
        
            for i, row in enumerate(header_rows[:60]):
                vals = [str(v).strip() for v in row if v is not None]
                row_text = " ".join(vals).lower()
                is_lsr_table_header = (
                    "обоснование" in row_text
                    and "наименование" in row_text
                    and ("работ" in row_text or "затрат" in row_text)
                    and ("единица" in row_text or "измерения" in row_text)
                    and "количество" in row_text
                )
                is_lsr_number_header = (
                    "№" in row_text
                    and "обоснование" in row_text
                    and "наименование" in row_text
                    and ("единица" in row_text or "количество" in row_text)
                )
                if "номер сметы" in row_text and "стоимость всего" in row_text and ("цена на единицу" in row_text or "цена на" in row_text):
                    data_start_row = i + 2
                    file_type = "contract"
                    break
                elif is_lsr_table_header or is_lsr_number_header:
                    data_start_row = i + 2
                    file_type = "lsr"
                    lsr_header_row_idx = i
                    break
                elif "наименование" in row_text and ("работ" in row_text or "затрат" in row_text) and ("единица" in row_text or "количество" in row_text) and ("сметн" in row_text or "стоим" in row_text):
                    data_start_row = i + 2
                    file_type = "lsr"
                    lsr_header_row_idx = i
                    break
                elif "наименование" in row_text and "ед" in row_text and "кол" in row_text and "обоснование" not in row_text:
                    data_start_row = i + 2
                    file_type = "defect"
                    break
                elif "обоснование" in row_text and "наименование" in row_text and "общее" in row_text:
                    data_start_row = i + 2
                    file_type = "vedomost"
                    break
        
            work_prefixes = ["ГЭСНм", "ФЕРм", "ТЕРм", "ГЭСНи", "ФЕРи", "ГЭСНр", "ФЕРр", "ТЕРр", "ГЭСН", "ФЕР", "ТЕР"]
            material_prefixes = ["ФСБЦ", "ФССЦ", "ТССЦ", "ТССЦпг", "ТЦ_", "ТЦ-", "КАЦ", "МАТ"]
            work_words = ("монтаж", "демонтаж", "установка", "устройство", "прокладка", "разбор", "разборка", "сборка", "замена", "подключение", "снятие", "очистка", "ремонт", "отбивка", "облицов", "окраск", "шпатлев", "шпаклев", "грунтов", "стяжк", "укладка")
            material_words = ("материал", "труба", "кабель", "провод", "смесь", "раствор", "ротбанд", "кнауф", "штукатурка", "штукатурк", "шпатлевка", "шпатлевк", "шпаклевка", "шпаклевк", "клей", "краска", "акрил", "грунтовка", "цемент", "бетон", "кирпич", "блок", "лист", "профиль", "саморез", "шуруп", "анкер", "болт", "гайк", "шайб", "плитка", "плитк", "керамическ", "керамогранит", "гранит", "livorno", "ливорно", "axima", "аксима", "линолеум", "арматур", "битум", "бризол", "лак", "мастик", "утеплитель", "рубероид", "пвх", "уголок", "уголк", "угол", "подвес", "маяк", "рейк", "лента", "серпян", "панель", "плинтус", "наличник", "дюбел", "втулк", "скреп", "крепеж", "металлочереп", "черепиц", "монтеррей", "кровл", "кроншт", "направл", "фитинг", "муфт", "клапан", "американ", "воздухоотвод", "радиатор", "светиль", "ламп", "розет", "выключ", "доска", "брус", "фанер", "осп", "osb", "гвозд", "шпил", "перфолента")
            lsr_service_tokens = (
                "итого", "всего", "в том числе", "объем=", "объём=",
                "фот", "средства на оплату труда", "нормативные затраты труда",
                "накладные расходы", "сметная прибыль", "индекс", "индексы",
                "коэффициент к итогам", "коэффициенты к итогам",
                "заготовительно-складские", "заготовительно складские",
                "справочно", "начисление", "начисления",
                "вспомогательные ненормируемые"
            )

            def _lsr_text_key(value):
                return str(value or "").lower().replace("ё", "е").replace("\xa0", " ").strip()

            def _detect_lsr_columns(rows, header_idx):
                columns = {
                    "num": 0,
                    "obosn": 1,
                    "name": 2,
                    "unit": 3,
                    "quantity_base": 4,
                    "quantity_coeff": 5,
                    "quantity_final": 6,
                }
                if header_idx is None:
                    return columns

                start = max(0, header_idx - 2)
                end = min(len(rows), header_idx + 6)
                scan_rows = rows[start:end]

                for row in scan_rows:
                    for idx, value in enumerate(row):
                        key = _lsr_text_key(value).replace("\n", " ")
                        if "обоснование" in key:
                            columns["obosn"] = idx
                        if "наименование" in key and ("работ" in key or "затрат" in key):
                            columns["name"] = idx
                        if "единица" in key and "измер" in key:
                            columns["unit"] = idx

                unit_idx = columns.get("unit")
                if unit_idx is not None:
                    columns.setdefault("quantity_base", unit_idx + 1)
                    columns.setdefault("quantity_coeff", unit_idx + 2)
                    columns.setdefault("quantity_final", unit_idx + 3)
                    columns["quantity_base"] = unit_idx + 1
                    columns["quantity_coeff"] = unit_idx + 2
                    columns["quantity_final"] = unit_idx + 3

                for row in scan_rows:
                    seen_cost_unit = False
                    for idx, value in enumerate(row):
                        key = _lsr_text_key(value).replace("\n", " ")
                        compact = re.sub(r"\s+", " ", key)
                        if "на единицу" in compact:
                            if idx <= columns.get("quantity_final", 6):
                                columns["quantity_base"] = idx
                            elif not seen_cost_unit:
                                columns["cost_unit"] = idx
                                seen_cost_unit = True
                        elif "всего" in compact:
                            if "учет" in compact and "коэффициент" in compact:
                                columns["quantity_final"] = idx
                            elif idx > columns.get("quantity_final", 6) and "cost_total" not in columns:
                                columns["cost_total"] = idx
                        elif "коэффициент" in compact and "сметн" not in compact:
                            if idx < columns.get("quantity_final", 6):
                                columns["quantity_coeff"] = idx
                            elif "cost_coeff" not in columns:
                                columns["cost_coeff"] = idx
                        elif "индекс" in compact:
                            columns["cost_index"] = idx

                qf = columns.get("quantity_final", 6)
                if "cost_unit" not in columns:
                    columns["cost_unit"] = qf + 1
                if "cost_coeff" not in columns:
                    columns["cost_coeff"] = columns["cost_unit"] + 1
                if "cost_total" not in columns:
                    columns["cost_total"] = columns["cost_coeff"] + 1
                if "cost_index" not in columns:
                    columns["cost_index"] = columns["cost_total"] + 1
                if "cost_current_total" not in columns:
                    columns["cost_current_total"] = columns["cost_index"] + 1
                return columns

            lsr_columns = _detect_lsr_columns(header_rows, lsr_header_row_idx) if file_type == "lsr" else {}

            def _is_lsr_service_row(name_value, obosn_value=""):
                name_key = _lsr_text_key(name_value)
                code_key = _lsr_text_key(obosn_value)
                if not name_key:
                    return True
                if "пр/" in code_key or "648/" in code_key:
                    return True
                if any(token in name_key for token in lsr_service_tokens):
                    return True
                return False

            def _lsr_item_type(obosn_value, name_value):
                code = str(obosn_value or "").strip()
                name_key = _lsr_text_key(name_value)
                if any(code.startswith(x) for x in work_prefixes):
                    return "work"
                if any(code.startswith(x) for x in material_prefixes):
                    return "material"
                if re.match(r"^\d{2,}[-/]\d+", code) or re.match(r"^\d{3,}$", code):
                    return "material"
                if any(word in name_key for word in work_words):
                    return "work"
                if any(word in name_key for word in material_words):
                    return "material"
                return "material"

            def _lsr_name_looks_resource(name_value):
                name_key = _lsr_text_key(name_value)
                if not name_key:
                    return False
                work_starts = (
                    "монтаж", "демонтаж", "устройство", "установка", "разбор",
                    "разборка", "снятие", "очистка", "отбивка", "облицовка",
                    "окраска", "грунтование", "штукатурка поверхностей", "стяжка",
                    "смена", "прокладка", "подключение", "ремонт"
                )
                if any(name_key.startswith(prefix) for prefix in work_starts):
                    return False
                return any(word in name_key for word in material_words)

            def _lsr_code_is_resource(obosn_value):
                code = str(obosn_value or "").strip()
                return (
                    any(code.startswith(x) for x in material_prefixes)
                    or re.match(r"^\d{2,}[-/]\d+", code)
                    or re.match(r"^\d{3,}$", code)
                )

            def _infer_lsr_unit(unit_value, name_value, item_type):
                raw = str(unit_value or "").strip()
                compact = raw.lower().replace(" ", "")
                if raw and compact not in ("1", "ед", "ед.", "единица"):
                    return _normalize_lsr_unit_text(raw)
                name_key = _lsr_text_key(name_value)
                if item_type == "material":
                    if any(w in name_key for w in ("смесь", "штукатурная", "шпатлевка", "шпатлевк", "шпаклевка", "шпаклевк", "клей", "затирка", "цемент", "пескобетон", "сухая смесь")):
                        return "кг"
                    if any(w in name_key for w in ("краска", "грунтовка", "эмаль", "лак", "акрил")):
                        return "кг"
                    if any(w in name_key for w in ("уголок", "уголк", "угол", "плинтус", "наличник", "кабель", "провод", "труба", "трубопровод", "лоток", "короб", "профиль", "маяк", "подоконник")):
                        return "м"
                    if any(w in name_key for w in ("плитк", "керамогранит", "гранит", "линолеум", "покрытие пола", "обои", "лист гипсокартон", "листы гипсокартон", "гкл", "панель", "плита")):
                        return "м2"
                    if any(w in name_key for w in ("кирпич", "блок", "саморез", "дюбель", "гвозд", "светильник", "розетка", "выключатель", "прибор", "радиатор", "решетк")):
                        return "шт"
                if any(w in name_key for w in ("розет", "выключател", "светильник", "табло", "прибор", "радиатор", "решетк", "унитаз", "раковин", "смесител", "коробка", "шкаф", "стол")):
                    return "шт"
                if any(w in name_key for w in ("уголок", "уголк", "угол", "плинтус", "наличник", "кабель", "провод", "труба", "трубопровод", "лоток", "короб", "профиль", "маяк", "подоконник")):
                    return "м"
                if any(w in name_key for w in ("бетон", "брус", "каркас из брус")):
                    return "м3"
                if item_type == "work" and any(w in name_key for w in ("поверхност", "фасад", "стен", "перегород", "потол", "обои", "облицов", "окраск", "штукатур", "шпатлев", "шпаклев", "грунтов", "плитк", "керамогранит", "гранит", "линолеум", "покрытие пола", "гкл", "гипсокартон", "сетка")):
                    return "100 м2"
                return "шт"

            def _row_float(value):
                if value is None:
                    return None
                if isinstance(value, (int, float)):
                    return float(value)
                text = str(value).strip().replace("\xa0", " ")
                if not text:
                    return None
                compact = text.replace(" ", "").replace(",", ".")
                if re.match(r"^\([+-]?\d+(?:\.\d+)?\)$", compact):
                    compact = "-" + compact.strip("()").lstrip("+")
                if not re.match(r"^[+-]?\d+(?:\.\d+)?$", compact):
                    return None
                try:
                    return float(compact)
                except Exception:
                    return None

            def _money_candidates_from_text(value):
                text = str(value or "").replace("\xa0", " ")
                candidates = []
                for match in re.finditer(r"(?<!\d)(?:\d[\d\s]{3,}\d)(?:[,.]\d{1,2})?(?!\d)", text):
                    raw = match.group(0).replace(" ", "").replace(",", ".")
                    try:
                        amount = abs(float(raw))
                    except Exception:
                        continue
                    if 1000 <= amount < 1000000000000:
                        candidates.append(amount)
                for match in re.finditer(r"(?<![\w-])(?:\d{1,3}(?:\s+\d{3})+|\d{5,})(?:[,.]\d{1,2})?(?![\w-])", text):
                    raw = match.group(0).replace(" ", "").replace(",", ".")
                    try:
                        amount = abs(float(raw))
                    except Exception:
                        continue
                    if 1000 <= amount < 1000000000000:
                        candidates.append(amount)
                return candidates

            def _extract_declared_estimate_total():
                file_candidates = _money_candidates_from_text(file.filename or "")
                if file_candidates:
                    return round(max(file_candidates), 2), "filename"
                summary_candidates = []
                summary_tokens = ("сметн", "итого", "всего", "стоимость", "стоимост", "руб")
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join(str(v) for v in row if v is not None).lower().replace("ё", "е")
                    if not any(token in row_text for token in summary_tokens):
                        continue
                    for value in row:
                        if isinstance(value, (int, float)) and 1000 <= abs(float(value)) < 1000000000000:
                            summary_candidates.append(abs(float(value)))
                        elif isinstance(value, str):
                            summary_candidates.extend(_money_candidates_from_text(value))
                if summary_candidates:
                    return round(max(summary_candidates), 2), "workbook"
                return 0, ""

            def _lsr_items_total(items):
                return round(sum(float(item.get("total") or 0) for item in items if item.get("type") in ("work", "material", "equipment", "transport", "overhead", "adjustment")), 2)

            def _declared_total_diagnostics(items, declared_total, declared_source):
                parsed_total = _lsr_items_total(items)
                factor = (declared_total / parsed_total) if declared_total and parsed_total > 0 else 1
                difference = (declared_total - parsed_total) if declared_total and parsed_total > 0 else 0
                mismatch = bool(declared_total and parsed_total > 0 and abs(difference) > max(1000, declared_total * 0.01))
                meta = {
                    "declaredTotal": declared_total,
                    "declaredTotalSource": declared_source,
                    "parsedTotal": parsed_total,
                    "difference": round(difference, 2),
                    "diagnosticFactor": round(factor, 8),
                    "totalMismatch": mismatch,
                    "scaleApplied": False,
                }
                return meta

            def _normalize_lsr_unit_text(value):
                text = str(value or "").strip().replace("\xa0", " ")
                text = re.sub(r"\s+", " ", text)
                text = text.replace("м²", "м2").replace("М²", "м2").replace("м³", "м3").replace("М³", "м3")
                compact = text.lower().replace(" ", "").replace(".", "")
                single_unit = re.match(r"^1\s*(шт\.?|м2|м3|м|кг|т|л)$", text.lower())
                if single_unit:
                    return _normalize_lsr_unit_text(single_unit.group(1))
                if not re.match(r"^\d", compact):
                    if compact.startswith("маш"):
                        return "маш-ч"
                    if compact.startswith("чел"):
                        return "чел-ч"
                    if compact.startswith("м2"):
                        return "м2"
                    if compact.startswith("м3"):
                        return "м3"
                    if compact.startswith(("мп", "пм", "м/п", "м")):
                        return "м"
                    if compact.startswith("шт"):
                        return "шт"
                    if compact.startswith("кг"):
                        return "кг"
                    if compact.startswith("т"):
                        return "т"
                    if compact.startswith("л"):
                        return "л"
                return text

            def _looks_lsr_unit(value):
                text = _normalize_lsr_unit_text(value).lower().replace(" ", "")
                if not text or text in ("1", "ед", "ед.", "единица"):
                    return False
                units = ("шт", "м", "м2", "м3", "кг", "т", "л", "рулон", "лист", "упак", "компл", "маш-ч", "чел-ч")
                if text in units:
                    return True
                return bool(re.match(r"^\d+(м2|м3|маш-ч|чел-ч|шт|кг|рулон|лист|упак|компл|м|т|л)", text))

            def _normalize_lsr_measure(qty, unit):
                raw_unit = _normalize_lsr_unit_text(unit)
                raw_qty = qty if qty is not None else 0
                m = re.match(r"^(\d{2,})\s*(.+)$", raw_unit)
                if not m:
                    return float(raw_qty or 0), raw_unit, 1
                factor = int(m.group(1) or 1)
                if factor < 10:
                    return float(raw_qty or 0), _normalize_lsr_unit_text(m.group(2)), 1
                return float(raw_qty or 0) * factor, _normalize_lsr_unit_text(m.group(2)), factor

            def _pick_lsr_unit_and_quantity(row, fallback_unit, fallback_qty, preferred_unit_idx=None,
                                            quantity_base_idx=None, quantity_coeff_idx=None, quantity_final_idx=None):
                found_unit = None
                found_idx = None
                quantity_base = None
                quantity_coeff = None
                quantity_final = None
                if preferred_unit_idx is not None and len(row) > preferred_unit_idx and _looks_lsr_unit(row[preferred_unit_idx]):
                    found_unit = _normalize_lsr_unit_text(row[preferred_unit_idx])
                    found_idx = preferred_unit_idx
                elif preferred_unit_idx is None:
                    # Если колонка единицы не определена, ищем только в начале строки ЛСР.
                    # В хвосте строки встречаются единицы ресурсов/труда ("т", "чел-ч"),
                    # их нельзя принимать за единицу самой работы.
                    for idx, value in enumerate(row[:10]):
                        if _looks_lsr_unit(value):
                            found_unit = _normalize_lsr_unit_text(value)
                            found_idx = idx
                            break
                unit = found_unit or fallback_unit
                qty = fallback_qty

                def _num_at(idx):
                    if idx is None or idx < 0 or len(row) <= idx:
                        return None
                    return _row_float(row[idx])

                quantity_base = _num_at(quantity_base_idx)
                quantity_coeff = _num_at(quantity_coeff_idx)
                quantity_final = _num_at(quantity_final_idx)
                if quantity_final is not None:
                    qty = quantity_final
                elif quantity_base is not None:
                    qty = quantity_base

                if quantity_base is None and quantity_final is None and found_idx is not None:
                    nums = []
                    for idx in range(found_idx + 1, min(len(row), found_idx + 6)):
                        n = _row_float(row[idx])
                        if n is not None:
                            nums.append(n)
                    quantity_base = nums[0] if len(nums) >= 1 else None
                    quantity_coeff = nums[1] if len(nums) >= 2 else None
                    quantity_final = nums[2] if len(nums) >= 3 else None
                    if len(nums) >= 3 and abs(nums[2]) > 0.0001:
                        qty = nums[2]
                    elif nums:
                        qty = nums[0]
                normalized_qty, normalized_unit, factor = _normalize_lsr_measure(qty, unit)
                return normalized_unit, normalized_qty, unit, qty, factor, found_idx, quantity_base, quantity_coeff, quantity_final

            def _pick_lsr_sum(row, preferred_indexes=(), unit_idx=None):
                max_line_total = 1000000000000
                if unit_idx is not None:
                    after_unit = []
                    for idx in range(unit_idx + 1, len(row)):
                        n = _row_float(row[idx])
                        if n is not None and 0.0001 < abs(n) < max_line_total:
                            after_unit.append((idx, n))
                    # В ЛСР после единицы обычно идут: объем на единицу, коэффициент,
                    # итоговый объем, затем денежная часть. Первые 3 числа не считаем суммой.
                    money_candidates = [item for item in after_unit[3:] if abs(item[1]) >= 1]
                    if money_candidates:
                        return round(max(money_candidates, key=lambda item: abs(item[1]))[1], 2)
                for idx in preferred_indexes:
                    if len(row) > idx:
                        n = _row_float(row[idx])
                        if n is not None and 0.0001 < abs(n) < max_line_total:
                            return round(n, 2)
                candidates = []
                start_idx = max(6, (unit_idx + 4) if unit_idx is not None else 6)
                for idx, value in enumerate(row):
                    if idx < start_idx:
                        continue
                    n = _row_float(value)
                    if n is not None and 0.0001 < abs(n) < max_line_total:
                        candidates.append((idx, n))
                if not candidates:
                    return 0
                money_candidates = [item for item in candidates if abs(item[1]) >= 1]
                if money_candidates:
                    return round(max(money_candidates, key=lambda item: abs(item[1]))[1], 2)
                return round(candidates[-1][1], 2)

            def _cell_float(row, idx):
                if idx is None or idx < 0 or len(row) <= idx:
                    return None
                return _row_float(row[idx])

            def _valid_lsr_money(value):
                return value is not None and 0.0001 < abs(float(value)) < 1000000000000

            def _pick_lsr_money(row, columns, unit_idx=None):
                base_unit_price = _cell_float(row, columns.get("cost_unit"))
                cost_coeff = _cell_float(row, columns.get("cost_coeff"))
                base_total = _cell_float(row, columns.get("cost_total"))
                cost_index = _cell_float(row, columns.get("cost_index"))
                current_total = _cell_float(row, columns.get("cost_current_total"))
                quantity_final = _cell_float(row, columns.get("quantity_final"))

                result = {
                    "baseUnitPrice": round(float(base_unit_price), 6) if base_unit_price is not None else None,
                    "costCoefficient": round(float(cost_coeff), 6) if cost_coeff is not None else None,
                    "baseTotal": round(float(base_total), 2) if base_total is not None else None,
                    "costIndex": round(float(cost_index), 6) if cost_index is not None else None,
                    "currentTotal": round(float(current_total), 2) if current_total is not None else None,
                    "lineTotalSource": "",
                    "lineTotal": 0,
                }

                if _valid_lsr_money(current_total):
                    if _valid_lsr_money(base_total) and cost_index not in (None, 0):
                        expected = float(base_total) * float(cost_index)
                        if abs(float(current_total) - expected) <= max(1, abs(expected) * 0.03):
                            result["lineTotal"] = round(float(current_total), 2)
                            result["lineTotalSource"] = "current_total"
                            return result
                    elif not _valid_lsr_money(base_total) or abs(float(current_total)) >= abs(float(base_total)):
                        result["lineTotal"] = round(float(current_total), 2)
                        result["lineTotalSource"] = "current_total"
                        return result

                if _valid_lsr_money(base_total) and cost_index not in (None, 0) and 0.0001 < abs(float(cost_index)) < 1000:
                    result["lineTotal"] = round(float(base_total) * float(cost_index), 2)
                    result["lineTotalSource"] = "base_total_x_index"
                    return result

                if _valid_lsr_money(base_total):
                    result["lineTotal"] = round(float(base_total), 2)
                    result["lineTotalSource"] = "base_total"
                    return result

                if base_unit_price is not None and quantity_final is not None:
                    coeff = float(cost_coeff) if cost_coeff not in (None, 0) else 1
                    result["lineTotal"] = round(float(base_unit_price) * coeff * float(quantity_final), 2)
                    result["lineTotalSource"] = "unit_price_x_quantity"
                    return result

                fallback = _pick_lsr_sum(row, (), unit_idx)
                result["lineTotal"] = fallback
                result["lineTotalSource"] = "fallback_scan" if fallback else ""
                return result

            def _lsr_summary_indexes(columns):
                idx_col = columns.get("cost_index", 12)
                name_idx = columns.get("name", 2)
                indexes = {}
                for row in ws.iter_rows(min_row=max(1, ws.max_row - 260), values_only=True):
                    name = _lsr_text_key(row[name_idx] if len(row) > name_idx else "")
                    idx_value = _row_float(row[idx_col]) if idx_col is not None and len(row) > idx_col else None
                    if idx_value is None or idx_value <= 1:
                        continue
                    if "материал" in name:
                        indexes["material"] = float(idx_value)
                    elif "эксплуатация машин" in name or "машин" in name or "механизм" in name:
                        indexes["machine"] = float(idx_value)
                        indexes.setdefault("transport", float(idx_value))
                    elif "оборудован" in name:
                        indexes["equipment"] = float(idx_value)
                    elif "перевоз" in name or "транспорт" in name:
                        indexes["transport"] = float(idx_value)
                return indexes

            lsr_indexes = _lsr_summary_indexes(lsr_columns) if file_type == "lsr" else {}

            def _lsr_cost_kind(name_value):
                key = _lsr_text_key(name_value)
                compact = key.replace(" ", "")
                if compact in ("от",):
                    return "labor"
                if compact in ("эм",):
                    return "machine"
                if compact == "м":
                    return "material_total"
                if key.startswith("нр "):
                    return "overhead"
                if key.startswith("сп "):
                    return "profit"
                if compact in ("зт", "зтм", "фот") or key.startswith("в т.ч.") or key.startswith("в том числе") or "итого по расценке" in key:
                    return "skip"
                return ""

            def _lsr_index_for_kind(kind, name_value="", code_value=""):
                text = _lsr_text_key(str(name_value or "") + " " + str(code_value or ""))
                if kind in ("machine",):
                    return lsr_indexes.get("machine") or lsr_indexes.get("transport") or 1
                if kind in ("equipment",):
                    return lsr_indexes.get("equipment") or lsr_indexes.get("material") or 1
                if kind in ("transport",):
                    return lsr_indexes.get("transport") or lsr_indexes.get("machine") or 1
                if kind in ("material", "material_total"):
                    return lsr_indexes.get("material") or 1
                if "эксплуатация машин" in text or "машин" in text or "механизм" in text:
                    return lsr_indexes.get("machine") or 1
                if "материал" in text:
                    return lsr_indexes.get("material") or 1
                return 1

            def _lsr_money_with_index(row, columns, unit_idx=None, kind="material", name_value="", code_value=""):
                money = _pick_lsr_money(row, columns, unit_idx)
                line = float(money.get("lineTotal") or 0)
                source = money.get("lineTotalSource") or ""
                base_total = money.get("baseTotal")
                if (not line or source in ("base_total", "fallback_scan", "")) and _valid_lsr_money(base_total):
                    idx = money.get("costIndex") or _lsr_index_for_kind(kind, name_value, code_value)
                    if idx and abs(float(idx)) > 1.0001:
                        money["lineTotal"] = round(float(base_total) * float(idx), 2)
                        money["currentTotal"] = money["lineTotal"]
                        money["costIndex"] = round(float(idx), 6)
                        money["lineTotalSource"] = "base_total_x_summary_index"
                    else:
                        money["lineTotal"] = round(float(base_total), 2)
                        money["lineTotalSource"] = source or "base_total"
                return money

            def _add_imported_work_amount(item_idx, amount, component_key):
                if item_idx is None or item_idx < 0 or item_idx >= len(results):
                    return
                amount = float(amount or 0)
                item = results[item_idx]
                item["totalWork"] = round(float(item.get("totalWork") or 0) + amount, 2)
                item["total"] = round(float(item.get("total") or 0) + amount, 2)
                item["lineTotal"] = round(float(item.get("lineTotal") or 0) + amount, 2)
                components = item.setdefault("costComponents", {})
                components[component_key] = round(float(components.get(component_key) or 0) + amount, 2)

            def _lsr_work_key(section, code, name, row_index):
                raw = "|".join([
                    str(section or ""),
                    str(code or ""),
                    str(name or ""),
                    str(row_index or ""),
                ]).lower().replace("ё", "е")
                return re.sub(r"[^a-zа-я0-9]+", "-", raw).strip("-")[:160]
        
            last_work_result_idx = None
            current_work_ref = None
            for i, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True)):
                try:
                    num_idx = lsr_columns.get("num", 0) if file_type == "lsr" else 0
                    obosn_idx = lsr_columns.get("obosn", 1) if file_type == "lsr" else 1
                    name_idx = lsr_columns.get("name", 2) if file_type == "lsr" else 2
                    first_val = str(row[0]).strip() if row[0] is not None else ""
                    name_col = str(row[name_idx]).strip() if len(row) > name_idx and row[name_idx] else ""
                    obosn = str(row[obosn_idx]).strip() if len(row) > obosn_idx and row[obosn_idx] else ""
                
                    if "Раздел" in first_val or "РАЗДЕЛ" in first_val:
                        current_section = first_val
                        continue
                
                    if file_type == "lsr":
                        num = row[num_idx] if len(row) > num_idx else None
                        num_is_main = False
                        if num not in (None, ""):
                            try:
                                float(str(num).strip())
                                num_is_main = True
                            except:
                                num_is_main = bool(re.match(r"^\s*\d+", str(num)))

                        if "Всего по позиции" in name_col:
                            if last_work_result_idx is not None and 0 <= last_work_result_idx < len(results):
                                try:
                                    v = _pick_lsr_sum(row, (13, 14, 15, 16, 17))
                                    if v:
                                        results[last_work_result_idx]["positionBaseTotal"] = round(float(v), 2)
                                except:
                                    pass
                            continue

                        cost_kind = _lsr_cost_kind(name_col)
                        if not num_is_main and current_work_ref and cost_kind:
                            if cost_kind == "skip":
                                continue
                            money = _lsr_money_with_index(row, lsr_columns, None, cost_kind, name_col, obosn)
                            line_money = float(money.get("lineTotal") or 0)
                            if cost_kind == "material_total":
                                if int(current_work_ref.get("resourceCount") or 0) == 0 and abs(line_money) > 0.0001:
                                    agg_item = {
                                        "section": current_section,
                                        "name": "Материалы по позиции: " + current_work_ref.get("workName", ""),
                                        "unit": "компл",
                                        "quantity": 1,
                                        "rawUnit": "компл",
                                        "rawQuantity": 1,
                                        "unitFactor": 1,
                                        "baseUnitPrice": money.get("baseUnitPrice"),
                                        "costCoefficient": money.get("costCoefficient"),
                                        "baseTotal": money.get("baseTotal"),
                                        "costIndex": money.get("costIndex"),
                                        "currentTotal": money.get("currentTotal"),
                                        "lineTotalSource": money.get("lineTotalSource"),
                                        "lineTotal": line_money,
                                        "total": line_money,
                                        "totalWork": 0,
                                        "totalMaterial": line_money,
                                        "type": "material",
                                        "sourceCode": obosn,
                                        "parentWorkKey": current_work_ref.get("workKey"),
                                        "parentWorkName": current_work_ref.get("workName"),
                                        "parentWorkSourceCode": current_work_ref.get("workSourceCode"),
                                        "resourceRole": "material",
                                        "importKind": "aggregate_material",
                                    }
                                    results.append(agg_item)
                                    current_work_ref["resourceCount"] = int(current_work_ref.get("resourceCount") or 0) + 1
                                continue
                            if abs(line_money) > 0.0001:
                                _add_imported_work_amount(last_work_result_idx, line_money, cost_kind)
                            continue

                        if not name_col:
                            continue

                        if not num_is_main:
                            # Дочерние ресурсные строки внутри позиции: материалы, оборудование, перевозка.
                            if not current_work_ref or _is_lsr_service_row(name_col, obosn) or not _lsr_code_is_resource(obosn):
                                continue
                            item_type = _lsr_item_type(obosn, name_col)
                            if item_type == "work":
                                if _lsr_name_looks_resource(name_col):
                                    item_type = "material"
                                else:
                                    continue
                        else:
                            if len(name_col) < 5 or _is_lsr_service_row(name_col, obosn):
                                continue
                            item_type = _lsr_item_type(obosn, name_col)

                        unit_idx_hint = lsr_columns.get("unit")
                        qty_base_idx = lsr_columns.get("quantity_base")
                        qty_coeff_idx = lsr_columns.get("quantity_coeff")
                        qty_final_idx = lsr_columns.get("quantity_final")
                        unit_raw = str(row[unit_idx_hint]).strip() if unit_idx_hint is not None and len(row) > unit_idx_hint and row[unit_idx_hint] else "шт"
                        inferred_unit = _infer_lsr_unit(unit_raw, name_col, item_type)
                        raw_qty = _row_float(row[qty_final_idx]) if qty_final_idx is not None and len(row) > qty_final_idx else None
                        if raw_qty is None and qty_base_idx is not None and len(row) > qty_base_idx:
                            raw_qty = _row_float(row[qty_base_idx])
                        raw_qty = raw_qty if raw_qty is not None else 0
                        unit, qty, raw_unit, raw_qty, unit_factor, unit_idx, quantity_base, quantity_coeff, quantity_final = _pick_lsr_unit_and_quantity(
                            row,
                            inferred_unit,
                            raw_qty,
                            unit_idx_hint,
                            qty_base_idx,
                            qty_coeff_idx,
                            qty_final_idx,
                        )

                        money = _lsr_money_with_index(row, lsr_columns, unit_idx, item_type, name_col, obosn)
                        line_money = float(money.get("lineTotal") or 0)
                        resource_item_type = item_type
                        is_negative_line = float(qty or 0) < -0.0001 or line_money < 0
                        is_resource_adjustment = is_negative_line and (
                            item_type in ("material", "equipment", "transport")
                            or (item_type == "work" and _lsr_name_looks_resource(name_col))
                        )
                        if is_resource_adjustment and item_type == "work":
                            resource_item_type = "material"
                        if is_resource_adjustment:
                            item_type = "adjustment"
                        work_total = line_money if item_type == "work" else 0
                        mat_total = line_money if item_type in ("material", "equipment", "transport") else 0
                        line_total = line_money if is_resource_adjustment else (mat_total if item_type in ("material", "equipment", "transport") else work_total)
                        if resource_item_type in ("material", "equipment", "transport") and not is_resource_adjustment and abs(float(qty or 0)) <= 0.0001 and abs(line_total) <= 0.0001:
                            continue

                        item = {
                            "section": current_section,
                            "name": name_col,
                            "unit": unit,
                            "quantity": round(qty, 4),
                            "rawUnit": raw_unit,
                            "rawQuantity": round(float(raw_qty or 0), 4),
                            "unitFactor": unit_factor,
                            "quantityBase": round(float(quantity_base), 4) if quantity_base is not None else None,
                            "quantityCoefficient": round(float(quantity_coeff), 6) if quantity_coeff is not None else None,
                            "quantityFinal": round(float(quantity_final), 4) if quantity_final is not None else None,
                            "baseUnitPrice": money.get("baseUnitPrice"),
                            "costCoefficient": money.get("costCoefficient"),
                            "baseTotal": money.get("baseTotal"),
                            "costIndex": money.get("costIndex"),
                            "currentTotal": money.get("currentTotal"),
                            "lineTotalSource": money.get("lineTotalSource"),
                            "lineTotal": line_total,
                            "total": line_total,
                            "totalWork": work_total if item_type == "work" else 0,
                            "totalMaterial": mat_total if item_type in ("material", "equipment", "transport") else 0,
                            "type": item_type,
                            "sourceCode": obosn
                        }
                        if item_type == "work":
                            work_key = _lsr_work_key(current_section, obosn, name_col, i)
                            current_work_ref = {
                                "workKey": work_key,
                                "workName": name_col,
                                "workSourceCode": obosn,
                                "resourceCount": 0,
                            }
                            item.update(current_work_ref)
                        elif resource_item_type in ("material", "equipment", "transport") and current_work_ref:
                            item.update({
                                "parentWorkKey": current_work_ref.get("workKey"),
                                "parentWorkName": current_work_ref.get("workName"),
                                "parentWorkSourceCode": current_work_ref.get("workSourceCode"),
                                "resourceRole": resource_item_type,
                                "importKind": "resource_adjustment" if is_resource_adjustment else "resource",
                            })
                            current_work_ref["resourceCount"] = int(current_work_ref.get("resourceCount") or 0) + 1
                        results.append(item)
                        if item_type == "work":
                            last_work_result_idx = len(results) - 1

                    elif file_type == "contract":
                        num = row[0] if len(row) > 0 else None
                        first_val = str(num or "").strip()
                        num_is_main = False
                        if num not in (None, ""):
                            try:
                                float(str(num).strip())
                                num_is_main = True
                            except:
                                num_is_main = bool(re.match(r"^\s*\d+", str(num)))

                        if not num_is_main:
                            label = first_val
                            if label and label not in ("1", "2", "3", "4", "5", "6", "7", "8") and "№" not in label and "наименование" not in label.lower():
                                current_section = label
                            continue

                        name = str((row[4] if len(row) > 4 and row[4] else (row[2] if len(row) > 2 else "")) or "").strip()
                        if not name or len(name) < 5:
                            continue
                        source_code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        raw_unit = str((row[7] if len(row) > 7 and row[7] else (row[3] if len(row) > 3 else "шт")) or "шт").strip()
                        raw_qty = _row_float(row[8]) if len(row) > 8 and row[8] is not None else (_row_float(row[4]) if len(row) > 4 else 0)
                        unit_qty, unit, unit_factor = _normalize_lsr_measure(raw_qty or 0, raw_unit)
                        unit_price = _row_float(row[9]) if len(row) > 9 and row[9] is not None else (_row_float(row[5]) if len(row) > 5 else None)
                        total = _row_float(row[10]) if len(row) > 10 and row[10] is not None else (_row_float(row[6]) if len(row) > 6 else None)
                        if total is None and unit_price is not None:
                            total = float(unit_price) * float(raw_qty or 0)
                        total = round(float(total or 0), 2)
                        item_type = "material" if _lsr_name_looks_resource(name) else "work"
                        is_resource_adjustment = item_type == "material" and (
                            float(unit_qty or 0) < -0.0001 or total < 0
                        )
                        if is_resource_adjustment:
                            item_type = "adjustment"
                        results.append({
                            "section": current_section,
                            "name": name,
                            "unit": unit,
                            "quantity": round(float(unit_qty or 0), 4),
                            "rawUnit": raw_unit,
                            "rawQuantity": round(float(raw_qty or 0), 4),
                            "unitFactor": unit_factor,
                            "quantityBase": round(float(raw_qty or 0), 4),
                            "quantityCoefficient": 1,
                            "quantityFinal": round(float(raw_qty or 0), 4),
                            "baseUnitPrice": round(float(unit_price), 6) if unit_price is not None else None,
                            "costCoefficient": None,
                            "baseTotal": total,
                            "costIndex": None,
                            "currentTotal": total,
                            "lineTotalSource": "contract_total",
                            "lineTotal": total,
                            "total": total,
                            "totalWork": total if item_type == "work" else 0,
                            "totalMaterial": total if item_type == "material" else 0,
                            "type": item_type,
                            "sourceCode": source_code,
                            "importKind": "resource_adjustment" if is_resource_adjustment else "contract_smeta",
                        })
                
                    elif file_type == "defect":
                        if all(v is None or (isinstance(v, (int,float)) and v < 10) for v in row[:5]):
                            continue
                        name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        unit = str(row[2]).strip() if len(row) > 2 and row[2] else "шт"
                        try:
                            qty = float(row[3]) if len(row) > 3 and row[3] and isinstance(row[3], (int,float)) else 0
                        except:
                            qty = 0
                        if name and len(name) > 5 and name not in ["Наименование", "2"] and not all(c.isdigit() or c == " " for c in name):
                            results.append({"section": current_section, "name": name, "unit": unit, "quantity": qty, "total": 0, "type": "work"})
                
                    elif file_type == "vedomost":
                        name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                        unit = str(row[3]).strip() if len(row) > 3 and row[3] else "шт"
                        try:
                            qty = float(row[4]) if len(row) > 4 and row[4] and isinstance(row[4], (int,float)) else 0
                            total = float(row[6]) if len(row) > 6 and row[6] and isinstance(row[6], (int,float)) else 0
                        except:
                            qty = total = 0
                        if name and len(name) > 5 and name not in ["Наименование", "3"]:
                            results.append({"section": current_section, "name": name, "unit": unit, "quantity": round(qty,4), "total": round(total,2), "type": "material"})
                except:
                    continue
        
            declared_total, declared_source = _extract_declared_estimate_total() if file_type in ("lsr", "contract") else (0, "")
            if file_type in ("lsr", "contract") and declared_total:
                parsed_before_reconciliation = _lsr_items_total(results)
                reconciliation_delta = round(float(declared_total) - float(parsed_before_reconciliation), 2)
                if abs(reconciliation_delta) > 1 and abs(reconciliation_delta) <= max(1000, float(declared_total) * 0.01):
                    results.append({
                        "section": "Итоговая сверка ЛСР",
                        "name": "Корректировка округлений и итоговых строк до ВСЕГО по смете",
                        "unit": "компл",
                        "quantity": 1,
                        "rawUnit": "компл",
                        "rawQuantity": 1,
                        "unitFactor": 1,
                        "baseUnitPrice": None,
                        "costCoefficient": None,
                        "baseTotal": None,
                        "costIndex": None,
                        "currentTotal": reconciliation_delta,
                        "lineTotalSource": "reconciliation_to_declared_total",
                        "lineTotal": reconciliation_delta,
                        "total": reconciliation_delta,
                        "totalWork": 0,
                        "totalMaterial": 0,
                        "type": "overhead",
                        "sourceCode": "ВСЕГО по смете",
                        "importKind": "reconciliation",
                    })
            meta = _declared_total_diagnostics(results, declared_total, declared_source) if file_type in ("lsr", "contract") else {}
            if file_type in ("lsr", "contract"):
                meta.update({
                    "parserVersion": "grand-contract-v1" if file_type == "contract" else "grand-lsr-v2",
                    "fileType": file_type,
                    "dataStartRow": data_start_row,
                    "lsrColumns": lsr_columns,
                    "lsrIndexes": lsr_indexes,
                })
            os.unlink(tmp_path)
            return {"items": results, "count": len(results), "meta": meta}
        except Exception as e:
            try:
                os.unlink(tmp_path)
            except:
                pass
            return {"error": str(e)}
